import csv
import datetime
import difflib
import ipaddress
import os
import re
import socket
import subprocess
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import sys
if sys.platform == "win32":
    import ctypes
    import ctypes.wintypes
import time
import tkinter as tk
from tkinter import messagebox, scrolledtext, filedialog, ttk, simpledialog

import paramiko
import pyte
from PIL import Image, ImageTk

# Regex para remoção de sequências ANSI / VT100 de saídas de shell
_ANSI_ESCAPE_RE = re.compile(r'(\x1b\[[0-9;]*[A-Za-z]|\x1b\].*?\x07|\x1b[()][AB012]|\x1b[=>]|\x08)')

def _limpar_ansi(texto):
    """Remove sequências ANSI/VT100 de uma string."""
    return _ANSI_ESCAPE_RE.sub('', texto)

# Caminho padrão do arquivo CSV (mesmo diretório do script)
DIRETORIO_BASE = os.path.dirname(os.path.abspath(__file__))
CSV_PADRAO = os.path.join(DIRETORIO_BASE, "roteadores.csv")

# Cores do tema Nokia
COR_AZUL_NOKIA = "#124191"
COR_AZUL_CLARO = "#1B5FC1"
COR_BRANCO = "#FFFFFF"
COR_CINZA_FUNDO = "#F0F2F5"
COR_CINZA_BORDA = "#D0D5DD"
COR_VERDE = "#12B76A"
COR_TEXTO = "#1D2939"
COR_TEXTO_CLARO = "#667085"
COR_VERMELHO = "#DC2626"
COR_AMARELO = "#D97706"
COR_AZUL_ESCURO = "#0D2B5E"
COR_AZUL_HOVER = "#1A4FAA"

# Padrões de erro Nokia SR OS
_NOKIA_ERROR_PATTERNS = re.compile(
    r"(?i)"
    r"(?:^MINOR:|^MAJOR:|^CRITICAL:|^WARNING:)"
    r"|(?:Error:)"
    r"|(?:Bad command)"
    r"|(?:Invalid parameter)"
    r"|(?:Syntax error)"
    r"|(?:Not allowed)"
    r"|(?:not found)"
    r"|(?:Cannot )"
    r"|(?:failed)"
    r"|(?:Aborted)"
    r"|(?:denied)"
    r"|(?:INFO: CLI Could not find)"
)

def detectar_erro_nokia(texto):
    """Verifica se a resposta do roteador contém erro. Retorna a linha de erro ou None."""
    for linha in texto.splitlines():
        if _NOKIA_ERROR_PATTERNS.search(linha.strip()):
            return linha.strip()
    return None


# Mapa de cores VT100 para emulador de terminal
_VT_FG = {
    "default": "#00FF00", "black": "#000000", "red": "#CC0000",
    "green": "#4AF626", "brown": "#CCCC00", "blue": "#5555FF",
    "magenta": "#CC00CC", "cyan": "#00CCCC", "white": "#AAAAAA",
}
_VT_FG_BOLD = {
    "default": "#55FF55", "black": "#555555", "red": "#FF5555",
    "green": "#55FF55", "brown": "#FFFF55", "blue": "#7777FF",
    "magenta": "#FF55FF", "cyan": "#55FFFF", "white": "#FFFFFF",
}
_VT_BG = {
    "default": "", "black": "#000000", "red": "#AA0000",
    "green": "#00AA00", "brown": "#AA5500", "blue": "#0000AA",
    "magenta": "#AA00AA", "cyan": "#00AAAA", "white": "#AAAAAA",
}


def carregar_csv(caminho_csv):
    """Carrega o CSV e retorna um dicionário {hostname: {ip, tipo}}."""
    roteadores = {}
    if not os.path.isfile(caminho_csv):
        return roteadores
    with open(caminho_csv, newline="", encoding="utf-8-sig") as f:
        leitor = csv.DictReader(f)
        for linha in leitor:
            hostname = linha.get("hostname", "").strip()
            if hostname:
                roteadores[hostname.upper()] = {
                    "ip": linha.get("ip", "").strip(),
                    "tipo": linha.get("tipo", "").strip(),
                    "timos": linha.get("TiMOS", "").strip(),
                }
    return roteadores


def parsear_system_info(texto):
    """Extrai System Name, System Type e System Version de 'show system information'."""
    info = {"hostname": "", "tipo": "", "timos": ""}
    for linha in texto.splitlines():
        linha_strip = linha.strip()
        if linha_strip.lower().startswith("system name"):
            partes = linha_strip.split(":", 1)
            if len(partes) == 2:
                info["hostname"] = partes[1].strip()
        elif linha_strip.lower().startswith("system type"):
            partes = linha_strip.split(":", 1)
            if len(partes) == 2:
                info["tipo"] = partes[1].strip()
        elif linha_strip.lower().startswith("system version"):
            partes = linha_strip.split(":", 1)
            if len(partes) == 2:
                info["timos"] = partes[1].strip()
    return info


def adicionar_ao_csv(caminho_csv, hostname, ip, tipo, timos):
    """Adiciona uma linha ao CSV de roteadores."""
    arquivo_existe = os.path.isfile(caminho_csv)
    # Garantir que o arquivo termina com newline antes de adicionar
    if arquivo_existe:
        with open(caminho_csv, "rb") as f:
            f.seek(0, 2)  # vai para o final
            if f.tell() > 0:
                f.seek(-1, 2)
                if f.read(1) != b"\n":
                    with open(caminho_csv, "a", encoding="utf-8") as fa:
                        fa.write("\n")
    with open(caminho_csv, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not arquivo_existe:
            writer.writerow(["hostname", "ip", "tipo", "TiMOS"])
        writer.writerow([hostname, ip, tipo, timos])


def gerar_script(hostname, ip, tipo, timos):
    """Gera um script de configuração Nokia SR OS de exemplo."""
    script = f"""\
#--------------------------------------------------
# Script de Configuração - Nokia {tipo}
# Hostname: {hostname}
# IP de Gerência: {ip}
# TiMOS: {timos}
#--------------------------------------------------

/environment no more
/configure system name "{hostname}"

/configure router interface "system"
    address {ip}/32
exit

/configure router
    autonomous-system 65000
    router-id {ip}
exit

/configure router ospf 0
    area 0.0.0.0
        interface "system"
            no shutdown
        exit
    exit
    no shutdown
exit

/configure router mpls
    interface "system"
        no shutdown
    exit
    no shutdown
exit

/configure router rsvp
    interface "system"
        no shutdown
    exit
    no shutdown
exit

/configure router ldp
    interface-parameters
        interface "system"
            no shutdown
        exit
    exit
    no shutdown
exit

/admin save
#--------------------------------------------------
# Fim do script para {hostname}
#--------------------------------------------------
"""
    return script


class GeradorScriptApp:
    def _piscar_icone(self):
        # Pisca o ícone da barra de tarefas se a janela não estiver em foco
        if sys.platform != "win32":
            return
        try:
            hwnd = self.root.winfo_id()
            # Verifica se a janela está minimizada ou não em foco
            is_iconic = ctypes.windll.user32.IsIconic(hwnd)
            fg_win = ctypes.windll.user32.GetForegroundWindow()
            if is_iconic or fg_win != hwnd:
                FLASHW_ALL = 3
                FLASHW_TIMERNOFG = 12
                class FLASHWINFO(ctypes.Structure):
                    _fields_ = [("cbSize", ctypes.wintypes.UINT),
                                ("hwnd", ctypes.wintypes.HWND),
                                ("dwFlags", ctypes.wintypes.DWORD),
                                ("uCount", ctypes.wintypes.UINT),
                                ("dwTimeout", ctypes.wintypes.DWORD)]
                fi = FLASHWINFO(ctypes.sizeof(FLASHWINFO), hwnd, FLASHW_ALL | FLASHW_TIMERNOFG, 5, 0)
                ctypes.windll.user32.FlashWindowEx(ctypes.byref(fi))
        except Exception:
            pass
    def __init__(self, root, only_content=False):
        self.root = root
        if not only_content:
            self.root.title("Nokia Router Script Generator - Eron Netto")
            self.root.geometry("1200x750")
            self.root.resizable(True, True)
            self.root.configure(bg=COR_CINZA_FUNDO)
            self.root.minsize(900, 600)

        self.roteadores = carregar_csv(CSV_PADRAO)
        self.ip_atual = None

        self._configurar_estilo()
        self._criar_widgets()

        if not only_content:
            # Centralizar janela na tela
            self.root.update_idletasks()
            w = self.root.winfo_width()
            h = self.root.winfo_height()
            x = (self.root.winfo_screenwidth() // 2) - (w // 2)
            y = (self.root.winfo_screenheight() // 2) - (h // 2)
            self.root.geometry(f"+{x}+{y}")

    def _configurar_estilo(self):
        style = ttk.Style()
        style.theme_use("clam")

        style.configure("Header.TFrame", background=COR_AZUL_NOKIA)
        style.configure(
            "Header.TLabel",
            background=COR_AZUL_NOKIA,
            foreground=COR_BRANCO,
            font=("Segoe UI", 16, "bold"),
        )
        style.configure(
            "SubHeader.TLabel",
            background=COR_AZUL_NOKIA,
            foreground="#A8C4E6",
            font=("Segoe UI", 9),
        )
        style.configure("Card.TFrame", background=COR_BRANCO)
        style.configure("Main.TFrame", background=COR_CINZA_FUNDO)
        style.configure(
            "CardTitle.TLabel",
            background=COR_BRANCO,
            foreground=COR_TEXTO,
            font=("Segoe UI", 10, "bold"),
        )
        style.configure(
            "CardText.TLabel",
            background=COR_BRANCO,
            foreground=COR_TEXTO_CLARO,
            font=("Segoe UI", 9),
        )
        style.configure(
            "Info.TLabel",
            background=COR_BRANCO,
            foreground=COR_AZUL_NOKIA,
            font=("Consolas", 10),
        )
        style.configure(
            "InfoOk.TLabel",
            background=COR_BRANCO,
            foreground=COR_VERDE,
            font=("Consolas", 10, "bold"),
        )

        # Botões estilizados
        style.configure(
            "Primary.TButton",
            background=COR_AZUL_NOKIA,
            foreground=COR_BRANCO,
            font=("Segoe UI", 8, "bold"),
            padding=(8, 4),
        )
        style.map(
            "Primary.TButton",
            background=[("active", COR_AZUL_CLARO), ("pressed", COR_AZUL_NOKIA)],
        )
        style.configure(
            "Secondary.TButton",
            background=COR_BRANCO,
            foreground=COR_TEXTO,
            font=("Segoe UI", 8),
            padding=(8, 4),
            borderwidth=1,
        )
        style.map(
            "Secondary.TButton",
            background=[("active", COR_CINZA_FUNDO)],
        )

        style.configure(
            "Danger.TButton",
            background="#DC2626",
            foreground=COR_BRANCO,
            font=("Segoe UI", 8, "bold"),
            padding=(8, 4),
        )
        style.map(
            "Danger.TButton",
            background=[("active", "#B91C1C")],
        )
        style.configure(
            "Success.TButton",
            background=COR_VERDE,
            foreground=COR_BRANCO,
            font=("Segoe UI", 8, "bold"),
            padding=(8, 4),
        )
        style.map(
            "Success.TButton",
            background=[("active", "#0E9F5A")],
        )

    def _criar_widgets(self):
        # --- Área principal ---
        self.main_frame = ttk.Frame(self.root, style="Main.TFrame")
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        self.var_csv = tk.StringVar(value=CSV_PADRAO)

        # --- Card: Hostname ---
        card_host = ttk.Frame(self.main_frame, style="Card.TFrame", relief="solid", borderwidth=1)
        card_host.pack(fill=tk.X, pady=(0, 10))

        self.host_inner = ttk.Frame(card_host, style="Card.TFrame")
        self.host_inner.pack(fill=tk.X, padx=12, pady=6)

        ttk.Label(self.host_inner, text="Hostname do Roteador", style="CardTitle.TLabel").pack(anchor=tk.W)

        host_row = ttk.Frame(self.host_inner, style="Card.TFrame")
        host_row.pack(fill=tk.X, pady=(4, 0))

        self.var_hostname = tk.StringVar()
        self.entry_host = tk.Entry(
            host_row, textvariable=self.var_hostname,
            font=("Segoe UI", 10), relief="solid", bd=1,
            highlightcolor=COR_AZUL_NOKIA, highlightthickness=2,
        )
        self.entry_host.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=3)
        self.entry_host.focus_set()

        # --- Credenciais SSH na mesma linha ---
        cred_row = ttk.Frame(self.host_inner, style="Card.TFrame")
        cred_row.pack(fill=tk.X, pady=(6, 0))

        ttk.Label(cred_row, text="Usuário:", style="CardText.TLabel").pack(side=tk.LEFT)
        self.var_usuario = tk.StringVar(value="admin")
        self.var_senha = tk.StringVar(value="admin")
        self.combo_usuario = ttk.Combobox(
            cred_row,
            textvariable=self.var_usuario,
            values=["admin", "93191142"],
            state="readonly",
            width=15,
            font=("Segoe UI", 9),
        )
        self.combo_usuario.pack(side=tk.LEFT, padx=(4, 12), ipady=2)
        def atualizar_senha_e_jump(event=None):
            usuario = self.var_usuario.get()
            if hasattr(self, 'combo_jump'):
                if usuario == "admin":
                    self.var_senha.set("admin")
                    # Preencher jumpserver para admin
                    self.combo_jump.set("192.168.0.131 (root)")
                    self.var_jump_ip.set("192.168.0.131")
                    self.var_jump_user.set("root")
                    self.var_jump_senha.set("root")
                elif usuario == "93191142":
                    self.var_senha.set("X%aA5&z3")
                    # Preencher jumpserver para 93191142
                    self.combo_jump.set("10.73.0.4 (supnokia)")
                    self.var_jump_ip.set("10.73.0.4")
                    self.var_jump_user.set("supnokia")
                    self.var_jump_senha.set("NokiaNsp1!")
                else:
                    self.var_senha.set("")
            else:
                # Só senha, jump ainda não existe
                if usuario == "admin":
                    self.var_senha.set("admin")
                elif usuario == "93191142":
                    self.var_senha.set("X%aA5&z3")
                else:
                    self.var_senha.set("")
        self.combo_usuario.bind("<<ComboboxSelected>>", atualizar_senha_e_jump)
        # Remover chamada aqui, será chamada após combo_jump existir

        ttk.Label(cred_row, text="Senha:", style="CardText.TLabel").pack(side=tk.LEFT)
        tk.Entry(
            cred_row, textvariable=self.var_senha, font=("Segoe UI", 9),
            relief="solid", bd=1, width=15, show="*",
        ).pack(side=tk.LEFT, padx=(4, 12), ipady=2)

        # --- Jumpserver ---
        jump_row = ttk.Frame(self.host_inner, style="Card.TFrame")
        jump_row.pack(fill=tk.X, pady=(6, 0))

        ttk.Label(jump_row, text="Jumpserver:", style="CardText.TLabel").pack(side=tk.LEFT)
        self.var_jump_ip = tk.StringVar()
        self.var_jump_user = tk.StringVar()
        self.var_jump_senha = tk.StringVar()
        jump_options = [
            "",
            "192.168.0.131 (root)",
            "10.73.0.4 (supnokia)"
        ]
        self.combo_jump = ttk.Combobox(
            jump_row,
            values=jump_options,
            state="readonly",
            width=20,
            font=("Segoe UI", 9),
        )
        self.combo_jump.pack(side=tk.LEFT, padx=(4, 12), ipady=2)
        def atualizar_jump(event=None):
            val = self.combo_jump.get()
            if val == "192.168.0.131 (root)":
                self.var_jump_ip.set("192.168.0.131")
                self.var_jump_user.set("root")
                self.var_jump_senha.set("root")
            elif val == "10.73.0.4 (supnokia)":
                self.var_jump_ip.set("10.73.0.4")
                self.var_jump_user.set("supnokia")
                self.var_jump_senha.set("NokiaNsp1!")
            else:
                self.var_jump_ip.set("")
                self.var_jump_user.set("")
                self.var_jump_senha.set("")
        self.combo_jump.bind("<<ComboboxSelected>>", atualizar_jump)
        self.combo_jump.current(0)
        atualizar_jump()
        # Agora que combo_jump existe, podemos chamar atualizar_senha_e_jump para garantir sincronismo
        atualizar_senha_e_jump()

        # Porta removida da interface do jumpserver (sempre 22)

        ttk.Label(jump_row, text="User:", style="CardText.TLabel").pack(side=tk.LEFT)
        tk.Entry(
            jump_row, textvariable=self.var_jump_user, font=("Segoe UI", 9),
            relief="solid", bd=1, width=12,
            state="readonly"
        ).pack(side=tk.LEFT, padx=(4, 12), ipady=2)

        ttk.Label(jump_row, text="Senha:", style="CardText.TLabel").pack(side=tk.LEFT)
        tk.Entry(
            jump_row, textvariable=self.var_jump_senha, font=("Segoe UI", 9),
            relief="solid", bd=1, width=12, show="*",
            state="readonly"
        ).pack(side=tk.LEFT, padx=(4, 0), ipady=2)

        btn_frame = ttk.Frame(self.host_inner, style="Card.TFrame")
        btn_frame.pack(fill=tk.X, pady=(8, 0))

        self.btn_baixar = ttk.Button(
            btn_frame, text="Baixar Configuração", style="Primary.TButton",
            command=self._baixar_config,
        )
        self.btn_baixar.pack(side=tk.LEFT)

        ttk.Button(
            btn_frame, text="Gerar Script", style="Primary.TButton",
            command=self._buscar_e_gerar,
        ).pack(side=tk.LEFT, padx=(8, 0))

        ttk.Button(
            btn_frame, text="Salvar Script", style="Secondary.TButton",
            command=self._salvar_script,
        ).pack(side=tk.LEFT, padx=(8, 0))

        ttk.Button(
            btn_frame, text="Limpar", style="Secondary.TButton",
            command=self._limpar,
        ).pack(side=tk.LEFT, padx=(8, 0))

        self.btn_enviar = ttk.Button(
            btn_frame, text="Conectar e Enviar via SSH", style="Success.TButton",
            command=self._enviar_ssh,
        )
        self.btn_enviar.pack(side=tk.LEFT, padx=(8, 0))

        # --- Info do roteador ---
        self.lbl_info = ttk.Label(
            self.host_inner, text="", style="Info.TLabel",
        )
        self.lbl_info.pack(fill=tk.X, pady=(6, 0))

        # --- Área inferior: Script Gerado (esquerda) + Log SSH (direita) ---
        self.bottom_frame = ttk.Frame(self.root, style="Main.TFrame")
        self.bottom_frame.pack(fill=tk.BOTH, expand=True)

        # PanedWindow permite redimensionar a divisão arrastando
        paned = tk.PanedWindow(self.bottom_frame, orient=tk.HORIZONTAL, sashwidth=6,
                       bg=COR_CINZA_FUNDO, relief="flat")
        paned.pack(fill=tk.BOTH, expand=True)

        # --- Painel esquerdo: Script Gerado ---
        card_script = ttk.Frame(paned, style="Card.TFrame", relief="solid", borderwidth=1)
        script_inner = ttk.Frame(card_script, style="Card.TFrame")
        script_inner.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        ttk.Label(script_inner, text="Script Gerado", style="CardTitle.TLabel").pack(anchor=tk.W)

        self.txt_script = scrolledtext.ScrolledText(
            script_inner, wrap=tk.WORD, font=("Consolas", 10),
            bg="#1E293B", fg="#E2E8F0", insertbackground=COR_BRANCO,
            relief="solid", bd=1, padx=10, pady=8,
        )
        self.txt_script.pack(fill=tk.BOTH, expand=True, pady=(4, 0))

        paned.add(card_script, stretch="always")

        # --- Painel direito: Log SSH ---
        card_log = ttk.Frame(paned, style="Card.TFrame", relief="solid", borderwidth=1)
        log_inner = ttk.Frame(card_log, style="Card.TFrame")
        log_inner.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        ttk.Label(log_inner, text="Log SSH", style="CardTitle.TLabel").pack(anchor=tk.W)

        self.txt_log = scrolledtext.ScrolledText(
            log_inner, wrap=tk.WORD, font=("Consolas", 9),
            bg="#0F172A", fg="#94A3B8", insertbackground=COR_BRANCO,
            relief="solid", bd=1, padx=8, pady=6,
        )
        self.txt_log.pack(fill=tk.BOTH, expand=True, pady=(4, 0))

        # Tags de cor para o log
        self.txt_log.tag_configure("erro", foreground="#EF4444", font=("Consolas", 9, "bold"))
        self.txt_log.tag_configure("aviso", foreground="#F59E0B")
        self.txt_log.tag_configure("ok", foreground="#34D399", font=("Consolas", 9, "bold"))

        paned.add(card_log, stretch="always")

    def _buscar_e_gerar(self):
        hostname = self.var_hostname.get().strip()
        if not hostname:
            messagebox.showwarning("Atenção", "Digite o hostname do roteador.")
            return

        # Recarrega CSV caso tenha mudado o caminho
        caminho_csv = self.var_csv.get()
        self.roteadores = carregar_csv(caminho_csv)

        chave = hostname.upper()
        if chave not in self.roteadores:
            self._hostname_nao_encontrado(hostname)
            return

        dados = self.roteadores[chave]
        ip = dados["ip"]
        tipo = dados["tipo"]
        timos = dados["timos"]

        self.lbl_info.config(
            text=f"Roteador: {hostname}  |  IP: {ip}  |  Tipo: {tipo}  |  TiMOS: {timos}",
            style="InfoOk.TLabel",
        )

        self.ip_atual = ip
        script = gerar_script(hostname, ip, tipo, timos)
        self.txt_script.delete("1.0", tk.END)
        self.txt_script.insert(tk.END, script)

    def _salvar_script(self):
        conteudo = self.txt_script.get("1.0", tk.END).strip()
        if not conteudo:
            messagebox.showwarning("Atenção", "Nenhum script para salvar.")
            return

        hostname = self.var_hostname.get().strip() or "script"
        caminho = filedialog.asksaveasfilename(
            title="Salvar Script",
            defaultextension=".txt",
            initialfile=f"script_{hostname}.txt",
            filetypes=[("Text files", "*.txt"), ("Todos", "*.*")],
            initialdir=DIRETORIO_BASE,
        )
        if caminho:
            with open(caminho, "w", encoding="utf-8") as f:
                f.write(conteudo)
            messagebox.showinfo("Salvo", f"Script salvo em:\n{caminho}")

    def _limpar(self):
        self.var_hostname.set("")
        self.ip_atual = None
        self.lbl_info.config(text="", style="Info.TLabel")
        self.txt_script.delete("1.0", tk.END)
        self.txt_log.delete("1.0", tk.END)
        self.entry_host.focus_set()

    def _hostname_nao_encontrado(self, hostname):
        """Hostname não está no CSV — resolve via DNS, conecta via SSH, extrai info e atualiza CSV."""
        import subprocess
        import re
        fqdn = f"{hostname}.embratel.net.br"

        try:
            # Executa ping para resolver o DNS
            proc = subprocess.run([
                "ping", fqdn, "-n", "1"
            ], capture_output=True, text=True, timeout=5)
            saida = proc.stdout
            # Extrai o IP do DNS resolvido
            match = re.search(r"\[(\d+\.\d+\.\d+\.\d+)\]", saida)
            if not match:
                match = re.search(r"(\d+\.\d+\.\d+\.\d+)", saida)
            if match:
                ip = match.group(1)
            else:
                # Não conseguiu resolver, perguntar ao usuário
                ip = tk.simpledialog.askstring(
                    "IP Manual",
                    f"Não foi possível resolver o IP de {fqdn} pelo DNS.\n\nDigite o IP manualmente:",
                    parent=self.root
                )
                if not ip:
                    messagebox.showwarning("Atenção", "IP não informado. Operação cancelada.")
                    return
        except Exception as e:
            # Falha ao tentar resolver, perguntar ao usuário
            ip = tk.simpledialog.askstring(
                "IP Manual",
                f"Falha ao tentar resolver {fqdn}: {e}\n\nDigite o IP manualmente:",
                parent=self.root
            )
            if not ip:
                messagebox.showwarning("Atenção", "IP não informado. Operação cancelada.")
                return

        usuario = self.var_usuario.get().strip()
        senha = self.var_senha.get()
        porta = 22

        if not usuario:
            messagebox.showwarning("Atenção", "Preencha o usuário SSH.")
            return

        self.txt_log.delete("1.0", tk.END)

        def ssh_extrair_info():
            try:
                client, jump_client = self._conectar_ssh(ip, porta, usuario, senha)
                self.root.after(0, self._log, f"[SSH] Conectado a {ip}. Executando 'show system information'...")
                shell = client.invoke_shell(width=512)
                time.sleep(1)
                # Descartar banner
                if shell.recv_ready():
                    shell.recv(65535)
                shell.send("environment no more\n")
                time.sleep(1)
                if shell.recv_ready():
                    shell.recv(65535)
                shell.send("show system information\n")
                saida = ""
                tentativas = 0
                while tentativas < 6:
                    time.sleep(1)
                    if shell.recv_ready():
                        bloco = shell.recv(65535).decode("utf-8", errors="replace")
                        saida += bloco
                        tentativas = 0
                    else:
                        tentativas += 1
                shell.close()
                self._fechar_ssh(client, jump_client)
                # Parsear informações
                info = parsear_system_info(saida)
                sys_name = info["hostname"] or hostname
                sys_type = info["tipo"]
                sys_version = info["timos"]
                # Adicionar ao CSV
                caminho_csv = self.var_csv.get()
                adicionar_ao_csv(caminho_csv, sys_name, ip, sys_type, sys_version)
                # Recarregar CSV
                self.roteadores = carregar_csv(caminho_csv)
                # Atualizar hostname no campo se diferente
                self.root.after(0, self.var_hostname.set, sys_name)
                self.ip_atual = ip
                self.root.after(0, self._log_ok,
                    f"\n[OK] Roteador descoberto e adicionado ao CSV:"
                    f"\n  System Name: {sys_name}"
                    f"\n  System Type: {sys_type}"
                    f"\n  System Version: {sys_version}"
                    f"\n  IP: {ip}")
                self.root.after(0, self.lbl_info.config,
                    {"text": f"Roteador: {sys_name}  |  IP: {ip}  |  Tipo: {sys_type}  |  TiMOS: {sys_version}",
                     "style": "InfoOk.TLabel"})
                self.root.after(0, lambda: messagebox.showinfo(
                    "Roteador Descoberto",
                    f"Informações obtidas com sucesso e salvas no CSV!\n\n"
                    f"System Name: {sys_name}\n"
                    f"System Type: {sys_type}\n"
                    f"System Version: {sys_version}\n"
                    f"IP: {ip}",
                ))
                # Após salvar no CSV, baixar a configuração automaticamente
                self.root.after(0, self._baixar_config)
            except paramiko.AuthenticationException:
                self.root.after(0, self._log_erro, "[ERRO] Falha de autenticação.")
                self.root.after(0, lambda: messagebox.showerror(
                    "Erro SSH", "Falha de autenticação.\nVerifique usuário e senha.",
                ))
            except paramiko.SSHException as e:
                self.root.after(0, self._log_erro, f"[ERRO] Erro SSH: {e}")
                self.root.after(0, lambda: messagebox.showerror("Erro SSH", str(e)))
            except OSError as e:
                self.root.after(0, self._log_erro, f"[ERRO] Não foi possível conectar: {e}")
                self.root.after(0, lambda: messagebox.showerror(
                    "Erro de Conexão", f"Não foi possível conectar a {ip}:{porta}\n{e}",
                ))

        thread = threading.Thread(
            target=ssh_extrair_info,
            daemon=True,
        )
        thread.start()

    def _descobrir_roteador(self, hostname_digitado, ip, porta, usuario, senha):
        """Conecta via SSH, executa show system information e atualiza o CSV."""
        try:
            client, jump_client = self._conectar_ssh(ip, porta, usuario, senha)

            self.root.after(0, self._log,
                f"[SSH] Conectado a {ip}. Executando 'show system information'...")

            shell = client.invoke_shell(width=512)
            time.sleep(1)

            # Descartar banner
            if shell.recv_ready():
                shell.recv(65535)

            shell.send("environment no more\n")
            time.sleep(1)
            if shell.recv_ready():
                shell.recv(65535)

            shell.send("show system information\n")

            saida = ""
            tentativas = 0
            while tentativas < 6:
                time.sleep(1)
                if shell.recv_ready():
                    bloco = shell.recv(65535).decode("utf-8", errors="replace")
                    saida += bloco
                    tentativas = 0
                else:
                    tentativas += 1

            shell.close()
            self._fechar_ssh(client, jump_client)

            # Parsear informações
            info = parsear_system_info(saida)
            sys_name = info["hostname"] or hostname_digitado
            sys_type = info["tipo"]
            sys_version = info["timos"]

            if not sys_name and not sys_type:
                self.root.after(0, self._log_erro,
                    "[ERRO] Não foi possível extrair informações do roteador.")
                self.root.after(0, lambda: messagebox.showerror(
                    "Erro", "Não foi possível extrair informações de 'show system information'.",
                ))
                return

            # Adicionar ao CSV
            caminho_csv = self.var_csv.get()
            adicionar_ao_csv(caminho_csv, sys_name, ip, sys_type, sys_version)

            # Recarregar CSV
            self.roteadores = carregar_csv(caminho_csv)

            # Atualizar hostname no campo se diferente
            self.root.after(0, self.var_hostname.set, sys_name)

            self.ip_atual = ip

            self.root.after(0, self._log_ok,
                f"\n[OK] Roteador descoberto e adicionado ao CSV:"
                f"\n  System Name: {sys_name}"
                f"\n  System Type: {sys_type}"
                f"\n  System Version: {sys_version}"
                f"\n  IP: {ip}")

            self.root.after(0, self.lbl_info.config,
                {"text": f"Roteador: {sys_name}  |  IP: {ip}  |  Tipo: {sys_type}  |  TiMOS: {sys_version}",
                 "style": "InfoOk.TLabel"})

            self.root.after(0, lambda: messagebox.showinfo(
                "Roteador Descoberto",
                f"Informações obtidas com sucesso e salvas no CSV!\n\n"
                f"System Name: {sys_name}\n"
                f"System Type: {sys_type}\n"
                f"System Version: {sys_version}\n"
                f"IP: {ip}",
            ))

        except paramiko.AuthenticationException:
            self.root.after(0, self._log_erro, "[ERRO] Falha de autenticação.")
            self.root.after(0, lambda: messagebox.showerror(
                "Erro SSH", "Falha de autenticação.\nVerifique usuário e senha.",
            ))
        except paramiko.SSHException as e:
            self.root.after(0, self._log_erro, f"[ERRO] Erro SSH: {e}")
            self.root.after(0, lambda: messagebox.showerror("Erro SSH", str(e)))
        except OSError as e:
            self.root.after(0, self._log_erro, f"[ERRO] Não foi possível conectar: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "Erro de Conexão", f"Não foi possível conectar a {ip}:{porta}\n{e}",
            ))

    def _log(self, msg, tag=None):
        if tag:
            self.txt_log.insert(tk.END, msg + "\n", tag)
        else:
            self.txt_log.insert(tk.END, msg + "\n")
        self.txt_log.see(tk.END)
        self.root.update_idletasks()

    def _log_erro(self, msg):
        self._log(msg, "erro")

    def _log_aviso(self, msg):
        self._log(msg, "aviso")

    def _log_ok(self, msg):
        self._log(msg, "ok")

    def _enviar_ssh(self):
        script = self.txt_script.get("1.0", tk.END).strip()
        if not script:
            messagebox.showwarning("Atenção", "Gere o script primeiro.")
            return
        if not self.ip_atual:
            messagebox.showwarning("Atenção", "Busque um hostname primeiro.")
            return

        ip = self.ip_atual
        usuario = self.var_usuario.get().strip()
        senha = self.var_senha.get()
        # Porta SSH fixa 22 (não mais editável pelo usuário)
        porta = 22

        if not usuario:
            messagebox.showwarning("Atenção", "Preencha o usuário SSH.")
            return

        # Confirmar antes de enviar
        hostname = self.var_hostname.get().strip()
        if not messagebox.askyesno(
            "Confirmar Envio",
            f"Conectar via SSH em {hostname} ({ip}:{porta}) "
            f"e enviar o script?\n\nUsuário: {usuario}",
        ):
            return

        # Desabilitar botão durante execução
        self.btn_enviar.config(state="disabled")
        self.txt_log.delete("1.0", tk.END)

        # Executar em thread separada para não travar a GUI
        thread = threading.Thread(
            target=self._executar_ssh,
            args=(ip, porta, usuario, senha, script),
            daemon=True,
        )
        thread.start()

    def _obter_jump_params(self):
        """Retorna parâmetros do jumpserver ou None se não configurado. Porta sempre 22."""
        jump_ip = self.var_jump_ip.get().strip()
        if not jump_ip:
            return None
        jump_user = self.var_jump_user.get().strip()
        jump_senha = self.var_jump_senha.get()
        return {"ip": jump_ip, "porta": 22, "user": jump_user, "senha": jump_senha}

    def _conectar_via_jump(self, ip, porta, usuario, senha, jump):
        """Conecta ao roteador via jumpserver. Retorna (client, jump_client)."""
        self.root.after(0, self._log,
            f"[SSH] Conectando ao jumpserver {jump['ip']}:{jump['porta']}...")

        jump_client = paramiko.SSHClient()
        jump_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        jump_client.connect(
            hostname=jump["ip"],
            port=jump["porta"],
            username=jump["user"],
            password=jump["senha"],
            timeout=15,
            look_for_keys=False,
            allow_agent=False,
        )

        self.root.after(0, self._log,
            f"[SSH] Jumpserver conectado. Abrindo túnel para {ip}:{porta}...")

        jump_transport = jump_client.get_transport()
        jump_channel = jump_transport.open_channel(
            "direct-tcpip", (ip, porta), ("127.0.0.1", 0),
        )

        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(
            hostname=ip,
            port=porta,
            username=usuario,
            password=senha,
            timeout=15,
            look_for_keys=False,
            allow_agent=False,
            sock=jump_channel,
        )

        self.root.after(0, self._log,
            f"[SSH] Conectado a {ip} via jumpserver {jump['ip']}.")
        return client, jump_client

    def _conectar_direto(self, ip, porta, usuario, senha):
        """Conecta diretamente ao roteador. Retorna (client, None)."""
        self.root.after(0, self._log, f"[SSH] Conectando a {ip}:{porta}...")

        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(
            hostname=ip,
            port=porta,
            username=usuario,
            password=senha,
            timeout=15,
            look_for_keys=False,
            allow_agent=False,
        )

        self.root.after(0, self._log, f"[SSH] Conectado a {ip}:{porta}.")
        return client, None

    def _conectar_ssh(self, ip, porta, usuario, senha):
        """Conecta ao roteador (direto ou via jump). Retorna (client, jump_client)."""
        jump = self._obter_jump_params()
        if jump:
            return self._conectar_via_jump(ip, porta, usuario, senha, jump)
        return self._conectar_direto(ip, porta, usuario, senha)

    def _fechar_ssh(self, client, jump_client):
        """Fecha conexões SSH."""
        try:
            client.close()
        except Exception:
            pass
        if jump_client:
            try:
                jump_client.close()
            except Exception:
                pass

    def _executar_ssh(self, ip, porta, usuario, senha, script):
        erros = []
        try:
            client, jump_client = self._conectar_ssh(ip, porta, usuario, senha)

            self.root.after(0, self._log, "[SSH] Abrindo shell...")

            shell = client.invoke_shell()
            time.sleep(1)

            # Ler banner inicial
            if shell.recv_ready():
                banner = shell.recv(65535).decode("utf-8", errors="replace")
                self.root.after(0, self._log, banner.strip())

            # Enviar comandos linha a linha (ignora comentários e linhas vazias)
            linhas = script.splitlines()
            total_cmds = 0
            for linha in linhas:
                linha_strip = linha.strip()
                if not linha_strip or linha_strip.startswith("#"):
                    continue
                total_cmds += 1
                shell.send(linha + "\n")
                time.sleep(0.5)

                # Ler resposta
                if shell.recv_ready():
                    resp = shell.recv(65535).decode("utf-8", errors="replace")
                    erro_detectado = detectar_erro_nokia(resp)
                    if erro_detectado:
                        erros.append({"cmd": linha_strip, "erro": erro_detectado})
                        self.root.after(0, self._log,
                            f">>> {linha_strip}")
                        self.root.after(0, self._log_erro,
                            f"[ERRO] {erro_detectado}")
                    else:
                        self.root.after(0, self._log, resp.strip())

            # Esperar resposta final
            time.sleep(2)
            if shell.recv_ready():
                resp_final = shell.recv(65535).decode("utf-8", errors="replace")
                erro_detectado = detectar_erro_nokia(resp_final)
                if erro_detectado:
                    erros.append({"cmd": "(resposta final)", "erro": erro_detectado})
                    self.root.after(0, self._log_erro, f"[ERRO] {erro_detectado}")
                else:
                    self.root.after(0, self._log, resp_final.strip())

            shell.close()
            self._fechar_ssh(client, jump_client)

            # Resumo final
            if erros:
                resumo = f"\n{'='*50}\n"
                resumo += f"[RESUMO] {len(erros)} erro(s) detectado(s) em {total_cmds} comando(s):\n"
                for i, e in enumerate(erros, 1):
                    resumo += f"  {i}. Comando: {e['cmd']}\n     Erro: {e['erro']}\n"
                resumo += f"{'='*50}"
                self.root.after(0, self._log_erro, resumo)
                self.root.after(0, lambda: messagebox.showwarning(
                    "Concluído com erros",
                    f"Script enviado para {ip}, mas {len(erros)} comando(s) "
                    f"retornaram erro.\nVeja o Log SSH para detalhes.",
                ))
            else:
                self.root.after(0, self._log_ok,
                    f"\n[SSH] Script enviado com sucesso. "
                    f"{total_cmds} comando(s) executados sem erros.")
                self.root.after(0, lambda: messagebox.showinfo(
                    "Sucesso", f"Script enviado para {ip} com sucesso!\n"
                    f"{total_cmds} comando(s) executados sem erros.",
                ))

        except paramiko.AuthenticationException:
            self.root.after(0, self._log, "[ERRO] Falha de autenticação. Verifique usuário/senha.")
            self.root.after(0, lambda: messagebox.showerror(
                "Erro SSH", "Falha de autenticação.\nVerifique usuário e senha."
            ))
        except paramiko.SSHException as e:
            self.root.after(0, self._log, f"[ERRO] Erro SSH: {e}")
            self.root.after(0, lambda: messagebox.showerror("Erro SSH", str(e)))
        except OSError as e:
            self.root.after(0, self._log, f"[ERRO] Não foi possível conectar: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "Erro de Conexão", f"Não foi possível conectar a {ip}:{porta}\n{e}"
            ))
        finally:
            self.root.after(0, lambda: self.btn_enviar.config(state="normal"))

    def _resolver_ip(self):
        """Busca hostname no CSV e retorna (hostname, ip) ou None."""
        hostname = self.var_hostname.get().strip()
        if not hostname:
            messagebox.showwarning("Aten\u00e7\u00e3o", "Digite o hostname do roteador.")
            return None

        caminho_csv = self.var_csv.get()
        self.roteadores = carregar_csv(caminho_csv)
        chave = hostname.upper()

        if chave not in self.roteadores:
            self._hostname_nao_encontrado(hostname)
            return None

        dados = self.roteadores[chave]
        ip = dados["ip"]
        tipo = dados["tipo"]
        timos = dados["timos"]

        self.ip_atual = ip
        self.lbl_info.config(
            text=f"Roteador: {hostname}  |  IP: {ip}  |  Tipo: {tipo}  |  TiMOS: {timos}",
            style="InfoOk.TLabel",
        )
        return hostname, ip

    def _baixar_config(self):
        resultado = self._resolver_ip()
        if not resultado:
            return

        hostname, ip = resultado
        usuario = self.var_usuario.get().strip()
        senha = self.var_senha.get()
        # Porta SSH fixa 22 (não mais editável pelo usuário)
        porta = 22
        if not usuario:
            messagebox.showwarning("Aten\u00e7\u00e3o", "Preencha o usu\u00e1rio SSH.")
            return

        self.btn_baixar.config(state="disabled")
        self.txt_log.delete("1.0", tk.END)
        self.txt_script.delete("1.0", tk.END)

        thread = threading.Thread(
            target=self._executar_baixar_config,
            args=(hostname, ip, porta, usuario, senha),
            daemon=True,
        )
        thread.start()

    def _executar_baixar_config(self, hostname, ip, porta, usuario, senha):
        import datetime
        try:
            client, jump_client = self._conectar_ssh(ip, porta, usuario, senha)

            self.root.after(0, self._log, "[SSH] Conectado. Baixando configuração...")

            shell = client.invoke_shell(width=512)
            time.sleep(1)

            # Descartar banner
            if shell.recv_ready():
                shell.recv(65535)

            # Enviar environment no more
            shell.send("environment no more\n")
            time.sleep(1)
            if shell.recv_ready():
                shell.recv(65535)

            # Enviar admin display-config
            shell.send("admin display-config\n")

            # Coletar toda a saída
            config_completa = ""
            tentativas_sem_dados = 0
            while tentativas_sem_dados < 6:
                time.sleep(1)
                if shell.recv_ready():
                    bloco = shell.recv(65535).decode("utf-8", errors="replace")
                    config_completa += bloco
                    tentativas_sem_dados = 0
                else:
                    tentativas_sem_dados += 1

            shell.close()
            client.close()

            # Limpar a saída
            linhas = config_completa.splitlines()
            config_limpa = []
            capturando = False
            for linha in linhas:
                if "admin display-config" in linha and not capturando:
                    capturando = True
                    continue
                if capturando:
                    config_limpa.append(linha)

            texto_final = "\n".join(config_limpa).strip() if config_limpa else config_completa.strip()

            # Salvar arquivo no diretório do hostname
            try:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                dir_hostname = os.path.join(base_dir, hostname)
                if not os.path.exists(dir_hostname):
                    os.makedirs(dir_hostname)
                datahora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                nome_arquivo = f"{hostname}_{datahora}.txt"
                caminho_arquivo = os.path.join(dir_hostname, nome_arquivo)
                with open(caminho_arquivo, "w", encoding="utf-8") as f:
                    f.write(texto_final)
                self.root.after(0, lambda: messagebox.showinfo(
                    "Configuração Salva",
                    f"Configuração salva em:\n{caminho_arquivo}"
                ))
            except Exception as e:
                self.root.after(0, self._log_erro, f"[ERRO] Falha ao salvar arquivo: {e}")

            def _atualizar_tela():
                self.txt_script.delete("1.0", tk.END)
                self.txt_script.insert(tk.END, texto_final)
                self._log(f"\n[SSH] Configuração de {hostname} ({ip}) baixada com sucesso.")

            self.root.after(0, _atualizar_tela)

        except paramiko.AuthenticationException:
            self.root.after(0, self._log, "[ERRO] Falha de autenticação. Verifique usuário/senha.")
            self.root.after(0, lambda: messagebox.showerror(
                "Erro SSH", "Falha de autenticação.\nVerifique usuário e senha."
            ))
        except paramiko.SSHException as e:
            self.root.after(0, self._log, f"[ERRO] Erro SSH: {e}")
            self.root.after(0, lambda: messagebox.showerror("Erro SSH", str(e)))
        except OSError as e:
            self.root.after(0, self._log, f"[ERRO] Não foi possível conectar: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "Erro de Conexão", f"Não foi possível conectar a {ip}:{porta}\n{e}"
            ))
        finally:
            self.root.after(0, lambda: self.btn_baixar.config(state="normal"))



# --- NOVA CLASSE: AppMaster para menu principal ---
class AppMaster:
    def __init__(self, root):
        self.root = root
        self.root.title("Nokia Tools - CH-Berde")
        self.root.state('zoomed')
        self.root.geometry("1200x750")
        self.root.resizable(True, True)
        self.root.configure(bg=COR_CINZA_FUNDO)
        self.root.minsize(900, 600)

        # ── Configure ttk styles ─────────────────────────────────
        _style = ttk.Style()
        _style.theme_use("clam")
        _style.configure("Header.TFrame", background=COR_AZUL_NOKIA)
        _style.configure("Header.TLabel", background=COR_AZUL_NOKIA,
                         foreground=COR_BRANCO, font=("Segoe UI", 16, "bold"))
        _style.configure("SubHeader.TLabel", background=COR_AZUL_NOKIA,
                         foreground="#A8C4E6", font=("Segoe UI", 9))

        # ── Header ───────────────────────────────────────────────
        self.header = ttk.Frame(self.root, style="Header.TFrame")
        self.header.pack(fill=tk.X, side=tk.TOP)
        # Thin accent line at bottom of header
        tk.Frame(self.header, height=2, bg="#2563EB").pack(fill=tk.X, side=tk.BOTTOM)
        header_inner = ttk.Frame(self.header, style="Header.TFrame")
        header_inner.pack(fill=tk.X, padx=20, pady=(14, 12))
        self._nokia_home_btn = tk.Button(
            header_inner, text="NOKIA",
            font=("Segoe UI", 20, "bold"),
            bg=COR_AZUL_NOKIA, fg=COR_BRANCO, bd=0, relief="flat",
            activebackground=COR_AZUL_NOKIA, activeforeground="#93C5FD",
            cursor="hand2", command=self._mostrar_boas_vindas,
        )
        self._nokia_home_btn.pack(side=tk.LEFT)
        self._nokia_home_btn.bind("<Enter>", lambda e: self._nokia_home_btn.config(fg="#93C5FD"))
        self._nokia_home_btn.bind("<Leave>", lambda e: self._nokia_home_btn.config(fg=COR_BRANCO))

        # "Projeto Fotônico" centered in header
        center_frame = ttk.Frame(header_inner, style="Header.TFrame")
        center_frame.pack(side=tk.LEFT, expand=True)
        tk.Label(
            center_frame, text="Projeto Fotônico",
            font=("Segoe UI", 16, "bold"),
            bg=COR_AZUL_NOKIA, fg=COR_BRANCO,
        ).pack(anchor=tk.CENTER)

        # Claro icon on the right (from claro.png)
        right_frame = ttk.Frame(header_inner, style="Header.TFrame")
        right_frame.pack(side=tk.RIGHT)
        try:
            _claro_path = os.path.join(DIRETORIO_BASE, "Claro.png")
            _claro_img = Image.open(_claro_path)
            # Redimensionar mantendo proporção para caber no cabeçalho (max 70px altura)
            _max_h = 70
            _orig_w, _orig_h = _claro_img.size
            _ratio = _max_h / _orig_h
            _new_w = int(_orig_w * _ratio)
            _claro_img = _claro_img.resize((_new_w, _max_h), Image.LANCZOS)
            self._claro_photo = ImageTk.PhotoImage(_claro_img)
            tk.Label(
                right_frame, image=self._claro_photo,
                bg=COR_AZUL_NOKIA, bd=0,
            ).pack(anchor=tk.E)
        except Exception:
            tk.Label(
                right_frame, text="claro",
                font=("Segoe UI", 10, "bold"),
                bg=COR_AZUL_NOKIA, fg="#EE1C25",
            ).pack(anchor=tk.E)

        # ── Main frame  ─────────────────────────────────────────
        self.main_frame = tk.Frame(self.root, bg=COR_CINZA_FUNDO)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # ── Sidebar (bookmark 3-D) ──────────────────────────────
        self.menu_frame = tk.Frame(self.main_frame, bg=COR_AZUL_NOKIA, width=230)
        # Don't pack yet – welcome page hides sidebar; _mostrar_sidebar() packs it when needed
        self.menu_frame.pack_propagate(False)

        # -- helper: bookmark-style button -----------------------
        def _bookmark_btn(parent, text, command, indent=0, is_sub=False):
            font = ("Segoe UI", 8) if is_sub else ("Segoe UI", 9, "bold")
            bg = COR_AZUL_ESCURO if is_sub else COR_AZUL_NOKIA
            padx = 18 + indent
            btn = tk.Button(
                parent, text=text, font=font,
                bg=bg, fg=COR_BRANCO, bd=0, relief="flat",
                activebackground=COR_AZUL_CLARO, activeforeground=COR_BRANCO,
                anchor="w", padx=padx, cursor="hand2",
                command=command,
            )
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg=COR_AZUL_HOVER) if b != getattr(self, '_menu_btn_ativo', None) else None)
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg=COR_AZUL_CLARO if b == getattr(self, '_menu_btn_ativo', None) else (COR_AZUL_ESCURO if is_sub else COR_AZUL_NOKIA)))
            return btn

        # 3-D separator
        def _sep(parent):
            container = tk.Frame(parent, bg=COR_AZUL_NOKIA)
            tk.Frame(container, height=1, bg="#0A2A5C").pack(fill=tk.X, padx=4)
            tk.Frame(container, height=1, bg="#2563EB").pack(fill=tk.X, padx=4)
            container.pack(fill=tk.X)
            return container

        # ── Gerador de script (parent) ──────────────────────────
        self.btn_gerador = _bookmark_btn(self.menu_frame, "  \U0001F4DD  Gerador de Script", self._toggle_gerador_menu)
        self.btn_gerador.pack(fill=tk.X, pady=(24, 0), padx=4, ipady=7)
        self._sep_gerador = _sep(self.menu_frame)

        # Sub-menu container (hidden by default)
        self._sub_gerador_frame = tk.Frame(self.menu_frame, bg=COR_AZUL_NOKIA)
        self._sub_gerador_visible = False

        self.btn_integracao = _bookmark_btn(self._sub_gerador_frame, "     ▸  Integração", lambda: self._mostrar_pagina("integracao"), indent=8, is_sub=True)
        self.btn_integracao.pack(fill=tk.X, padx=4, ipady=4)
        self.btn_swap = _bookmark_btn(self._sub_gerador_frame, "     ▸  SWAP", lambda: self._mostrar_pagina("swap"), indent=8, is_sub=True)
        self.btn_swap.pack(fill=tk.X, padx=4, ipady=4)
        self.btn_ampliacao = _bookmark_btn(self._sub_gerador_frame, "     ▸  Ampliação de placa", lambda: self._mostrar_pagina("ampliacao"), indent=8, is_sub=True)
        self.btn_ampliacao.pack(fill=tk.X, padx=4, ipady=4)
        self.btn_pcep = _bookmark_btn(self._sub_gerador_frame, "     ▸  Script PCEP", lambda: self._mostrar_pagina("pcep"), indent=8, is_sub=True)
        self.btn_pcep.pack(fill=tk.X, padx=4, ipady=4)
        self.btn_rollback = _bookmark_btn(self._sub_gerador_frame, "     ▸  Alterar Config (Via Rollback)", lambda: self._mostrar_pagina("rollback"), indent=8, is_sub=True)
        self.btn_rollback.pack(fill=tk.X, padx=4, ipady=4)

        # ── Upgrade de software (parent) ─────────────────────────
        self.btn_upgrade = _bookmark_btn(self.menu_frame, "  \U0001F504  Upgrade de Software", self._toggle_upgrade_menu)
        self.btn_upgrade.pack(fill=tk.X, pady=(2, 0), padx=4, ipady=7)
        self._sep_upgrade = _sep(self.menu_frame)

        # Sub-menu container (hidden by default)
        self._sub_upgrade_frame = tk.Frame(self.menu_frame, bg=COR_AZUL_NOKIA)
        self._sub_upgrade_visible = False

        self.btn_upgrade_7x50 = _bookmark_btn(self._sub_upgrade_frame, "     ▸  Upgrade 7x50", lambda: self._mostrar_pagina("upgrade_7x50"), indent=8, is_sub=True)
        self.btn_upgrade_7x50.pack(fill=tk.X, padx=4, ipady=4)
        self.btn_upgrade_vrr = _bookmark_btn(self._sub_upgrade_frame, "     ▸  Upgrade vRR ou RR", lambda: self._mostrar_pagina("upgrade_vrr"), indent=8, is_sub=True)
        self.btn_upgrade_vrr.pack(fill=tk.X, padx=4, ipady=4)

        # Map of sub-buttons for highlighting
        self._sub_buttons = {
            "integracao": self.btn_integracao,
            "swap": self.btn_swap,
            "ampliacao": self.btn_ampliacao,
            "pcep": self.btn_pcep,
            "rollback": self.btn_rollback,
            "upgrade_7x50": self.btn_upgrade_7x50,
            "upgrade_vrr": self.btn_upgrade_vrr,
        }

        # ── Content area ─────────────────────────────────────────
        self.content_frame = tk.Frame(self.main_frame, bg=COR_CINZA_FUNDO)
        self.content_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ── Page cache ───────────────────────────────────────────
        self._pagina_atual = None
        self._menu_btn_ativo = None
        self.page_gerador = None    # Integração content
        self.page_upgrade = None    # Upgrade 7x50 content
        self.page_swap = None       # SWAP content
        self._pages_em_breve = {}   # placeholder pages
        self._wcard_gerador = None  # welcome card refs
        self._wcard_upgrade = None

        # ── Footer (not shown on welcome page) ───────────────────
        self.footer = tk.Frame(self.root, bg=COR_AZUL_NOKIA, height=24)
        self.footer.pack(fill=tk.X, side=tk.BOTTOM)
        self.footer.pack_propagate(False)
        tk.Label(
            self.footer,
            text="Developed by Eron Netto - ION IP Resident Engineer - Nokia 2026",
            font=("Segoe UI", 7),
            bg=COR_AZUL_NOKIA, fg="#FFFFFF",
        ).pack(expand=True)

        # ── Show welcome page ────────────────────────────────────
        self._mostrar_boas_vindas()

    # ─────────────────────────────────────────────────────────────
    #  Welcome page
    # ─────────────────────────────────────────────────────────────
    def _mostrar_boas_vindas(self):
        # Close any open welcome cards
        self._welcome_close_cards()
        self._limpar_content()
        self._pagina_atual = "welcome"
        # Reset active button highlight
        self._reset_menu_highlight()
        # Collapse submenus and hide sidebar
        if self._sub_gerador_visible:
            self._sub_gerador_frame.pack_forget()
            self._sub_gerador_visible = False
        if self._sub_upgrade_visible:
            self._sub_upgrade_frame.pack_forget()
            self._sub_upgrade_visible = False
        self.menu_frame.pack_forget()
        # Hide header and footer on welcome page
        self.header.pack_forget()
        if hasattr(self, 'footer'):
            self.footer.pack_forget()

        page = tk.Frame(self.content_frame, bg="#D6E4F0")
        page.pack(fill=tk.BOTH, expand=True)
        self._page_welcome = page

        # ── Background canvas ─────────────────────────────────────
        canvas = tk.Canvas(page, highlightthickness=0, bg="#D6E4F0")
        canvas.pack(fill=tk.BOTH, expand=True)

        def _draw_welcome(event=None):
            canvas.delete("all")
            cw, ch = canvas.winfo_width(), canvas.winfo_height()
            if cw < 10 or ch < 10:
                return
            cx, cy = cw // 2, ch // 2

            # Lighter Nokia-themed gradient background
            gradients = [
                ("#D6E4F0", 1.0), ("#C3D6EC", 0.92), ("#A8C4E6", 0.82),
                ("#8BB0DB", 0.70), ("#6E9BD0", 0.55), ("#5A8AC5", 0.38),
                ("#4A7AB8", 0.25),
            ]
            for color, scale in gradients:
                w2, h2 = int(cw * scale / 2), int(ch * scale / 2)
                canvas.create_rectangle(cx - w2, cy - h2, cx + w2, cy + h2,
                                        fill=color, outline="", width=0)

            # Glow ring (lighter tones)
            for i in range(8):
                r = 160 + i * 12
                alpha_hex = f"#{80 + i * 8:02x}{120 + i * 10:02x}{220 - i * 5:02x}"
                canvas.create_oval(cx - r, cy - r - 60, cx + r, cy + r - 60,
                                   outline=alpha_hex, width=2)

            # ── Nokia logo (text-based) ──────────────────────────
            canvas.create_text(cx, cy - 120, text="N O K I A",
                               font=("Segoe UI", 52, "bold"), fill=COR_AZUL_NOKIA,
                               anchor="center")
            # Underline bar
            canvas.create_rectangle(cx - 180, cy - 78, cx + 180, cy - 74,
                                    fill=COR_AZUL_NOKIA, outline="")
            canvas.create_rectangle(cx - 176, cy - 76, cx + 176, cy - 73,
                                    fill=COR_AZUL_CLARO, outline="")

            # Welcome message
            canvas.create_text(cx, cy - 30,
                               text="Welcome to CH-Berde Tools",
                               font=("Segoe UI", 24, "bold"), fill=COR_AZUL_NOKIA,
                               anchor="center")
            canvas.create_text(cx, cy + 10,
                               text="Empowering network engineers with automation",
                               font=("Segoe UI", 12), fill="#3D5A80",
                               anchor="center", justify="center")

            # ── Rounded rectangle helper ──────────────────────────
            def _rr(x1, y1, x2, y2, r=14, **kw):
                pts = [x1+r,y1, x2-r,y1, x2,y1, x2,y1+r,
                       x2,y2-r, x2,y2, x2-r,y2, x1+r,y2,
                       x1,y2, x1,y2-r, x1,y1+r, x1,y1]
                return canvas.create_polygon(pts, smooth=True, **kw)

            # ── Two main buttons (Gerador / Upgrade) ─
            btn_w, btn_h = 200, 42
            gap = 32
            total_w = 2 * btn_w + gap
            start_x = cx - total_w // 2
            by1 = cy + 88
            bx1 = start_x                           # Gerador
            bx2 = start_x + btn_w + gap             # Upgrade

            # Outer glow rings (mesma cor azul para os botões)
            for i in range(3):
                off = 4 + i * 3
                c_glow = f"#{60+i*10:02x}{100+i*15:02x}{180+i*15:02x}"
                for bx in (bx1, bx2):
                    _rr(bx-off, by1-off, bx+btn_w+off, by1+btn_h+off,
                        r=12+off, fill="", outline=c_glow, width=1)

            # Shadows
            for bx in (bx1, bx2):
                _rr(bx+3, by1+4, bx+btn_w+3, by1+btn_h+4, r=12,
                    fill="#8BA5C0", outline="")

            # Gerador / Upgrade button bodies
            _btn1_id = _rr(bx1, by1, bx1+btn_w, by1+btn_h, r=12,
                           fill="#1B5FC1", outline="#3B82F6", width=2)
            _txt1_id = canvas.create_text(bx1+btn_w//2, by1+btn_h//2,
                text="\U0001F4DD  Gerador de Script",
                font=("Segoe UI", 9, "bold"), fill=COR_BRANCO)

            _btn2_id = _rr(bx2, by1, bx2+btn_w, by1+btn_h, r=12,
                           fill="#1B5FC1", outline="#3B82F6", width=2)
            _txt2_id = canvas.create_text(bx2+btn_w//2, by1+btn_h//2,
                text="\U0001F504  Upgrade de Software",
                font=("Segoe UI", 9, "bold"), fill=COR_BRANCO)

            # Hint text
            canvas.create_text(cx, by1 + btn_h + 22,
                text="Click to explore the available tools",
                font=("Segoe UI", 9, "italic"), fill="#4A6A8A")

            # Hover / click events
            def _in_rect(x, y, rx, ry, rw, rh):
                return rx <= x <= rx + rw and ry <= y <= ry + rh

            def _on_move(e):
                hov_1 = _in_rect(e.x, e.y, bx1, by1, btn_w, btn_h)
                hov_2 = _in_rect(e.x, e.y, bx2, by1, btn_w, btn_h)
                canvas.itemconfig(_btn1_id, fill="#2563EB" if hov_1 else "#1B5FC1")
                canvas.itemconfig(_btn2_id, fill="#2563EB" if hov_2 else "#1B5FC1")
                canvas.config(cursor="hand2" if (hov_1 or hov_2) else "")

            def _on_click(e):
                if _in_rect(e.x, e.y, bx1, by1, btn_w, btn_h):
                    self._welcome_toggle_card("gerador")
                elif _in_rect(e.x, e.y, bx2, by1, btn_w, btn_h):
                    self._welcome_toggle_card("upgrade")

            canvas.bind("<Motion>", _on_move)
            canvas.bind("<Button-1>", _on_click)

        canvas.bind("<Configure>", _draw_welcome)

    # ─────────────────────────────────────────────────────────────
    #  Sidebar toggle helpers
    # ─────────────────────────────────────────────────────────────
    def _mostrar_sidebar(self):
        """Ensure sidebar, header and footer are visible."""
        if not self.header.winfo_ismapped():
            self.header.pack(fill=tk.X, side=tk.TOP, before=self.main_frame)
        if hasattr(self, 'footer') and not self.footer.winfo_ismapped():
            self.footer.pack(fill=tk.X, side=tk.BOTTOM)
        if not self.menu_frame.winfo_ismapped():
            self.menu_frame.pack(side=tk.LEFT, fill=tk.Y, before=self.content_frame)

    def _toggle_gerador_menu(self):
        """Expand / collapse Gerador de Script sub-menu."""
        if self._sub_upgrade_visible:
            self._sub_upgrade_frame.pack_forget()
            self._sub_upgrade_visible = False
        if not self._sub_gerador_visible:
            self._sub_gerador_frame.pack(fill=tk.X, after=self._sep_gerador)
            self._sub_gerador_visible = True
        else:
            self._sub_gerador_frame.pack_forget()
            self._sub_gerador_visible = False

    def _toggle_upgrade_menu(self):
        """Expand / collapse Upgrade de Software sub-menu."""
        if self._sub_gerador_visible:
            self._sub_gerador_frame.pack_forget()
            self._sub_gerador_visible = False
        if not self._sub_upgrade_visible:
            self._sub_upgrade_frame.pack(fill=tk.X, after=self._sep_upgrade)
            self._sub_upgrade_visible = True
        else:
            self._sub_upgrade_frame.pack_forget()
            self._sub_upgrade_visible = False

    def _welcome_toggle_card(self, grupo):
        """Toggle an animated submenu card on the welcome page."""
        attr = '_wcard_' + grupo
        card = getattr(self, attr, None)
        # If card visible, animate out
        if card and card.winfo_exists() and card.winfo_ismapped():
            self._welcome_animate_out(card, grupo)
            return
        # Destroy stale ref
        if card and card.winfo_exists():
            card.destroy()
        card = self._welcome_create_card(grupo)
        setattr(self, attr, card)
        self._welcome_animate_in(card, grupo)

    def _welcome_create_card(self, grupo):
        """Build a styled submenu card widget."""
        page = self._page_welcome

        # Outer glow border
        card = tk.Frame(page, bg="#2563EB", bd=0)
        inner = tk.Frame(card, bg=COR_AZUL_ESCURO)
        inner.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        # Header
        header = tk.Frame(inner, bg="#0A1829")
        header.pack(fill=tk.X)

        if grupo == "gerador":
            icon, title = "\U0001F4DD", "Gerador de Script"
            items = [
                ("Integração", "integracao"),
                ("SWAP", "swap"),
                ("Ampliação de placa", "ampliacao"),
                ("Script PCEP", "pcep"),
                ("Alterar Config (Via Rollback)", "rollback"),
            ]
            toggle_fn = self._toggle_gerador_menu
        else:
            icon, title = "\U0001F504", "Upgrade de Software"
            items = [
                ("Upgrade 7x50", "upgrade_7x50"),
                ("Upgrade vRR ou RR", "upgrade_vrr"),
            ]
            toggle_fn = self._toggle_upgrade_menu

        tk.Label(header, text=f"  {icon}  {title}",
                 font=("Segoe UI", 12, "bold"),
                 bg="#0A1829", fg=COR_BRANCO,
                 anchor="w").pack(fill=tk.X, padx=12, pady=(14, 10))

        # Accent line
        tk.Frame(inner, height=2, bg="#2563EB").pack(fill=tk.X, padx=10)

        # Body
        body = tk.Frame(inner, bg=COR_AZUL_ESCURO)
        body.pack(fill=tk.X, pady=(8, 14))

        item_btns = []
        for label, nome in items:
            def _cmd(n=nome, t=toggle_fn):
                self._welcome_close_cards()
                self._mostrar_sidebar()
                t()
                self._mostrar_pagina(n)
            btn = tk.Button(
                body, text=f"    ▸  {label}",
                font=("Segoe UI", 10),
                bg=COR_AZUL_ESCURO, fg=COR_AZUL_ESCURO,  # starts invisible
                bd=0, relief="flat",
                activebackground=COR_AZUL_CLARO,
                activeforeground=COR_BRANCO,
                anchor="w", padx=8, cursor="hand2",
                command=_cmd,
            )
            btn.pack(fill=tk.X, ipady=7, padx=8, pady=1)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg=COR_AZUL_HOVER, fg=COR_BRANCO))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg=COR_AZUL_ESCURO, fg="#CBD5E1"))
            item_btns.append(btn)
        card._item_btns = item_btns
        return card

    def _welcome_animate_in(self, card, grupo):
        """Slide card in from side with ease-out + staggered item reveal."""
        relx = 0.01 if grupo == "gerador" else 0.99
        anchor = "nw" if grupo == "gerador" else "ne"
        start_rely, end_rely = 0.80, 0.35
        steps = 14
        card.place(relx=relx, rely=start_rely, anchor=anchor, width=280)
        card.lift()

        def _ease_out(t):
            return 1 - (1 - t) ** 3

        def _step(i):
            if not card.winfo_exists():
                return
            if i > steps:
                card.place_configure(rely=end_rely)
                # Stagger-reveal items
                for idx, btn in enumerate(card._item_btns):
                    card.after(80 + idx * 70, lambda b=btn: b.config(fg="#CBD5E1"))
                return
            t = _ease_out(i / steps)
            card.place_configure(rely=start_rely + (end_rely - start_rely) * t)
            card.after(18, lambda: _step(i + 1))

        _step(0)

    def _welcome_animate_out(self, card, grupo):
        """Slide card out downward with ease-in."""
        start_rely, end_rely = 0.35, 0.80
        steps = 10

        def _ease_in(t):
            return t ** 2

        def _step(i):
            if not card.winfo_exists():
                return
            if i > steps:
                card.place_forget()
                card.destroy()
                setattr(self, '_wcard_' + grupo, None)
                return
            t = _ease_in(i / steps)
            card.place_configure(rely=start_rely + (end_rely - start_rely) * t)
            card.after(18, lambda: _step(i + 1))

        _step(0)

    def _welcome_close_cards(self):
        """Destroy all welcome-page cards immediately."""
        for g in ("gerador", "upgrade"):
            card = getattr(self, '_wcard_' + g, None)
            if card and card.winfo_exists():
                card.destroy()
            setattr(self, '_wcard_' + g, None)

    def _reset_menu_highlight(self):
        """Remove highlight from all sidebar buttons."""
        for _nome, btn in self._sub_buttons.items():
            btn.config(bg=COR_AZUL_ESCURO)
        self.btn_gerador.config(bg=COR_AZUL_NOKIA)
        self.btn_upgrade.config(bg=COR_AZUL_NOKIA)
        self._menu_btn_ativo = None

    # ─────────────────────────────────────────────────────────────
    #  Page navigation
    # ─────────────────────────────────────────────────────────────
    def _mostrar_pagina(self, nome):
        """Route to the correct page by name."""
        if self._pagina_atual == nome:
            return
        self._pagina_atual = nome

        # Ensure sidebar is visible for content pages
        self._mostrar_sidebar()

        # Highlight active submenu button
        self._reset_menu_highlight()
        if nome in self._sub_buttons:
            self._sub_buttons[nome].config(bg=COR_AZUL_CLARO)
            self._menu_btn_ativo = self._sub_buttons[nome]

        if nome == "upgrade_7x50":
            self.mostrar_upgrade()
        elif nome == "swap":
            self.mostrar_swap()
        else:
            self._mostrar_em_breve(nome)

    def _mostrar_em_breve(self, nome):
        """Show a styled 'Coming Soon' placeholder page."""
        self._limpar_content()
        if nome in self._pages_em_breve:
            self._pages_em_breve[nome].pack(fill=tk.BOTH, expand=True)
            return

        page = tk.Frame(self.content_frame, bg=COR_CINZA_FUNDO)
        page.pack(fill=tk.BOTH, expand=True)
        self._pages_em_breve[nome] = page

        # Card with shadow
        shadow = tk.Frame(page, bg="#CBD5E1")
        shadow.place(relx=0.502, rely=0.452, anchor="center", width=382, height=282)
        inner = tk.Frame(page, bg=COR_BRANCO, highlightbackground=COR_AZUL_NOKIA,
                         highlightthickness=2)
        inner.place(relx=0.5, rely=0.45, anchor="center", width=380, height=280)

        # Top accent bar
        tk.Frame(inner, height=4, bg=COR_AZUL_NOKIA).pack(fill=tk.X)

        tk.Label(inner, text="\U0001F6A7", font=("Segoe UI", 44),
                 bg=COR_BRANCO).pack(pady=(24, 4))
        tk.Label(inner, text="Em breve", font=("Segoe UI", 20, "bold"),
                 bg=COR_BRANCO, fg=COR_AZUL_NOKIA).pack()
        nomes_amigaveis = {
            "integracao": "Integração",
            "ampliacao": "Ampliação de Placa",
            "pcep": "Script PCEP",
            "rollback": "Alterar Config (Via Rollback)",
            "upgrade_vrr": "Upgrade vRR ou RR",
        }
        titulo = nomes_amigaveis.get(nome, nome)
        tk.Frame(inner, height=1, bg=COR_CINZA_BORDA).pack(fill=tk.X, padx=40, pady=6)
        tk.Label(inner, text=titulo, font=("Segoe UI", 13, "bold"),
                 bg=COR_BRANCO, fg=COR_TEXTO).pack(pady=(2, 2))
        tk.Label(inner, text="This feature is under development.\nStay tuned!",
                 font=("Segoe UI", 10), bg=COR_BRANCO, fg=COR_TEXTO_CLARO,
                 justify="center").pack(pady=(6, 20), padx=40)

    # ─────────────────────────────────────────────────────────────
    #  Actual content pages
    # ─────────────────────────────────────────────────────────────
    def mostrar_gerador(self):
        self._limpar_content()
        if not self.page_gerador:
            self.page_gerador = tk.Frame(self.content_frame, bg=COR_CINZA_FUNDO)
            self.page_gerador.pack(fill=tk.BOTH, expand=True)
            GeradorScriptApp(self.page_gerador, only_content=True)
        else:
            self.page_gerador.pack(fill=tk.BOTH, expand=True)

    # ─────────────────────────────────────────────────────────────
    #  SWAP page (DE/PARA)
    # ─────────────────────────────────────────────────────────────
    def mostrar_swap(self):
        self._limpar_content()
        if self.page_swap:
            self.page_swap.pack(fill=tk.BOTH, expand=True)
            return

        page = tk.Frame(self.content_frame, bg=COR_CINZA_FUNDO)
        page.pack(fill=tk.BOTH, expand=True)
        self.page_swap = page

        # Estado SWAP
        self._swap_csv_path = tk.StringVar()      # caminho completo (uso interno)
        self._swap_csv_display = tk.StringVar()   # apenas nome do arquivo (UI)
        self._swap_rows = []   # lista de dicts: de_host, de_portas, para_host, para_portas
        self._swap_info_vars = {}  # {hostname_lower: {"hostname":SV,"ip":SV,"chassis":SV,"timos":SV,"sat":SV}}
        self._swap_cards_de_frame = None
        self._swap_cards_para_frame = None
        self._swap_cards_vizinho_frame = None
        self._swap_vizinho_outer = None
        # StringVars dos cards de vizinhos: {hostname_lower: {"hostname","ip"}}
        self._swap_vizinho_vars = {}
        # Cache de roteadores — vive durante toda a sessão do plugin.
        # Limpar NÃO zera; armazena o que conseguir até mesmo em aborto parcial.
        self._swap_cache_hosts = {}
        # Cache de IPs de vizinhos resolvidos via DNS — preservada entre
        # importacoes (mesma logica do _swap_cache_hosts).
        # {hostname_lower: ip_str}  (ip pode ser '' quando DNS falhou)
        self._swap_cache_vizinhos = {}
        # Contador de geração: ao abortar/limpar, incrementa.
        # Workers em background descartam resultados de gerações antigas.
        self._swap_gen = 0
        self._swap_abortar = False
        self._swap_coleta_em_andamento = False
        self._swap_script_em_andamento = False
        # Mapa (host_lower, porta) → [iids] populado em _swap_importar_csv.
        # Inicializado aqui para evitar AttributeError se workers verificarem
        # antes de qualquer import.
        self._swap_iid_by_de_port = {}

        # ── Barra superior: importar planilha + botões (compacta) ─
        topo = tk.Frame(page, bg=COR_CINZA_FUNDO)
        topo.pack(fill=tk.X, padx=10, pady=(6, 4))

        tk.Label(topo, text="Planilha DE/PARA:",
                 font=("Segoe UI", 8, "bold"),
                 bg=COR_CINZA_FUNDO, fg=COR_TEXTO).pack(side=tk.LEFT)
        tk.Entry(topo, textvariable=self._swap_csv_display, font=("Segoe UI", 8),
                 width=55, state="readonly", readonlybackground="#FFFFFF",
                 fg=COR_TEXTO).pack(side=tk.LEFT, padx=(4, 4), ipady=1)
        tk.Button(topo, text="Importar...", font=("Segoe UI", 8, "bold"),
                  bg=COR_AZUL_NOKIA, fg=COR_BRANCO, bd=0, relief="flat",
                  activebackground=COR_AZUL_CLARO, activeforeground=COR_BRANCO,
                  cursor="hand2", padx=8, pady=1,
                  command=self._swap_importar_csv).pack(side=tk.LEFT, padx=(0, 4))
        tk.Button(topo, text="Script_SWAP", font=("Segoe UI", 8, "bold"),
                  bg=COR_VERDE, fg=COR_BRANCO, bd=0, relief="flat",
                  activebackground="#0E9F5B", activeforeground=COR_BRANCO,
                  cursor="hand2", padx=8, pady=1,
                  command=self._swap_gerar_script).pack(side=tk.LEFT)
        tk.Button(topo, text="Limpar", font=("Segoe UI", 8, "bold"),
                  bg="#B71C1C", fg=COR_BRANCO, bd=0, relief="flat",
                  activebackground="#E53935", activeforeground=COR_BRANCO,
                  cursor="hand2", padx=8, pady=1,
                  command=self._swap_limpar).pack(side=tk.LEFT, padx=(4, 0))

        # ═══════════════════════════════════════════════════════════
        # CONTAINER PRINCIPAL: PanedWindow horizontal
        #   ┌─ Quadro 1 ─┬─ PanedWindow vertical ─┐
        #   │            │  ┌─ Quadro 3 ─┐         │
        #   │  (DE/PARA) │  └────────────┘         │
        #   │            │  ┌─ Quadro 2 ─┐         │
        #   │            │  └────────────┘         │
        #   └────────────┴─────────────────────────┘
        # Bordas arrastáveis (sashpad/sashwidth visíveis).
        # ═══════════════════════════════════════════════════════════
        paned_h = tk.PanedWindow(
            page, orient=tk.HORIZONTAL, bg=COR_CINZA_FUNDO,
            sashwidth=6, sashrelief="raised", sashpad=0, bd=0,
            opaqueresize=True,
        )
        paned_h.pack(fill=tk.BOTH, expand=True, padx=10, pady=(2, 6))

        # ── QUADRO 1 — Informações dos roteadores (DE em cima, PARA embaixo)
        quadro1 = tk.LabelFrame(
            paned_h, text=" Informações dos roteadores ",
            font=("Segoe UI", 9, "bold"),
            bg=COR_CINZA_FUNDO, fg=COR_AZUL_NOKIA,
            bd=1, relief="groove", padx=4, pady=2,
        )
        # minsize garante que o painel n\u00e3o desapare\u00e7a ao arrastar
        paned_h.add(quadro1, minsize=200, width=440, stretch="always")

        # Container scroll\u00e1vel verticalmente: quando muitos cards
        # forem exibidos (DE + PARA + VIZINHO) e n\u00e3o couberem na
        # altura do painel, uma barra de rolagem vertical aparece
        # automaticamente.
        q1_canvas = tk.Canvas(
            quadro1, bg=COR_CINZA_FUNDO, bd=0, highlightthickness=0,
        )
        q1_vscroll = ttk.Scrollbar(
            quadro1, orient="vertical", command=q1_canvas.yview,
        )
        q1_canvas.configure(yscrollcommand=q1_vscroll.set)
        q1_vscroll.pack(side=tk.RIGHT, fill=tk.Y)
        q1_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        q1_inner = tk.Frame(q1_canvas, bg=COR_CINZA_FUNDO)
        q1_window = q1_canvas.create_window(
            (0, 0), window=q1_inner, anchor="nw",
        )

        def _q1_on_canvas_configure(event):
            # Largura do inner = largura vis\u00edvel do canvas (cards
            # continuam expandindo horizontalmente para preencher).
            try:
                q1_canvas.itemconfigure(q1_window, width=event.width)
            except Exception:
                pass

        def _q1_on_inner_configure(_evt=None):
            try:
                bbox = q1_canvas.bbox("all")
                if bbox:
                    q1_canvas.configure(scrollregion=bbox)
            except Exception:
                pass

        def _q1_on_mousewheel(event):
            try:
                q1_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        q1_canvas.bind("<Configure>", _q1_on_canvas_configure)
        q1_inner.bind("<Configure>", _q1_on_inner_configure)
        # Roda do mouse rola a lista de cards quando o cursor estiver
        # sobre o canvas.
        q1_canvas.bind(
            "<Enter>",
            lambda _e: q1_canvas.bind_all("<MouseWheel>", _q1_on_mousewheel),
        )
        q1_canvas.bind(
            "<Leave>",
            lambda _e: q1_canvas.unbind_all("<MouseWheel>"),
        )
        # Guardar refer\u00eancias para uso futuro
        self._swap_q1_canvas = q1_canvas
        self._swap_q1_inner = q1_inner
        self._swap_q1_window = q1_window

        # Lado DE (em cima)
        de_outer = tk.Frame(q1_inner, bg=COR_CINZA_FUNDO)
        de_outer.pack(fill=tk.BOTH, expand=True, pady=(0, 2))
        tk.Label(de_outer, text="DE", font=("Segoe UI", 10, "bold"),
                 bg="#0D47A1", fg=COR_BRANCO).pack(fill=tk.X, pady=(0, 2))
        self._swap_cards_de_frame = tk.Frame(de_outer, bg=COR_CINZA_FUNDO)
        self._swap_cards_de_frame.pack(fill=tk.BOTH, expand=True)

        # Lado PARA (embaixo)
        para_outer = tk.Frame(q1_inner, bg=COR_CINZA_FUNDO)
        para_outer.pack(fill=tk.BOTH, expand=True, pady=(2, 0))
        tk.Label(para_outer, text="PARA", font=("Segoe UI", 10, "bold"),
                 bg="#1B5E20", fg=COR_BRANCO).pack(fill=tk.X, pady=(0, 2))
        self._swap_cards_para_frame = tk.Frame(para_outer, bg=COR_CINZA_FUNDO)
        self._swap_cards_para_frame.pack(fill=tk.BOTH, expand=True)

        # Lado VIZINHO (mais embaixo) — só populado quando há roteador
        # diretamente conectado extraído da description das portas DE.
        self._swap_vizinho_outer = tk.Frame(q1_inner, bg=COR_CINZA_FUNDO)
        # NOTA: pack/pack_forget é controlado dinamicamente por
        # _swap_render_cards_vizinhos. Nasce escondido.
        self._swap_vizinho_label = tk.Label(
            self._swap_vizinho_outer, text="VIZINHO",
            font=("Segoe UI", 10, "bold"),
            bg="#6A1B9A", fg=COR_BRANCO,
        )
        self._swap_vizinho_label.pack(fill=tk.X, pady=(0, 2))
        self._swap_cards_vizinho_frame = tk.Frame(
            self._swap_vizinho_outer, bg=COR_CINZA_FUNDO)
        self._swap_cards_vizinho_frame.pack(fill=tk.BOTH, expand=True)

        # Placeholder inicial
        self._swap_render_cards_placeholder()

        # ── PanedWindow vertical à direita: Quadro 3 (cima) / Quadro 2 (baixo)
        paned_v = tk.PanedWindow(
            paned_h, orient=tk.VERTICAL, bg=COR_CINZA_FUNDO,
            sashwidth=6, sashrelief="raised", sashpad=0, bd=0,
            opaqueresize=True,
        )
        paned_h.add(paned_v, minsize=400, stretch="always")

        # ══════════════════════════════════════════════════════════
        # QUADRO 3 — Saída do Script_SWAP (em cima na direita)
        # ══════════════════════════════════════════════════════════
        quadro3 = tk.LabelFrame(
            paned_v, text=" Script SWAP — Coleta da configuração ",
            font=("Segoe UI", 9, "bold"),
            bg=COR_CINZA_FUNDO, fg=COR_AZUL_NOKIA,
            bd=1, relief="groove", padx=6, pady=3,
        )
        paned_v.add(quadro3, minsize=120, height=380, stretch="always")

        q3_inner = tk.Frame(quadro3, bg=COR_CINZA_FUNDO)
        q3_inner.pack(fill=tk.BOTH, expand=True)
        q3_scroll = tk.Scrollbar(q3_inner, orient="vertical")
        q3_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self._swap_script_text = tk.Text(
            q3_inner, bg="#111", fg="#00FF00",
            font=("Lucida Console", 8), insertbackground="#00FF00",
            wrap="none", bd=0, relief="flat",
            yscrollcommand=q3_scroll.set,
        )
        self._swap_script_text.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        q3_scroll.config(command=self._swap_script_text.yview)

        # ═══════════════════════════════════════════════════════════
        # QUADRO 2 — Mapeamento de portas (embaixo na direita)
        # ═══════════════════════════════════════════════════════════
        quadro2 = tk.LabelFrame(
            paned_v, text=" Mapeamento de portas DE / PARA ",
            font=("Segoe UI", 9, "bold"),
            bg=COR_CINZA_FUNDO, fg=COR_AZUL_NOKIA,
            bd=1, relief="groove", padx=6, pady=3,
        )
        paned_v.add(quadro2, minsize=120, height=260, stretch="always")
        # Guarda refer\u00eancia para atualizar o t\u00edtulo (com qtd de portas) ap\u00f3s import
        self._swap_quadro2_frame = quadro2

        # Cabeçalho de DOIS NÍVEIS:
        #   Linha 1: [JM (rowspan=2)] [DE (colspan=3)] [PARA (colspan=2)]
        #   Linha 2:                  [Hostname][Porta][Description][Hostname][Porta]
        # A linha 1 é desenhada num Canvas acima da Treeview e é mantida
        # alinhada/sincronizada com as colunas e o scroll horizontal.
        # As colunas do Treeview continuam redimensionáveis (Excel-like).
        cols = ("jm", "de_host", "de_porta", "de_desc", "para_host", "para_porta")
        col_titles = {
            "jm": "",  # JM aparece só na linha 1 (visual de mesclagem)
            "de_host": "Hostname",
            "de_porta": "Porta",
            "de_desc": "Description",
            "para_host": "Hostname",
            "para_porta": "Porta",
        }
        col_widths = {
            "jm": 60,
            "de_host": 130, "de_porta": 85, "de_desc": 280,
            "para_host": 130, "para_porta": 95,
        }
        col_anchor = {
            "jm": "center",
            "de_host": "center", "de_porta": "center", "de_desc": "w",
            "para_host": "center", "para_porta": "center",
        }
        # Mapeia cada grupo às suas colunas (na ordem)
        grupos = [
            ("JM",   ["jm"]),
            ("DE",   ["de_host", "de_porta", "de_desc"]),
            ("PARA", ["para_host", "para_porta"]),
        ]

        _style_sw = ttk.Style()
        try:
            _style_sw.configure("SwapTV.Treeview",
                                font=("Segoe UI", 9), rowheight=20)
            _style_sw.configure("SwapTV.Treeview.Heading",
                                font=("Segoe UI", 9, "bold"),
                                background=COR_AZUL_NOKIA,
                                foreground=COR_BRANCO,
                                relief="raised")
            _style_sw.map("SwapTV.Treeview.Heading",
                          background=[("active", COR_AZUL_CLARO)])
        except Exception:
            pass

        tabela_inner = tk.Frame(quadro2, bg=COR_CINZA_FUNDO)
        tabela_inner.pack(fill=tk.BOTH, expand=True)
        tabela_inner.grid_rowconfigure(1, weight=1)
        tabela_inner.grid_columnconfigure(0, weight=1)

        # ── Linha 1 do cabeçalho: Canvas com os rótulos de grupo ──
        HEADER_H = 22
        # Altura da linha de heading nativa do Treeview (depende do tema).
        # Medida em runtime via _atualizar_grupos para evitar emendas visuais.
        TREE_HEAD_H = 24
        # Cor por grupo (PARA em verde para destaque)
        cor_grupo = {
            "JM":   COR_AZUL_NOKIA,
            "DE":   COR_AZUL_NOKIA,
            "PARA": COR_VERDE,
        }
        self._swap_cor_grupo = cor_grupo
        header_canvas = tk.Canvas(
            tabela_inner, height=HEADER_H,
            bg=COR_AZUL_NOKIA, highlightthickness=0, bd=0,
        )
        header_canvas.grid(row=0, column=0, sticky="ew")
        self._swap_header_canvas = header_canvas
        self._swap_header_items = {}  # nome_grupo -> (rect_id, text_id)
        for nome, _ in grupos:
            if nome == "JM":
                # JM é desenhado como Label sobreposto cobrindo as 2 linhas
                continue
            cor = cor_grupo.get(nome, COR_AZUL_NOKIA)
            rid = header_canvas.create_rectangle(
                0, 0, 0, HEADER_H,
                fill=cor, outline=COR_BRANCO, width=1,
            )
            tid = header_canvas.create_text(
                0, HEADER_H // 2, text=nome, fill=COR_BRANCO,
                font=("Segoe UI", 9, "bold"), anchor="center",
            )
            self._swap_header_items[nome] = (rid, tid)

        scroll_y = tk.Scrollbar(tabela_inner, orient="vertical")
        scroll_y.grid(row=1, column=1, sticky="ns")
        scroll_x = tk.Scrollbar(tabela_inner, orient="horizontal")
        scroll_x.grid(row=2, column=0, sticky="ew")

        self._swap_tree = ttk.Treeview(
            tabela_inner, columns=cols, show="headings",
            style="SwapTV.Treeview",
            yscrollcommand=scroll_y.set,
        )
        # stretch=False mantém a largura escolhida pelo usuário (Excel-like)
        for c in cols:
            self._swap_tree.heading(c, text=col_titles[c], anchor="center")
            self._swap_tree.column(
                c,
                width=col_widths[c],
                minwidth=40,
                anchor=col_anchor[c],
                stretch=False,
            )
        # Tags coloridas para diferenciar DE / PARA em linhas alternadas
        try:
            self._swap_tree.tag_configure("par", background="#F5F5F5")
            self._swap_tree.tag_configure("impar", background="#FFFFFF")
        except Exception:
            pass

        self._swap_tree.grid(row=1, column=0, sticky="nsew")
        scroll_y.config(command=self._swap_tree.yview)

        # ── Label "JM" sobreposto cobrindo as DUAS linhas do cabeçalho ──
        self._swap_jm_label = tk.Label(
            tabela_inner, text="JM",
            bg=COR_AZUL_NOKIA, fg=COR_BRANCO,
            font=("Segoe UI", 9, "bold"),
            bd=0, relief="flat",
        )
        # Posicionado via place() em _atualizar_grupos
        self._swap_jm_label.place(x=0, y=0, width=col_widths["jm"],
                                  height=HEADER_H + TREE_HEAD_H)

        # ── Sincroniza Canvas (linha 1) com colunas e xview do tree ──
        self._swap_grupos = grupos

        def _atualizar_grupos(*_a):
            try:
                # Largura útil do tree e offset da rolagem horizontal
                tree_w = self._swap_tree.winfo_width()
                if tree_w <= 1:
                    return
                widths = {c: int(self._swap_tree.column(c, "width")) for c in cols}
                total = sum(widths.values())
                # x0 = quanto a tree está rolada para a esquerda (em px)
                xv = self._swap_tree.xview()
                x_off = -int(xv[0] * total) if total else 0
                # Atualiza scrollregion para o canvas refletir rolagem
                header_canvas.configure(scrollregion=(0, 0, total, HEADER_H))
                # Altura real do heading nativo (mede pela 1ª linha visível)
                head_h = TREE_HEAD_H
                try:
                    children = self._swap_tree.get_children()
                    if children:
                        bb = self._swap_tree.bbox(children[0])
                        if bb:
                            head_h = max(TREE_HEAD_H, int(bb[1]))
                except Exception:
                    pass
                # Posiciona cada grupo (exceto JM, que é Label sobreposto)
                x_acc = 0
                jm_x = 0
                jm_w = widths.get("jm", 0)
                for nome, gcols in grupos:
                    largura = sum(widths[c] for c in gcols)
                    x1 = x_acc
                    x2 = x_acc + largura
                    if nome == "JM":
                        jm_x = x1
                        jm_w = largura
                    else:
                        rid, tid = self._swap_header_items[nome]
                        header_canvas.coords(rid, x1, 0, x2, HEADER_H)
                        header_canvas.coords(tid, (x1 + x2) / 2, HEADER_H / 2)
                    x_acc = x2
                # Aplica o scroll horizontal correspondente ao canvas
                if total > 0:
                    header_canvas.xview_moveto(xv[0])
                # Reposiciona o Label JM acompanhando rolagem horizontal.
                # Altura = HEADER_H + altura real do heading + 1px para
                # encobrir a linha separadora inferior do heading nativo.
                self._swap_jm_label.place_configure(
                    x=jm_x + x_off, y=0,
                    width=jm_w, height=HEADER_H + head_h + 1,
                )
                # Garante que o Label fique acima da Treeview
                self._swap_jm_label.lift()
            except Exception:
                pass
        self._swap_atualizar_grupos = _atualizar_grupos

        def _xscroll_set(lo, hi):
            scroll_x.set(lo, hi)
            _atualizar_grupos()
        self._swap_tree.configure(xscrollcommand=_xscroll_set)

        def _xscroll_cmd(*args):
            self._swap_tree.xview(*args)
            _atualizar_grupos()
        scroll_x.config(command=_xscroll_cmd)

        # Reposiciona quando colunas são redimensionadas / janela muda
        self._swap_tree.bind("<Configure>", lambda e: _atualizar_grupos())
        self._swap_tree.bind("<B1-Motion>", lambda e: _atualizar_grupos(), add="+")
        self._swap_tree.bind("<ButtonRelease-1>", lambda e: _atualizar_grupos(), add="+")
        header_canvas.bind("<Configure>", lambda e: _atualizar_grupos())
        # Primeira pintura
        self.root.after(50, _atualizar_grupos)

        # Ctrl+roda do mouse aumenta/diminui altura das linhas
        def _ajustar_rowheight(delta):
            try:
                cur = int(_style_sw.lookup("SwapTV.Treeview", "rowheight") or 20)
            except Exception:
                cur = 20
            novo = max(14, min(60, cur + delta))
            try:
                _style_sw.configure("SwapTV.Treeview", rowheight=novo)
            except Exception:
                pass

        def _on_ctrl_wheel(ev):
            _ajustar_rowheight(2 if ev.delta > 0 else -2)
            return "break"

        self._swap_tree.bind("<Control-MouseWheel>", _on_ctrl_wheel)

    # ────────────────────────────────────────────────────────────
    #  Limpar SWAP — reseta tela e TODO o estado interno.
    #  Tamb\u00e9m aborta qualquer opera\u00e7\u00e3o em andamento.
    # ────────────────────────────────────────────────────────────
    def _swap_limpar(self):
        """Aborta a opera\u00e7\u00e3o em curso e zera completamente o plugin SWAP.

        Comportamento: ap\u00f3s clicar em Limpar, o plugin volta ao estado
        inicial (como na primeira execu\u00e7\u00e3o) \u2014 inclusive todas as
        caches em mem\u00f3ria (roteadores DE/PARA e vizinhos) s\u00e3o
        zeradas, for\u00e7ando nova resolu\u00e7\u00e3o DNS / coleta SSH na
        pr\u00f3xima importa\u00e7\u00e3o.
        """
        # Sinaliza aborto \u2014 workers em paralelo verificam essa flag
        self._swap_abortar = True
        # Invalida workers em background (gera\u00e7\u00e3o nova). Eles checam
        # a gera\u00e7\u00e3o antes de aplicar resultados ou submeter mais tarefas.
        self._swap_gen = getattr(self, "_swap_gen", 0) + 1
        # Libera flags imediatamente para n\u00e3o bloquear nova opera\u00e7\u00e3o
        self._swap_coleta_em_andamento = False
        self._swap_script_em_andamento = False
        try:
            self._swap_script_text.insert(
                tk.END,
                "\n[ABORT] Limpar acionado \u2014 abortando opera\u00e7\u00e3o...\n")
            self._swap_script_text.see(tk.END)
        except Exception:
            pass

        try:
            self._swap_csv_path.set("")
            self._swap_csv_display.set("")
        except Exception:
            pass
        self._swap_rows = []
        self._swap_info_vars = {}
        # Estado interno auxiliar
        self._swap_vizinho_vars = {}
        self._swap_portas_por_de = {}
        self._swap_portas_por_para = {}
        self._swap_jm = ""
        self._swap_vpn_avisado = False
        # Caches em mem\u00f3ria \u2014 zeradas para que o pr\u00f3ximo import
        # comporte-se como uma execu\u00e7\u00e3o nova (re-resolve DNS,
        # re-coleta via SSH, re-pergunta IP de vizinhos IPRAN).
        self._swap_cache_hosts = {}
        self._swap_cache_vizinhos = {}
        # Tabela Quadro 2
        try:
            for iid in self._swap_tree.get_children():
                self._swap_tree.delete(iid)
        except Exception:
            pass
        try:
            self._swap_iid_by_de_port = {}
        except Exception:
            pass
        # Cards Quadro 1 (DE/PARA volta ao placeholder, VIZINHO some)
        try:
            self._swap_render_cards_placeholder()
        except Exception:
            pass
        # Log Quadro 3
        try:
            self._swap_script_text.delete("1.0", tk.END)
        except Exception:
            pass
        # T\u00edtulo Quadro 2 (se foi atualizado com qtd de portas)
        try:
            if getattr(self, "_swap_quadro2_frame", None) is not None:
                self._swap_quadro2_frame.configure(
                    text=" Mapeamento de portas DE / PARA ")
        except Exception:
            pass

    # ────────────────────────────────────────────────────────────
    #  Abrir PuTTY conectado a um roteador via jumpserver
    # ────────────────────────────────────────────────────────────
    def _swap_abrir_putty(self, hostname, ip_atual="", lado="de"):
        """Abre PuTTY no roteador usando plink como proxy via jumpserver.

        Requer putty.exe e plink.exe no PATH. Se ip não estiver resolvido,
        tenta resolver via DNS antes de abrir.

        `lado` em {"de", "para", "vizinho"}. Para vizinhos, o
        jumpserver alternativo (10.119.175.4 / nokiasupp) s\u00f3 \u00e9
        usado quando o hostname casa com o padr\u00e3o de roteador
        IPRAN (ex.: ``TOEBT01-RMC01``, ``PEEPH90-RMC02``). Vizinhos
        n\u00e3o-IPRAN usam o mesmo jump dos roteadores DE/PARA.
        Senha do roteador continua a mesma (X%aA5&z3).
        """
        import shutil
        import subprocess

        # Padr\u00e3o IPRAN: <PREFIXO>-RMC<n>
        re_ipran = re.compile(r"^[A-Z0-9]+-RMC\d+$", re.IGNORECASE)
        eh_ipran = (lado == "vizinho"
                    and bool(re_ipran.match((hostname or "").strip())))

        if eh_ipran:
            jump_ip = "10.119.175.4"
            jump_user = "nokiasupp"
            jump_pass = "c@rd0s0BN2018"
        else:
            jump_ip = "10.73.0.4"
            jump_user = "supnokia"
            jump_pass = "NokiaNsp1!"
        router_user = "93191142"
        router_pass = "X%aA5&z3"

        ip = (ip_atual or "").strip()
        if not ip or ip == "—":
            try:
                ip = self._swap_resolver_dns(hostname) or ""
            except Exception:
                ip = ""
        alvo = ip or hostname

        putty = shutil.which("putty") or shutil.which("putty.exe")
        plink = shutil.which("plink") or shutil.which("plink.exe")
        if not putty:
            messagebox.showerror(
                "PuTTY não encontrado",
                "putty.exe não foi encontrado no PATH do sistema.\n"
                "Instale o PuTTY ou adicione-o ao PATH.",
            )
            return
        if not plink:
            messagebox.showerror(
                "plink não encontrado",
                "plink.exe não foi encontrado no PATH do sistema.\n"
                "É necessário para fazer o proxy via jumpserver.\n"
                "Ele acompanha a instalação do PuTTY.",
            )
            return

        proxy_cmd = (
            f'"{plink}" -ssh -batch -pw {jump_pass} '
            f'{jump_user}@{jump_ip} -nc %host:%port'
        )
        cmd = [
            putty,
            "-ssh",
            f"{router_user}@{alvo}",
            "-pw", router_pass,
            "-P", "22",
            "-proxycmd", proxy_cmd,
        ]
        try:
            subprocess.Popen(cmd, shell=False)
        except Exception as e:
            messagebox.showerror(
                "Erro ao abrir PuTTY",
                f"Falha ao iniciar PuTTY para {hostname} ({alvo}):\n{e}",
            )

    def _swap_render_cards_placeholder(self):
        """Mensagem inicial quando ainda não importou planilha."""
        for parent in (self._swap_cards_de_frame, self._swap_cards_para_frame):
            for w in parent.winfo_children():
                w.destroy()
            tk.Label(parent, text="(importe a planilha)",
                     font=("Segoe UI", 9, "italic"),
                     bg=COR_CINZA_FUNDO, fg=COR_TEXTO_CLARO).pack(pady=10)
        # Esconde a sec\u00e7\u00e3o VIZINHO ate haver vizinhos a exibir
        try:
            if self._swap_cards_vizinho_frame is not None:
                for w in self._swap_cards_vizinho_frame.winfo_children():
                    w.destroy()
            if self._swap_vizinho_outer is not None:
                self._swap_vizinho_outer.pack_forget()
            self._swap_vizinho_vars = {}
        except Exception:
            pass

    def _swap_criar_card(self, parent, hostname, lado):
        """Cria um card de roteador. Retorna dict de StringVars.

        `lado` em {"de", "para", "vizinho"}. Para "vizinho", o card é
        simplificado: apenas Hostname + IP (sem Chassis/TiMOS/Satellite).
        """
        cor_borda = {
            "de": "#0D47A1",
            "para": "#1B5E20",
            "vizinho": "#6A1B9A",
        }.get(lado, "#0D47A1")
        simples = (lado == "vizinho")
        card = tk.Frame(parent, bg=cor_borda, bd=0)
        card.pack(fill=tk.X, padx=2, pady=3)
        inner = tk.Frame(card, bg=COR_BRANCO)
        inner.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        sv_host = tk.StringVar(value=hostname)
        sv_ip = tk.StringVar(value="—")
        sv_chassis = tk.StringVar(value="—")
        sv_timos = tk.StringVar(value="—")
        sv_sat = tk.StringVar(value="—")

        if simples:
            linhas = [("Hostname:", sv_host), ("IP:", sv_ip)]
        else:
            linhas = [
                ("Hostname:", sv_host),
                ("IP:", sv_ip),
                ("Chassis:", sv_chassis),
                ("TiMOS:", sv_timos),
                ("Satelitte:", sv_sat),
            ]

        # Bot\u00e3o SSH \u2014 lado direito, centralizado verticalmente.
        # Abre PuTTY no roteador via jumpserver. No card "vizinho"
        # (apenas 2 linhas) usamos vers\u00e3o compacta, em uma s\u00f3
        # linha e sem padding vertical, para n\u00e3o esticar as
        # linhas (Hostname/IP) do card.
        if simples:
            btn_ssh = tk.Button(
                inner, text="SSH", font=("Segoe UI", 8, "bold"),
                bg=cor_borda, fg=COR_BRANCO, bd=0, relief="flat",
                activebackground="#37474F", activeforeground=COR_BRANCO,
                cursor="hand2", padx=6, pady=0, width=4,
                command=lambda h=hostname, sv=sv_ip, l=lado: self._swap_abrir_putty(h, sv.get(), l),
            )
            btn_ssh.grid(
                row=0, column=2, rowspan=len(linhas), sticky="ns",
                padx=(4, 4), pady=0,
            )
        else:
            btn_ssh = tk.Button(
                inner, text="🖥\nSSH", font=("Segoe UI", 9, "bold"),
                bg=cor_borda, fg=COR_BRANCO, bd=0, relief="flat",
                activebackground="#37474F", activeforeground=COR_BRANCO,
                cursor="hand2", padx=8, pady=4, width=4,
                command=lambda h=hostname, sv=sv_ip, l=lado: self._swap_abrir_putty(h, sv.get(), l),
            )
            btn_ssh.grid(
                row=0, column=2, rowspan=len(linhas), sticky="ns",
                padx=(4, 4), pady=2,
            )

        for i, (lbl, sv) in enumerate(linhas):
            tk.Label(inner, text=lbl, font=("Segoe UI", 7, "bold"),
                     bg=COR_BRANCO, fg=COR_TEXTO, anchor="e", width=11
                     ).grid(row=i, column=0, sticky="e", padx=(2, 2), pady=0)
            tk.Label(inner, textvariable=sv, font=("Segoe UI", 7),
                     bg=COR_BRANCO, fg=COR_AZUL_NOKIA, anchor="w"
                     ).grid(row=i, column=1, sticky="w", padx=(0, 4), pady=0)
        inner.columnconfigure(1, weight=1)

        return {
            "hostname": sv_host, "ip": sv_ip, "chassis": sv_chassis,
            "timos": sv_timos, "sat": sv_sat,
        }

    def _swap_render_cards(self, hosts_de, hosts_para):
        """Renderiza cards para os hostnames únicos. hosts_* = lista preservando ordem."""
        for w in self._swap_cards_de_frame.winfo_children():
            w.destroy()
        for w in self._swap_cards_para_frame.winfo_children():
            w.destroy()
        self._swap_info_vars = {}

        if not hosts_de:
            tk.Label(self._swap_cards_de_frame, text="(sem hostname DE)",
                     font=("Segoe UI", 9, "italic"),
                     bg=COR_CINZA_FUNDO, fg=COR_TEXTO_CLARO).pack(pady=10)
        for h in hosts_de:
            self._swap_info_vars[h.lower()] = self._swap_criar_card(
                self._swap_cards_de_frame, h, "de")

        if not hosts_para:
            tk.Label(self._swap_cards_para_frame, text="(sem hostname PARA)",
                     font=("Segoe UI", 9, "italic"),
                     bg=COR_CINZA_FUNDO, fg=COR_TEXTO_CLARO).pack(pady=10)
        for h in hosts_para:
            # Pode haver host comum a DE e PARA — cria card PARA separado.
            # O worker atualiza ambos via sufixos "" e "::para".
            chave = h.lower()
            card_para = self._swap_criar_card(
                self._swap_cards_para_frame, h, "para")
            if chave in self._swap_info_vars:
                self._swap_info_vars[chave + "::para"] = card_para
            else:
                self._swap_info_vars[chave] = card_para

    # ────────────────────────────────────────────────────────────
    #  Vizinhos diretamente conectados (extra\u00eddos das descriptions
    #  das portas DE).
    # ────────────────────────────────────────────────────────────
    def _swap_extrair_vizinho_da_desc(self, desc):
        """Extrai o hostname do roteador vizinho de uma description.

        Padr\u00e3o esperado nas descriptions Embratel:
            ``Link XYZ | 10G | ... | (core02.bsa 2/x1/1/c5/4) | (lag-134)``

        A regra: dentro de par\u00eanteses, encontrar o primeiro grupo cujo
        primeiro token tenha cara de hostname (cont\u00e9m '.' ou letras+
        d\u00edgitos) seguido por uma porta no formato ``N/...``. Ignora
        grupos do tipo ``(lag-NNN)``.

        Retorna o hostname (string) ou ''.
        """
        if not desc:
            return ""
        # Captura grupos parentizados; primeiro token = candidato a hostname
        for grupo in re.findall(r"\(([^()]+)\)", desc):
            partes = grupo.strip().split()
            if len(partes) < 2:
                continue
            host = partes[0].strip().rstrip(",;")
            porta = partes[1].strip()
            # Hostname valido: nao comeca com 'lag', tem letras
            if not host or host.lower().startswith("lag"):
                continue
            if not re.search(r"[A-Za-z]", host):
                continue
            # Porta deve ter cara de identificador SR OS (ex.: 1/1/3,
            # 2/x1/1/c5/4, lag-134 nao). Aceita digito ou 'x' no inicio.
            if not re.match(r"^[0-9]+/", porta) and not re.match(
                    r"^[A-Za-z]+\d+/", porta):
                continue
            return host
        return ""

    def _swap_render_cards_vizinhos(self, vizinhos):
        """Renderiza cards de roteadores vizinhos.

        `vizinhos`: lista de tuplas (hostname, ip) preservando ordem.
        Quando vazia, esconde toda a se\u00e7\u00e3o do Quadro 1.
        """
        if self._swap_cards_vizinho_frame is None:
            return
        for w in self._swap_cards_vizinho_frame.winfo_children():
            w.destroy()
        self._swap_vizinho_vars = {}

        if not vizinhos:
            try:
                if self._swap_vizinho_outer is not None:
                    self._swap_vizinho_outer.pack_forget()
            except Exception:
                pass
            return

        # Garante que a sec\u00e7\u00e3o esteja visivel
        try:
            if self._swap_vizinho_outer is not None:
                self._swap_vizinho_outer.pack(
                    fill=tk.BOTH, expand=True, pady=(2, 0))
        except Exception:
            pass

        for host, ip in vizinhos:
            vars_ = self._swap_criar_card(
                self._swap_cards_vizinho_frame, host, "vizinho")
            try:
                vars_["ip"].set(ip or "—")
            except Exception:
                pass
            self._swap_vizinho_vars[host.lower()] = vars_

    def _swap_processar_vizinhos(self):
        """Varre as descriptions das portas DE coletadas, extrai hostnames
        de vizinhos diretamente conectados, resolve via DNS (anexando
        ``.embratel.net.br``) e renderiza os cards no Quadro 1.

        Reaproveita ``self._swap_cache_vizinhos`` entre execu\u00e7\u00f5es:
        IP j\u00e1 conhecido n\u00e3o \u00e9 reresolvido. Resultados novos s\u00e3o
        gravados na cache para uso futuro.
        """
        try:
            cache_hosts = self._swap_cache_hosts or {}
            portas_por_de = getattr(self, "_swap_portas_por_de", {}) or {}
        except Exception:
            return
        # Conjunto de hostnames de vizinhos preservando ordem de descoberta
        seen = set()
        ordem = []
        for de_h in portas_por_de.keys():
            cache = cache_hosts.get(de_h, {}) or {}
            descs = cache.get("descs") or {}
            for porta, desc in descs.items():
                viz = self._swap_extrair_vizinho_da_desc(desc)
                if not viz:
                    continue
                k = viz.lower()
                # Evita listar como vizinho um roteador que j\u00e1 \u00e9 DE/PARA
                if k in self._swap_info_vars:
                    continue
                if k in seen:
                    continue
                seen.add(k)
                ordem.append(viz)

        if not ordem:
            # Nada a exibir \u2014 esconde a sec\u00e7\u00e3o
            self._swap_render_cards_vizinhos([])
            return

        # Resolve via DNS (apenas o que ainda nao esta em cache)
        novos = [h for h in ordem if h.lower() not in self._swap_cache_vizinhos]
        if novos:
            self._swap_log_msg(
                f"[INFO] Resolvendo via DNS {len(novos)} vizinho(s) "
                f"diretamente conectado(s)..."
            )

        # Padr\u00e3o de hostname de roteador IPRAN (ex.: TOEBT01-RMC01,
        # PEEPH90-RMC02). IPRAN n\u00e3o tem DNS \u2014 o IP \u00e9 perguntado
        # ao usu\u00e1rio via dialog.
        re_ipran = re.compile(r"^[A-Z0-9]+-RMC\d+$", re.IGNORECASE)

        def _ask_ip_ipran(hostname):
            """Abre dialog (na main thread) pedindo o IP de um vizinho
            IPRAN. Retorna IP digitado ou ''. Bloqueia o caller via
            Event ate o usuario responder.
            """
            import threading as _th
            from tkinter import simpledialog as _sd
            ev = _th.Event()
            resp = {"ip": ""}

            def _ask():
                try:
                    v = _sd.askstring(
                        "IP do vizinho IPRAN",
                        f"O vizinho {hostname} \u00e9 um roteador IPRAN "
                        f"(sem DNS).\n\n"
                        f"Digite o IP manualmente:",
                        parent=self.root,
                    )
                    if v:
                        resp["ip"] = v.strip()
                finally:
                    ev.set()

            try:
                self.root.after(0, _ask)
                ev.wait(timeout=600)
            except Exception:
                pass
            return resp["ip"]

        def _resolver_em_thread():
            for h in novos:
                if getattr(self, "_swap_abortar", False):
                    return
                if re_ipran.match(h):
                    ip = _ask_ip_ipran(h)
                    self._swap_cache_vizinhos[h.lower()] = ip or ""
                    if ip:
                        self._swap_log_msg(
                            f"[OK]   vizinho IPRAN {h} \u2192 {ip} "
                            f"(informado pelo usu\u00e1rio)"
                        )
                    else:
                        self._swap_log_msg(
                            f"[WARN] vizinho IPRAN {h}: IP n\u00e3o informado"
                        )
                    continue
                ip = self._swap_resolver_dns(h)
                self._swap_cache_vizinhos[h.lower()] = ip or ""
                if ip:
                    self._swap_log_msg(
                        f"[OK]   vizinho {h} \u2192 {ip}"
                    )
                else:
                    self._swap_log_msg(
                        f"[WARN] vizinho {h}: DNS n\u00e3o resolveu"
                    )
            # Atualiza UI no thread principal
            vizinhos_finais = [
                (h, self._swap_cache_vizinhos.get(h.lower(), ""))
                for h in ordem
            ]
            try:
                self.root.after(
                    0, self._swap_render_cards_vizinhos, vizinhos_finais)
            except Exception:
                pass

        if novos:
            threading.Thread(target=_resolver_em_thread, daemon=True).start()
        else:
            # Tudo em cache \u2014 renderiza imediatamente
            vizinhos_finais = [
                (h, self._swap_cache_vizinhos.get(h.lower(), ""))
                for h in ordem
            ]
            self._swap_render_cards_vizinhos(vizinhos_finais)

    def _swap_log_msg(self, msg):
        # Espelha a mensagem em stdout (debug) e no Quadro 3 (Script_SWAP),
        # para exibir em TEMPO REAL toda a atividade SSH.
        # Thread-safe: se chamado fora da main thread, agenda no event loop.
        try:
            print(f"[SWAP] {msg}")
        except Exception:
            pass
        def _do():
            try:
                txt = getattr(self, "_swap_script_text", None)
                if txt is not None:
                    txt.insert(tk.END, msg + "\n")
                    txt.see(tk.END)
            except Exception:
                pass
        try:
            if threading.current_thread() is threading.main_thread():
                _do()
            else:
                self.root.after(0, _do)
        except Exception:
            pass

    def _swap_importar_csv(self):
        """Importa planilha DE/PARA (Excel .xlsx) no formato:

            JM;DE;;PARA;
            JM1;Hostname;Porta;Hostname;Porta
            JM1;agg01.ula;1/1/3;agg03.ula;2/1/c4/1
            ...

        As duas primeiras linhas são de cabeçalho (grupo + colunas).
        Cada linha de dado é um mapeamento de 1 porta DE → 1 porta PARA.
        """
        from tkinter import filedialog
        caminho = filedialog.askopenfilename(
            title="Selecionar planilha DE/PARA (Excel .xlsx)",
            filetypes=[("Excel 2007+", "*.xlsx")],
            initialdir=DIRETORIO_BASE,
        )
        if not caminho:
            return
        try:
            # Lista de tuplas: (jm, de_host, de_porta, para_host, para_porta)
            linhas_quadro2 = []
            ext = os.path.splitext(caminho)[1].lower()
            if ext != ".xlsx":
                raise ValueError(
                    "Formato não suportado. Selecione um arquivo .xlsx."
                )
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise RuntimeError(
                    "Pacote 'openpyxl' não instalado. Execute:\n"
                    "  pip install openpyxl"
                )
            todas = []
            wb = load_workbook(caminho, data_only=True, read_only=True)
            try:
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    linha_vals = []
                    for c in row:
                        if c is None:
                            linha_vals.append("")
                        elif isinstance(c, float) and c.is_integer():
                            linha_vals.append(str(int(c)))
                        else:
                            linha_vals.append(str(c))
                    todas.append(linha_vals)
            finally:
                try:
                    wb.close()
                except Exception:
                    pass

            if len(todas) < 3:
                raise ValueError(
                    "Planilha deve conter pelo menos 2 linhas de cabeçalho e 1 de dados.\n"
                    "Formato esperado:\n"
                    "  JM;DE;;PARA;\n"
                    "  JM1;Hostname;Porta;Hostname;Porta\n"
                    "  JM1;agg01.ula;1/1/3;agg03.ula;2/1/c4/1"
                )

            # Valida as 2 linhas de cabeçalho (tolerante a maiúsc/minúsc e espaços)
            def _norm(s):
                return (s or "").strip().lower()

            l1 = [_norm(x) for x in todas[0]]
            l2 = [_norm(x) for x in todas[1]]
            tem_jm = ("jm" in l1)
            tem_de_para = ("de" in l1) and ("para" in l1)
            tem_host_porta = (l2.count("hostname") >= 2 and l2.count("porta") >= 2)
            if not (tem_jm and tem_de_para and tem_host_porta):
                raise ValueError(
                    "Cabeçalho inválido. Esperado:\n"
                    "  Linha 1: JM;DE;;PARA;\n"
                    "  Linha 2: JM1;Hostname;Porta;Hostname;Porta"
                )

            # Parse dados: 0=JM, 1=de_host, 2=de_porta, 3=para_host, 4=para_porta
            for row in todas[2:]:
                if not row or all((c or "").strip() == "" for c in row):
                    continue
                row = row + ["", "", "", "", ""]  # pad
                jm = (row[0] or "").strip()
                de_h = (row[1] or "").strip()
                de_p = (row[2] or "").strip()
                para_h = (row[3] or "").strip()
                para_p = (row[4] or "").strip()
                if not (de_h or de_p or para_h or para_p):
                    continue
                linhas_quadro2.append((jm, de_h, de_p, para_h, para_p))

            if not linhas_quadro2:
                messagebox.showwarning("Atenção", "Planilha sem linhas de dados.")
                return
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao ler planilha:\n{e}")
            return

        # Agrega linhas por (de_host, para_host) para manter `_swap_rows`
        # compatível com o coletor (hostnames únicos).
        from collections import OrderedDict
        agreg = OrderedDict()
        for jm, de_h, de_p, para_h, para_p in linhas_quadro2:
            chave = (de_h.lower(), para_h.lower())
            if chave not in agreg:
                agreg[chave] = {
                    "jm": jm,
                    "de_host": de_h, "de_portas": [],
                    "para_host": para_h, "para_portas": [],
                }
            if de_p:
                agreg[chave]["de_portas"].append(de_p)
            if para_p:
                agreg[chave]["para_portas"].append(para_p)
        linhas = []
        for v in agreg.values():
            linhas.append({
                "jm": v["jm"],
                "de_host": v["de_host"],
                "de_portas": " ".join(v["de_portas"]),
                "para_host": v["para_host"],
                "para_portas": " ".join(v["para_portas"]),
            })

        self._swap_csv_path.set(caminho)
        self._swap_csv_display.set(os.path.basename(caminho))
        self._swap_rows = linhas

        # JM "global" (primeiro encontrado) para nome de pasta
        self._swap_jm = linhas[0].get("jm", "") if linhas else ""

        # Atualiza t\u00edtulo do Quadro 2 com a quantidade de portas
        try:
            self._swap_quadro2_frame.config(
                text=f" Mapeamento de portas DE / PARA \u2014 {len(linhas_quadro2)} portas "
            )
        except Exception:
            pass

        # ── Popular Quadro 2 ─ uma linha por porta ───────────
        for iid in self._swap_tree.get_children():
            self._swap_tree.delete(iid)
        # Mapa (host_lower, porta) → iid, usado depois para preencher a Description
        self._swap_iid_by_de_port = {}
        for idx, (jm, de_h, de_p, para_h, para_p) in enumerate(linhas_quadro2):
            tag = "par" if (idx % 2 == 0) else "impar"
            iid = self._swap_tree.insert(
                "", tk.END,
                values=(jm, de_h, de_p, "Coletando...", para_h, para_p),
                tags=(tag,),
            )
            if de_h and de_p:
                self._swap_iid_by_de_port.setdefault(
                    (de_h.lower(), de_p), []).append(iid)

        # ── Popular Quadro 1 ─ cards por hostname único (preserva ordem) ──
        seen_de = set()
        hosts_de = []
        for _jm, de_h, _dp, _ph, _pp in linhas_quadro2:
            if de_h and de_h.lower() not in seen_de:
                seen_de.add(de_h.lower())
                hosts_de.append(de_h)
        seen_para = set()
        hosts_para = []
        for _jm, _dh, _dp, para_h, _pp in linhas_quadro2:
            if para_h and para_h.lower() not in seen_para:
                seen_para.add(para_h.lower())
                hosts_para.append(para_h)
        self._swap_render_cards(hosts_de, hosts_para)

        self._swap_log_msg(
            f"[OK] {len(linhas_quadro2)} mapeamento(s) importado(s) | "
            f"{len(hosts_de)} roteador(es) DE / {len(hosts_para)} roteador(es) PARA"
        )

        # Pergunta inicial sobre VPN e dispara a coleta. O fluxo entra
        # em loop automático: se algum roteador falhar a conexão SSH,
        # o worker de coleta agenda uma nova chamada (variante de
        # falha) ao final.
        # A pergunta inicial de VPN ja foi feita no inicio (antes do
        # file dialog). Aqui apenas dispara a coleta — o cache ainda
        # esta vazio, entao tudo sera resolvido por DNS.
        self._swap_vpn_avisado = False
        self._swap_coletar_info()

    def _swap_prompt_vpn_inicial(self):
        """Aviso inicial de VPN, exibido ao clicar em Importar.

        Apenas informativo (botao OK). Garante que o usuario tenha
        ciencia de que precisa estar conectado a VPN antes do file
        dialog ser aberto. Sempre retorna True.
        """
        try:
            messagebox.showinfo(
                "Conexão VPN",
                "Para prosseguir conecte a VPN.",
                parent=self.root,
            )
        except Exception:
            pass
        return True

    def _swap_perguntar_vpn_e_coletar(self, motivo="inicial"):
        """Loop de VPN apos falha de conexao a roteador.

        Sempre pergunta *"Você conectou a VPN ?"* (Sim/Não):
        - Sim  : reinicia a logica desde o DNS (cache preserva sucessos).
        - Não  : exibe "Então conecta a VPN." (OK) e ao confirmar
                 reinicia desde o DNS.

        O proprio worker de coleta agenda nova chamada deste metodo
        sempre que houver falha de conexao, mantendo o loop ate que
        todos os roteadores conectem com sucesso.
        """
        conectou = messagebox.askyesno(
            "Conexão VPN",
            "Você conectou a VPN ?",
            parent=self.root,
        )
        if not conectou:
            messagebox.showinfo(
                "Conectar VPN",
                "Então conecta a VPN.",
                parent=self.root,
            )
        # Reseta o flag para que o aviso de timeout de VPN possa voltar
        # a aparecer em uma nova rodada de coleta, se necessario.
        self._swap_vpn_avisado = False
        # Reinicia a logica a partir do DNS (cache preserva sucessos).
        self._swap_coletar_info()

    def _swap_coletar_info(self):
        if not self._swap_rows:
            return
        # Reseta flag de aborto e cria nova geração.
        # Workers de gerações anteriores se descartam sozinhos.
        self._swap_abortar = False
        self._swap_gen = getattr(self, "_swap_gen", 0) + 1
        my_gen = self._swap_gen
        self._swap_coleta_em_andamento = True
        # Marca todos os cards como "coletando..." para feedback visual
        for chave, vars_ in self._swap_info_vars.items():
            try:
                vars_["ip"].set("coletando...")
                vars_["chassis"].set("coletando...")
                vars_["timos"].set("coletando...")
                vars_["sat"].set("coletando...")
            except Exception:
                pass
        # Coletar em thread para não travar UI
        def _runner():
            try:
                self._swap_worker_coletar(my_gen)
            finally:
                # Só limpa a flag se ainda for a geração ativa
                if getattr(self, "_swap_gen", 0) == my_gen:
                    self._swap_coleta_em_andamento = False
        threading.Thread(target=_runner, daemon=True).start()

    def _swap_resolver_dns(self, hostname):
        """Resolve hostname via ping, sempre anexando o sufixo .embratel.net.br.

        Exemplo: agg01.ula  →  agg01.ula.embratel.net.br
        Se o hostname já terminar com .embratel.net.br, usa como está.
        Retorna o IP resolvido ou ''.
        """
        if not hostname:
            return ""
        h = hostname.strip().rstrip(".")
        sufixo = ".embratel.net.br"
        if h.lower().endswith(sufixo):
            fqdn = h
        else:
            fqdn = h + sufixo
        try:
            proc = subprocess.run(
                ["ping", fqdn, "-n", "1", "-w", "2000"],
                capture_output=True, text=True, timeout=8,
            )
            saida = proc.stdout or ""
            m = re.search(r"\[(\d+\.\d+\.\d+\.\d+)\]", saida)
            if not m:
                m = re.search(r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})", saida)
            if m:
                return m.group(1)
        except Exception:
            pass
        return ""

    def _swap_resolver_dns_ou_perguntar(self, hostname):
        """Tenta resolver via DNS; se falhar, abre uma caixa pedindo o
        IP manualmente ao usu\u00e1rio (thread-safe). Retorna IP ou ''.
        """
        ip = self._swap_resolver_dns(hostname)
        if ip:
            return ip
        # Pode estar rodando em thread worker — usa Event para
        # esperar a resposta do dialog na main thread.
        import threading as _th
        from tkinter import simpledialog as _sd
        ev = _th.Event()
        resp = {"ip": ""}

        def _ask():
            try:
                v = _sd.askstring(
                    "IP manual",
                    f"N\u00e3o foi poss\u00edvel resolver o IP de "
                    f"{hostname} pelo DNS.\n\n"
                    f"Digite o IP manualmente:",
                    parent=self.root,
                )
                if v:
                    resp["ip"] = v.strip()
            finally:
                ev.set()

        try:
            self.root.after(0, _ask)
            ev.wait(timeout=300)
        except Exception:
            pass
        return resp["ip"]

    def _swap_avisar_vpn_timeout(self, hostname, ip):
        """No-op: a pergunta de VPN apos falha de conexao e centralizada\n        em ``_swap_perguntar_vpn_e_coletar(motivo='falha')``, chamada\n        no FIM do worker (uma vez por rodada). Esta funcao foi mantida\n        apenas para nao quebrar chamadas existentes; nao exibe mais\n        nenhuma caixa para evitar duplicacao do dialogo.\n        """
        return

    def _swap_coletar_um(self, hostname, portas=None, log_cb=None):
        """Coleta info de UM roteador.

        Retorna dict {ip, chassis, timos, sat, descs, speeds, erro}.
        `portas` (opcional): lista de portas (ex.: ["1/1/3", "1/1/4"]).
                             Se fornecido, executa `show port X description`
                             para cada uma e preenche `descs` = {porta: texto}.
                             Também executa `show port X` para extrair
                             `Oper Speed` em `speeds` = {porta: valor}.
        `log_cb` (opcional): callable(str) chamado em TEMPO REAL para cada
                             comando enviado e cada bloco de bytes recebido
                             do SSH. Use `lambda m: root.after(0, log, m)`
                             ao chamar a partir de uma thread worker.
        """
        info = {"ip": "", "chassis": "", "timos": "", "sat": "",
                "descs": {}, "speeds": {},
                "location_raw": "", "loc_code": "", "erro": ""}
        if not hostname:
            info["erro"] = "hostname vazio"
            return info

        ip = self._swap_resolver_dns_ou_perguntar(hostname)
        if not ip:
            info["erro"] = "DNS não resolveu e IP não informado"
            if log_cb:
                try: log_cb(f"[{hostname}] DNS: (falhou)")
                except Exception: pass
            return info
        info["ip"] = ip
        if log_cb:
            try: log_cb(f"[{hostname}] DNS → {ip}")
            except Exception: pass

        def _emit(prefix, texto):
            if not log_cb or not texto:
                return
            for ln in texto.splitlines():
                try:
                    log_cb(f"[{hostname}] {prefix} {ln}")
                except Exception:
                    pass

        jump_ip = "10.73.0.4"
        jump_user = "supnokia"
        jump_pass = "NokiaNsp1!"
        router_user = "93191142"
        router_pass = "X%aA5&z3"

        jump_client = None
        client = None
        shell = None
        try:
            jump_client = paramiko.SSHClient()
            jump_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            jump_client.connect(
                hostname=jump_ip, port=22, username=jump_user,
                password=jump_pass, timeout=15,
                look_for_keys=False, allow_agent=False,
            )
            transport = jump_client.get_transport()
            channel = transport.open_channel(
                "direct-tcpip", (ip, 22), ("127.0.0.1", 0),
            )
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(
                hostname=ip, port=22,
                username=router_user, password=router_pass,
                timeout=15, look_for_keys=False, allow_agent=False,
                sock=channel,
            )
            shell = client.invoke_shell(width=512)
            if log_cb:
                try: log_cb(f"[{hostname}] SSH conectado em {ip}")
                except Exception: pass
            time.sleep(1)
            if shell.recv_ready():
                shell.recv(65535)

            def _run_cmd(cmd, ciclos=6, intervalo=1.0):
                """Envia cmd e coleta resposta, transmitindo tudo em
                tempo real para `log_cb` (linha a linha por chunk)."""
                if log_cb:
                    try: log_cb(f"[{hostname}] >>> {cmd}")
                    except Exception: pass
                try:
                    if shell.recv_ready():
                        shell.recv(65535)
                except Exception:
                    pass
                shell.send(cmd + "\n")
                buf = ""
                pend = ""  # parcial sem '\n' ainda
                tent = 0
                while tent < ciclos:
                    time.sleep(intervalo)
                    if shell.recv_ready():
                        try:
                            chunk = shell.recv(65535).decode(
                                "utf-8", errors="replace")
                        except Exception:
                            break
                        if chunk:
                            buf += chunk
                            limpo = _limpar_ansi(chunk)
                            pend += limpo
                            if "\n" in pend:
                                partes = pend.split("\n")
                                pend = partes[-1]
                                for ln in partes[:-1]:
                                    if log_cb:
                                        try: log_cb(
                                            f"[{hostname}] {ln}")
                                        except Exception: pass
                        tent = 0
                    else:
                        tent += 1
                if pend.strip() and log_cb:
                    try: log_cb(f"[{hostname}] {pend}")
                    except Exception: pass
                return _limpar_ansi(buf)

            _run_cmd("environment no more", ciclos=2, intervalo=0.5)

            # show system information
            saida = _run_cmd("show system information", ciclos=6, intervalo=1.0)
            sys_info = parsear_system_info(saida)
            info["chassis"] = sys_info.get("tipo", "")
            info["timos"] = sys_info.get("timos", "")

            # System Location → código curto (apenas tokens MAIÚSCULOS)
            location_raw = ""
            for linha in saida.splitlines():
                s = linha.strip()
                if s.lower().startswith("system location"):
                    partes = s.split(":", 1)
                    if len(partes) == 2:
                        location_raw = partes[1].strip().strip('"').strip("'")
                    break
            info["location_raw"] = location_raw
            tokens = re.findall(r"\b[A-Z]{2,5}\b", location_raw)
            info["loc_code"] = "".join(tokens)

            # show system satellite
            saida_sat = _run_cmd(
                "show system satellite", ciclos=6, intervalo=1.0)
            qtd = "0"
            for linha in saida_sat.splitlines():
                m_q = re.search(r"No\.\s*of\s*Satellites:\s*(\d+)", linha, re.IGNORECASE)
                if m_q:
                    qtd = m_q.group(1)
                    break
            info["sat"] = qtd

            # show port X description (uma por uma para cada porta DE)
            if portas:
                for porta in portas:
                    porta = (porta or "").strip()
                    if not porta:
                        continue
                    saida_p = _run_cmd(
                        f"show port {porta} description",
                        ciclos=4, intervalo=0.7)
                    desc = ""
                    # Formato t\u00edpico do SR OS (a description pode ocupar
                    # v\u00e1rias linhas, com wrap por COLUNA fixa):
                    #   ===============================================================
                    #   Port Descriptions on Slot N
                    #   ===============================================================
                    #   Port          Description
                    #   ---------------------------------------------------------------
                    #   1/1/4         Link ULA-BSA | 10G | ... | (core02.bsa 2/x1/
                    #                 1/c5/4) | (lag-134) | SSD 12345
                    #   ===============================================================
                    #
                    # Estrat\u00e9gia (preserva exatamente o que est\u00e1 no
                    # roteador, inclusive espa\u00e7os internos):
                    #   1. Localiza a linha que come\u00e7a com o nome da
                    #      porta e calcula `desc_col` (coluna onde a
                    #      description inicia, ap\u00f3s o nome da porta e
                    #      os espa\u00e7os de alinhamento).
                    #   2. Para cada linha seguinte (continua\u00e7\u00e3o do
                    #      wrap), pega ``linha[desc_col:].rstrip()`` SEM
                    #      strip do in\u00edcio \u2014 assim mantemos espa\u00e7os
                    #      internos. Concatena sem inserir separador
                    #      (o wrap do SR OS \u00e9 por coluna, n\u00e3o quebra
                    #      adicionando espa\u00e7os).
                    #   3. Para a continua\u00e7\u00e3o quando encontra
                    #      separador (===/---), prompt, eco de comando,
                    #      cabe\u00e7alho ou linha em branco.
                    linhas_p = saida_p.splitlines()
                    idx_inicio = -1
                    desc_col = 0
                    re_porta = re.compile(
                        r"^(" + re.escape(porta) + r")(\s+)(.+)$")
                    for idx, linha in enumerate(linhas_p):
                        # N\u00e3o usamos strip aqui pra preservar a
                        # coluna original; mas pulamos linhas
                        # claramente decorativas.
                        s_strip = linha.strip()
                        if (not s_strip
                                or s_strip.startswith("-")
                                or s_strip.startswith("=")):
                            continue
                        if (s_strip.lower().startswith("port")
                                and "description" in s_strip.lower()):
                            continue
                        m_pd = re_porta.match(linha)
                        if m_pd:
                            # Coluna inicial da description = len(porta) +
                            # len(espa\u00e7os de alinhamento). Usado para
                            # extrair as continua\u00e7\u00f5es preservando
                            # espa\u00e7os internos da description original.
                            desc_col = len(m_pd.group(1)) + len(m_pd.group(2))
                            desc = m_pd.group(3).rstrip()
                            idx_inicio = idx
                            break

                    if idx_inicio >= 0:
                        # Linhas seguintes: continua\u00e7\u00e3o at\u00e9
                        # separador/prompt/vazia.
                        for linha in linhas_p[idx_inicio + 1:]:
                            s = linha.strip()
                            if not s:
                                break
                            if s.startswith("=") or s.startswith("-"):
                                break
                            # Prompt do roteador (ex.: "A:hostname#" ou
                            # "*A:hostname#")
                            if re.match(r"^\*?[A-Za-z]:\S+[#>]", s):
                                break
                            # Eco de comando ou novo bloco "Port
                            # Descriptions"
                            if (s.lower().startswith("show port")
                                    or s.lower().startswith(
                                        "port descriptions")):
                                break
                            # Cabe\u00e7alho repetido "Port ... Description"
                            if (s.lower().startswith("port")
                                    and "description" in s.lower()):
                                break
                            # Continua\u00e7\u00e3o de wrap por coluna: pega o
                            # texto a partir de `desc_col` SEM strip do
                            # in\u00edcio (preserva espa\u00e7os internos).
                            # Se a linha for mais curta que desc_col
                            # (improv\u00e1vel mas seguro), usa o que tiver
                            # ap\u00f3s lstrip apenas dos espa\u00e7os de
                            # alinhamento.
                            if len(linha) >= desc_col:
                                cont = linha[desc_col:].rstrip()
                            else:
                                cont = linha.lstrip().rstrip()
                            # Concatena sem separador \u2014 o wrap por
                            # coluna do SR OS n\u00e3o introduz espa\u00e7os
                            # extras: o caractere quebrado em uma linha
                            # continua diretamente na pr\u00f3xima.
                            desc += cont
                    info["descs"][porta] = desc

            # show port X (sem 'description') — extrai Oper Speed
            # Os valores possíveis no SR OS são SEMPRE:
            #   1 Gbps, 10 Gbps, 100 Gbps, 400 Gbps
            # Importante: no output do `show port X`, o campo "Oper Speed"
            # não fica no início da linha — costuma vir como segunda coluna,
            # ex.: "Interface : 1/1/1                Oper Speed       : 10 Gbps"
            # Por isso buscamos o padrão em qualquer posição da saída.
            if portas:
                for porta in portas:
                    porta = (porta or "").strip()
                    if not porta:
                        continue
                    saida_s = _run_cmd(
                        f"show port {porta}", ciclos=4, intervalo=0.7)
                    speed_val = ""
                    m_sp = re.search(
                        r"Oper\s*Speed\s*:\s*(\d+\s*Gbps)",
                        saida_s, re.IGNORECASE)
                    if m_sp:
                        # Normaliza espaços internos: "10  Gbps" -> "10 Gbps"
                        speed_val = re.sub(
                            r"\s+", " ", m_sp.group(1).strip())
                    info["speeds"][porta] = speed_val
                    if log_cb:
                        try:
                            log_cb(
                                f"[{hostname}] Oper Speed porta {porta}: "
                                f"{speed_val or '(não encontrado)'}")
                        except Exception:
                            pass
        except Exception as e:
            info["erro"] = str(e)
            if log_cb:
                try: log_cb(f"[{hostname}] [ERRO] {e}")
                except Exception: pass
            # Se a falha foi timeout de SSH (DNS resolveu mas conex\u00e3o
            # nao completou), pergunta UMA VEZ se a VPN esta conectada.
            msg_low = str(e).lower()
            if (info.get("ip")
                    and ("timed out" in msg_low
                         or "timeout" in msg_low
                         or "10060" in msg_low
                         or "unreachable" in msg_low)):
                self._swap_avisar_vpn_timeout(hostname, info["ip"])
        finally:
            try:
                if shell:
                    shell.close()
            except Exception:
                pass
            try:
                if client:
                    client.close()
            except Exception:
                pass
            try:
                if jump_client:
                    jump_client.close()
            except Exception:
                pass
        return info

    def _swap_worker_coletar(self, my_gen=None):
        """Thread worker: coleta info dos hostnames únicos e atualiza cards.

        `my_gen` (opcional) — geração capturada no início. Se a geração
        global avançar (Limpar foi clicado), o worker se descarta.
        """
        def _gen_invalida():
            return (my_gen is not None
                    and getattr(self, "_swap_gen", 0) != my_gen)
        # Coletar hostnames únicos (DE + PARA), preservando origem.
        # Coletamos descriptions tanto das portas DE quanto das portas
        # PARA — usadas para alertar o usuário que a porta PARA está em uso.
        unicos = []        # lista de hostnames na ordem de 1ª ocorrência
        seen = set()
        portas_por_de = {}    # {host_lower: [portas DE]}
        portas_por_para = {}  # {host_lower: [portas PARA]}
        for r in self._swap_rows:
            de_h = r.get("de_host", "")
            para_h = r.get("para_host", "")
            if de_h:
                key = de_h.lower()
                if key not in seen:
                    seen.add(key); unicos.append(de_h)
                portas = [p for p in (r.get("de_portas") or "").split() if p]
                portas_por_de.setdefault(key, []).extend(portas)
            if para_h:
                key = para_h.lower()
                if key not in seen:
                    seen.add(key); unicos.append(para_h)
                portas_p = [p for p in (r.get("para_portas") or "").split() if p]
                portas_por_para.setdefault(key, []).extend(portas_p)
        # Guarda referência das portas PARA para usar em _aplicar_resultado
        self._swap_portas_por_para = portas_por_para
        self._swap_portas_por_de = portas_por_de

        if not unicos:
            self.root.after(0, self._swap_log_msg, "[AVISO] Nenhum hostname para coletar.")
            return

        self.root.after(0, self._swap_log_msg,
                        f"[INFO] Iniciando coleta de {len(unicos)} roteador(es) em paralelo...")

        # Garante o cache para uso pelo Script_SWAP
        if not hasattr(self, "_swap_cache_hosts"):
            self._swap_cache_hosts = {}

        # Callback de log SSH em tempo real (thread-safe via root.after)
        def _log_cb(m):
            try:
                self.root.after(0, self._swap_log_msg, m)
            except Exception:
                pass

        def _processar(host):
            """Executa a coleta de um roteador e devolve (host, res)."""
            chave = host.lower()
            # Aborto solicitado ou geração invalidada?
            if getattr(self, "_swap_abortar", False) or _gen_invalida():
                return host, {"ip": "", "chassis": "", "timos": "", "sat": "",
                              "descs": {}, "speeds": {},
                              "location_raw": "", "loc_code": "",
                              "erro": "abortado"}
            # Cache hit \u2014 reaproveita resultado de coleta anterior
            cache = self._swap_cache_hosts.get(chave)
            if cache and not cache.get("erro"):
                self.root.after(
                    0, self._swap_log_msg,
                    f"[CACHE] {host} \u2192 reutilizando dados em mem\u00f3ria",
                )
                # Mesmo com cache, recoleta descrições das portas (DE+PARA)
                portas_de = portas_por_de.get(chave, [])
                portas_pa = portas_por_para.get(chave, [])
                seen_p = set(); portas_uniq = []
                for p in (portas_de + portas_pa):
                    if p and p not in seen_p:
                        seen_p.add(p); portas_uniq.append(p)
                precisa_recoleta = (
                    portas_uniq and (not (cache.get("descs") or {})
                                     or not (cache.get("speeds") or {}))
                )
                if precisa_recoleta:
                    try:
                        res2 = self._swap_coletar_um(
                            host, portas=portas_uniq, log_cb=_log_cb)
                        if not res2.get("erro"):
                            if not (cache.get("descs") or {}):
                                cache["descs"] = res2.get("descs") or {}
                            if not (cache.get("speeds") or {}):
                                cache["speeds"] = res2.get("speeds") or {}
                            self._swap_cache_hosts[chave] = cache
                    except Exception as e:
                        self.root.after(
                            0, self._swap_log_msg,
                            f"[WARN] {host}: recoleta de cache falhou: {e}",
                        )
                return host, cache

            portas_de = portas_por_de.get(chave, [])
            portas_pa = portas_por_para.get(chave, [])
            seen_p = set(); portas_uniq = []
            for p in (portas_de + portas_pa):
                if p and p not in seen_p:
                    seen_p.add(p); portas_uniq.append(p)

            self.root.after(
                0, self._swap_log_msg,
                f"[INFO] Coletando: {host}"
                + (f" ({len(portas_uniq)} porta(s))" if portas_uniq else ""),
            )
            try:
                res = self._swap_coletar_um(
                    host, portas=portas_uniq, log_cb=_log_cb)
            except Exception as e:
                res = {"ip": "", "chassis": "", "timos": "", "sat": "",
                       "descs": {}, "speeds": {},
                       "location_raw": "", "loc_code": "",
                       "erro": f"exce\u00e7\u00e3o: {e}"}
            return host, res

        def _aplicar_resultado(host, res):
            # Descarta se a geração foi invalidada (Limpar)
            if _gen_invalida():
                return
            chave = host.lower()
            erro = res.get("erro") or ""
            # Atualiza cache: NUNCA sobrescreve um cache bom com resultado
            # com erro (mantém o que conseguimos antes).
            if erro == "abortado":
                # Aborto: não atualiza nada na cache nem na UI
                return
            existing = self._swap_cache_hosts.get(chave)
            if not (existing and not existing.get("erro") and erro):
                self._swap_cache_hosts[chave] = res
            if erro:
                self._swap_log_msg(f"[ERRO] {host}: {erro}")
            else:
                self._swap_log_msg(
                    f"[OK]   {host} \u2192 IP={res['ip']} Chassis={res['chassis']} "
                    f"TiMOS={res['timos']} Sat={res['sat']}"
                )
            erro = res.get("erro") or ""
            for sufixo in ("", "::para"):
                vars_ = self._swap_info_vars.get(chave + sufixo)
                if not vars_:
                    continue
                if erro:
                    vars_["ip"].set(res["ip"] or "(falhou)")
                    vars_["chassis"].set(f"erro: {erro[:40]}")
                    vars_["timos"].set("\u2014")
                    vars_["sat"].set("\u2014")
                else:
                    vars_["ip"].set(res["ip"] or "\u2014")
                    vars_["chassis"].set(res["chassis"] or "\u2014")
                    vars_["timos"].set(res["timos"] or "\u2014")
                    vars_["sat"].set(res["sat"] or "\u2014")
            descs = res.get("descs") or {}
            iid_map = getattr(self, "_swap_iid_by_de_port", {})
            if erro:
                for (hkey, _p), iids in iid_map.items():
                    if hkey != chave:
                        continue
                    for iid in iids:
                        try:
                            self._swap_tree.set(iid, "de_desc", "(erro)")
                        except Exception:
                            pass
            else:
                for porta, desc in descs.items():
                    for iid in iid_map.get((chave, porta), []):
                        try:
                            self._swap_tree.set(iid, "de_desc", desc or "")
                        except Exception:
                            pass

            # ── Aviso: portas PARA em uso ────────────────────────
            # Verifica description não vazia, ignorando descriptions
            # iniciadas com 1-Gig, 10-Gig, 100-Gig, 10/100/Gig, 10/100.
            if not erro:
                portas_pa = (getattr(self, "_swap_portas_por_para", {})
                             .get(chave, []))
                re_desc_ignore = re.compile(
                    r"^(?:1-Gig|10-Gig|100-Gig|10/100/Gig|10/100)\b",
                    re.IGNORECASE,
                )
                seen_w = set()
                for p in portas_pa:
                    if not p or p in seen_w:
                        continue
                    seen_w.add(p)
                    desc_p = (descs.get(p) or "").strip()
                    if desc_p and not re_desc_ignore.match(desc_p):
                        msg = (
                            f"[AVISO] Porta PARA {host} {p} está EM USO: "
                            f"{desc_p}"
                        )
                        try:
                            self._swap_script_log(msg)
                        except Exception:
                            pass
                        try:
                            self._swap_log_msg(msg)
                        except Exception:
                            pass

        # Executa todos em paralelo. Limita workers para n\u00e3o saturar o jumpserver.
        max_workers = min(len(unicos), 12) or 1
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futuros = [pool.submit(_processar, h) for h in unicos]
            for fut in as_completed(futuros):
                if getattr(self, "_swap_abortar", False) or _gen_invalida():
                    if not _gen_invalida():
                        self.root.after(
                            0, self._swap_log_msg,
                            "[ABORT] Coleta abortada pelo usu\u00e1rio.",
                        )
                    for f in futuros:
                        f.cancel()
                    break
                try:
                    host, res = fut.result()
                except Exception as e:
                    self.root.after(
                        0, self._swap_log_msg,
                        f"[ERRO] Falha inesperada na coleta paralela: {e}",
                    )
                    continue
                # Atualiza UI no thread principal
                self.root.after(0, _aplicar_resultado, host, res)

        if not (getattr(self, "_swap_abortar", False) or _gen_invalida()):
            # ── Comparação de banda (Oper Speed) DE × PARA ──────
            # Pareamento posicional: 1ª porta DE ↔ 1ª PARA, etc.
            # Comparação direta (case-insensitive, espaços colapsados).
            def _clean_speed(v):
                return re.sub(r"\s+", " ", (v or "").strip()).lower()

            cache = self._swap_cache_hosts or {}
            divergencias = []  # lista de info_txt para o diálogo
            for r in self._swap_rows:
                de_h = (r.get("de_host") or "").strip()
                para_h = (r.get("para_host") or "").strip()
                if not de_h or not para_h:
                    continue
                ports_de = [p for p in (r.get("de_portas") or "").split() if p]
                ports_pa = [p for p in (r.get("para_portas") or "").split() if p]
                de_cache = cache.get(de_h.lower(), {}) or {}
                pa_cache = cache.get(para_h.lower(), {}) or {}
                de_err = de_cache.get("erro") or ""
                pa_err = pa_cache.get("erro") or ""
                if de_err or pa_err:
                    quem = de_h if de_err else para_h
                    erro = de_err or pa_err
                    self.root.after(
                        0, self._swap_log_msg,
                        f"[AVISO] Check de banda pulado para par "
                        f"{de_h} ↔ {para_h}: {quem} com erro ({erro})",
                    )
                    continue
                spd_de_map = de_cache.get("speeds") or {}
                spd_pa_map = pa_cache.get("speeds") or {}
                for i in range(min(len(ports_de), len(ports_pa))):
                    pd = ports_de[i]; pp = ports_pa[i]
                    sd_raw = (spd_de_map.get(pd) or "").strip()
                    sp_raw = (spd_pa_map.get(pp) or "").strip()
                    if not sd_raw or not sp_raw:
                        falt = []
                        if not sd_raw:
                            falt.append(f"DE {de_h} {pd}")
                        if not sp_raw:
                            falt.append(f"PARA {para_h} {pp}")
                        self.root.after(
                            0, self._swap_log_msg,
                            "[AVISO] Oper Speed não extraído para: "
                            + ", ".join(falt),
                        )
                        continue
                    # Comparação direta: se iguais, ignora.
                    if _clean_speed(sd_raw) == _clean_speed(sp_raw):
                        continue
                    info_txt = (
                        f"DE  : {de_h}  porta {pd}  →  Oper Speed: {sd_raw}\n"
                        f"PARA: {para_h}  porta {pp}  →  Oper Speed: {sp_raw}"
                    )
                    self.root.after(
                        0, self._swap_log_msg,
                        f"[AVISO] Velocidade diferente: {de_h} {pd} "
                        f"({sd_raw}) ≠ {para_h} {pp} ({sp_raw})",
                    )
                    divergencias.append(info_txt)
            if divergencias:
                self.root.after(
                    0, self._swap_dialog_velocidade_queue, divergencias)
            self.root.after(0, self._swap_log_msg, "[OK] Coleta finalizada.")

            # ── Vizinhos diretamente conectados (extraidos das
            # descriptions das portas DE) ──
            try:
                self.root.after(0, self._swap_processar_vizinhos)
            except Exception:
                pass

            # ── Loop de VPN: se algum host falhou conexão, repete ──
            # Detecta erros relacionados a conexão (timeout, unreachable,
            # connection refused, DNS não resolveu, etc.) e, se houver
            # qualquer um, pergunta novamente sobre a VPN e refaz toda a
            # lógica desde a resolução DNS.
            try:
                hosts_falha = []
                cache_final = self._swap_cache_hosts or {}
                re_conn = re.compile(
                    r"timed out|timeout|10060|10061|unreachable|"
                    r"no route|refused|getaddrinfo|name or service|"
                    r"could not resolve|sem ip|não resolveu",
                    re.IGNORECASE,
                )
                for h in unicos:
                    info_h = cache_final.get(h.lower()) or {}
                    err_h = (info_h.get("erro") or "").strip()
                    ip_h = (info_h.get("ip") or "").strip()
                    # Falha de conexão: erro casa o regex OU não
                    # conseguimos sequer um IP (DNS falhou).
                    if (err_h and re_conn.search(err_h)) or (not ip_h):
                        hosts_falha.append(h)
                if hosts_falha:
                    # Limpa entradas de cache com falha para que a nova
                    # rodada realmente tente reconectar (cache hit só
                    # acontece em entradas SEM erro).
                    for h in hosts_falha:
                        try:
                            del self._swap_cache_hosts[h.lower()]
                        except Exception:
                            pass
                    self.root.after(
                        0, self._swap_log_msg,
                        f"[VPN] Falha de conexão em {len(hosts_falha)} "
                        f"roteador(es): {', '.join(hosts_falha)}. "
                        f"Vou perguntar sobre a VPN e tentar novamente.",
                    )
                    self.root.after(
                        100,
                        lambda: self._swap_perguntar_vpn_e_coletar(
                            motivo="falha"))
            except Exception as _e_loop:
                self.root.after(
                    0, self._swap_log_msg,
                    f"[WARN] Falha ao avaliar loop de VPN: {_e_loop}",
                )

    # ─────────────────────────────────────────────────────────────
    #  Diálogo "esquisito" — alerta de velocidade DE × PARA
    # ─────────────────────────────────────────────────────────────
    def _swap_dialog_velocidade_queue(self, lista):
        """Mostra os diálogos de divergência sequencialmente, um por vez,
        para evitar várias janelas se sobrepondo.
        `lista` é uma lista de `info_txt` (strings já formatadas)."""
        if not lista:
            return
        info_txt = lista[0]
        rest = lista[1:]
        # Ao fechar este, dispara o próximo (se houver)
        on_close = None
        if rest:
            def on_close():
                self.root.after(50, self._swap_dialog_velocidade_queue, rest)
        self._swap_dialog_velocidade(info_txt, on_close=on_close)

    def _swap_dialog_velocidade(self, info_txt, on_close=None):
        """Abre uma janela ESQUISITA avisando divergência de Oper Speed."""
        try:
            top = tk.Toplevel(self.root)
            top.title("⚠ ATENÇÃO ⚠ ⚠ ATENÇÃO ⚠")
            top.configure(bg="#ff00aa")
            top.attributes("-topmost", True)
            top.resizable(False, False)
            try:
                top.transient(self.root)
            except Exception:
                pass

            def _fechar():
                try:
                    top.grab_release()
                except Exception:
                    pass
                try:
                    top.destroy()
                except Exception:
                    pass
                if callable(on_close):
                    try:
                        on_close()
                    except Exception:
                        pass

            top.protocol("WM_DELETE_WINDOW", _fechar)

            # Borda ciano berrante
            outer = tk.Frame(top, bg="#00ffff", bd=0)
            outer.pack(padx=8, pady=8)
            inner = tk.Frame(outer, bg="#ffff00", bd=0)
            inner.pack(padx=10, pady=10)

            # Título zigue-zague (cores alternadas por letra)
            titulo = "P o R t A   c O m   v E l O c I d A d E   d I f E r E n T e !"
            cores = ["#ff0066", "#0066ff", "#00aa00", "#aa00ff",
                     "#ff6600", "#cc0000", "#006699"]
            faixa = tk.Frame(inner, bg="#ffff00")
            faixa.pack(pady=(4, 8))
            for i, ch in enumerate(titulo):
                lbl = tk.Label(
                    faixa, text=ch,
                    font=("Comic Sans MS", 18, "bold italic"),
                    fg=cores[i % len(cores)], bg="#ffff00",
                )
                lbl.pack(side="left")

            # Mensagem principal — fonte enorme em Comic Sans
            tk.Label(
                inner,
                text="Porta do novo sistema com velecidade diferente,",
                font=("Comic Sans MS", 14, "bold"),
                fg="#cc0066", bg="#ffff00",
            ).pack(pady=(0, 0))
            tk.Label(
                inner,
                text="Está correto?",
                font=("Comic Sans MS", 22, "bold italic"),
                fg="#0000aa", bg="#ffff00",
            ).pack(pady=(0, 10))

            # Detalhes
            det = tk.Label(
                inner, text=info_txt,
                font=("Courier New", 11, "bold"),
                fg="#003300", bg="#ccffcc",
                justify="left", padx=10, pady=8, bd=2, relief="ridge",
            )
            det.pack(padx=8, pady=8)

            # Botões berrantes
            barra = tk.Frame(inner, bg="#ffff00")
            barra.pack(pady=(6, 4))
            tk.Button(
                barra, text="✔ SIM, está correto",
                font=("Comic Sans MS", 12, "bold"),
                bg="#00cc00", fg="white", activebackground="#008800",
                activeforeground="white", padx=14, pady=6, bd=4,
                relief="raised",
                command=_fechar,
            ).pack(side="left", padx=8)
            tk.Button(
                barra, text="✘ NÃO!",
                font=("Comic Sans MS", 12, "bold"),
                bg="#cc0000", fg="white", activebackground="#880000",
                activeforeground="white", padx=14, pady=6, bd=4,
                relief="raised",
                command=_fechar,
            ).pack(side="left", padx=8)

            # Posiciona com leve "tremida"
            top.update_idletasks()
            w = top.winfo_width(); h = top.winfo_height()
            sw = top.winfo_screenwidth(); sh = top.winfo_screenheight()
            import random as _rnd
            x = (sw - w) // 2 + _rnd.randint(-30, 30)
            y = (sh - h) // 2 + _rnd.randint(-30, 30)
            top.geometry(f"+{max(0, x)}+{max(0, y)}")
            try:
                top.bell()
            except Exception:
                pass
            top.lift()
            top.focus_force()
            try:
                top.grab_set()
            except Exception:
                pass
        except Exception as e:
            try:
                messagebox.showwarning(
                    "Velocidade diferente",
                    "Porta do novo sistema com velecidade diferente, "
                    f"Está correto?\n\n{info_txt}\n\n({e})",
                )
            except Exception:
                pass
            if callable(on_close):
                try:
                    on_close()
                except Exception:
                    pass

    # ─────────────────────────────────────────────────────────────
    #  Script_SWAP — coleta show port / show service sap-using
    # ─────────────────────────────────────────────────────────────
    def _swap_gerar_script(self):
        """Botão Script_SWAP: dispara coleta em background."""
        if not self._swap_rows:
            messagebox.showwarning("Atenção", "Importe a planilha primeiro.")
            return
        # Reseta flag de aborto e cria nova geração
        self._swap_abortar = False
        self._swap_gen = getattr(self, "_swap_gen", 0) + 1
        my_gen = self._swap_gen
        self._swap_script_em_andamento = True
        try:
            self._swap_script_text.delete("1.0", tk.END)
            self._swap_script_text.insert(tk.END, "[INFO] Iniciando Script_SWAP...\n")
        except Exception:
            pass
        def _runner():
            try:
                self._swap_worker_script(my_gen)
            finally:
                if getattr(self, "_swap_gen", 0) == my_gen:
                    self._swap_script_em_andamento = False
        threading.Thread(target=_runner, daemon=True).start()

    def _swap_script_log(self, msg):
        # Thread-safe: agenda no event loop se chamado fora da main thread.
        def _do():
            try:
                self._swap_script_text.insert(tk.END, msg + "\n")
                self._swap_script_text.see(tk.END)
            except Exception:
                pass
        try:
            if threading.current_thread() is threading.main_thread():
                _do()
            else:
                self.root.after(0, _do)
        except Exception:
            pass

    # ═════════════════════════════════════════════════════════════
    #  Script_SWAP — coleta admin display-config + extração de
    #               toda a configuração relacionada à porta DE
    # ═════════════════════════════════════════════════════════════
    def _swap_worker_script(self, my_gen=None):
        """Fluxo do botão Script_SWAP:

          1. Cria pasta `<JM>_SWAP_<loc>` ao lado do .xlsx importado.
          2. Em PARALELO, conecta via SSH em cada roteador (DE+PARA),
             executa `environment no more` + `admin display-config` e
             salva `<host>_config_<dd-mm-yyyy>.txt` na pasta.
          3. Para cada par (DE→PARA), lê o config do roteador DE e
             extrai TODA a configuração relacionada às portas DE:
               - stanza `port <X>`
               - stanza `lag <N>` (se a porta participa de um LAG)
               - toda `interface "Y"` (router Base ou serviço) cujo
                 corpo referencia a porta ou `lag-N`
               - todo serviço (vpls/vprn/ies/epipe/ipipe/apipe/...)
                 cujo corpo referencia a porta ou `lag-N`
             Salva em `SWAP_<de_short>_<para_short>.txt`.
        """
        def _gen_invalida():
            return (my_gen is not None
                    and getattr(self, "_swap_gen", 0) != my_gen)

        def _abortado():
            return getattr(self, "_swap_abortar", False) or _gen_invalida()

        if not self._swap_rows:
            self.root.after(0, self._swap_script_log,
                            "[AVISO] Nenhum par DE/PARA.")
            return

        # ── 1) Pasta de destino ──────────────────────────────────
        xlsx_path = self._swap_csv_path.get() or ""
        base_dir = (os.path.dirname(xlsx_path)
                    if xlsx_path else DIRETORIO_BASE)
        jm = (getattr(self, "_swap_jm", "") or "JM").strip() or "JM"

        cache = getattr(self, "_swap_cache_hosts", {}) or {}
        loc = ""
        for r in self._swap_rows:
            de_h = (r.get("de_host") or "").strip().lower()
            if de_h and cache.get(de_h, {}).get("loc_code"):
                loc = cache[de_h]["loc_code"]
                break
        if not loc:
            loc = "LOC"
            self.root.after(
                0, self._swap_script_log,
                "[AVISO] loc_code não disponível no cache; usando 'LOC' "
                "no nome da pasta. Recomendo rodar 'Coletar info' antes.",
            )

        nome_pasta = f"{jm}_SWAP_{loc}"
        dir_jm = os.path.join(base_dir, nome_pasta)
        try:
            os.makedirs(dir_jm, exist_ok=True)
            self.root.after(
                0, self._swap_script_log, f"[OK] Pasta: {dir_jm}")
        except Exception as e:
            self.root.after(
                0, self._swap_script_log,
                f"[ERRO] Não foi possível criar pasta {dir_jm}: {e}",
            )
            return

        data_str = datetime.datetime.now().strftime("%d-%m-%Y")

        # ── 2) Hostnames únicos e portas DE por host ─────────────
        unique_hosts = []
        seen = set()
        de_ports_by_host = {}  # host_lower → [portas]
        for r in self._swap_rows:
            for chave_host in ("de_host", "para_host"):
                h = (r.get(chave_host) or "").strip()
                if h and h.lower() not in seen:
                    seen.add(h.lower())
                    unique_hosts.append(h)
            de_h = (r.get("de_host") or "").strip()
            if de_h:
                lst = de_ports_by_host.setdefault(de_h.lower(), [])
                for p in (r.get("de_portas") or "").split():
                    if p and p not in lst:
                        lst.append(p)

        # ── 3) Worker por host: dump admin display-config ────────
        configs_por_host = {}  # host_lower → texto do config
        ips_por_host = {}      # host_lower → ip resolvido

        def _baixar_config(host):
            chave = host.lower()
            ip = self._swap_resolver_dns_ou_perguntar(host)
            if not ip:
                self.root.after(
                    0, self._swap_script_log,
                    f"[ERRO] {host}: DNS falhou e IP não informado",
                )
                return host, "", ""

            jump_client = client = shell = None
            try:
                jump_client, client, shell = self._swap_ssh_abrir(ip)
            except Exception as e:
                self.root.after(
                    0, self._swap_script_log,
                    f"[ERRO] {host}: SSH: {e}",
                )
                return host, "", ip

            try:
                self._swap_exec_shell(
                    shell, "environment no more",
                    ciclos=2, intervalo=0.4, aborted=_abortado,
                )
                texto = self._swap_exec_shell(
                    shell, "admin display-config",
                    ciclos=30, intervalo=0.6, aborted=_abortado,
                )
                texto = self._swap_strip_eco_prompt(
                    texto, "admin display-config")

                nome_cfg = self._swap_safe_filename(
                    f"{host}_config_{data_str}.txt")
                caminho_cfg = os.path.join(dir_jm, nome_cfg)
                try:
                    with open(caminho_cfg, "w", encoding="utf-8") as f:
                        f.write(texto.rstrip() + "\n")
                    self.root.after(
                        0, self._swap_script_log,
                        f"[OK] Config salva: {caminho_cfg} "
                        f"({len(texto.splitlines())} linhas)",
                    )
                except Exception as e:
                    self.root.after(
                        0, self._swap_script_log,
                        f"[ERRO] {host}: Salvar config: {e}",
                    )
                return host, texto, ip
            except Exception as e:
                self.root.after(
                    0, self._swap_script_log,
                    f"[ERRO] {host}: {e}",
                )
                return host, "", ip
            finally:
                for obj in (shell, client, jump_client):
                    try:
                        if obj:
                            obj.close()
                    except Exception:
                        pass

        # ── 4) Executa download em paralelo ──────────────────────
        self.root.after(
            0, self._swap_script_log,
            f"\n=== Baixando admin display-config de "
            f"{len(unique_hosts)} roteador(es) em paralelo ===",
        )
        max_w = min(len(unique_hosts), 12) or 1
        with ThreadPoolExecutor(max_workers=max_w) as pool:
            futuros = {pool.submit(_baixar_config, h): h
                       for h in unique_hosts}
            for fut in as_completed(futuros):
                if _abortado():
                    self.root.after(
                        0, self._swap_script_log,
                        "[ABORT] Script_SWAP abortado pelo usuário.",
                    )
                    for f in futuros:
                        f.cancel()
                    return
                try:
                    host, cfg_text, ip_host = fut.result()
                except Exception as e:
                    self.root.after(
                        0, self._swap_script_log,
                        f"[ERRO] Falha inesperada na coleta: {e}",
                    )
                    continue
                if cfg_text:
                    configs_por_host[host.lower()] = cfg_text
                if ip_host:
                    ips_por_host[host.lower()] = ip_host

        if _abortado():
            return

        # ── 4b) Vizinhos diretamente conectados (Quadro 1) ───────
        # Tamb\u00e9m baixa admin display-config dos roteadores vizinhos
        # j\u00e1 descobertos/resolvidos. Vizinhos IPRAN usam jumpserver
        # alternativo. Falhas de "Access denied" s\u00e3o silenciosamente
        # ignoradas (alguns roteadores podem n\u00e3o aceitar nosso login).
        vizinhos_cache = getattr(self, "_swap_cache_vizinhos", {}) or {}
        # Preserva ordem dos cards renderizados
        vizinho_hosts = list(getattr(self, "_swap_vizinho_vars", {}).keys())
        # _swap_vizinho_vars usa hostname_lower; precisamos do hostname
        # original (com case) para gerar nome de arquivo. Reconstruimos
        # via _swap_cache_hosts se houver, sen\u00e3o usamos a chave mesmo.
        # _swap_extrair_vizinho_da_desc preservou case; obtemos a forma
        # original a partir das descs em cache.
        viz_orig_by_lower = {}
        try:
            for de_chave in (getattr(self, "_swap_portas_por_de", {}) or {}):
                cache_de = (self._swap_cache_hosts or {}).get(de_chave, {}) or {}
                for desc in (cache_de.get("descs") or {}).values():
                    viz = self._swap_extrair_vizinho_da_desc(desc)
                    if viz:
                        viz_orig_by_lower.setdefault(viz.lower(), viz)
        except Exception:
            pass

        re_ipran = re.compile(r"^[A-Z0-9]+-RMC\d+$", re.IGNORECASE)

        def _baixar_config_vizinho(host_lower):
            host = viz_orig_by_lower.get(host_lower, host_lower)
            ip = (vizinhos_cache.get(host_lower) or "").strip()
            if not ip:
                self.root.after(
                    0, self._swap_script_log,
                    f"[SKIP] vizinho {host}: sem IP resolvido",
                )
                return host, "", ""

            eh_ipran = bool(re_ipran.match(host))
            if eh_ipran:
                jip, juser, jpass = (
                    "10.119.175.4", "nokiasupp", "c@rd0s0BN2018")
            else:
                jip = juser = jpass = None  # usa o jump padr\u00e3o

            jump_client = client = shell = None
            try:
                jump_client, client, shell = self._swap_ssh_abrir(
                    ip, jump_ip=jip, jump_user=juser, jump_pass=jpass)
            except Exception as e:
                msg = str(e)
                # "Access denied" (alguns roteadores n\u00e3o aceitam
                # nosso login): apenas ignora silenciosamente.
                if "access denied" in msg.lower() \
                        or "authentication failed" in msg.lower():
                    self.root.after(
                        0, self._swap_script_log,
                        f"[SKIP] vizinho {host}: Access denied (ignorado)",
                    )
                    return host, "", ip
                self.root.after(
                    0, self._swap_script_log,
                    f"[ERRO] vizinho {host}: SSH: {e}",
                )
                return host, "", ip

            try:
                self._swap_exec_shell(
                    shell, "environment no more",
                    ciclos=2, intervalo=0.4, aborted=_abortado,
                )
                texto = self._swap_exec_shell(
                    shell, "admin display-config",
                    ciclos=30, intervalo=0.6, aborted=_abortado,
                )
                texto = self._swap_strip_eco_prompt(
                    texto, "admin display-config")

                nome_cfg = self._swap_safe_filename(
                    f"{host}_config_{data_str}.txt")
                caminho_cfg = os.path.join(dir_jm, nome_cfg)
                try:
                    with open(caminho_cfg, "w", encoding="utf-8") as f:
                        f.write(texto.rstrip() + "\n")
                    self.root.after(
                        0, self._swap_script_log,
                        f"[OK] Config vizinho salva: {caminho_cfg} "
                        f"({len(texto.splitlines())} linhas)",
                    )
                except Exception as e:
                    self.root.after(
                        0, self._swap_script_log,
                        f"[ERRO] vizinho {host}: Salvar config: {e}",
                    )
                return host, texto, ip
            except Exception as e:
                self.root.after(
                    0, self._swap_script_log,
                    f"[ERRO] vizinho {host}: {e}",
                )
                return host, "", ip
            finally:
                for obj in (shell, client, jump_client):
                    try:
                        if obj:
                            obj.close()
                    except Exception:
                        pass

        if vizinho_hosts:
            self.root.after(
                0, self._swap_script_log,
                f"\n=== Baixando admin display-config de "
                f"{len(vizinho_hosts)} vizinho(s) ===",
            )
            max_w_v = min(len(vizinho_hosts), 8) or 1
            with ThreadPoolExecutor(max_workers=max_w_v) as pool:
                futuros_v = {
                    pool.submit(_baixar_config_vizinho, h): h
                    for h in vizinho_hosts
                }
                for fut in as_completed(futuros_v):
                    if _abortado():
                        self.root.after(
                            0, self._swap_script_log,
                            "[ABORT] Script_SWAP abortado pelo usu\u00e1rio.",
                        )
                        for f in futuros_v:
                            f.cancel()
                        return
                    try:
                        host, cfg_text, ip_host = fut.result()
                    except Exception as e:
                        self.root.after(
                            0, self._swap_script_log,
                            f"[ERRO] Falha inesperada na coleta "
                            f"de vizinho: {e}",
                        )
                        continue
                    if cfg_text:
                        configs_por_host[host.lower()] = cfg_text
                    if ip_host:
                        ips_por_host[host.lower()] = ip_host

        if _abortado():
            return

        # ── 5) Para cada par DE/PARA, extrai blocos da porta DE ──
        self.root.after(
            0, self._swap_script_log,
            "\n=== Extraindo configuração relacionada às portas DE ===",
        )
        for r in self._swap_rows:
            if _abortado():
                return
            de_h = (r.get("de_host") or "").strip()
            para_h = (r.get("para_host") or "").strip()
            if not de_h:
                continue
            portas = [p for p in (r.get("de_portas") or "").split() if p]
            cfg_de = configs_por_host.get(de_h.lower(), "")

            de_short = de_h.split(".")[0]
            para_short = para_h.split(".")[0] if para_h else "X"
            nome_arq = self._swap_safe_filename(
                f"SWAP_{de_short}_{para_short}.txt")
            caminho_arq = os.path.join(dir_jm, nome_arq)

            if not cfg_de:
                try:
                    with open(caminho_arq, "w", encoding="utf-8") as f:
                        f.write(
                            f"# [ERRO] Config do roteador DE '{de_h}' "
                            f"não disponível (falha na coleta).\n"
                        )
                    self.root.after(
                        0, self._swap_script_log,
                        f"[AVISO] {caminho_arq} gerado vazio "
                        f"(sem config DE)",
                    )
                except Exception as e:
                    self.root.after(
                        0, self._swap_script_log,
                        f"[ERRO] Salvar {caminho_arq}: {e}",
                    )
                continue

            if _abortado():
                return
            cfg_para = configs_por_host.get(para_h.lower(), "")
            portas_para = [
                p for p in (r.get("para_portas") or "").split() if p]

            # Detecta o router-id do PARA (system-IP) para
            # correlacionar VPLS sem SAP via SDP→PARA.
            para_rid = self._swap_extract_router_id(cfg_para)

            # Callback sincrono para o dialogo CHAMATIVO de VPLS
            # sem SAP. Roda no worker thread; despacha o dialogo
            # na GUI thread via queue e bloqueia ate a resposta.
            def _vpls_alert_cb(v_id, v_name, sdp_id,
                               _de=de_h, _para=para_h):
                import queue as _q_vpls
                resp_q = _q_vpls.Queue()
                self.root.after(
                    0,
                    self._swap_dialog_vpls_sem_sap,
                    v_id, v_name, _de, _para, sdp_id, resp_q,
                )
                try:
                    return bool(resp_q.get(timeout=600))
                except Exception:
                    return False

            (texto_saida, lags_detectados, num_blocos,
             texto_shutdown_de, texto_testes_de,
             texto_testes_para, texto_fallback_de,
             texto_fallback_para) = (
                self._swap_extrair_config_portas(
                    cfg_de, portas, cfg_para=cfg_para,
                    portas_para=portas_para,
                    para_router_id=para_rid,
                    vpls_alert_cb=_vpls_alert_cb,
                    de_h=de_h, para_h=para_h)
            )
            self.root.after(
                0, self._swap_script_log,
                f"-- {de_h} :: portas={portas} :: "
                f"LAGs={lags_detectados or '(nenhum)'} :: "
                f"{num_blocos} bloco(s) extraído(s)",
            )

            # Check de policies: extrai do DE as policies referenciadas
            # no texto, verifica se existem no PARA, e anexa as que
            # FALTAM ao final do .txt gerado.
            try:
                # Resolve vizinho do PARA (agg03 <=> agg04) para
                # reescrever far-end/description do `sdp 1`.
                vizinho_h = self._swap_vizinho_para(para_h)
                vizinho_ip = ""
                if vizinho_h:
                    vizinho_ip = ips_por_host.get(
                        vizinho_h.lower(), "")
                    if not vizinho_ip:
                        vizinho_ip = self._swap_resolver_dns(
                            vizinho_h) or ""
                bloco_policies = self._swap_montar_bloco_policies(
                    cfg_de, cfg_para, texto_saida,
                    vizinho_host=vizinho_h,
                    vizinho_ip=vizinho_ip)
                if bloco_policies:
                    texto_saida = (texto_saida.rstrip() + "\n\n"
                                   + bloco_policies)
                    n_pol = bloco_policies.count("\n        ")
                    self.root.after(
                        0, self._swap_script_log,
                        f"   [POLICIES] {n_pol} policy(ies) faltavam "
                        f"no PARA '{para_h}' e foram anexadas",
                    )
                else:
                    self.root.after(
                        0, self._swap_script_log,
                        f"   [POLICIES] Nenhuma policy faltando no "
                        f"PARA '{para_h}'",
                    )
            except Exception as e:
                self.root.after(
                    0, self._swap_script_log,
                    f"   [AVISO] Falha no check de policies: {e}",
                )

            try:
                with open(caminho_arq, "w", encoding="utf-8") as f:
                    # ── Banner: configuracao a aplicar no roteador PARA ──
                    banner_para = (
                        "#" + "=" * 76 + "\n"
                        "# CONFIGURACAO PARA APLICAR NO ROTEADOR: "
                        f"{para_h or '(?)'}\n"
                        f"# Origem do SWAP: {de_h} -> {para_h or '(?)'}\n"
                        f"# Portas DE  : {' '.join(portas) or '(nenhuma)'}\n"
                        f"# Portas PARA: {' '.join(portas_para) or '(nenhuma)'}\n"
                        f"# Gerado em : {data_str}\n"
                        + "#" + "=" * 76 + "\n"
                    )
                    f.write(banner_para)
                    f.write(texto_saida.rstrip() + "\n")

                    # ── Banner + bloco de SHUTDOWN no roteador DE ──
                    if texto_shutdown_de and texto_shutdown_de.strip():
                        banner_de = (
                            "\n\n"
                            "#" + "=" * 76 + "\n"
                            "# CONFIGURACAO PARA EXECUTAR NO ROTEADOR: "
                            f"{de_h}\n"
                            "# Objetivo: shutdown das portas/LAGs/interfaces/"
                            "static-routes/BGP\n"
                            "#           envolvidos no SWAP, isolando o "
                            "trafego do roteador DE\n"
                            "#           apos o roteador PARA assumir.\n"
                            "# Modo    : classic CLI ('/configure ...' "
                            "uma linha por comando)\n"
                            + "#" + "=" * 76 + "\n"
                        )
                        f.write(banner_de)
                        f.write(texto_shutdown_de.rstrip() + "\n")

                    # ── Banner + bloco de TESTES no roteador DE ──
                    if texto_testes_de and texto_testes_de.strip():
                        banner_testes_de = (
                            "\n\n"
                            "#" + "=" * 76 + "\n"
                            "# TESTES A SEREM EXECUTADOS NO ROTEADOR: "
                            f"{de_h}\n"
                            "# Objetivo: validar conectividade (ping) e "
                            "estado dos protocolos\n"
                            "#           (ospf/mpls/rsvp/ldp/pim/bgp/arp) "
                            "ANTES do SWAP,\n"
                            "#           gerando baseline para comparar com "
                            "o roteador PARA.\n"
                            "# Escopo  : interfaces da router Base + "
                            "ifaces/SAPs dos servicos\n"
                            "#           (vpls/epipe/ies/vprn) que "
                            "envolvem as portas DE.\n"
                            + "#" + "=" * 76 + "\n"
                        )
                        f.write(banner_testes_de)
                        f.write(texto_testes_de.rstrip() + "\n")

                    # ── Banner + bloco de TESTES no roteador PARA ──
                    if texto_testes_para and texto_testes_para.strip():
                        banner_testes_para = (
                            "\n\n"
                            "#" + "=" * 76 + "\n"
                            "# TESTES A SEREM EXECUTADOS NO ROTEADOR: "
                            f"{para_h or '(?)'}\n"
                            "# Objetivo: validar conectividade (ping) e "
                            "estado dos protocolos\n"
                            "#           (ospf/mpls/rsvp/ldp/pim/bgp/arp) "
                            "APOS o SWAP,\n"
                            "#           comparando com o baseline "
                            "coletado no roteador DE.\n"
                            "# Escopo  : mesmo conjunto do roteador DE, "
                            "porem com substituicao\n"
                            "#           DE -> PARA nos nomes/SAPs que "
                            "carregam o slot/MDA.\n"
                            + "#" + "=" * 76 + "\n"
                        )
                        f.write(banner_testes_para)
                        f.write(texto_testes_para.rstrip() + "\n")

                    # ── Banner + FALLBACK no roteador DE ───────
                    if (texto_fallback_de
                            and texto_fallback_de.strip()):
                        banner_fb_de = (
                            "\n\n"
                            "#" + "=" * 76 + "\n"
                            "# FALLBACK A EXECUTAR NO ROTEADOR: "
                            f"{de_h}\n"
                            "# Objetivo: REVERTER o shutdown e "
                            "restaurar a configuracao\n"
                            "#           original (no shutdown + "
                            "description original)\n"
                            "#           de portas, LAGs, "
                            "interfaces, static-routes\n"
                            "#           e BGP neighbors.\n"
                            + "#" + "=" * 76 + "\n"
                        )
                        f.write(banner_fb_de)
                        f.write(texto_fallback_de.rstrip() + "\n")

                    # ── Banner + FALLBACK no roteador PARA ─────
                    if (texto_fallback_para
                            and texto_fallback_para.strip()):
                        banner_fb_para = (
                            "\n\n"
                            "#" + "=" * 76 + "\n"
                            "# FALLBACK A EXECUTAR NO ROTEADOR: "
                            f"{para_h or '(?)'}\n"
                            "# Objetivo: ISOLAR o trafego do "
                            "roteador PARA via shutdown\n"
                            "#           de portas, LAGs e "
                            "interfaces criadas pelo SWAP.\n"
                            + "#" + "=" * 76 + "\n"
                        )
                        f.write(banner_fb_para)
                        f.write(texto_fallback_para.rstrip() + "\n")
                self.root.after(
                    0, self._swap_script_log,
                    f"[OK] Arquivo gerado: {caminho_arq}",
                )
            except Exception as e:
                self.root.after(
                    0, self._swap_script_log,
                    f"[ERRO] Salvar {caminho_arq}: {e}",
                )

        # ── 6) Para cada porta DE, encontra contexto correspondente
        # no roteador VIZINHO via peer-IP da interface ──────────
        # A porta DE est\u00e1 ligada a uma interface (router Base ou
        # vprn/ies/...). O endere\u00e7o dessa interface tem um peer
        # (outra ponta do /30 ou /31). No cfg do vizinho, procuramos
        # a interface com esse peer-IP e geramos um bloco SIMPLIFICADO
        # (apenas description de port/lag + recriacao da interface
        # com nome substituido DE -> PARA), apendado ao arquivo
        # SWAP_<de>_<para>.txt ja gerado na etapa 5.
        if not _abortado() and configs_por_host:
            self.root.after(
                0, self._swap_script_log,
                "\n=== Anexando configura\u00e7\u00e3o do roteador VIZINHO "
                "(via peer-IP) ===",
            )
            cache_hosts_local = self._swap_cache_hosts or {}
            # Acumula por (de_h_lower, para_h_lower, viz_h_lower) ─
            # cada bucket gera 1 bloco no arquivo do par DE/PARA.
            vizinho_buckets = {}
            for r in self._swap_rows:
                if _abortado():
                    return
                de_h = (r.get("de_host") or "").strip()
                para_h = (r.get("para_host") or "").strip()
                if not de_h or not para_h:
                    continue
                cfg_de_local = configs_por_host.get(de_h.lower(), "")
                if not cfg_de_local:
                    continue
                portas_de = [
                    p for p in (r.get("de_portas") or "").split() if p]
                portas_para = [
                    p for p in (r.get("para_portas") or "").split() if p]
                # Mapa posicional de_porta -> para_porta
                mapa_pp = {}
                for idx_p, dp in enumerate(portas_de):
                    if idx_p < len(portas_para):
                        mapa_pp[dp] = portas_para[idx_p]
                descs_de = (
                    (cache_hosts_local.get(de_h.lower(), {}) or {})
                    .get("descs") or {}
                )
                for porta_de in portas_de:
                    if _abortado():
                        return
                    desc = descs_de.get(porta_de) or ""
                    viz_h = self._swap_extrair_vizinho_da_desc(desc)
                    if not viz_h:
                        continue
                    cfg_viz = configs_por_host.get(viz_h.lower(), "")
                    if not cfg_viz:
                        self.root.after(
                            0, self._swap_script_log,
                            f"[SKIP] vizinho {viz_h} (porta DE "
                            f"{de_h} {porta_de}): config n\u00e3o "
                            f"dispon\u00edvel",
                        )
                        continue
                    try:
                        vports, peers_ok = (
                            self._swap_descobrir_porta_vizinho(
                                cfg_de_local, porta_de, cfg_viz)
                        )
                    except Exception as e:
                        self.root.after(
                            0, self._swap_script_log,
                            f"[ERRO] descoberta vizinho {viz_h} "
                            f"para {de_h} {porta_de}: {e}",
                        )
                        continue
                    if not vports:
                        self.root.after(
                            0, self._swap_script_log,
                            f"[AVISO] vizinho {viz_h}: nenhuma "
                            f"porta encontrada casando peer-IP "
                            f"da {de_h} {porta_de}",
                        )
                        continue
                    chave = (de_h.lower(), para_h.lower(),
                             viz_h.lower())
                    bucket = vizinho_buckets.setdefault(chave, {
                        "de_h": de_h, "para_h": para_h,
                        "vizinho_h": viz_h,
                        "vizinho_ports": [], "peer_ips": [],
                        "mapa_porta": {},
                    })
                    for vp in vports:
                        if vp not in bucket["vizinho_ports"]:
                            bucket["vizinho_ports"].append(vp)
                    for pi in peers_ok:
                        if pi not in bucket["peer_ips"]:
                            bucket["peer_ips"].append(pi)
                    # mapa_porta acumulado
                    for k_pp, v_pp in mapa_pp.items():
                        bucket["mapa_porta"].setdefault(k_pp, v_pp)
                    self.root.after(
                        0, self._swap_script_log,
                        f"-- vizinho {viz_h}: porta DE {de_h} "
                        f"{porta_de} -> peer-IP "
                        f"{peers_ok[0] if peers_ok else '?'} "
                        f"-> porta(s) {vports}",
                    )

            # Apenda bloco por (DE, PARA, vizinho)
            for chave, info in vizinho_buckets.items():
                if _abortado():
                    return
                de_h = info["de_h"]; para_h = info["para_h"]
                viz_h = info["vizinho_h"]
                vports = info["vizinho_ports"]
                cfg_viz = configs_por_host.get(viz_h.lower(), "")
                try:
                    bloco_v = self._swap_montar_bloco_vizinho_swap(
                        cfg_viz, vports, de_h, para_h,
                        info["mapa_porta"])
                except Exception as e:
                    self.root.after(
                        0, self._swap_script_log,
                        f"[ERRO] montar bloco vizinho {viz_h}: {e}",
                    )
                    continue
                if not bloco_v.strip():
                    self.root.after(
                        0, self._swap_script_log,
                        f"[AVISO] vizinho {viz_h}: bloco vazio "
                        f"(sem description/interface a substituir)",
                    )
                    continue
                de_short = de_h.split(".")[0]
                para_short = para_h.split(".")[0] if para_h else "X"
                nome_arq = self._swap_safe_filename(
                    f"SWAP_{de_short}_{para_short}.txt")
                cam_arq = os.path.join(dir_jm, nome_arq)
                try:
                    banner_v = (
                        "\n\n"
                        "#" + "=" * 76 + "\n"
                        "# CONFIGURACAO PARA APLICAR NO ROTEADOR "
                        f"VIZINHO: {viz_h}\n"
                        f"# Origem: porta(s) DE em {de_h} -> "
                        f"PARA {para_h}\n"
                        f"# Peer-IPs casados: "
                        f"{' '.join(info['peer_ips']) or '(?)'}\n"
                        f"# Portas no vizinho: "
                        f"{' '.join(vports)}\n"
                        "# Acoes : description (port/lag) + "
                        "shutdown/no/recria interface\n"
                        f"# Gerado em : {data_str}\n"
                        + "#" + "=" * 76 + "\n"
                    )
                    with open(cam_arq, "a", encoding="utf-8") as f:
                        f.write(banner_v)
                        f.write(bloco_v.rstrip() + "\n")
                    self.root.after(
                        0, self._swap_script_log,
                        f"[OK] Anexado bloco vizinho {viz_h} "
                        f"em: {cam_arq}",
                    )
                except Exception as e:
                    self.root.after(
                        0, self._swap_script_log,
                        f"[ERRO] Anexar em {cam_arq}: {e}",
                    )
                    continue

                # ── FALLBACK do vizinho (rollback do bloco) ──
                try:
                    bloco_v_fb = (
                        self._swap_montar_bloco_vizinho_swap(
                            cfg_viz, vports, de_h, para_h,
                            info["mapa_porta"], rollback=True)
                    )
                except Exception as e:
                    self.root.after(
                        0, self._swap_script_log,
                        f"[ERRO] montar fallback vizinho "
                        f"{viz_h}: {e}",
                    )
                    bloco_v_fb = ""
                if bloco_v_fb and bloco_v_fb.strip():
                    try:
                        banner_v_fb = (
                            "\n\n"
                            "#" + "=" * 76 + "\n"
                            "# FALLBACK A EXECUTAR NO ROTEADOR "
                            f"VIZINHO: {viz_h}\n"
                            f"# Objetivo: REVERTER a configuracao "
                            "aplicada (description\n"
                            "#           original + recriar "
                            "interface ORIGINAL com\n"
                            "#           protocolos e description "
                            "originais).\n"
                            + "#" + "=" * 76 + "\n"
                        )
                        with open(cam_arq, "a",
                                  encoding="utf-8") as f:
                            f.write(banner_v_fb)
                            f.write(bloco_v_fb.rstrip() + "\n")
                        self.root.after(
                            0, self._swap_script_log,
                            f"[OK] Anexado FALLBACK vizinho "
                            f"{viz_h} em: {cam_arq}",
                        )
                    except Exception as e:
                        self.root.after(
                            0, self._swap_script_log,
                            f"[ERRO] Anexar FALLBACK em "
                            f"{cam_arq}: {e}",
                        )

        self.root.after(
            0, self._swap_script_log,
            "\n[OK] Script_SWAP finalizado.",
        )

    # ─────────────────────────────────────────────────────────────
    #  Helpers SSH e parsing
    # ─────────────────────────────────────────────────────────────
    def _swap_exec_shell(self, shell, cmd, ciclos=8, intervalo=0.6,
                         aborted=None):
        """Envia `cmd` no shell e coleta a saída até `ciclos` iterações
        consecutivas sem novos dados. Retorna texto bruto (ANSI removido)."""
        try:
            if shell.recv_ready():
                shell.recv(65535)
        except Exception:
            pass
        shell.send(cmd + "\n")
        buf = ""
        tent = 0
        while tent < ciclos:
            if callable(aborted) and aborted():
                break
            time.sleep(intervalo)
            if shell.recv_ready():
                try:
                    buf += shell.recv(65535).decode(
                        "utf-8", errors="replace")
                except Exception:
                    break
                tent = 0
            else:
                tent += 1
        return _limpar_ansi(buf)

    def _swap_ssh_abrir(self, ip, jump_ip=None, jump_user=None,
                        jump_pass=None):
        """Abre sessão SSH no roteador via jumpserver (jump → router).

        Retorna (jump_client, client, shell). Lança exceção em falha.
        Caller é responsável por fechar tudo.

        Par\u00e2metros opcionais ``jump_*`` permitem usar um jumpserver
        alternativo (ex.: rede IPRAN). Quando omitidos, usa o jump
        padr\u00e3o (10.73.0.4 / supnokia).
        """
        jump_ip = jump_ip or "10.73.0.4"
        jump_user = jump_user or "supnokia"
        jump_pass = jump_pass or "NokiaNsp1!"
        router_user = "93191142"
        router_pass = "X%aA5&z3"

        jump_client = paramiko.SSHClient()
        jump_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        jump_client.connect(
            hostname=jump_ip, port=22, username=jump_user,
            password=jump_pass, timeout=15,
            look_for_keys=False, allow_agent=False,
        )
        channel = None
        client = None
        try:
            transport = jump_client.get_transport()
            channel = transport.open_channel(
                "direct-tcpip", (ip, 22), ("127.0.0.1", 0),
            )
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(
                paramiko.AutoAddPolicy())
            client.connect(
                hostname=ip, port=22,
                username=router_user, password=router_pass,
                timeout=20, look_for_keys=False, allow_agent=False,
                sock=channel,
            )
            shell = client.invoke_shell(width=512, height=2000)
            time.sleep(1)
            if shell.recv_ready():
                shell.recv(65535)
            return jump_client, client, shell
        except Exception:
            # Garante cleanup do channel/client se algo falhar entre
            # abrir o channel e completar a conexão ao roteador.
            try:
                if client is not None:
                    client.close()
            except Exception:
                pass
            try:
                if channel is not None:
                    channel.close()
            except Exception:
                pass
            try:
                jump_client.close()
            except Exception:
                pass
            raise

    def _swap_safe_filename(self, nome):
        """Sanitiza nome de arquivo para Windows."""
        return re.sub(r'[<>:"/\\|?*]', "_", nome)

    # ─────────────────────────────────────────────────────────────
    #  Diálogo CHAMATIVO de VPLS sem SAP detectada via SDP→PARA
    # ─────────────────────────────────────────────────────────────
    def _swap_dialog_vpls_sem_sap(self, vpls_id, vpls_name,
                                   de_h, para_h, sdp_id, queue_resp):
        """Mostra um dialogo modal MUITO chamativo (vermelho/amarelo,
        piscando) avisando que foi detectada uma VPLS sem SAP no
        roteador DE que parece pertencer ao escopo da migracao,
        pois usa um SDP cujo far-end e o IP do roteador PARA.

        O usuario decide entre INCLUIR ou IGNORAR. A resposta
        (`True` para incluir, `False` para ignorar) e enviada via
        `queue_resp.put(...)`.
        """
        try:
            dlg = tk.Toplevel(self.root)
            dlg.title("⚠ ATENÇÃO — VPLS sem SAP detectada")
            dlg.transient(self.root)
            dlg.grab_set()
            dlg.configure(bg="#B00020")
            dlg.resizable(False, False)
            try:
                dlg.attributes("-topmost", True)
            except Exception:
                pass

            # Borda externa amarela
            outer = tk.Frame(dlg, bg="#FFD700", padx=4, pady=4)
            outer.pack(fill="both", expand=True)
            inner = tk.Frame(outer, bg="#B00020", padx=18, pady=14)
            inner.pack(fill="both", expand=True)

            tk.Label(
                inner,
                text="⚠  ATENÇÃO  ⚠",
                font=("Segoe UI", 22, "bold"),
                bg="#B00020", fg="#FFD700",
            ).pack(pady=(0, 6))

            tk.Label(
                inner,
                text=("Detectei uma VPLS SEM SAP no roteador DE\n"
                      "que parece fazer parte do escopo da migração."),
                font=("Segoe UI", 12, "bold"),
                bg="#B00020", fg="white",
                justify="center",
            ).pack(pady=(0, 10))

            box = tk.Frame(inner, bg="#222222", padx=12, pady=10)
            box.pack(fill="x", pady=(0, 10))
            for label, val in (
                ("VPLS service-id", str(vpls_id)),
                ("VPLS name", vpls_name or "(sem name)"),
                ("DE", de_h or "(?)"),
                ("PARA", para_h or "(?)"),
                ("SDP correlato", str(sdp_id)),
            ):
                row = tk.Frame(box, bg="#222222")
                row.pack(fill="x", pady=1)
                tk.Label(row, text=f"{label}:",
                         font=("Consolas", 10, "bold"),
                         bg="#222222", fg="#FFD700",
                         width=18, anchor="w").pack(side="left")
                tk.Label(row, text=val,
                         font=("Consolas", 10),
                         bg="#222222", fg="white",
                         anchor="w").pack(side="left")

            tk.Label(
                inner,
                text=("A spoke-sdp aponta para o SDP cujo far-end\n"
                      "é o IP do roteador PARA — logo a VPLS\n"
                      "provavelmente também precisa ser migrada."),
                font=("Segoe UI", 10),
                bg="#B00020", fg="white",
                justify="center",
            ).pack(pady=(0, 12))

            tk.Label(
                inner,
                text="Deseja INCLUIR esta VPLS no script?",
                font=("Segoe UI", 12, "bold"),
                bg="#B00020", fg="#FFD700",
            ).pack(pady=(0, 10))

            btnf = tk.Frame(inner, bg="#B00020")
            btnf.pack(pady=(0, 4))

            resp = {"v": False}

            def _ok():
                resp["v"] = True
                try:
                    dlg.destroy()
                except Exception:
                    pass

            def _no():
                resp["v"] = False
                try:
                    dlg.destroy()
                except Exception:
                    pass

            tk.Button(btnf, text="✓  INCLUIR VPLS",
                      font=("Segoe UI", 11, "bold"),
                      bg="#28a745", fg="white",
                      activebackground="#1e7e34",
                      activeforeground="white",
                      width=18, command=_ok).pack(
                side="left", padx=8)
            tk.Button(btnf, text="✗  IGNORAR",
                      font=("Segoe UI", 11, "bold"),
                      bg="#6c757d", fg="white",
                      activebackground="#495057",
                      activeforeground="white",
                      width=18, command=_no).pack(
                side="left", padx=8)

            # Efeito de piscar (alterna borda amarela/preta)
            blink_state = {"on": True, "alive": True}

            def _blink():
                if not blink_state["alive"]:
                    return
                try:
                    if not dlg.winfo_exists():
                        blink_state["alive"] = False
                        return
                    cor = "#FFD700" if blink_state["on"] else "#000000"
                    outer.configure(bg=cor)
                    blink_state["on"] = not blink_state["on"]
                    dlg.after(500, _blink)
                except Exception:
                    blink_state["alive"] = False

            dlg.after(500, _blink)

            # Centraliza
            dlg.update_idletasks()
            w = dlg.winfo_width()
            h = dlg.winfo_height()
            sw = dlg.winfo_screenwidth()
            sh = dlg.winfo_screenheight()
            x = max(0, (sw - w) // 2)
            y = max(0, (sh - h) // 3)
            dlg.geometry(f"+{x}+{y}")

            dlg.protocol("WM_DELETE_WINDOW", _no)
            dlg.wait_window()
            blink_state["alive"] = False
            try:
                queue_resp.put(bool(resp["v"]))
            except Exception:
                pass
        except Exception as e:
            try:
                queue_resp.put(False)
            except Exception:
                pass

    def _swap_vizinho_para(self, para_host):
        """Devolve o hostname vizinho do roteador PARA.

        Regra fixa do projeto: agg03 <=> agg04. Para qualquer outro
        prefixo, devolve string vazia.
        """
        h = (para_host or "").strip()
        if not h:
            return ""
        partes = h.split(".", 1)
        short = partes[0]
        rest = "." + partes[1] if len(partes) > 1 else ""
        sl = short.lower()
        if sl.startswith("agg03"):
            novo = "agg04" + short[5:]
        elif sl.startswith("agg04"):
            novo = "agg03" + short[5:]
        else:
            return ""
        return novo + rest

    def _swap_compute_ag_extras_de_lsp(self, lsp_extras, cfg_de,
                                       cfg_para):
        """Para cada `include "X"` em uma stanza de lsp-template em
        `lsp_extras`, busca a definicao `admin-group "X" value N` no
        cfg_de e retorna [(nome, valor)] que NAO existem no cfg_para.
        """
        if not lsp_extras or not cfg_de:
            return []
        rx_inc = re.compile(r'^\s*include\s+"([^"]+)"')
        ags_ref = set()
        for _nm, st in lsp_extras:
            for ln in st:
                m = rx_inc.match(ln)
                if m:
                    ags_ref.add(m.group(1))
        if not ags_ref:
            return []
        rx_def = re.compile(
            r'^\s*admin-group\s+"([^"]+)"\s+value\s+(\d+)')
        ags_no_para = set()
        if cfg_para:
            for ln in cfg_para.splitlines():
                m = rx_def.match(ln)
                if m:
                    ags_no_para.add(m.group(1))
        out = []
        vistos = set()
        for ln in cfg_de.splitlines():
            m = rx_def.match(ln)
            if not m:
                continue
            nome = m.group(1)
            if (nome in ags_ref
                    and nome not in ags_no_para
                    and nome not in vistos):
                out.append((nome, m.group(2)))
                vistos.add(nome)
        return out

    def _swap_filter_proto_for_targets(self, stanza, target_iface_names):
        """Filtra stanza de protocolo (ospf/ospf3/isis) mantendo
        apenas `interface "X"` cujo X esteja em target_iface_names.

        - Sub-blocos `area Y` sao recursivamente filtrados; areas que
          ficarem sem nenhuma `interface` sao DESCARTADAS por inteiro.
        - Linhas soltas no nivel raiz da stanza (preference, export,
          import, no shutdown, etc.) sao preservadas.
        - Retorna a stanza filtrada (lista de linhas) ou None se nao
          sobrar nenhum bloco util.
        """
        if not stanza or not target_iface_names:
            return stanza
        rx_iface = re.compile(
            r'^(\s*)(?:network-)?interface\s+"([^"]+)"')
        rx_area = re.compile(r'^(\s*)area\s+\S+')

        def _walk(stz):
            """Devolve (stz_filtrada, qtd_ifaces_mantidas).

            Mantem todas as linhas/sub-stanzas exceto:
              - `interface "X"` substanza onde X NAO esta em
                target_iface_names => removida.
              - `area Y` substanza cujo recurse retornar 0 ifaces
                mantidas => removida.
            """
            if not stz:
                return [], 0
            head = stz[0]
            base_ind = len(head) - len(head.lstrip(" "))
            child_ind = base_ind + 4
            out_lines = [head]
            keep_iface_count = 0
            j = 1
            n_loc = len(stz)
            while j < n_loc:
                ln = stz[j]
                if not ln.strip():
                    out_lines.append(ln)
                    j += 1
                    continue
                ind_ln = len(ln) - len(ln.lstrip(" "))
                if (ind_ln == base_ind
                        and ln.lstrip().startswith("exit")):
                    out_lines.append(ln)
                    j += 1
                    while j < n_loc:
                        out_lines.append(stz[j])
                        j += 1
                    break
                if ind_ln != child_ind:
                    out_lines.append(ln)
                    j += 1
                    continue
                # Detecta substanza
                j2 = j + 1
                while j2 < n_loc and not stz[j2].strip():
                    j2 += 1
                eh_stanza = (
                    j2 < n_loc
                    and (len(stz[j2]) - len(stz[j2].lstrip(" ")))
                    > ind_ln)
                if not eh_stanza:
                    out_lines.append(ln)
                    j += 1
                    continue
                # Extrai sub
                sub_end = j + 1
                while sub_end < n_loc:
                    ln2 = stz[sub_end]
                    if not ln2.strip():
                        sub_end += 1
                        continue
                    ind2 = len(ln2) - len(ln2.lstrip(" "))
                    if (ind2 == ind_ln
                            and ln2.lstrip().startswith("exit")):
                        sub_end += 1
                        break
                    if ind2 < ind_ln:
                        break
                    sub_end += 1
                sub = stz[j:sub_end]
                m_if = rx_iface.match(ln)
                m_ar = rx_area.match(ln)
                if m_if:
                    if m_if.group(2) in target_iface_names:
                        out_lines.extend(sub)
                        keep_iface_count += 1
                    # senao descarta a substanza inteira
                elif m_ar:
                    sub_filt, kf = _walk(sub)
                    if kf > 0:
                        out_lines.extend(sub_filt)
                        keep_iface_count += kf
                    # senao descarta area vazia
                else:
                    out_lines.extend(sub)
                j = sub_end
            return out_lines, keep_iface_count

        filt, kf_total = _walk(stanza)
        if kf_total == 0:
            return None
        return filt

    def _swap_vprn_tem_blackhole_no_para(self, cfg_para, vprn_id,
                                         prefix):
        """Verifica se o cfg_para ja contem, dentro do bloco
        `vprn <vprn_id>`, uma `static-route-entry <prefix>` com
        `black-hole`.
        """
        if not cfg_para or not vprn_id or not prefix:
            return False
        rx_svc = re.compile(
            rf'^(\s*)vprn\s+{re.escape(str(vprn_id))}\b')
        rx_sr = re.compile(
            rf'^\s*static-route-entry\s+{re.escape(prefix)}\b')
        para_lines = cfg_para.splitlines()
        jp = 0
        n_p = len(para_lines)
        while jp < n_p:
            if rx_svc.match(para_lines[jp]):
                stz_p, end_p = self._swap_extract_stanza(
                    para_lines, jp)
                # Procura static-route-entry <prefix> e verifica se
                # tem `black-hole` no body imediato.
                k = 0
                while k < len(stz_p):
                    if rx_sr.match(stz_p[k]):
                        ind_sr = (
                            len(stz_p[k])
                            - len(stz_p[k].lstrip(" ")))
                        kk = k + 1
                        while kk < len(stz_p):
                            ln_b = stz_p[kk]
                            if not ln_b.strip():
                                kk += 1
                                continue
                            ind_b = (len(ln_b)
                                     - len(ln_b.lstrip(" ")))
                            if ind_b <= ind_sr:
                                break
                            if re.match(r'^\s*black-hole\b', ln_b):
                                return True
                            kk += 1
                    k += 1
                jp = end_p
                continue
            jp += 1
        return False

    def _swap_strip_eco_prompt(self, texto, cmd=""):
        """Remove eco do comando, prompt do roteador e linhas vazias finais."""
        out = []
        for linha in texto.splitlines():
            s = linha.rstrip()
            if not s.strip():
                out.append(s)
                continue
            # Prompt: "*A:hostname# " ou "A:hostname> "
            if re.match(r"^\*?[A-Za-z]:\S+\s*[#>]\s*", s):
                continue
            # Eco do comando enviado
            if cmd and s.strip().endswith(cmd.strip()):
                continue
            out.append(s)
        return "\n".join(out).rstrip()

    def _swap_extract_stanza(self, lines, idx):
        """Dado um índice de header, retorna (lista_de_linhas, end_exclusive).

        A stanza vai do header até o `exit` na mesma indentação (inclusive),
        ou até a próxima linha de indentação <= base (exclusive).
        """
        n = len(lines)
        if idx >= n:
            return [], idx
        header = lines[idx]
        base = len(header) - len(header.lstrip(" "))
        body = [header]
        i = idx + 1
        while i < n:
            s = lines[i]
            if not s.strip():
                body.append(s); i += 1; continue
            ind = len(s) - len(s.lstrip(" "))
            if ind > base:
                body.append(s); i += 1
            elif ind == base and s.lstrip().startswith("exit"):
                body.append(s); i += 1
                return body, i
            else:
                return body, i
        return body, i

    # ─────────────────────────────────────────────────────────────
    #  Extrai o "system-IP" do roteador (router-id ou system iface)
    # ─────────────────────────────────────────────────────────────
    def _swap_extract_router_id(self, cfg_text):
        """Extrai o IP do roteador (system-IP) usado por SDPs como
        far-end. Tenta primeiro `router-id` dentro de `router Base`,
        em seguida `interface "system" address X.X.X.X/32`.
        Retorna None se nao encontrar.
        """
        if not cfg_text:
            return None
        lines = cfg_text.split("\n")
        rx_rb = re.compile(r'^\s{4}router(?:\s+Base)?\s*$')
        rx_rid = re.compile(r'^\s*router-id\s+(\S+)')
        rx_sys_if = re.compile(
            r'^\s*interface\s+"system"\s*$', re.IGNORECASE)
        rx_addr = re.compile(r'^\s*address\s+(\S+?)(?:/\d+)?\s*$')
        n = len(lines)
        for i, ln in enumerate(lines):
            if not rx_rb.match(ln):
                continue
            stz, _ = self._swap_extract_stanza(lines, i)
            # 1) router-id
            for sub in stz[1:]:
                m = rx_rid.match(sub)
                if m:
                    return m.group(1)
            # 2) interface "system" address X/32
            j = 0
            while j < len(stz):
                if rx_sys_if.match(stz[j]):
                    sub_st, _ = self._swap_extract_stanza(stz, j)
                    for sub in sub_st[1:]:
                        m_a = rx_addr.match(sub)
                        if m_a:
                            return m_a.group(1)
                    break
                j += 1
        return None

    # ─────────────────────────────────────────────────────────────
    #  Descoberta da porta do vizinho via peer-IP da interface
    # ─────────────────────────────────────────────────────────────
    def _swap_descobrir_porta_vizinho(self, cfg_de, porta_de,
                                       cfg_vizinho):
        """Descobre a(s) porta(s) no roteador VIZINHO correspondentes
        a `porta_de`, atrav\u00e9s do peer-IP da interface ligada
        \u00e0 `porta_de` no roteador DE.

        Algoritmo:
          1. No `cfg_de`, identifica LAG (se houver) cujos membros
             incluem `porta_de`.
          2. Localiza interfaces (router Base ou servi\u00e7os) cujo
             body cont\u00e9m `port <porta_de>` ou `port lag-N`.
             Coleta todos os endere\u00e7os IPv4/IPv6.
          3. Para cada `address X/N`, calcula candidatos a peer-IP
             (vizinhan\u00e7a da subrede, limitada a 256 endere\u00e7os).
          4. No `cfg_vizinho`, varre todas as interfaces e seleciona
             as que t\u00eam `address Y/N` com Y em algum dos peer-IPs.
          5. Para cada interface casada, extrai `port X` do body. Se
             for `port lag-N`, expande para os membros do LAG.

        Retorna (lista_portas_vizinho, lista_peer_ips_casados).
        """
        if not cfg_de or not porta_de or not cfg_vizinho:
            return [], []

        lines_de = cfg_de.splitlines()
        rx_lag_hdr = re.compile(r"^(\s*)lag\s+(\d+)\s*$")
        rx_port_in_lag = re.compile(r"^\s*port\s+(\S+)")
        rx_iface_hdr = re.compile(
            r'^(\s*)(?:network-)?interface\s+"([^"]+)"')

        # --- 1) LAG do porta_de em cfg_de (se membro) -----------
        lag_de = ""
        for i, s in enumerate(lines_de):
            m = rx_lag_hdr.match(s)
            if not m:
                continue
            stz, _ = self._swap_extract_stanza(lines_de, i)
            for ln in stz[1:]:
                mp = rx_port_in_lag.match(ln)
                if mp and mp.group(1) == porta_de:
                    lag_de = m.group(2)
                    break
            if lag_de:
                break

        targets_port_lines = {f"port {porta_de}"}
        if lag_de:
            targets_port_lines.add(f"port lag-{lag_de}")

        # --- 2) Endere\u00e7os das interfaces de cfg_de ligadas -
        enderecos = []  # lista de (ip_str, prefixlen)
        for i, s in enumerate(lines_de):
            if not rx_iface_hdr.match(s):
                continue
            stz, _ = self._swap_extract_stanza(lines_de, i)
            body_strip = [ln.strip() for ln in stz[1:]]
            bound = False
            for b in body_strip:
                for t in targets_port_lines:
                    if b == t or (b.startswith(t)
                                  and len(b) > len(t)
                                  and b[len(t)].isspace()):
                        bound = True
                        break
                if bound:
                    break
            if not bound:
                continue
            for ln in stz[1:]:
                m4 = re.search(
                    r"^\s*address\s+(\d+\.\d+\.\d+\.\d+)/(\d+)\b", ln)
                if m4:
                    enderecos.append((m4.group(1), int(m4.group(2))))
                    continue
                m6 = re.search(
                    r"^\s*address\s+([0-9a-fA-F:]+)/(\d+)\b", ln)
                if m6 and ":" in m6.group(1):
                    enderecos.append(
                        (m6.group(1), int(m6.group(2))))

        if not enderecos:
            return [], []

        # --- 3) Peer-IPs candidatos -----------------------------
        peer_candidates = []
        seen_pc = set()
        for ip_str, plen in enderecos:
            try:
                iface = ipaddress.ip_interface(f"{ip_str}/{plen}")
            except ValueError:
                continue
            net = iface.network
            if net.num_addresses > 256:
                # Para subredes grandes, considera apenas o par /31
                # ou /30 mais pr\u00f3ximo. Evita explos\u00e3o.
                continue
            for h in net:
                s_h = str(h)
                if s_h == ip_str:
                    continue
                if s_h not in seen_pc:
                    seen_pc.add(s_h)
                    peer_candidates.append(s_h)

        if not peer_candidates:
            return [], []

        # --- 4) Em cfg_vizinho: interfaces com address em peers -
        lines_v = cfg_vizinho.splitlines()
        peer_set = set(peer_candidates)
        matched_ports = []
        matched_peers = []
        matched_lags = []
        for i, s in enumerate(lines_v):
            if not rx_iface_hdr.match(s):
                continue
            stz, _ = self._swap_extract_stanza(lines_v, i)
            casou = ""
            for ln in stz[1:]:
                m4 = re.search(
                    r"^\s*address\s+(\d+\.\d+\.\d+\.\d+)/\d+", ln)
                if m4 and m4.group(1) in peer_set:
                    casou = m4.group(1)
                    break
                m6 = re.search(
                    r"^\s*address\s+([0-9a-fA-F:]+)/\d+", ln)
                if (m6 and ":" in m6.group(1)
                        and m6.group(1) in peer_set):
                    casou = m6.group(1)
                    break
            if not casou:
                continue
            if casou not in matched_peers:
                matched_peers.append(casou)
            for ln in stz[1:]:
                mp = re.match(r"^\s*port\s+(\S+)", ln)
                if not mp:
                    continue
                pname = mp.group(1)
                m_lag = re.match(r"^lag-(\d+)$", pname)
                if m_lag:
                    if m_lag.group(1) not in matched_lags:
                        matched_lags.append(m_lag.group(1))
                else:
                    if pname not in matched_ports:
                        matched_ports.append(pname)
                break  # primeiro `port X` apenas

        # --- 5) Expande LAGs do vizinho em portas membros -------
        for lag_id in matched_lags:
            rx_lag_hdr_v = re.compile(rf"^(\s*)lag\s+{lag_id}\s*$")
            for i, s in enumerate(lines_v):
                if not rx_lag_hdr_v.match(s):
                    continue
                stz, _ = self._swap_extract_stanza(lines_v, i)
                for ln in stz[1:]:
                    mpv = rx_port_in_lag.match(ln)
                    if mpv:
                        pn = mpv.group(1)
                        if pn not in matched_ports:
                            matched_ports.append(pn)
                break

        return matched_ports, matched_peers

    # ─────────────────────────────────────────────────────────────
    #  Bloco de SWAP do roteador VIZINHO (apenas description e
    #  recriacao de interface com nome substituido DE -> PARA)
    # ─────────────────────────────────────────────────────────────
    def _swap_montar_bloco_vizinho_swap(self, cfg_viz, vports,
                                          de_h, para_h, mapa_porta,
                                          rollback=False):
        """Monta o bloco simplificado do roteador VIZINHO para o
        arquivo SWAP_<de>_<para>.txt.

        Conteudo gerado (modo forward):
          - Para cada porta fisica do vizinho: 1 linha
            `/configure port <X> description "<desc_substituida>"`.
          - Para cada LAG do vizinho que contem alguma vport: 1 linha
            `/configure lag <N> description "<desc_substituida>"`.
          - Bloco `router` com:
              * Para cada interface (router Base) ligada ao port/LAG:
                shutdown + `no interface "OLD"` + recria com nome
                substituido (DE -> PARA), corpo intacto.
              * Para cada protocolo (mpls/rsvp/ldp/pim) com referencia
                a essa interface, recria a interface com nome
                substituido (corpo intacto).

        mapa_porta: dict {de_porta: para_porta} (substituicao 1:1).

        rollback=True: gera o bloco INVERSO (FALLBACK do vizinho):
          - description ORIGINAL (sem substituicao) em port/lag.
          - shutdown da interface NOVA (substituida) + `no interface
            "NOVA"` + recriacao da interface ORIGINAL (com body e
            descriptions originais).
          - Protocolos (mpls/rsvp/ldp/pim) recriam interfaces com
            o NOME ORIGINAL.

        Substituicao (somente no modo forward):
          - hostname DE (lower e UPPER) -> hostname PARA.
          - cada de_porta (lower e UPPER) -> respectivo para_porta.
        """
        if not cfg_viz or not vports:
            return ""

        def _subst(text):
            out = text
            if de_h and para_h and de_h.lower() != para_h.lower():
                out = out.replace(de_h.lower(), para_h.lower())
                out = out.replace(de_h.upper(), para_h.upper())
            for dp, pp in (mapa_porta or {}).items():
                if not dp or not pp or dp == pp:
                    continue
                out = out.replace(dp.lower(), pp.lower())
                if dp.upper() != dp.lower():
                    out = out.replace(dp.upper(), pp.upper())
            return out

        # `admin display-config` insere `echo "..."` (indent 0) e
        # linhas separadoras `#---...` / `# TiMOS-...` (tambem indent
        # 0) DENTRO do bloco `router Base`. Sem filtrar,
        # _swap_extract_stanza trata a primeira como fim do router
        # e perde os blocos mpls/rsvp/ldp/pim. Filtra essas linhas
        # antes de extrair.
        lines = [
            ln for ln in cfg_viz.splitlines()
            if not (
                re.match(r'^echo\s+"', ln)
                or re.match(r'^#', ln)
            )
        ]
        rx_port_hdr = re.compile(r'^(\s*)port\s+(\S+)\s*$')
        rx_lag_hdr = re.compile(r'^(\s*)lag\s+(\d+)\s*$')
        rx_iface_hdr = re.compile(
            r'^(\s*)interface\s+"([^"]+)"')
        rx_desc = re.compile(r'^\s*description\s+(.+)$')
        rx_port_in = re.compile(r'^\s*port\s+(\S+)')
        rx_router_base = re.compile(
            r'^\s{4}router(?:\s+Base)?\s*$')

        # ── 1) LAGs do vizinho que contem alguma vport ───────
        lags_info = []   # (n, desc_raw, [membros])
        lag_targets = set()
        for i, s in enumerate(lines):
            m = rx_lag_hdr.match(s)
            if not m:
                continue
            nlag = m.group(2)
            stz, _ = self._swap_extract_stanza(lines, i)
            membros = []
            desc_raw = ""
            for ln in stz[1:]:
                mp = rx_port_in.match(ln)
                if mp:
                    membros.append(mp.group(1))
                md = rx_desc.match(ln)
                if md and not desc_raw:
                    desc_raw = md.group(1).strip()
            if any(p in vports for p in membros):
                lags_info.append((nlag, desc_raw, membros))
                lag_targets.add(nlag)

        # ── 2) Descricoes das portas vizinhas ────────────────
        port_descs = {}  # vport -> desc_raw
        for i, s in enumerate(lines):
            m = rx_port_hdr.match(s)
            if not m:
                continue
            pname = m.group(2)
            if pname not in vports:
                continue
            stz, _ = self._swap_extract_stanza(lines, i)
            for ln in stz[1:]:
                md = rx_desc.match(ln)
                if md:
                    port_descs[pname] = md.group(1).strip()
                    break

        # ── 3) Localiza stanza `router` (Base) ───────────────
        # `admin display-config` pode emitir MULTIPLAS stanzas
        # `router Base` (Network Side, Service Side, etc.).
        # Concatena os corpos para pegar todos os protocolos.
        router_bodies = []
        for i, s in enumerate(lines):
            if rx_router_base.match(s):
                stz, _ = self._swap_extract_stanza(lines, i)
                router_bodies.extend(stz[1:])
        router_stanza = ([lines[0]] if False else [""]) + router_bodies
        # primeiro elemento e dummy; o walker comeca em i=1

        targets_in_iface = {f"port {p}" for p in vports}
        for n in lag_targets:
            targets_in_iface.add(f"port lag-{n}")

        iface_stanzas = []   # [(nome_old, [linhas])]
        # Aceita MULTIPLAS stanzas por protocolo (ex.: mpls aparece
        # duas vezes em SR-OS: interfaces + lsps). Junta todas.
        proto_stanzas = {"mpls": [], "rsvp": [], "ldp": [], "pim": []}

        if len(router_stanza) > 1:
            i = 1
            n = len(router_stanza)
            while i < n:
                ln = router_stanza[i]
                if not ln.strip():
                    i += 1; continue
                ind = len(ln) - len(ln.lstrip(" "))
                if ind != 8:
                    i += 1; continue
                head = ln.strip()
                if head.startswith('interface "'):
                    sub, end = self._swap_extract_stanza(
                        router_stanza, i)
                    bound = False
                    for body in sub[1:]:
                        bs = body.strip()
                        for t in targets_in_iface:
                            if bs == t or bs.startswith(t + " "):
                                bound = True; break
                        if bound: break
                    if bound:
                        m = rx_iface_hdr.match(sub[0])
                        if m:
                            iface_stanzas.append(
                                (m.group(2), sub))
                    i = end
                elif head in ("mpls", "rsvp", "ldp", "pim"):
                    sub, end = self._swap_extract_stanza(
                        router_stanza, i)
                    proto_stanzas[head].append(sub)
                    i = end
                else:
                    i += 1

        nomes_old = {nome for nome, _ in iface_stanzas}

        # Em rollback, descritions e nomes recriados ficam ORIGINAIS
        # (sem subst); apenas os nomes a "shutdown" e "no interface"
        # usam o nome NOVO (substituido).
        def _subst_desc(text):
            return text if rollback else _subst(text)

        out_parts = []

        # ── 4) Linhas de description (port/lag) ──────────────
        for vp in vports:
            d = port_descs.get(vp)
            if not d:
                continue
            out_parts.append(
                f'/configure port {vp} description {_subst_desc(d)}')
        for nlag, desc_raw, _membros in lags_info:
            if not desc_raw:
                continue
            out_parts.append(
                f'/configure lag {nlag} description '
                f'{_subst_desc(desc_raw)}')

        # ── 5) Bloco router (interfaces e protocolos) ────────
        if iface_stanzas:
            out_parts.append("")
            out_parts.append("/configure")
            out_parts.append("    router")
            for nome_old, sub in iface_stanzas:
                nome_new = _subst(nome_old)
                # Forward: shutdown OLD, no OLD, recria como NEW.
                # Rollback: shutdown NEW, no NEW, recria como OLD.
                nome_remove = nome_new if rollback else nome_old
                nome_create = nome_old if rollback else nome_new
                out_parts.append(
                    f'        interface "{nome_remove}" shutdown')
                out_parts.append(
                    f'        no interface "{nome_remove}"')
                new_header = sub[0].replace(
                    f'"{nome_old}"', f'"{nome_create}"', 1)
                out_parts.append(new_header.rstrip())
                for body_ln in sub[1:]:
                    raw = body_ln.rstrip()
                    if re.match(r'^\s*description\s+', raw):
                        raw = _subst_desc(raw)
                    out_parts.append(raw)

            # Protocolos: recria interface com nome substituido
            for proto in ("mpls", "rsvp", "ldp", "pim"):
                lst_sub = proto_stanzas.get(proto) or []
                if not lst_sub:
                    continue
                # ifaces no proto: ind 12 (mpls/rsvp/pim) ou
                # ind 16 (ldp dentro de interface-parameters)
                if proto == "ldp":
                    indent_iface = 16
                else:
                    indent_iface = 12
                sub_ifaces = []
                for sub_p in lst_sub:
                    i = 1
                    while i < len(sub_p):
                        ln = sub_p[i]
                        if not ln.strip():
                            i += 1; continue
                        ind = len(ln) - len(ln.lstrip(" "))
                        stripped = ln.strip()
                        if (ind == indent_iface
                                and stripped.startswith(
                                    'interface "')):
                            m = rx_iface_hdr.match(ln)
                            if m and m.group(2) in nomes_old:
                                ssub, end = (
                                    self._swap_extract_stanza(
                                        sub_p, i))
                                sub_ifaces.append(ssub)
                                i = end
                                continue
                        i += 1
                if not sub_ifaces:
                    continue
                out_parts.append(f"        {proto}")
                if proto == "ldp":
                    out_parts.append(
                        "            interface-parameters")
                for ssub in sub_ifaces:
                    m = rx_iface_hdr.match(ssub[0])
                    n_old = m.group(2)
                    n_new = _subst(n_old)
                    n_create = n_old if rollback else n_new
                    new_h = ssub[0].replace(
                        f'"{n_old}"', f'"{n_create}"', 1)
                    out_parts.append(new_h.rstrip())
                    for body_ln in ssub[1:]:
                        out_parts.append(body_ln.rstrip())
                if proto == "ldp":
                    out_parts.append("            exit")
                out_parts.append("        exit")
            out_parts.append("    exit")
            out_parts.append("exit all")

        return "\n".join(out_parts)

    # ─────────────────────────────────────────────────────────────
    #  Policies referenciadas + check no roteador PARA
    # ─────────────────────────────────────────────────────────────
    def _swap_policies_referenciadas(self, texto_extraido):
        """Varre o texto extraido (saida de _swap_extrair_config_portas)
        e devolve um dict {kind: set(nomes)} com as policies citadas.

        Tipos identificados (kind):
          - 'qos'              → numericas (sap-ingress N) e nomeadas
                                  (port-scheduler-policy, queue-group,
                                  network-policy, etc.).
          - 'filter'           → ip-filter / mac-filter / ipv6-filter.
          - 'policy_statement' → vrf-import / vrf-export / import / export.
          - 'community'        → community "X".
          - 'prefix_list'      → prefix-list "X".
          - 'as_path'          → as-path "X".
          - 'as_path_group'    → as-path-group "X".
          - 'slope_policy'     → wred-queue policy "X" (em queue-group).
          - 'sdp'              → spoke-sdp N:M / mesh-sdp N:M (extrai N).
          - 'ip_filter_num'    → filter ip <N>.
          - 'ipv6_filter_num'  → filter ipv6 <N>.
          - 'ip_prefix_list'   → ip-prefix-list "X" (em filter entries).
        """
        ref = {"qos": set(), "filter": set(),
               "policy_statement": set(), "community": set(),
               "prefix_list": set(), "slope_policy": set(),
               "sdp": set(), "ip_filter_num": set(),
               "ipv6_filter_num": set(), "ip_prefix_list": set(),
               "as_path": set(), "as_path_group": set()}
        for ln in texto_extraido.splitlines():
            s = ln.strip()
            m = re.match(r"^qos\s+(\d+)\b", s)
            if m:
                ref["qos"].add(m.group(1))
                # NAO faz continue: a mesma linha pode tambem trazer
                # `egress-port-redirect-group "X"` inline.

            # `wred-queue policy "X" mode ...` → slope-policy X
            m_wq = re.match(
                r'^wred-queue\s+policy\s+"([^"]+)"', s)
            if m_wq:
                ref["slope_policy"].add(m_wq.group(1))

            # `spoke-sdp N:M ...` ou `mesh-sdp N:M ...` → sdp N
            m_sp = re.match(
                r'^(?:spoke-sdp|mesh-sdp)\s+(\d+):\d+\b', s)
            if m_sp:
                ref["sdp"].add(m_sp.group(1))

            # `filter ip <N>` / `filter ipv6 <N>` (numerico, em sap)
            m_fi = re.match(r'^filter\s+ip\s+(\d+)\b', s)
            if m_fi:
                ref["ip_filter_num"].add(m_fi.group(1))
            m_fi6 = re.match(r'^filter\s+ipv6\s+(\d+)\b', s)
            if m_fi6:
                ref["ipv6_filter_num"].add(m_fi6.group(1))

            # rVPLS bind em interface IES/VPRN:
            #   `v4-routed-override-filter <N>`  -> ip-filter <N>
            #   `v6-routed-override-filter <N>`  -> ipv6-filter <N>
            m_v4o = re.match(
                r'^v4-routed-override-filter\s+(\d+)\b', s)
            if m_v4o:
                ref["ip_filter_num"].add(m_v4o.group(1))
            m_v6o = re.match(
                r'^v6-routed-override-filter\s+(\d+)\b', s)
            if m_v6o:
                ref["ipv6_filter_num"].add(m_v6o.group(1))

            # `(src-ip|dst-ip) ip-prefix-list "X"` em filter entries.
            for nm in re.findall(
                    r'\bip-prefix-list\s+"([^"]+)"', s):
                ref["ip_prefix_list"].add(nm)

            # `community expression "([X]OR[Y]AND[Z])"` — extrai os
            # nomes dentro de [...] como communities.
            m_ce = re.match(r'^community\s+expression\s+"([^"]+)"', s)
            if m_ce:
                for nm in re.findall(r'\[([^\]]+)\]', m_ce.group(1)):
                    ref["community"].add(nm)
            elif re.match(r'^community\b', s):
                # Qualquer outra forma: `community "X"`,
                # `community add "X" "Y"`, `community replace "X"`,
                # `community <verbo> ... "X"` — captura TODOS os nomes
                # entre aspas como communities. Ignora o que vem entre
                # `community` e o nome.
                for nm in re.findall(r'"([^"]+)"', s):
                    ref["community"].add(nm)

            # `route-exists "[X]"` (em conditional-expression) — X e
            # uma prefix-list.
            m_re = re.match(r'^route-exists\s+"\[([^\]]+)\]"', s)
            if m_re:
                ref["prefix_list"].add(m_re.group(1))
            # `import "A" "B" ...` / `export "A" "B" ..."` /
            # `vrf-import "A" "B" ...` / `vrf-export "A" "B" ...`:
            # SR-OS aceita multiplas policies na mesma linha; capturamos
            # TODAS as strings entre aspas.
            if re.match(
                    r'^(?:vrf-import|vrf-export|import|export)\b', s):
                for nm in re.findall(r'"([^"]+)"', s):
                    ref["policy_statement"].add(nm)

            for kind, rx in [
                ("qos",
                 r'^(?:port-scheduler-policy|scheduler-policy|'
                 r'queue-group|network-policy|network-queue|'
                 r'queue-policy|sap-ingress-policy|'
                 r'sap-egress-policy|egress-scheduler-policy|'
                 r'ingress-scheduler-policy|policer-control-policy|'
                 r'queue-group-redirect-list|'
                 r'egress-port-redirect-group|'
                 r'ingress-port-redirect-group)\s+"([^"]+)"'),
                # Caso inline em interface:
                #   `qos 1003 egress-port-redirect-group "SaidaIntRMS" ...`
                ("qos",
                 r'.*\begress-port-redirect-group\s+"([^"]+)"'),
                ("qos",
                 r'.*\bingress-port-redirect-group\s+"([^"]+)"'),
                ("filter",
                 r'^(?:ip-filter|mac-filter|ipv6-filter)\s+"([^"]+)"'),
                ("filter",
                 r'^filter\s+(?:ip|mac|ipv6)\s+"([^"]+)"'),
                ("policy_statement",
                 r'^(?:vrf-import|vrf-export|import|export)\s+"([^"]+)"'),
                ("community", r'^community\s+"([^"]+)"'),
                ("prefix_list", r'^prefix-list\s+"([^"]+)"'),
                ("as_path", r'^as-path\s+"([^"]+)"'),
                ("as_path_group", r'^as-path-group\s+"([^"]+)"'),
            ]:
                m = re.match(rx, s)
                if m:
                    ref[kind].add(m.group(1))
        return ref

    def _swap_buscar_policy_no_config(self, cfg_text, kind, nome):
        """Procura no `cfg_text` a definicao da policy `nome` do tipo
        `kind` e retorna a stanza completa (lista de linhas) ou [] se
        nao encontrada."""
        if not cfg_text or not nome:
            return []
        lines = cfg_text.splitlines()
        if kind == "qos":
            if nome.isdigit():
                heads = [
                    rf'^\s*sap-ingress\s+{re.escape(nome)}\b',
                    rf'^\s*sap-egress\s+{re.escape(nome)}\b',
                    rf'^\s*network\s+{re.escape(nome)}\b',
                ]
            else:
                heads = [
                    rf'^\s*(?:port-scheduler-policy|scheduler-policy|'
                    rf'queue-group|network-policy|network-queue|'
                    rf'sap-ingress-policy|'
                    rf'sap-egress-policy|egress-scheduler-policy|'
                    rf'ingress-scheduler-policy|policer-control-policy|'
                    rf'queue-group-redirect-list)'
                    rf'\s+"{re.escape(nome)}"',
                ]
        elif kind == "filter":
            heads = [
                rf'^\s*(?:ip-filter|mac-filter|ipv6-filter)'
                rf'\s+"{re.escape(nome)}"',
            ]
        elif kind == "policy_statement":
            heads = [rf'^\s*policy-statement\s+"{re.escape(nome)}"']
        elif kind == "community":
            heads = [rf'^\s*community\s+"{re.escape(nome)}"']
        elif kind == "prefix_list":
            heads = [rf'^\s*prefix-list\s+"{re.escape(nome)}"']
        elif kind == "as_path":
            heads = [rf'^\s*as-path\s+"{re.escape(nome)}"']
        elif kind == "as_path_group":
            heads = [rf'^\s*as-path-group\s+"{re.escape(nome)}"']
        elif kind == "slope_policy":
            heads = [rf'^\s*slope-policy\s+"{re.escape(nome)}"']
        elif kind == "sdp":
            heads = [rf'^\s*sdp\s+{re.escape(nome)}\b']
        elif kind == "ip_filter_num":
            heads = [rf'^\s*ip-filter\s+{re.escape(nome)}\b']
        elif kind == "ipv6_filter_num":
            heads = [rf'^\s*ipv6-filter\s+{re.escape(nome)}\b']
        elif kind == "ip_prefix_list":
            heads = [rf'^\s*ip-prefix-list\s+"{re.escape(nome)}"']
        else:
            return []
        rx_heads = [re.compile(p) for p in heads]
        for i, s in enumerate(lines):
            for rx in rx_heads:
                if rx.match(s):
                    stanza, _ = self._swap_extract_stanza(lines, i)
                    if len(stanza) >= 2:
                        return stanza
        return []

    def _swap_montar_bloco_policies(self, cfg_de, cfg_para,
                                    texto_extraido,
                                    vizinho_host=None,
                                    vizinho_ip=None):
        """Identifica as policies referenciadas em `texto_extraido`,
        compara contra `cfg_para` e devolve um trecho de configuracao
        (string ja indentada) com APENAS as policies que faltam no
        PARA. Devolve string vazia se nada faltar.

        Resolu\u00e7\u00e3o transitiva: ap\u00f3s incluir uma policy-statement no
        bloco, varre seu body atr\u00e1s de `community \"X\"` / `prefix-list
        \"X\"` e tamb\u00e9m as inclui (caso n\u00e3o existam no PARA).
        """
        ref = self._swap_policies_referenciadas(texto_extraido)
        faltam = {"qos": [], "filter": [],
                  "policy_statement": [], "community": [],
                  "prefix_list": [], "slope_policy": [],
                  "sdp": [], "ip_filter_num": [],
                  "ipv6_filter_num": [], "ip_prefix_list": [],
                  "as_path": [], "as_path_group": []}
        # Sets para evitar duplicatas em resolu\u00e7\u00e3o transitiva
        ja_visto = {k: set() for k in faltam}
        # Nomes (por kind) que existem no PARA mas com body DIFERENTE
        # do DE \u2014 ser\u00e3o emitidos com `no <kind> "X"` antes da nova
        # defini\u00e7\u00e3o. Aplica-se a policy-statement, community,
        # as-path, as-path-group, prefix-list e ip-prefix-list.
        replace_kinds = {
            "policy_statement", "community", "as_path",
            "as_path_group", "prefix_list", "ip_prefix_list",
        }
        replace_marcados = {k: set() for k in replace_kinds}
        # Mapeia kind -> token literal do `no <token> "X"` em SR-OS.
        no_cmd_token = {
            "policy_statement": "policy-statement",
            "community": "community",
            "as_path": "as-path",
            "as_path_group": "as-path-group",
            "prefix_list": "prefix-list",
            "ip_prefix_list": "ip-prefix-list",
        }

        def _stanzas_iguais(st_a, st_b):
            """Compara duas stanzas ignorando indenta\u00e7\u00e3o e linhas
            vazias. Retorna True se forem semanticamente iguais."""
            def _norm(st):
                return [ln.strip() for ln in st if ln.strip()]
            return _norm(st_a) == _norm(st_b)

        def _buscar_por_head_rx(cfg_text, rx_head_str):
            if not cfg_text:
                return []
            ls = cfg_text.splitlines()
            rx = re.compile(rx_head_str)
            for i, s in enumerate(ls):
                if rx.match(s):
                    st, _ = self._swap_extract_stanza(ls, i)
                    if len(st) >= 2:
                        return st
            return []

        def _add_se_faltar(kind, nome):
            if not nome or nome in ja_visto[kind]:
                return None
            # QoS numerica: coleta SEPARADAMENTE sap-ingress, sap-egress
            # e network (pois a mesma numeracao pode existir em mais de
            # um tipo e cada uma deve ser copiada se faltar no PARA).
            if kind == "qos" and nome.isdigit():
                ja_visto[kind].add(nome)
                achou = None
                for rx_head in (
                        rf'^\s*sap-ingress\s+{re.escape(nome)}\b',
                        rf'^\s*sap-egress\s+{re.escape(nome)}\b',
                        rf'^\s*network\s+{re.escape(nome)}\b'):
                    if _buscar_por_head_rx(cfg_para, rx_head):
                        continue
                    st = _buscar_por_head_rx(cfg_de, rx_head)
                    if st:
                        faltam["qos"].append((nome, st))
                        achou = st
                return achou
            st_para = self._swap_buscar_policy_no_config(
                cfg_para, kind, nome)
            stanza = self._swap_buscar_policy_no_config(
                cfg_de, kind, nome)
            if st_para:
                # Para os kinds em `replace_kinds`: se o body diverge
                # entre DE e PARA, marca para REPLACE (no <kind> "X"
                # + nova defini\u00e7\u00e3o do DE). Para os demais kinds:
                # se existe no PARA, ignora.
                if kind in replace_kinds and stanza:
                    if _stanzas_iguais(stanza, st_para):
                        ja_visto[kind].add(nome)
                        return None
                    faltam[kind].append((nome, stanza))
                    replace_marcados[kind].add(nome)
                    ja_visto[kind].add(nome)
                    return stanza
                ja_visto[kind].add(nome)
                return None
            if stanza:
                faltam[kind].append((nome, stanza))
                ja_visto[kind].add(nome)
                return stanza
            ja_visto[kind].add(nome)
            return None

        # 1\u00aa rodada: refer\u00eancias diretas no texto extra\u00eddo
        for kind, nomes in ref.items():
            for nome in sorted(nomes):
                _add_se_faltar(kind, nome)

        # 2\u00aa rodada (transitiva): varre policy-statements inclu\u00eddos
        # atr\u00e1s de community/prefix-list refs. Repete enquanto houver
        # novidades (cobre tamb\u00e9m policy-statement-in-policy-statement).
        rx_comm = re.compile(r'^\s*community\s+"([^"]+)"')
        rx_pfx = re.compile(r'^\s*prefix-list\s+"([^"]+)"')
        rx_aspath = re.compile(r'^\s*as-path\s+"([^"]+)"')
        rx_aspath_grp = re.compile(
            r'^\s*as-path-group\s+"([^"]+)"')
        rx_ps_ref = re.compile(
            r'^\s*(?:from\s+)?policy\s+"([^"]+)"')
        novidade = True
        while novidade:
            novidade = False
            # Snapshot atual das stanzas inclu\u00eddas (qualquer kind cujo
            # body possa referenciar communities/prefix-lists/policies
            # ou \u2014 no caso de qos \u2014 wred-queue \u2192 slope-policy).
            todas = (list(faltam["policy_statement"])
                     + list(faltam["qos"])
                     + list(faltam["filter"])
                     + list(faltam["ip_filter_num"])
                     + list(faltam["ipv6_filter_num"]))
            for _nome, stanza in todas:
                # Determina kind do owner para resolu\u00e7\u00e3o de embed-filter
                # (ip-filter \u2192 ip_filter_num; ipv6-filter \u2192
                # ipv6_filter_num).
                hdr0 = stanza[0].strip() if stanza else ""
                owner_kind_emb = None
                if hdr0.startswith("ip-filter"):
                    owner_kind_emb = "ip_filter_num"
                elif hdr0.startswith("ipv6-filter"):
                    owner_kind_emb = "ipv6_filter_num"
                for ln in stanza:
                    s = ln.strip()
                    # embed-filter <N> [offset M] (apenas n\u00fameros) \u2192
                    # adiciona o ip-filter/ipv6-filter <N> recursivamente.
                    if owner_kind_emb:
                        m_emb = re.match(
                            r'^embed-filter\s+(\d+)\b', s)
                        if m_emb and m_emb.group(1) not in ja_visto[
                                owner_kind_emb]:
                            if _add_se_faltar(
                                    owner_kind_emb, m_emb.group(1)):
                                novidade = True
                    # wred-queue policy "X" \u2192 slope-policy X
                    mwq = re.match(
                        r'^wred-queue\s+policy\s+"([^"]+)"', s)
                    if mwq and mwq.group(1) not in ja_visto[
                            "slope_policy"]:
                        if _add_se_faltar(
                                "slope_policy", mwq.group(1)):
                            novidade = True
                    # ip-prefix-list "X" (em entry de ip-filter)
                    for nm in re.findall(
                            r'\bip-prefix-list\s+"([^"]+)"', s):
                        if nm not in ja_visto["ip_prefix_list"]:
                            if _add_se_faltar(
                                    "ip_prefix_list", nm):
                                novidade = True
                    m = rx_comm.match(s)
                    if m and m.group(1) not in ja_visto["community"]:
                        if _add_se_faltar("community", m.group(1)):
                            novidade = True
                    # `community expression "([X]OR[Y])"` — extrai
                    # nomes dentro de [...]
                    mce = re.match(
                        r'^community\s+expression\s+"([^"]+)"', s)
                    if mce:
                        for nm in re.findall(
                                r'\[([^\]]+)\]', mce.group(1)):
                            if nm not in ja_visto["community"]:
                                if _add_se_faltar("community", nm):
                                    novidade = True
                    elif re.match(r'^community\b', s):
                        # Qualquer outra forma: `community "X"`,
                        # `community add "X" "Y"`,
                        # `community replace "X"`,
                        # `community <verbo> ... "X"` \u2014 captura TODOS
                        # os nomes entre aspas. Ignora o que vem entre
                        # `community` e o nome.
                        for nm in re.findall(r'"([^"]+)"', s):
                            if nm not in ja_visto["community"]:
                                if _add_se_faltar("community", nm):
                                    novidade = True
                    m = rx_pfx.match(s)
                    if m and m.group(1) not in ja_visto["prefix_list"]:
                        if _add_se_faltar("prefix_list", m.group(1)):
                            novidade = True
                    m = rx_aspath.match(s)
                    if m and m.group(1) not in ja_visto["as_path"]:
                        if _add_se_faltar("as_path", m.group(1)):
                            novidade = True
                    m = rx_aspath_grp.match(s)
                    if m and m.group(1) not in ja_visto[
                            "as_path_group"]:
                        if _add_se_faltar(
                                "as_path_group", m.group(1)):
                            novidade = True
                    # `route-exists "[X]"` → prefix-list X
                    mre = re.match(
                        r'^route-exists\s+"\[([^\]]+)\]"', s)
                    if mre and mre.group(1) not in ja_visto[
                            "prefix_list"]:
                        if _add_se_faltar(
                                "prefix_list", mre.group(1)):
                            novidade = True
                    m = rx_ps_ref.match(s)
                    if m and m.group(1) not in ja_visto[
                            "policy_statement"]:
                        if _add_se_faltar(
                                "policy_statement", m.group(1)):
                            novidade = True

        if not any(faltam[k] for k in faltam):
            return ""

        def reindent(stanza, base):
            if not stanza:
                return []
            header = stanza[0]
            orig = len(header) - len(header.lstrip(" "))
            out = []
            for line in stanza:
                if not line.strip():
                    out.append("")
                    continue
                if line[:orig] == " " * orig:
                    rest = line[orig:]
                else:
                    rest = line.lstrip()
                out.append(base + rest)
            return out

        out = ["",
               "# ===== Policies que NAO existem no roteador PARA "
               "(extraidas do DE) =====",
               "exit all", "configure"]
        if faltam["sdp"]:
            # sdp \u00e9 filho direto de `service` \u2014 emitido PRIMEIRO
            # (antes das demais policies).
            out.append("    service")
            for _nome, st in faltam["sdp"]:
                # Para `sdp 1`, reescreve `far-end` e `description`
                # apontando para o roteador VIZINHO do PARA
                # (regra: agg03 <=> agg04). Aplica somente quando
                # vizinho_host/vizinho_ip foram resolvidos.
                if (str(_nome) == "1" and vizinho_ip
                        and vizinho_host):
                    viz_short = vizinho_host.split(".")[0]
                    viz_full = vizinho_host
                    desc_nova = f'description "TO_{viz_full.upper()}"'
                    far_nova = f"far-end {vizinho_ip}"
                    st_mod = []
                    for ln in st:
                        s = ln.strip()
                        ind = ln[: len(ln) - len(ln.lstrip(" "))]
                        if s.startswith("description "):
                            st_mod.append(ind + desc_nova)
                        elif s.startswith("far-end "):
                            st_mod.append(ind + far_nova)
                        else:
                            st_mod.append(ln)
                    out.extend(reindent(st_mod, "        "))
                else:
                    out.extend(reindent(st, "        "))
            out.append("    exit")
        if faltam["qos"]:
            out.append("    qos")
            for _nome, st in faltam["qos"]:
                out.extend(reindent(st, "        "))
            out.append("    exit")
        if faltam["slope_policy"]:
            # slope-policy mora em `qos`; emite em bloco próprio se
            # não houver outras qos.
            out.append("    qos")
            for _nome, st in faltam["slope_policy"]:
                out.extend(reindent(st, "        "))
            out.append("    exit")
        if (faltam["filter"] or faltam["ip_filter_num"]
                or faltam["ipv6_filter_num"]
                or faltam["ip_prefix_list"]):
            out.append("    filter")
            # match-list \u2192 ip-prefix-list (vai PRIMEIRO porque os
            # ip-filters dependem dele).
            if faltam["ip_prefix_list"]:
                out.append("        match-list")
                for _nome, st in faltam["ip_prefix_list"]:
                    if _nome in replace_marcados["ip_prefix_list"]:
                        out.append(
                            "            "
                            + f'no ip-prefix-list "{_nome}"')
                    out.extend(reindent(st, "            "))
                out.append("        exit")
            for _nome, st in faltam["filter"]:
                out.extend(reindent(st, "        "))
            for _nome, st in faltam["ip_filter_num"]:
                out.extend(reindent(st, "        "))
            for _nome, st in faltam["ipv6_filter_num"]:
                out.extend(reindent(st, "        "))
            out.append("    exit")
        po = (faltam["policy_statement"] + faltam["community"]
              + faltam["prefix_list"] + faltam["as_path"]
              + faltam["as_path_group"])
        if po:
            out.append("    router")
            out.append("        policy-options")
            out.append("            begin")
            # Para cada kind: se o nome ja existe no PARA com body
            # diferente, emite `no <kind> "X"` antes da nova
            # defini\u00e7\u00e3o (substitui\u00e7\u00e3o atomica dentro do
            # begin/commit).
            def _emit_po_kind(kind):
                token = no_cmd_token[kind]
                for _n, st in faltam[kind]:
                    if _n in replace_marcados[kind]:
                        out.append(
                            "            "
                            + f'no {token} "{_n}"')
                    out.extend(reindent(st, "            "))
            _emit_po_kind("prefix_list")
            _emit_po_kind("as_path")
            _emit_po_kind("as_path_group")
            _emit_po_kind("community")
            _emit_po_kind("policy_statement")
            out.append("            commit")
            out.append("        exit")
            out.append("    exit")
        out.append("exit all")
        return "\n".join(out) + "\n"

    def _swap_reorganizar_texto(self, texto):
        """Reorganiza o texto SR-OS em sec\u00f5es na ordem solicitada:
            1 QoS (slope, scheduler, policer, network-policy,
                   network-queue, queue-group, sap-ingress, sap-egress,
                   demais qos)
            2 Portas (f\u00edsicas + system esat)
            3 LAGs
            4 Pol\u00edticas de roteamento (prefix-list, community,
                   demais, policy-statement)
            5 Filtros IP (match-list, ip-filter, ipv6/mac-filter)
            6 MPLS
            7 Router Base (static, interfaces, protocolos n\u00e3o-MPLS,
                   demais)
            8 Servi\u00e7os (sdp, customer, demais)
        Cada se\u00e7\u00e3o vira um bloco
        `exit all` / `configure` ... `exit all`.
        Comandos avulsos `/configure ...` (uma linha s\u00f3) s\u00e3o
        preservados ao final.
        """
        if not texto or not texto.strip():
            return texto

        linhas = texto.splitlines()
        # 1) Extrai os \u201cblocos configure\u201d e os comandos avulsos
        #    `/configure ...` (linhas isoladas come\u00e7ando com /).
        blocos_cfg = []   # cada um = list[str] (linhas internas, indent>=4)
        comandos_avulsos = []   # linhas come\u00e7ando com `/configure`
        comentarios = []  # linhas de coment\u00e1rio top-level

        i = 0
        n = len(linhas)
        while i < n:
            ln = linhas[i]
            s = ln.strip()
            if not s:
                i += 1
                continue
            if s.startswith("/configure"):
                comandos_avulsos.append(s)
                i += 1
                continue
            if s.startswith("#"):
                comentarios.append(s)
                i += 1
                continue
            if s == "exit all":
                i += 1
                continue
            if s == "configure":
                # Coleta at\u00e9 o pr\u00f3ximo `exit all` (ou EOF)
                i += 1
                bloco = []
                while i < n:
                    ln2 = linhas[i]
                    s2 = ln2.strip()
                    if s2 == "exit all":
                        i += 1
                        break
                    if s2 == "configure":
                        # configure aninhado: ignora e continua
                        i += 1
                        continue
                    bloco.append(ln2)
                    i += 1
                if bloco:
                    blocos_cfg.append(bloco)
                continue
            i += 1

        # 2) Para cada bloco, extrai filhos diretos do `configure`
        #    (indent == 4) e classifica em buckets.
        # Buckets de saida (cada um e' uma lista de stanzas; cada
        # stanza e' uma lista de linhas internas com indent original
        # rebaixado para 0 = inserido sem indent extra).
        buckets = {
            # QoS
            "qos_slope": [], "qos_sched": [], "qos_policer": [],
            "qos_netpol": [], "qos_netq": [], "qos_qg": [],
            "qos_sapi": [], "qos_sape": [], "qos_outros": [],
            # Portas
            "port_fis": [], "system": [],
            # dist-cpu-protection (system > security > dist-cpu-protection)
            "dcp": [],
            # LAGs
            "lags": [],
            # Politicas de roteamento (router > policy-options)
            "polit_pfx": [], "polit_com": [], "polit_outros": [],
            "polit_ps": [],
            # Filtros
            "filt_match": [], "filt_ip": [], "filt_ipv6": [],
            # MPLS (vai contido em router > mpls)
            "mpls_blocos": [],   # list of stanza (filho do mpls)
            # Router Base demais
            "rb_static": [], "rb_iface": [], "rb_proto": [],
            "rb_outros": [],
            # Servi\u00e7os
            "svc_sdp": [], "svc_cust": [], "svc_outros": [],
            # Demais (n\u00e3o classificados)
            "fallback": [],
        }
        # Comandos avulsos no final
        # comandos_avulsos j\u00e1 coletados acima

        def _extrair_stanza_local(ls, idx):
            """Devolve (stanza_lines, end_idx). A stanza inclui o header
            em ls[idx] e todas as linhas com indent > do header at\u00e9
            o `exit` no mesmo n\u00edvel (inclusive)."""
            hdr = ls[idx]
            base_ind = len(hdr) - len(hdr.lstrip(" "))
            out = [hdr]
            j = idx + 1
            while j < len(ls):
                ln_ = ls[j]
                if not ln_.strip():
                    out.append(ln_)
                    j += 1
                    continue
                ind_ = len(ln_) - len(ln_.lstrip(" "))
                strip_ = ln_.strip()
                if ind_ < base_ind:
                    break
                if ind_ == base_ind:
                    # `exit` fecha; outro header no mesmo nivel encerra
                    if strip_.startswith("exit"):
                        out.append(ln_)
                        j += 1
                        break
                    break
                out.append(ln_)
                j += 1
            return out, j

        def _classificar_qos(stanza):
            """Devolve key do bucket QoS para a stanza."""
            if not stanza:
                return "qos_outros"
            s = stanza[0].lstrip()
            if s.startswith("slope-policy"):
                return "qos_slope"
            if (s.startswith("scheduler-policy")
                    or s.startswith("port-scheduler-policy")
                    or s.startswith("egress-scheduler-policy")
                    or s.startswith("ingress-scheduler-policy")):
                return "qos_sched"
            if s.startswith("policer-control-policy"):
                return "qos_policer"
            if (s.startswith("network-policy")
                    or re.match(r"^network\s+\d", s)):
                return "qos_netpol"
            if (s.startswith("network-queue")
                    or s.startswith("network-queue-policy")):
                return "qos_netq"
            if (s.startswith("queue-group")
                    or s.startswith("queue-group-redirect-list")):
                return "qos_qg"
            if s.startswith("sap-ingress"):
                return "qos_sapi"
            if s.startswith("sap-egress"):
                return "qos_sape"
            return "qos_outros"

        def _classificar_polit(stanza):
            if not stanza:
                return "polit_outros"
            s = stanza[0].lstrip()
            if s.startswith("prefix-list"):
                return "polit_pfx"
            if s.startswith("community"):
                return "polit_com"
            if s.startswith("policy-statement"):
                return "polit_ps"
            return "polit_outros"

        def _classificar_filter(stanza):
            if not stanza:
                return "filt_ip"
            s = stanza[0].lstrip()
            if s.startswith("match-list"):
                return "filt_match"
            if s.startswith("ipv6-filter"):
                return "filt_ipv6"
            return "filt_ip"  # ip-filter, mac-filter

        def _classificar_svc(stanza):
            if not stanza:
                return "svc_outros"
            s = stanza[0].lstrip()
            if s.startswith("sdp "):
                return "svc_sdp"
            if s.startswith("customer "):
                return "svc_cust"
            return "svc_outros"

        for bloco in blocos_cfg:
            i = 0
            while i < len(bloco):
                ln = bloco[i]
                if not ln.strip():
                    i += 1
                    continue
                ind = len(ln) - len(ln.lstrip(" "))
                if ind != 4:
                    i += 1
                    continue
                strip_ = ln.strip()
                stanza, end = _extrair_stanza_local(bloco, i)
                hdr0 = strip_.split()[0] if strip_ else ""
                if hdr0 == "qos":
                    # filhos diretos do qos
                    j = 1
                    while j < len(stanza) - 1:
                        ln2 = stanza[j]
                        if not ln2.strip():
                            j += 1
                            continue
                        ind2 = (len(ln2) - len(ln2.lstrip(" ")))
                        strip2 = ln2.strip()
                        if ind2 != 8 or strip2 == "exit":
                            j += 1
                            continue
                        sub, sub_end = _extrair_stanza_local(stanza, j)
                        buckets[_classificar_qos(sub)].append(sub)
                        j = sub_end
                elif hdr0 == "filter":
                    j = 1
                    while j < len(stanza) - 1:
                        ln2 = stanza[j]
                        if not ln2.strip():
                            j += 1
                            continue
                        ind2 = (len(ln2) - len(ln2.lstrip(" ")))
                        strip2 = ln2.strip()
                        if ind2 != 8 or strip2 == "exit":
                            j += 1
                            continue
                        sub, sub_end = _extrair_stanza_local(stanza, j)
                        buckets[_classificar_filter(sub)].append(sub)
                        j = sub_end
                elif hdr0 == "lag":
                    buckets["lags"].append(stanza)
                elif hdr0 == "port":
                    buckets["port_fis"].append(stanza)
                elif hdr0 == "system":
                    # Detecta sub-bloco dist-cpu-protection
                    is_dcp = any(
                        ln2.strip().startswith("dist-cpu-protection")
                        for ln2 in stanza)
                    if is_dcp:
                        # Extrai apenas as `policy "X" create` filhas
                        for j in range(1, len(stanza)):
                            if (stanza[j].lstrip()
                                    .startswith("dist-cpu-protection")):
                                dcp_blk, _ = _extrair_stanza_local(
                                    stanza, j)
                                k = 1
                                while k < len(dcp_blk) - 1:
                                    ln3 = dcp_blk[k]
                                    if not ln3.strip():
                                        k += 1
                                        continue
                                    if (ln3.lstrip()
                                            .startswith("policy ")):
                                        sub, sub_end = (
                                            _extrair_stanza_local(
                                                dcp_blk, k))
                                        buckets["dcp"].append(sub)
                                        k = sub_end
                                        continue
                                    k += 1
                                break
                    else:
                        buckets["system"].append(stanza)
                elif hdr0 == "service":
                    j = 1
                    while j < len(stanza) - 1:
                        ln2 = stanza[j]
                        if not ln2.strip():
                            j += 1
                            continue
                        ind2 = (len(ln2) - len(ln2.lstrip(" ")))
                        strip2 = ln2.strip()
                        if ind2 != 8 or strip2 == "exit":
                            j += 1
                            continue
                        sub, sub_end = _extrair_stanza_local(stanza, j)
                        buckets[_classificar_svc(sub)].append(sub)
                        j = sub_end
                elif hdr0 == "router":
                    # router Base ou router (policy-options)
                    # Detecta se e' "router" puro com policy-options
                    is_polopt = False
                    for ln2 in stanza[1:4]:
                        if "policy-options" in ln2:
                            is_polopt = True
                            break
                    if is_polopt:
                        # navega router > policy-options > filhos
                        # localiza policy-options
                        for j in range(1, len(stanza)):
                            if (stanza[j].lstrip()
                                    .startswith("policy-options")):
                                po, _po_end = _extrair_stanza_local(
                                    stanza, j)
                                k = 1
                                while k < len(po) - 1:
                                    ln3 = po[k]
                                    if not ln3.strip():
                                        k += 1
                                        continue
                                    ind3 = (len(ln3)
                                            - len(ln3.lstrip(" ")))
                                    strip3 = ln3.strip()
                                    if (ind3 != 12
                                            or strip3 in (
                                                "exit", "begin",
                                                "commit")):
                                        k += 1
                                        continue
                                    sub, sub_end = _extrair_stanza_local(
                                        po, k)
                                    buckets[_classificar_polit(sub)
                                            ].append(sub)
                                    k = sub_end
                                break
                    else:
                        # router Base: separa filhos diretos
                        j = 1
                        while j < len(stanza) - 1:
                            ln2 = stanza[j]
                            if not ln2.strip():
                                j += 1
                                continue
                            ind2 = (len(ln2)
                                    - len(ln2.lstrip(" ")))
                            strip2 = ln2.strip()
                            if ind2 != 8 or strip2 == "exit":
                                j += 1
                                continue
                            sub, sub_end = _extrair_stanza_local(
                                stanza, j)
                            shdr = strip2.split()[0]
                            if (strip2.startswith("static-route")
                                    or strip2.startswith(
                                        "static-routes")
                                    or strip2.startswith(
                                        "static-route-entry")):
                                buckets["rb_static"].append(sub)
                            elif strip2.startswith("interface "):
                                buckets["rb_iface"].append(sub)
                            elif shdr == "mpls":
                                # Coleta filhos diretos do mpls como
                                # \u201cmpls_blocos\u201d (preservados como
                                # est\u00e3o, ja com indent +12).
                                buckets["mpls_blocos"].append(sub)
                            elif shdr in ("ospf", "ospf3", "isis",
                                          "rsvp", "ldp", "pim",
                                          "bgp"):
                                buckets["rb_proto"].append(sub)
                            else:
                                buckets["rb_outros"].append(sub)
                            j = sub_end
                else:
                    buckets["fallback"].append(stanza)
                i = end

        # 3) Renderiza na ordem solicitada.
        def _emit_bloco(linhas_dentro):
            """Devolve list[str] envolvendo `linhas_dentro` com
            `exit all` / `configure` / `exit all`. Retorna [] se nao
            houver nenhum conteudo significativo."""
            if not linhas_dentro:
                return []
            # ignora se s\u00f3 cont\u00e9m exits/strings vazias
            sig = [x for x in linhas_dentro
                   if x.strip()
                   and x.strip() not in ("exit", "exit all")]
            if not sig:
                return []
            return ["exit all", "configure"] + linhas_dentro \
                + ["exit all"]

        def _stz_lines(stz_list):
            out = []
            for st in stz_list:
                out.extend(st)
            return out

        secoes = []

        # 1) QoS
        qos_inner = []
        for k in ("qos_slope", "qos_sched", "qos_policer",
                  "qos_netpol", "qos_netq", "qos_qg",
                  "qos_sapi", "qos_sape", "qos_outros"):
            qos_inner.extend(_stz_lines(buckets[k]))
        if qos_inner:
            secoes.append(["    qos"] + qos_inner + ["    exit"])

        # 1.b dist-cpu-protection (logo apos o QoS)
        if buckets["dcp"]:
            dcp_inner = []
            for st in buckets["dcp"]:
                dcp_inner.extend(st)
            secoes.append([
                "    system",
                "        security",
                "            dist-cpu-protection"]
                + dcp_inner
                + ["            exit",
                   "        exit",
                   "    exit"])

        # 2) Portas (f\u00edsicas + system esat)
        portas_inner = _stz_lines(buckets["port_fis"]) \
            + _stz_lines(buckets["system"])
        if portas_inner:
            secoes.append(portas_inner)

        # 3) LAGs
        lags_inner = _stz_lines(buckets["lags"])
        if lags_inner:
            secoes.append(lags_inner)

        # 4) Pol\u00edticas de roteamento
        pol_inner = []
        for k in ("polit_pfx", "polit_com", "polit_outros",
                  "polit_ps"):
            pol_inner.extend(_stz_lines(buckets[k]))
        if pol_inner:
            secoes.append([
                "    router",
                "        policy-options",
                "            begin"]
                + pol_inner
                + ["            commit",
                   "        exit",
                   "    exit"])

        # 5) Filtros IP
        filt_inner = []
        if buckets["filt_match"]:
            filt_inner.append("        match-list")
            for st in buckets["filt_match"]:
                # se st j\u00e1 e' header match-list (caso raro), ignora
                if st and st[0].lstrip().startswith("match-list"):
                    # pega filhos diretos
                    for ln2 in st[1:-1]:
                        filt_inner.append(ln2)
                else:
                    for ln2 in st:
                        filt_inner.append(ln2)
            filt_inner.append("        exit")
        for k in ("filt_ip", "filt_ipv6"):
            filt_inner.extend(_stz_lines(buckets[k]))
        if filt_inner:
            secoes.append(["    filter"] + filt_inner + ["    exit"])

        # 6) MPLS (router > mpls)
        if buckets["mpls_blocos"]:
            mpls_inner = []
            # Cada `st` em mpls_blocos e a stanza completa do
            # `mpls` (com header e exit). Extraimos apenas os
            # filhos para evitar duplicar `mpls`/`exit` no wrapper.
            for st in buckets["mpls_blocos"]:
                if not st:
                    continue
                # remove primeira linha (header `mpls`) e ultima
                # linha quando for o `exit` correspondente.
                body = st[1:]
                if body and body[-1].strip() == "exit":
                    body = body[:-1]
                mpls_inner.extend(body)
            secoes.append([
                "    router",
                "        mpls"]
                + mpls_inner
                + ["        exit",
                   "    exit"])

        # 7) Router Base (sem MPLS) na ordem static \u2192 iface \u2192
        #    protocolos \u2192 demais
        rb_inner = []
        for k in ("rb_static", "rb_iface", "rb_proto", "rb_outros"):
            rb_inner.extend(_stz_lines(buckets[k]))
        if rb_inner:
            secoes.append([
                "    router"]
                + rb_inner
                + ["    exit"])

        # 8) Servi\u00e7os: sdp \u2192 customer \u2192 demais
        svc_inner = []
        for k in ("svc_sdp", "svc_cust", "svc_outros"):
            svc_inner.extend(_stz_lines(buckets[k]))
        if svc_inner:
            secoes.append(["    service"] + svc_inner + ["    exit"])

        # Fallback (n\u00e3o classificado): emite no final como bloco
        if buckets["fallback"]:
            fb_inner = _stz_lines(buckets["fallback"])
            if fb_inner:
                secoes.append(fb_inner)

        # 4) Monta texto final
        out_final = []
        if comentarios:
            out_final.extend(comentarios)
        for sec in secoes:
            bloco = _emit_bloco(sec)
            if bloco:
                out_final.append("")
                out_final.extend(bloco)
        # Comandos avulsos `/configure ...` ao final
        if comandos_avulsos:
            out_final.append("")
            out_final.extend(comandos_avulsos)
        return "\n".join(out_final).strip() + "\n"

    def _swap_pedir_renomear_lag(self, lid_atual, lid_sugerido):
        """Mostra um dialog estilizado perguntando se o usuario quer
        substituir o ID de um LAG que ja existe no roteador PARA.
        Retorna True se aceitar (renomear), False caso contrario.

        Pode ser chamado de uma worker thread: a UI sera criada na
        main thread via `self.root.after`, e o resultado e devolvido
        atraves de uma queue."""
        import queue as _q_lag
        resp_q = _q_lag.Queue()

        def _build():
            dlg = tk.Toplevel(self.root)
            dlg.title("Conflito de LAG ID")
            dlg.resizable(False, False)
            dlg.attributes("-topmost", True)
            dlg.configure(bg="#1a1a2e")

            outer = tk.Frame(dlg, bg="#1a1a2e", bd=4, relief="raised")
            outer.pack(fill=tk.BOTH, expand=True)
            mid = tk.Frame(outer, bg="#16213e", bd=3, relief="ridge")
            mid.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)
            inner = tk.Frame(mid, bg="#0f3460", bd=2, relief="groove")
            inner.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)

            tk.Label(
                inner,
                text="⚠  CONFLITO DE LAG ID  ⚠",
                font=("Segoe UI", 14, "bold"),
                bg="#0f3460", fg="#FFD43B",
            ).pack(padx=30, pady=(18, 6))

            tk.Frame(inner, bg="#e94560", height=2).pack(
                fill=tk.X, padx=20, pady=6)

            tk.Label(
                inner,
                text=(f"O LAG  lag {lid_atual}  já existe no "
                      f"roteador PARA."),
                font=("Segoe UI", 11),
                bg="#0f3460", fg="white",
                wraplength=460, justify="center",
            ).pack(padx=30, pady=(8, 4))

            tk.Label(
                inner,
                text=f"Sugestão: substituir por  lag {lid_sugerido}",
                font=("Segoe UI", 12, "bold"),
                bg="#0f3460", fg="#7CFC00",
            ).pack(padx=30, pady=(4, 8))

            tk.Label(
                inner,
                text=(
                    f"• SIM  → todas as ocorrências de "
                    f"`lag {lid_atual}` e `lag-{lid_atual}` no .txt "
                    f"serão substituídas por `lag {lid_sugerido}` / "
                    f"`lag-{lid_sugerido}`.\n"
                    f"• NÃO → o LAG é salvo no .txt como "
                    f"`lag {lid_atual}` (sem alteração)."
                ),
                font=("Segoe UI", 9),
                bg="#0f3460", fg="#cccccc",
                wraplength=460, justify="left",
            ).pack(padx=30, pady=(2, 12))

            btn_frame = tk.Frame(inner, bg="#0f3460")
            btn_frame.pack(pady=(2, 16))

            resultado = {"v": False}

            def _sim():
                resultado["v"] = True
                dlg.destroy()

            def _nao():
                resultado["v"] = False
                dlg.destroy()

            tk.Button(
                btn_frame, text="SIM, substituir",
                font=("Segoe UI", 10, "bold"),
                bg="#28a745", fg="white",
                activebackground="#1e7e34",
                activeforeground="white",
                width=18, bd=2, relief="raised",
                command=_sim,
            ).pack(side=tk.LEFT, padx=10)

            tk.Button(
                btn_frame, text="NÃO, manter",
                font=("Segoe UI", 10, "bold"),
                bg="#dc3545", fg="white",
                activebackground="#a71d2a",
                activeforeground="white",
                width=18, bd=2, relief="raised",
                command=_nao,
            ).pack(side=tk.LEFT, padx=10)

            dlg.update_idletasks()
            w = dlg.winfo_reqwidth()
            h = dlg.winfo_reqheight()
            try:
                x = (self.root.winfo_x()
                     + (self.root.winfo_width() - w) // 2)
                y = (self.root.winfo_y()
                     + (self.root.winfo_height() - h) // 2)
                dlg.geometry(f"+{x}+{y}")
            except Exception:
                pass
            dlg.focus_set()
            dlg.grab_set()
            dlg.protocol("WM_DELETE_WINDOW", _nao)
            dlg.wait_window()
            resp_q.put(resultado["v"])

        try:
            self.root.after(0, _build)
            # Timeout defensivo: se a UI nao responder em 5 minutos,
            # assume "NAO" (mantem ID original) e libera a worker.
            return bool(resp_q.get(timeout=300))
        except Exception:
            return False

    def _swap_extrair_config_portas(self, cfg_text, portas,
                                    cfg_para=None, portas_para=None,
                                    para_router_id=None,
                                    vpls_alert_cb=None,
                                    de_h="", para_h=""):
        """Extrai TODA a configuração relacionada às portas DE da linha,
        organizada em hierarquia SR-OS, sem duplicação.

        O que é extraído:
          - `port <X>` para cada porta (1x cada).
          - `lag <N>` para cada LAG cujas portas membros estão na lista
            (1x cada).
          - Em `router Base`:
              * `interface "Y"` cujo body contém `port X` ou `port lag-N`
                ou `sap X:`/`sap lag-N:`.
              * `interface "Y"` correspondentes dentro dos protocolos
                ospf/ospf3/isis/mpls/rsvp/ldp/pim (com toda a hierarquia
                de containers — area, interface-parameters, etc.).
              * `static-route(-entry)` cujo `next-hop` casa o IP do peer
                (calculado a partir do `address X.X.X.X/N` de cada
                interface extraída, para máscaras /30 e /31).
              * `bgp` reduzido apenas ao(s) `neighbor <peer_ip>` (com
                hierarquia `group "..."` preservada quando aplicável).
          - Em cada serviço (vprn/ies/vpls/epipe/...):
              * Configuração geral (description, autonomous-system,
                router-id, vrf-import, vrf-export, route-distinguisher,
                auto-bind-tunnel, ecmp, service-name, etc.) — todos os
                filhos diretos do serviço que NÃO são interface/sap/bgp/
                static-route.
              * `interface "Y"` e `sap <X>:` que usam a porta/LAG.
              * Para vprn: `bgp` filtrado por peer-IP e
                `static-route(-entry)` filtrado por peer-IP, igual ao
                router Base.
          - Saída inicia com `exit all`, `configure`, e termina com
            `exit all`. Hierarquia em 4 espaços por nível.

        Retorna (texto_saida, lista_lags, num_blocos_extraidos,
                 texto_shutdown_de, texto_testes_de, texto_testes_para).
        """
        if not cfg_text or not portas:
            return ("", [], 0, "", "", "")

        lines = cfg_text.splitlines()
        n = len(lines)
        portas_set = set(portas)
        # Mapa DE\u2192PARA por posi\u00e7\u00e3o na lista (pareamento posicional
        # com a lista de portas_para vinda do xlsx). Se as listas n\u00e3o
        # tiverem o mesmo tamanho, faz-se best-effort: portas que
        # excederem o tamanho da outra lista n\u00e3o ser\u00e3o renomeadas.
        portas_para = portas_para or []
        de_to_para_port = {}
        for _i_p, _de_p in enumerate(portas):
            if _i_p < len(portas_para) and portas_para[_i_p]:
                de_to_para_port[_de_p] = portas_para[_i_p]

        # ── 1) Mapeia LAGs membros das portas ────────────────
        rx_lag_header = re.compile(r"^(\s*)lag\s+(\d+)\s*$")
        lags_alvo = []
        for i, s in enumerate(lines):
            m = rx_lag_header.match(s)
            if not m:
                continue
            stanza, _ = self._swap_extract_stanza(lines, i)
            for x in stanza[1:]:
                mp = re.match(r"^\s*port\s+(\S+)", x)
                if mp and mp.group(1) in portas_set:
                    lid = m.group(2)
                    if lid not in lags_alvo:
                        lags_alvo.append(lid)
                    break

        # ── 1.5) Conflito de LAG ID contra cfg_para ──────────
        # Para cada LAG alvo, verifica se o ID ja existe no PARA.
        # Se sim, sugere o proximo ID livre (lid+1, +2, ...) via
        # caixa de dialogo. Se o usuario aceitar, renomeia em
        # massa todas as ocorrencias `lag <old>` e `lag-<old>` no
        # cfg_text antes de prosseguir com a extracao.
        if cfg_para and lags_alvo:
            rx_lag_para = re.compile(r"^\s*lag\s+(\d+)\s*$")
            lags_no_para = set()
            for ln_p in cfg_para.splitlines():
                m_lp = rx_lag_para.match(ln_p)
                if m_lp:
                    lags_no_para.add(m_lp.group(1))

            rename_map = {}  # {old_lid: new_lid}
            usados = set(lags_no_para) | set(lags_alvo)
            for lid in list(lags_alvo):
                if lid not in lags_no_para:
                    continue
                novo = int(lid) + 1
                while str(novo) in usados:
                    novo += 1
                novo_str = str(novo)
                resp = self._swap_pedir_renomear_lag(lid, novo_str)
                if resp:
                    rename_map[lid] = novo_str
                    usados.add(novo_str)

            if rename_map:
                # Substitui em DUAS PASSADAS para evitar colisao
                # quando um novo ID coincide com um old ainda nao
                # processado (ex: rename {163:164, 164:165} ->
                # primeiro renomeia 163 para 164 e depois a segunda
                # passada renomearia esses 164 novos para 165).
                # Solucao: marca com placeholder unico e depois
                # converte para o final.
                novo_txt = cfg_text
                placeholders = {}
                for i_r, (_old, _new) in enumerate(rename_map.items()):
                    ph = f"__SWAPLAG_PH_{i_r}__"
                    placeholders[ph] = _new
                    novo_txt = re.sub(
                        rf"\blag\s+{re.escape(_old)}\b",
                        f"lag {ph}", novo_txt)
                    novo_txt = re.sub(
                        rf"\blag-{re.escape(_old)}\b",
                        f"lag-{ph}", novo_txt)
                for ph, _new in placeholders.items():
                    novo_txt = novo_txt.replace(
                        f"lag {ph}", f"lag {_new}")
                    novo_txt = novo_txt.replace(
                        f"lag-{ph}", f"lag-{_new}")
                cfg_text = novo_txt
                lines = cfg_text.splitlines()
                n = len(lines)
                lags_alvo = [
                    rename_map.get(l, l) for l in lags_alvo]

        # ── 2) Tokens de "uso" (matchers de linha) ───────────
        alvos_porta = list(portas) + [f"lag-{l}" for l in lags_alvo]
        rx_uso_port_loose = re.compile(
            r"^\s*port\s+(?:"
            + "|".join(re.escape(t) for t in alvos_porta)
            + r")\b"
        )
        rx_uso_sap_loose = re.compile(
            r"^\s*sap\s+(?:"
            + "|".join(re.escape(t) for t in alvos_porta)
            + r")(?::|\s|$)"
        )

        def stanza_usa_porta(stanza_lines):
            for x in stanza_lines:
                if rx_uso_port_loose.match(x) or rx_uso_sap_loose.match(x):
                    return True
                # Sat\u00e9lite: pode haver `port esat-N/...` ou
                # `sap esat-N/...` no body do servi\u00e7o/interface.
                if esat_sap_prefixes:
                    m_e = re.match(
                        r"^\s*(?:port|sap)\s+(esat-\S+)", x)
                    if m_e and m_e.group(1).startswith(
                            esat_sap_prefixes):
                        return True
            return False

        rx_sap_header = re.compile(r"^(\s*)sap\s+(\S+)")

        def sap_id_eh_alvo(sap_id):
            base = sap_id.split(":")[0]
            if base in alvos_porta:
                return True
            # SAPs em portas de satélite (esat-N/...) são alvo quando
            # o satélite N foi detectado nas descriptions das portas.
            if esat_sap_prefixes and base.startswith(esat_sap_prefixes):
                return True
            return False

        # ── 3) Coleta de stanzas básicas (port + lag) ────────
        port_stanzas = []
        seen_ports = set()
        for i, s in enumerate(lines):
            m = re.match(r"^(\s*)port\s+(\S+)\s*$", s)
            if not m:
                continue
            p = m.group(2)
            if p in portas_set and p not in seen_ports:
                stanza, _ = self._swap_extract_stanza(lines, i)
                port_stanzas.append((p, stanza))
                seen_ports.add(p)

        lag_stanzas = []
        seen_lags = set()
        for i, s in enumerate(lines):
            m = rx_lag_header.match(s)
            if not m:
                continue
            lid = m.group(2)
            if lid in lags_alvo and lid not in seen_lags:
                stanza, _ = self._swap_extract_stanza(lines, i)
                # Remove `administrative-key <N>` do LAG. Pode aparecer
                # como linha pr\u00f3pria OU inline ap\u00f3s `lacp active`
                # (ex.: "lacp active administrative-key 32768").
                stanza_filt = []
                for ln in stanza:
                    # Linha apenas com administrative-key → descarta
                    if re.match(r"^\s*administrative-key\b", ln):
                        continue
                    # Inline → remove o trecho `administrative-key <N>`
                    ln2 = re.sub(
                        r"\s+administrative-key\s+\S+", "", ln)
                    stanza_filt.append(ln2)
                lag_stanzas.append((lid, stanza_filt))
                seen_lags.add(lid)

        # ── 3.1) Detecta satélites (esat-N) nas descriptions ─
        # Se a description de uma porta extraída contiver "esat-N"
        # (ex.: "Ligacao AGG01.ULAMA-ESAT-1 | ..."), copiaremos toda a
        # configuração relacionada ao satélite N.
        esat_ids = []
        rx_esat_in_desc = re.compile(r"esat-(\d+)", re.IGNORECASE)
        for _p, _stz in port_stanzas:
            for ln in _stz:
                m_d = re.match(r'^\s*description\s+"([^"]*)"', ln)
                if not m_d:
                    continue
                for m_e in rx_esat_in_desc.finditer(m_d.group(1)):
                    eid = m_e.group(1)
                    if eid not in esat_ids:
                        esat_ids.append(eid)
        # SAPs em portas esat-N/... também são alvo (para arrastar
        # serviços associados ao satélite).
        esat_sap_prefixes = tuple(f"esat-{n}/" for n in esat_ids)

        # ── 3.2) Coleta portas f\u00edsicas dos sat\u00e9lites detectados ─
        # Para cada esat-N detectado, varre o cfg em busca de
        # `port esat-N/M/X` e adiciona ao port_stanzas (se ainda
        # n\u00e3o presente).
        if esat_ids:
            rx_port_esat_hdr = re.compile(
                r'^(\s*)port\s+(esat-(\d+)/\S+)\s*$')
            ports_ja = {p for p, _ in port_stanzas}
            for i, ln in enumerate(lines):
                m_pe = rx_port_esat_hdr.match(ln)
                if not m_pe:
                    continue
                eid = m_pe.group(3)
                if eid not in esat_ids:
                    continue
                pname = m_pe.group(2)
                if pname in ports_ja:
                    continue
                stz, _ = self._swap_extract_stanza(lines, i)
                port_stanzas.append((pname, stz))
                ports_ja.add(pname)

        # ── 4) Walker com pilha de contextos ─────────────────
        # Aceita `interface "X"` E `network-interface "X"` (VPRN CSC).
        rx_iface_header = re.compile(
            r'^(\s*)(?:network-)?interface\s+"([^"]+)"')
        rx_router_base = re.compile(r'^(\s*)router\s+Base\s*$')
        rx_service_header = re.compile(
            r'^(\s*)(vpls|vprn|ies|epipe|ipipe|apipe|fpipe|cpipe|'
            r'mirror-dest|pw-template)\s+(\S+)'
        )
        rx_proto_header = re.compile(
            # Aceita instance-id numerico (ex.: "ospf 0") OU em formato
            # IP/router-id (ex.: "ospf 10.26.40.73") OU sem instance.
            # Antes, '\d+' deixava de fora as instancias com IP, fazendo
            # com que o protocolo nao virasse contexto e as interfaces
            # OSPF dentro de VPRN nao fossem coletadas em proto_path_ifaces.
            r'^(\s*)(ospf3?|isis|mpls|rsvp|ldp|pim)(?:\s+\S+)?\s*$'
        )
        # Linhas que serão tratadas em pass dedicado de peer-IP — NÃO
        # devem entrar nos "extras" do serviço.
        rx_bgp_or_static_header = re.compile(
            r'^\s*(bgp|static-route|static-route-entry|static-routes)\b'
        )

        rb_interfaces = []
        servicos = []
        nomes_int_alvo = set()
        rb_bgp_raw_holder = []      # [stanza completa de bgp do router Base]
        rb_static_raw_holder = []   # [stanzas de static-* do router Base]

        contexto = []

        def _ctx_top_kind(*kinds):
            for c in reversed(contexto):
                if c["kind"] in kinds:
                    return c
            return None

        def _ctx_get_service():
            return _ctx_top_kind("service")

        i = 0
        while i < n:
            s = lines[i]
            if not s.strip():
                i += 1
                continue
            stripped = s.lstrip()
            # Pula coment\u00e1rios/echo/exit ANTES do pop de contexto. SR-OS
            # emite separadores `#----` e `echo "..."` na coluna 0, que
            # NUNCA representam fim de bloco — se o pop fosse aplicado
            # primeiro, a pilha de contexto (router Base, vprn N, ...)
            # seria zerada a cada coment\u00e1rio, perdendo o escopo dos
            # blocos seguintes (ex.: BGP do router Base).
            if (stripped.startswith("#") or stripped.startswith("echo ")
                    or stripped.startswith("exit")):
                i += 1
                continue
            ind = len(s) - len(s.lstrip(" "))
            while contexto and ind <= contexto[-1]["indent"]:
                contexto.pop()

            m_rb = rx_router_base.match(s)
            m_svc = rx_service_header.match(s)
            m_iface = rx_iface_header.match(s)
            m_proto = rx_proto_header.match(s)
            m_sap = rx_sap_header.match(s)

            if m_iface:
                nome = m_iface.group(2)
                stanza, end_idx = self._swap_extract_stanza(lines, i)
                if stanza_usa_porta(stanza[1:]):
                    nomes_int_alvo.add(nome)
                    pai = (contexto[-1] if contexto else None)
                    if pai and pai["kind"] == "router_base":
                        if not any(x[0] == nome for x in rb_interfaces):
                            rb_interfaces.append((nome, stanza))
                    elif pai and pai["kind"] == "service":
                        svc = pai["data"]
                        if not any(x[0] == nome for x in svc["ifaces"]):
                            svc["ifaces"].append((nome, stanza))
                            if svc not in servicos:
                                servicos.append(svc)
                i = end_idx
                continue

            if m_sap:
                sap_id = m_sap.group(2)
                stanza, end_idx = self._swap_extract_stanza(lines, i)
                if sap_id_eh_alvo(sap_id):
                    pai = _ctx_get_service()
                    if pai:
                        svc = pai["data"]
                        if not any(x[0] == sap_id for x in svc["saps"]):
                            svc["saps"].append((sap_id, stanza))
                            if svc not in servicos:
                                servicos.append(svc)
                # Consome a stanza do SAP (alvo ou não) para evitar que
                # caia no coletor de "extras" do serviço pai.
                i = end_idx
                continue

            if m_rb:
                contexto.append({"kind": "router_base",
                                 "indent": ind, "header": s, "data": None})
                i += 1
                continue

            if m_svc:
                svc_type = m_svc.group(2).lower()
                svc_id = m_svc.group(3)
                svc = next(
                    (x for x in servicos
                     if x["type"] == svc_type and x["id"] == svc_id),
                    None,
                )
                if svc is None:
                    svc = {"type": svc_type, "id": svc_id,
                           "header": s, "ifaces": [], "saps": [],
                           "extras_lines": [], "extras_stanzas": [],
                           "bgp_full": None, "statics_full": [],
                           "indent": ind}
                contexto.append({"kind": "service", "indent": ind,
                                 "header": s, "data": svc})
                i += 1
                continue

            if m_proto:
                proto_name = m_proto.group(2).lower()
                contexto.append({"kind": "protocol", "indent": ind,
                                 "header": s, "data": proto_name})
                i += 1
                continue

            # Coleta direta de bgp/static-* como filho direto de
            # serviço OU router Base (mais robusto que segunda passagem).
            pai_svc = _ctx_get_service()
            pai_rb = _ctx_top_kind("router_base")
            if rx_bgp_or_static_header.match(s):
                pai = pai_svc or pai_rb
                if pai and pai["indent"] + 4 == ind:
                    stanza, end_idx = self._swap_extract_stanza(lines, i)
                    strip_ln = s.lstrip()
                    if pai_svc:
                        svc = pai_svc["data"]
                        if (strip_ln == "bgp"
                                or strip_ln.startswith("bgp ")):
                            if svc["bgp_full"] is None:
                                svc["bgp_full"] = stanza
                        else:
                            svc["statics_full"].append(stanza)
                    else:
                        # router_base: armazena raw para filtragem
                        # posterior por peer-IP
                        if (strip_ln == "bgp"
                                or strip_ln.startswith("bgp ")):
                            rb_bgp_raw_holder.append(stanza)
                        else:
                            rb_static_raw_holder.append(stanza)
                    i = end_idx
                    continue

            # Filhos diretos de SERVIÇO que não são iface/sap/bgp/static:
            # coletamos como extras (config geral do serviço, ex.: VPRN).
            if (pai_svc and pai_svc["indent"] + 4 == ind
                    and not rx_bgp_or_static_header.match(s)):
                svc = pai_svc["data"]
                # Linha tem filhos? (próxima não-vazia com indent maior)
                j2 = i + 1
                while j2 < n and not lines[j2].strip():
                    j2 += 1
                if j2 < n and (
                        len(lines[j2]) - len(lines[j2].lstrip(" "))) > ind:
                    stanza, end_idx = self._swap_extract_stanza(lines, i)
                    # Dedup por header. NÃO adiciona o serviço a
                    # `servicos` aqui — só ifaces/saps relacionadas à
                    # porta marcam o serviço como relevante.
                    h = s.strip()
                    if not any(x[0].strip() == h
                               for x in svc["extras_stanzas"]):
                        svc["extras_stanzas"].append(stanza)
                    i = end_idx
                    continue
                else:
                    if s.strip() not in svc["extras_lines"]:
                        svc["extras_lines"].append(s.strip())
                    i += 1
                    continue

            # Container genérico — push só se essa linha tiver filhos
            j = i + 1
            while j < n and not lines[j].strip():
                j += 1
            if j < n:
                next_ind = len(lines[j]) - len(lines[j].lstrip(" "))
                if next_ind > ind:
                    contexto.append({"kind": "container", "indent": ind,
                                     "header": s, "data": None})
            i += 1

        # ── 5) Pass 2: interfaces dentro de protocolos ───────
        proto_path_ifaces = {}
        if nomes_int_alvo:
            contexto = []
            i = 0
            while i < n:
                s = lines[i]
                if not s.strip():
                    i += 1
                    continue
                stripped = s.lstrip()
                # Pula coment\u00e1rios/echo/exit ANTES do pop de contexto
                # (mesmo motivo do Pass 1: separadores SR-OS na coluna 0
                # n\u00e3o representam fim de bloco).
                if (stripped.startswith("#") or stripped.startswith("echo ")
                        or stripped.startswith("exit")):
                    i += 1
                    continue
                ind = len(s) - len(s.lstrip(" "))
                while contexto and ind <= contexto[-1]["indent"]:
                    contexto.pop()

                m_proto = rx_proto_header.match(s)
                m_iface = rx_iface_header.match(s)
                m_rb = rx_router_base.match(s)
                m_svc = rx_service_header.match(s)

                if m_iface:
                    nome = m_iface.group(2)
                    if nome in nomes_int_alvo:
                        idx_proto = -1
                        for k in range(len(contexto) - 1, -1, -1):
                            if contexto[k]["kind"] == "protocol":
                                idx_proto = k
                                break
                        if idx_proto >= 0:
                            # Encontra o parent (router_base ou service)
                            # ANTES do protocolo mais externo. Isso
                            # garante que path[0] = "router Base" ou
                            # "vprn N"/"ies N"/..., permitindo que o
                            # filtro de escopo (Pass 3 e emissao) saiba
                            # se o protocolo pertence ao router Base ou
                            # a um servi\u00e7o.
                            idx_parent = -1
                            for k in range(idx_proto - 1, -1, -1):
                                if contexto[k]["kind"] in (
                                        "router_base", "service"):
                                    idx_parent = k
                                    break
                            stanza, end_idx = self._swap_extract_stanza(
                                lines, i)
                            inicio = (idx_parent
                                      if idx_parent >= 0
                                      else idx_proto)
                            path = tuple(
                                c["header"].strip()
                                for c in contexto[inicio:]
                            )
                            lst = proto_path_ifaces.setdefault(path, [])
                            if not any(x[0] == nome for x in lst):
                                lst.append((nome, stanza))
                            i = end_idx
                            continue
                    _, end_idx = self._swap_extract_stanza(lines, i)
                    i = end_idx
                    continue

                if m_rb:
                    contexto.append({"kind": "router_base", "indent": ind,
                                     "header": s, "data": None})
                    i += 1
                    continue
                if m_svc:
                    contexto.append({"kind": "service", "indent": ind,
                                     "header": s, "data": None})
                    i += 1
                    continue
                if m_proto:
                    contexto.append({"kind": "protocol", "indent": ind,
                                     "header": s, "data": s.strip()})
                    i += 1
                    continue

                j = i + 1
                while j < n and not lines[j].strip():
                    j += 1
                if j < n:
                    next_ind = len(lines[j]) - len(lines[j].lstrip(" "))
                    if next_ind > ind:
                        contexto.append({"kind": "container", "indent": ind,
                                         "header": s, "data": None})
                i += 1

        # ── 6) Peer IPs a partir das interfaces extraídas ────
        def _peer_ips_de_stanza(stanza):
            """Devolve lista de IPs (str) para busca de BGP/static
            relacionada a esta interface: o próprio IP da interface
            + TODOS os endereços da mesma subrede (incluindo network
            e broadcast).

            Para /30 com IP .25 → retorna [.25, .24, .26, .27].
            Para /31 com IP .49 → retorna [.49, .48, .49] (deduplicado).
            Para /29 retorna 8 IPs. Para máscaras ≥ /24 limita a 256
            para evitar explosão.
            """
            ips = []
            for ln in stanza:
                m = re.search(
                    r"^\s*address\s+(\d+\.\d+\.\d+\.\d+)/(\d+)\b",
                    ln,
                )
                if not m:
                    # Tenta IPv6: `address <ipv6>/<plen>` (geralmente
                    # dentro de um sub-bloco `ipv6 { ... }`).
                    m6 = re.search(
                        r"^\s*address\s+([0-9a-fA-F:]+)/(\d+)\b",
                        ln,
                    )
                    if not m6 or ":" not in m6.group(1):
                        continue
                    try:
                        iface6 = ipaddress.ip_interface(
                            f"{m6.group(1)}/{m6.group(2)}")
                    except ValueError:
                        continue
                    ip6_str = str(iface6.ip)
                    if ip6_str not in ips:
                        ips.append(ip6_str)
                    net6 = iface6.network
                    # Para IPv6, limita a 256 enderecos para evitar
                    # explosao em prefixos curtos. /127 e /126 sao
                    # casos comuns em P2P.
                    if net6.num_addresses > 256:
                        continue
                    for h in net6:
                        s_h = str(h)
                        if s_h not in ips:
                            ips.append(s_h)
                    continue
                try:
                    iface = ipaddress.ip_interface(
                        f"{m.group(1)}/{m.group(2)}")
                except ValueError:
                    continue
                ip_str = str(iface.ip)
                if ip_str not in ips:
                    ips.append(ip_str)
                net = iface.network
                # Limita a 256 enderecos para evitar explosao em
                # prefixos curtos (/24 ou maiores). Verifica ANTES
                # de materializar a lista.
                if net.num_addresses > 256:
                    continue
                for h in net:
                    s_h = str(h)
                    if s_h not in ips:
                        ips.append(s_h)
            return ips

        # Peer IPs por escopo: 'rb' e ('vprn', svc_id)
        peers_por_escopo = {"rb": []}
        for nome, stanza in rb_interfaces:
            for ip in _peer_ips_de_stanza(stanza):
                if ip not in peers_por_escopo["rb"]:
                    peers_por_escopo["rb"].append(ip)
        # Tamb\u00e9m considera ifaces extra\u00eddas dos protocolos. Cada path
        # come\u00e7a com o cabe\u00e7alho do escopo ("router Base" ou
        # "vprn N"/"ies N"/...). Isolamos os peer-IPs no escopo correto:
        # protocolo dentro de vprn N -> escopo ("vprn", "N");
        # protocolo dentro de router Base -> escopo "rb"; servi\u00e7os
        # n\u00e3o-vprn (ies/vpls/...) seguem o tratamento direto j\u00e1 feito
        # abaixo (sem polui\u00e7\u00e3o cruzada com rb).
        rx_path_svc = re.compile(
            r'^(vprn|ies|vpls|epipe|ipipe|apipe|fpipe|cpipe)\s+(\S+)',
            re.IGNORECASE)
        for path, lst in proto_path_ifaces.items():
            head0 = path[0] if path else ""
            m_sc = rx_path_svc.match(head0.strip())
            if m_sc:
                tipo = m_sc.group(1).lower()
                sid = m_sc.group(2)
                escopo = (tipo, sid)
            else:
                escopo = "rb"
            if escopo not in peers_por_escopo:
                peers_por_escopo[escopo] = []
            for nome, stanza in lst:
                for ip in _peer_ips_de_stanza(stanza):
                    if ip not in peers_por_escopo[escopo]:
                        peers_por_escopo[escopo].append(ip)
        for svc in servicos:
            chave = (svc["type"], svc["id"])
            peers_por_escopo[chave] = []
            for nome, stanza in svc["ifaces"]:
                for ip in _peer_ips_de_stanza(stanza):
                    if ip not in peers_por_escopo[chave]:
                        peers_por_escopo[chave].append(ip)
                    # Servi\u00e7os IES (e demais n\u00e3o-VPRN) costumam ter
                    # static-route-entry NA router base referenciando o
                    # next-hop da interface. Adiciona esses peer-IPs ao
                    # escopo rb para que o filtro abaixo os capture.
                    if (svc["type"] != "vprn"
                            and ip not in peers_por_escopo["rb"]):
                        peers_por_escopo["rb"].append(ip)

        # ── 6.45) VPLS sem SAP correlacionada por SDP→PARA ────
        # Detecta servicos VPLS no cfg_de que NAO tem SAP nas
        # portas alvo (e portanto nao foram coletados acima),
        # mas que possuem `spoke-sdp X:M` onde o SDP `X` tem
        # `far-end <ip>` igual ao router-id do roteador PARA.
        # Para cada VPLS detectada, dispara `vpls_alert_cb` (que
        # mostra um dialogo chamativo na GUI). Se aceito, adiciona
        # o servico a `servicos` para que a etapa 6.5 (rVPLS)
        # cascade pegue automaticamente as interfaces VPRN/IES
        # ligadas a essa VPLS.
        if vpls_alert_cb and para_router_id:
            rx_sdp_hdr = re.compile(r'^\s*sdp\s+(\d+)\b')
            rx_far_end = re.compile(r'^\s*far-end\s+(\S+)')
            sdp_far_end = {}
            for i_s in range(n):
                m_sdp = rx_sdp_hdr.match(lines[i_s])
                if not m_sdp:
                    continue
                sid = m_sdp.group(1)
                if sid in sdp_far_end:
                    continue
                stz_s, _ = self._swap_extract_stanza(lines, i_s)
                for ln_b in stz_s[1:]:
                    m_fe = rx_far_end.match(ln_b)
                    if m_fe:
                        sdp_far_end[sid] = m_fe.group(1)
                        break
            sdp_alvos = {sid for sid, fe in sdp_far_end.items()
                         if fe == para_router_id}
            if sdp_alvos:
                rx_vpls_hdr = re.compile(r'^(\s*)vpls\s+(\S+)')
                rx_spoke = re.compile(
                    r'^\s*spoke-sdp\s+(\d+):\d+')
                ja_extr = {(s["type"], s["id"]) for s in servicos}
                for i_v in range(n):
                    ln_v = lines[i_v]
                    m_vh = rx_vpls_hdr.match(ln_v)
                    if not m_vh:
                        continue
                    v_ind = len(m_vh.group(1))
                    v_id = m_vh.group(2)
                    if ("vpls", v_id) in ja_extr:
                        continue
                    stz_v, _ = self._swap_extract_stanza(lines, i_v)
                    sdp_match = None
                    for ln_b in stz_v[1:]:
                        m_ss = rx_spoke.match(ln_b)
                        if m_ss and m_ss.group(1) in sdp_alvos:
                            sdp_match = m_ss.group(1)
                            break
                    if not sdp_match:
                        continue
                    # Extrai name (do header ou do body)
                    v_name = ""
                    m_n = re.search(
                        r'\bname\s+"([^"]+)"', ln_v)
                    if m_n:
                        v_name = m_n.group(1)
                    else:
                        for ln_b in stz_v[1:]:
                            m_n = re.search(
                                r'\bname\s+"([^"]+)"', ln_b)
                            if m_n:
                                v_name = m_n.group(1)
                                break
                    # Pergunta ao usuario via callback
                    try:
                        aceitar = bool(vpls_alert_cb(
                            v_id, v_name, sdp_match))
                    except Exception:
                        aceitar = False
                    if not aceitar:
                        continue
                    ja_extr.add(("vpls", v_id))
                    # Constroi o svc populando extras a partir do
                    # body bruto (a etapa 6.5 le novamente o body
                    # do cfg pelo header do svc para detectar
                    # allow-ip-int-bind).
                    svc_full = {
                        "type": "vpls", "id": v_id,
                        "header": ln_v,
                        "ifaces": [], "saps": [],
                        "extras_lines": [], "extras_stanzas": [],
                        "bgp_full": None, "statics_full": [],
                        "indent": v_ind,
                    }
                    base_filho = v_ind + 4
                    j_b = 1
                    while j_b < len(stz_v):
                        ln_b = stz_v[j_b]
                        if not ln_b.strip():
                            j_b += 1; continue
                        strip_b = ln_b.lstrip()
                        if (strip_b.startswith("exit")
                                or strip_b.startswith("#")
                                or strip_b.startswith("echo ")):
                            j_b += 1; continue
                        ind_b = (len(ln_b)
                                 - len(ln_b.lstrip(" ")))
                        if ind_b != base_filho:
                            j_b += 1; continue
                        # Tem filhos?
                        j2 = j_b + 1
                        while (j2 < len(stz_v)
                               and not stz_v[j2].strip()):
                            j2 += 1
                        tem_filhos = (
                            j2 < len(stz_v)
                            and (len(stz_v[j2])
                                 - len(stz_v[j2].lstrip(" ")))
                                 > ind_b)
                        if tem_filhos:
                            sub_st, sub_end = (
                                self._swap_extract_stanza(
                                    stz_v, j_b))
                            svc_full["extras_stanzas"].append(
                                sub_st)
                            j_b = sub_end
                        else:
                            svc_full["extras_lines"].append(
                                strip_b)
                            j_b += 1
                    servicos.append(svc_full)

        # ── 6.5) rVPLS: coleta interfaces IES/VPRN bound a VPLS ──
        # Para cada VPLS-alvo (que já está em `servicos`) detecta se
        # é uma rVPLS (`allow-ip-int-bind`) e captura seu `name "X"`.
        # Depois varre o cfg_text procurando interfaces IES/VPRN
        # cujo body contenha `vpls "X"` — essas interfaces sao
        # adicionadas ao servico IES/VPRN correspondente (criando-o
        # se ainda nao existir em `servicos`). Pula caso a interface
        # ja exista no cfg_para sob o mesmo servico (verificacao
        # textual).
        rvpls_names = []  # [(vpls_svc, nome_X)]
        rx_vpls_name = re.compile(r'\bname\s+"([^"]+)"')
        for svc in list(servicos):
            if svc["type"] != "vpls":
                continue
            m_nm = rx_vpls_name.search(svc.get("header", ""))
            if not m_nm:
                continue
            nome_x = m_nm.group(1)
            # Detecta `allow-ip-int-bind` no body real do VPLS.
            tem_bind = False
            hdr_str = svc.get("header", "").strip()
            for i_v, ln_v in enumerate(lines):
                if ln_v.strip() != hdr_str:
                    continue
                stz_v, _ = self._swap_extract_stanza(lines, i_v)
                for ln_b in stz_v[1:]:
                    if re.match(
                            r'^\s*allow-ip-int-bind\b', ln_b):
                        tem_bind = True
                        break
                break
            if tem_bind:
                rvpls_names.append((svc, nome_x))

        if rvpls_names:
            rx_iface_hdr_local = re.compile(
                r'^(\s*)(?:network-)?interface\s+"([^"]+)"')
            rx_vpls_bind_local = re.compile(
                r'^\s*vpls\s+"([^"]+)"')
            rx_svc_hdr_local = re.compile(
                r'^(\s*)(vprn|ies)\s+(\S+)')
            # Indexa cfg_para para checagem de "ja existe".
            para_lines = (cfg_para.splitlines()
                          if cfg_para else [])

            def _iface_existe_no_para(svc_type, svc_id, iface_nome):
                """Retorna True se interface "iface_nome" existir
                dentro do bloco <svc_type> <svc_id> em cfg_para."""
                if not para_lines:
                    return False
                rx_svc = re.compile(
                    rf'^(\s*){re.escape(svc_type)}\s+'
                    rf'{re.escape(svc_id)}\b')
                rx_if = re.compile(
                    rf'^\s*(?:network-)?interface\s+'
                    rf'"{re.escape(iface_nome)}"')
                jp = 0
                while jp < len(para_lines):
                    if rx_svc.match(para_lines[jp]):
                        stz_p, end_p = self._swap_extract_stanza(
                            para_lines, jp)
                        for ln_p in stz_p[1:]:
                            if rx_if.match(ln_p):
                                return True
                        jp = end_p
                        continue
                    jp += 1
                return False

            # Varre cfg_text procurando interfaces que fazem bind
            # com algum dos nomes de rVPLS coletados.
            nomes_alvo = {nm: vpls_svc
                          for vpls_svc, nm in rvpls_names}
            i_v = 0
            ctx_svc = None  # (type, id, header, indent) atual
            while i_v < n:
                ln_v = lines[i_v]
                if not ln_v.strip():
                    i_v += 1
                    continue
                stripped_v = ln_v.lstrip()
                if (stripped_v.startswith("#")
                        or stripped_v.startswith("echo ")
                        or stripped_v.startswith("exit")):
                    i_v += 1
                    continue
                ind_v = len(ln_v) - len(ln_v.lstrip(" "))
                if ctx_svc and ind_v <= ctx_svc[3]:
                    ctx_svc = None
                m_svc_hdr = rx_svc_hdr_local.match(ln_v)
                if m_svc_hdr:
                    ctx_svc = (
                        m_svc_hdr.group(2).lower(),
                        m_svc_hdr.group(3),
                        ln_v,
                        ind_v,
                    )
                    i_v += 1
                    continue
                m_if = rx_iface_hdr_local.match(ln_v)
                if m_if and ctx_svc:
                    iface_nome = m_if.group(2)
                    stz_if, end_if = self._swap_extract_stanza(
                        lines, i_v)
                    # procura `vpls "X"` direto no body
                    bind_nome = None
                    for ln_b in stz_if[1:]:
                        m_b = rx_vpls_bind_local.match(ln_b)
                        if m_b:
                            bind_nome = m_b.group(1)
                            break
                    if bind_nome and bind_nome in nomes_alvo:
                        svc_t, svc_id = ctx_svc[0], ctx_svc[1]
                        # Pula se ja existir no PARA
                        if not _iface_existe_no_para(
                                svc_t, svc_id, iface_nome):
                            # Encontra/cria o servico em `servicos`
                            svc_dest = next(
                                (x for x in servicos
                                 if x["type"] == svc_t
                                 and x["id"] == svc_id),
                                None,
                            )
                            if svc_dest is None:
                                svc_dest = {
                                    "type": svc_t, "id": svc_id,
                                    "header": ctx_svc[2],
                                    "ifaces": [], "saps": [],
                                    "extras_lines": [],
                                    "extras_stanzas": [],
                                    "bgp_full": None,
                                    "statics_full": [],
                                    "indent": ctx_svc[3],
                                }
                                servicos.append(svc_dest)
                            if not any(x[0] == iface_nome
                                       for x in svc_dest["ifaces"]):
                                svc_dest["ifaces"].append(
                                    (iface_nome, stz_if))
                                # Atualiza peers_por_escopo para
                                # capturar bgp/static-route do
                                # peer-IP dessa nova interface.
                                chave_d = (svc_t, svc_id)
                                if chave_d not in peers_por_escopo:
                                    peers_por_escopo[chave_d] = []
                                for ip_d in _peer_ips_de_stanza(
                                        stz_if):
                                    if (ip_d not in
                                            peers_por_escopo[
                                                chave_d]):
                                        peers_por_escopo[
                                            chave_d].append(ip_d)
                                    if (svc_t != "vprn"
                                            and ip_d not in
                                            peers_por_escopo["rb"]):
                                        peers_por_escopo[
                                            "rb"].append(ip_d)
                                nomes_int_alvo.add(iface_nome)
                    i_v = end_if
                    continue
                i_v += 1

        # ── 7) Pass 3: BGP e static-route por peer-IP ────────
        # Para cada escopo, procura a stanza pai (router Base ou vprn N)
        # e dentro dela procura `bgp`, `static-route-entry`, `static-route`,
        # `static-routes` filtrando por peer-IP.
        rb_bgp_neighbors = []   # [(group_header_or_None, neighbor_stanza)]
        rb_static_routes = []   # [stanza]
        svc_bgp_neighbors = {}  # (type,id) → list
        svc_static_routes = {}  # (type,id) → list

        def _filtra_bgp(bgp_stanza, peer_ips):
            """Devolve lista de tuplas (group_header_or_None, neighbor_stanza)
            casando neighbor <ip> OU `local-address <ip>` (ipv4/ipv6)
            dentro do body do neighbor — `local-address` corresponde
            ao proprio IP da interface do escopo."""
            result = []
            if not peer_ips:
                return result
            ips_set = set(peer_ips)
            blines = bgp_stanza
            base_ind = len(blines[0]) - len(blines[0].lstrip(" "))
            j = 1
            cur_group_header = None
            cur_group_indent = None
            while j < len(blines):
                ln = blines[j]
                if not ln.strip():
                    j += 1
                    continue
                ind_ln = len(ln) - len(ln.lstrip(" "))
                if cur_group_indent is not None and ind_ln <= cur_group_indent:
                    cur_group_header = None
                    cur_group_indent = None
                strip_ln = ln.lstrip()
                if strip_ln.startswith("group "):
                    cur_group_header = ln
                    cur_group_indent = ind_ln
                    j += 1
                    continue
                m_nb = re.match(
                    r"^\s*neighbor\s+(\S+)", ln)
                if m_nb:
                    # Pode ser linha única OU stanza com filhos
                    j2 = j + 1
                    while (j2 < len(blines) and not blines[j2].strip()):
                        j2 += 1
                    tem_filhos = (
                        j2 < len(blines)
                        and (len(blines[j2])
                             - len(blines[j2].lstrip(" "))) > ind_ln)
                    if tem_filhos:
                        st, end = self._swap_extract_stanza(blines, j)
                    else:
                        st = [ln]
                        end = j + 1
                    casa = m_nb.group(1) in ips_set
                    if not casa:
                        # Inspeciona body por `local-address <ip>`
                        # (o ip da interface = local-address do
                        # neighbor, valido p/ ipv4 e ipv6).
                        for ln_b in st[1:] if tem_filhos else []:
                            m_la = re.match(
                                r'^\s*local-address\s+(\S+)', ln_b)
                            if m_la and m_la.group(1) in ips_set:
                                casa = True
                                break
                    if casa:
                        result.append((cur_group_header, st))
                    j = end
                    continue
                j += 1
            return result

        def _filtra_static(stanza_pai, peer_ips):
            """Procura dentro de stanza_pai por static-route(-entry)/
            static-routes/route que tenham next-hop nos peer-IPs."""
            result = []
            if not peer_ips:
                return result
            ips_set = set(peer_ips)
            j = 1
            while j < len(stanza_pai):
                ln = stanza_pai[j]
                if not ln.strip():
                    j += 1
                    continue
                strip_ln = ln.lstrip()
                # Caso 1: `static-route <prefix> next-hop <ip>` (linha única)
                m1 = re.match(
                    r"^\s*static-route\s+\S+\s+next-hop\s+(\S+)",
                    ln)
                if m1 and m1.group(1) in ips_set:
                    result.append([ln])
                    j += 1
                    continue
                # Caso 2: `static-route-entry <prefix>` com bloco
                if (strip_ln.startswith("static-route-entry ")
                        or strip_ln.startswith("static-route ")):
                    st, end = self._swap_extract_stanza(stanza_pai, j)
                    body = "\n".join(st)
                    if any(re.search(
                            rf"next-hop\s+{re.escape(ip)}\b", body)
                           for ip in peer_ips):
                        result.append(st)
                    j = end
                    continue
                # Caso 3: `static-routes` (block) com `route <prefix> next-hop <ip>`
                if strip_ln.startswith("static-routes"):
                    st, end = self._swap_extract_stanza(stanza_pai, j)
                    base_ind_sr = len(st[0]) - len(st[0].lstrip(" "))
                    # Para cada `route ... next-hop <ip>` casando peer-IP,
                    # extrai a stanza COMPLETA (pode ter no shutdown,
                    # preference, metric, etc. como filhos).
                    rotas_filt = []
                    k = 1
                    while k < len(st):
                        ln2 = st[k]
                        if not ln2.strip():
                            k += 1
                            continue
                        m = re.match(
                            r"^\s*route\s+\S+\s+next-hop\s+(\S+)",
                            ln2)
                        if m and m.group(1) in ips_set:
                            sub, sub_end = self._swap_extract_stanza(st, k)
                            rotas_filt.extend(sub)
                            k = sub_end
                            continue
                        k += 1
                    if rotas_filt:
                        result.append(
                            [st[0]] + rotas_filt
                            + [" " * base_ind_sr + "exit"])
                    j = end
                    continue
                j += 1
            return result

        # Acha o router Base completo e cada vprn completo.
        # IMPORTANTE: SR-OS dump pode ter MULTIPLOS blocos `router Base`
        # (um por se\u00e7\u00e3o: Interface, OSPF, BGP, Static Route, Filter,
        # etc.). Coleta TODOS para garantir que static-route-entry,
        # bgp e demais filhos diretos sejam achados.
        rb_full = None
        rb_blocks = []  # lista de stanzas de cada `router Base`
        svc_full = {}   # (type,id) → stanza UNIFICADA (m\u00faltiplos
                        # blocos do mesmo svc no dump s\u00e3o mesclados
                        # em um \u00fanico body, na ordem em que aparecem,
                        # para que description, vrf-target, bgp,
                        # interfaces, mvpn, etc., sejam todos vis\u00edveis
                        # ao walker que preserva a ordem natural).
        svc_blocks = {}  # (type,id) → list[stanza]
        for i, s in enumerate(lines):
            if rx_router_base.match(s):
                st_rb, _ = self._swap_extract_stanza(lines, i)
                rb_blocks.append(st_rb)
                rb_full = st_rb  # mant\u00e9m \u00faltimo para compat
            m_svc = rx_service_header.match(s)
            if m_svc:
                t = m_svc.group(2).lower()
                sid = m_svc.group(3)
                st_svc, _ = self._swap_extract_stanza(lines, i)
                svc_blocks.setdefault((t, sid), []).append(st_svc)

        # Mescla todos os blocos do mesmo svc num \u00fanico body,
        # preservando a ordem original. Header = header do primeiro
        # bloco; corpo = uni\u00e3o dos corpos (sem o `exit` final de
        # cada um) + um `exit` final.
        for chave_svc, blks in svc_blocks.items():
            if not blks:
                continue
            base_ind_b = (
                len(blks[0][0]) - len(blks[0][0].lstrip(" ")))
            unified = [blks[0][0]]
            seen_simple = set()  # dedup linhas simples (description, etc.)
            for blk in blks:
                # corpo (descarta ultimo `exit` no nivel do header)
                k = 1
                end_k = len(blk)
                while end_k > 1:
                    last = blk[end_k - 1]
                    if last.strip() == "":
                        end_k -= 1
                        continue
                    ind_last = len(last) - len(last.lstrip(" "))
                    if (ind_last == base_ind_b
                            and last.lstrip().startswith("exit")):
                        end_k -= 1
                    break
                while k < end_k:
                    ln_b = blk[k]
                    if not ln_b.strip():
                        k += 1
                        continue
                    ind_ln_b = len(ln_b) - len(ln_b.lstrip(" "))
                    if ind_ln_b == base_ind_b + 4:
                        # filho direto: extrai stanza completa
                        sub_b, sub_end_b = (
                            self._swap_extract_stanza(blk, k))
                        # checa se eh stanza ou linha simples
                        eh_st = (sub_end_b - k) > 1
                        if not eh_st:
                            # linha simples: dedup
                            key_s = ln_b.strip()
                            if key_s not in seen_simple:
                                unified.append(ln_b)
                                seen_simple.add(key_s)
                            k = sub_end_b
                            continue
                        # stanza: dedup por header
                        hdr_b = sub_b[0].strip()
                        ja_existe = False
                        for ex_ln in unified[1:]:
                            if ex_ln.strip() == hdr_b:
                                ja_existe = True
                                break
                        if not ja_existe:
                            unified.extend(sub_b)
                        k = sub_end_b
                        continue
                    k += 1
            # exit final no nivel do header
            unified.append(" " * base_ind_b + "exit")
            svc_full[chave_svc] = unified

        # Re-escaneia TODOS os blocos `router Base` para coletar
        # static-route(-entry|s) e bgp como filhos diretos. Isso \u00e9
        # mais robusto do que depender s\u00f3 do walker (que pode perder
        # blocos quando h\u00e1 m\u00faltiplos `router Base`).
        rb_static_raw_holder = []
        rb_bgp_raw_holder_full = list(rb_bgp_raw_holder)
        for st_rb in rb_blocks:
            base_ind_rb = (
                len(st_rb[0]) - len(st_rb[0].lstrip(" ")))
            kk = 1
            while kk < len(st_rb):
                ln = st_rb[kk]
                if not ln.strip():
                    kk += 1
                    continue
                ind_ln = len(ln) - len(ln.lstrip(" "))
                if ind_ln != base_ind_rb + 4:
                    kk += 1
                    continue
                strip_ln = ln.lstrip()
                if (strip_ln.startswith("static-route-entry ")
                        or strip_ln.startswith("static-routes")
                        or strip_ln.startswith("static-route ")):
                    sub, sub_end = self._swap_extract_stanza(st_rb, kk)
                    rb_static_raw_holder.append(sub)
                    kk = sub_end
                    continue
                if (strip_ln == "bgp"
                        or strip_ln.startswith("bgp ")):
                    sub, sub_end = self._swap_extract_stanza(st_rb, kk)
                    if sub not in rb_bgp_raw_holder_full:
                        rb_bgp_raw_holder_full.append(sub)
                    kk = sub_end
                    continue
                kk += 1
        rb_bgp_raw_holder = rb_bgp_raw_holder_full

        # ── 7.b) Conjuntos de groups BGP existentes no PARA ─
        # Para cada escopo (router base e cada vprn), descobre quais
        # nomes de `group "X"` j\u00e1 existem em cfg_para. Quando o group
        # existe → emite somente o neighbor (sem repetir os
        # par\u00e2metros do group). Quando n\u00e3o existe → emite o group
        # completo (com peer-as, family, auth-key, etc.).
        bgp_grupos_no_para = {"rb": set()}

        def _coleta_grupos_bgp_em(stanza_pai, ind_filho_bgp):
            """Procura `bgp` filho direto da stanza_pai (ind_filho_bgp
            \u00e9 o indent dos filhos diretos) e devolve os nomes de
            `group "X"` declarados nele."""
            grupos = set()
            kk = 1
            base_ind_pai = (
                len(stanza_pai[0]) - len(stanza_pai[0].lstrip(" "))
                if stanza_pai else 0)
            while kk < len(stanza_pai):
                ln = stanza_pai[kk]
                if not ln.strip():
                    kk += 1
                    continue
                ind_ln = len(ln) - len(ln.lstrip(" "))
                strip_ln = ln.lstrip()
                if (ind_ln == ind_filho_bgp
                        and (strip_ln == "bgp"
                             or strip_ln.startswith("bgp "))):
                    bgp_st, end = self._swap_extract_stanza(
                        stanza_pai, kk)
                    for ln2 in bgp_st:
                        m_g = re.match(
                            r'^\s*group\s+"([^"]+)"', ln2)
                        if m_g:
                            grupos.add(m_g.group(1))
                    kk = end
                    continue
                kk += 1
            return grupos

        if cfg_para:
            p_lines = cfg_para.splitlines()
            # router Base no PARA: pode ter v\u00e1rios blocos
            for ii_p, ll_p in enumerate(p_lines):
                if rx_router_base.match(ll_p):
                    rb_st_p, _ = self._swap_extract_stanza(
                        p_lines, ii_p)
                    base_ind_rb_p = (
                        len(rb_st_p[0])
                        - len(rb_st_p[0].lstrip(" ")))
                    bgp_grupos_no_para["rb"].update(
                        _coleta_grupos_bgp_em(
                            rb_st_p, base_ind_rb_p + 4))
            # cada servi\u00e7o (vprn) no PARA
            for ii_p, ll_p in enumerate(p_lines):
                m_svc_p = rx_service_header.match(ll_p)
                if not m_svc_p:
                    continue
                t_p = m_svc_p.group(2).lower()
                sid_p = m_svc_p.group(3)
                if t_p != "vprn":
                    continue
                svc_st_p, _ = self._swap_extract_stanza(p_lines, ii_p)
                base_ind_svc_p = (
                    len(svc_st_p[0])
                    - len(svc_st_p[0].lstrip(" ")))
                chave_p = ("vprn", sid_p)
                bgp_grupos_no_para[chave_p] = _coleta_grupos_bgp_em(
                    svc_st_p, base_ind_svc_p + 4)

        # Router Base: filtra bgp por peer-IP e static por peer-IP a
        # partir do que foi coletado no walker.
        if peers_por_escopo.get("rb"):
            for bgp_st in rb_bgp_raw_holder:
                rb_bgp_neighbors.extend(
                    _filtra_bgp(bgp_st, peers_por_escopo["rb"]))

        # Router Base: filtra static-route-entry por peer-IP — só
        # traz rotas cujo `next-hop` esteja na MESMA SUBREDE de
        # alguma interface do escopo (ex.: IES com 200.243.225.1/31
        # casa next-hop 200.243.225.0). Não é para copiar tudo.
        rb_peers = peers_por_escopo.get("rb", [])
        if rb_peers:
            for st_raw in rb_static_raw_holder:
                if not st_raw:
                    continue
                pseudo_pai = ["pseudo"] + st_raw
                for st_filt in _filtra_static(pseudo_pai, rb_peers):
                    if st_filt not in rb_static_routes:
                        rb_static_routes.append(st_filt)

        # FALLBACK GLOBAL: alguns dumps SR-OS emitem a se\u00e7\u00e3o
        # "Static Route Configuration" SEM um header `router Base`
        # vis\u00edvel envolvendo as `static-route-entry` (ou com
        # separadores `exit all` / `configure` que quebram o
        # extract_stanza). Aqui varremos TODO o cfg em busca de
        # qualquer `static-route-entry` ou `static-routes` que tenha
        # ficado de fora — APLICANDO o mesmo filtro por peer-IP
        # (next-hop tem que estar na mesma subrede da interface).
        # IMPORTANTE: NAO interrompe a varredura ao achar uma entry
        # que casa o peer-IP — o mesmo peer pode estar referenciado
        # em multiplos `static-route-entry <prefix>` (inclusive com
        # o MESMO prefix em blocos separados, sintaxe SR-OS valida).
        rx_sr_any = re.compile(
            r"^(\s*)(static-route-entry\s+\S+|static-routes|"
            r"static-route\s+\S+)\s*$")
        i_g = 0
        while i_g < n:
            ln = lines[i_g]
            m_sr = rx_sr_any.match(ln)
            if not m_sr:
                i_g += 1
                continue
            sub, sub_end = self._swap_extract_stanza(lines, i_g)
            if sub and rb_peers:
                pseudo_pai = ["pseudo"] + sub
                for st_filt in _filtra_static(pseudo_pai, rb_peers):
                    # Dedup por CONTEUDO completo (nao por header):
                    # mantem multiplos blocos com mesmo prefix se o
                    # body diferir (ex.: next-hop diferente).
                    if st_filt not in rb_static_routes:
                        rb_static_routes.append(st_filt)
            i_g = sub_end

        # VPRN: filtra bgp_full e statics_full pelo peer-IP da
        # interface relacionada — só traz neighbors/rotas que casam
        # endereço da subnet da iface extraída. Exceção: a rota default
        # `static-route-entry 0.0.0.0/0` (geralmente black-hole) é
        # SEMPRE copiada, mesmo sem peer-IP casado.
        for svc in servicos:
            chave = (svc["type"], svc["id"])
            if svc["type"] != "vprn":
                continue
            peers = peers_por_escopo.get(chave, [])
            if svc.get("bgp_full") and peers:
                nbs = _filtra_bgp(svc["bgp_full"], peers)
                if nbs:
                    svc_bgp_neighbors[chave] = nbs
            if svc.get("statics_full"):
                srs_filt = []
                for st in svc["statics_full"]:
                    # Sempre inclui as rotas default 0.0.0.0/0 (IPv4)
                    # e ::/0 (IPv6) — independente de peer-IP. Mas
                    # se o cfg_para ja tiver a vprn com a mesma
                    # static-route-entry black-hole, ignora.
                    m_def = re.match(
                        r"^\s*static-route-entry\s+"
                        r"(0\.0\.0\.0/0|::/0)\b",
                        st[0]) if st else None
                    if m_def:
                        prefix = m_def.group(1)
                        # so consideramos black-hole — outras default
                        # routes (next-hop) seguem regra de peer-IP.
                        eh_blackhole = any(
                            re.match(r'^\s*black-hole\b', ln)
                            for ln in st[1:])
                        if eh_blackhole:
                            if not self._swap_vprn_tem_blackhole_no_para(
                                    cfg_para, svc["id"], prefix):
                                srs_filt.append(st)
                            continue
                        srs_filt.append(st)
                        continue
                    if peers:
                        pseudo_pai = ["pseudo"] + st
                        srs_filt.extend(
                            _filtra_static(pseudo_pai, peers))
                if srs_filt:
                    svc_static_routes[chave] = srs_filt

        # ── 7.1) Coleta de blocos `system` para satélites ────
        # Para cada esat_id detectado nas descriptions das portas,
        # extrai do config completo:
        #   - system { software-repository "esat-N" { ... } }
        #   - system { satellite { eth-sat N create { ... } } }
        #   - system { port-topology { port X to esat-N/... create } }
        esat_sw_repos = {}      # esat_id → stanza software-repository
        esat_eth_sats = {}      # esat_id → stanza eth-sat N
        esat_port_topo = {}     # esat_id → list de linhas port ... to esat-N/...

        if esat_ids:
            rx_system_hdr = re.compile(r"^(\s*)system\s*$")
            rx_satellite_hdr = re.compile(r"^(\s*)satellite\s*$")
            rx_port_topo_hdr = re.compile(r"^(\s*)port-topology\s*$")
            i_sys = 0
            while i_sys < n:
                if not rx_system_hdr.match(lines[i_sys]):
                    i_sys += 1
                    continue
                sys_stanza, sys_end = self._swap_extract_stanza(
                    lines, i_sys)
                # Itera filhos diretos do system
                k = 1
                while k < len(sys_stanza):
                    ln = sys_stanza[k]
                    if not ln.strip():
                        k += 1
                        continue
                    strip_ln = ln.lstrip()
                    # software-repository "esat-N" create
                    m_sr = re.match(
                        r'^\s*software-repository\s+"esat-(\d+)"',
                        ln)
                    if m_sr:
                        eid = m_sr.group(1)
                        sub, sub_end = self._swap_extract_stanza(
                            sys_stanza, k)
                        if eid in esat_ids and eid not in esat_sw_repos:
                            esat_sw_repos[eid] = sub
                        k = sub_end
                        continue
                    # satellite { eth-sat N ... } — pode aparecer
                    # MULTIPLAS vezes (ex.: bloco 'phase 1' com create
                    # e 'phase 2' so com port-maps). Mergeia tudo.
                    if rx_satellite_hdr.match(ln):
                        sat_st, sat_end = self._swap_extract_stanza(
                            sys_stanza, k)
                        kk = 1
                        while kk < len(sat_st):
                            ln2 = sat_st[kk]
                            if not ln2.strip():
                                kk += 1
                                continue
                            m_es = re.match(
                                r"^\s*eth-sat\s+(\d+)\b", ln2)
                            if m_es:
                                eid = m_es.group(1)
                                sub2, sub2_end = (
                                    self._swap_extract_stanza(
                                        sat_st, kk))
                                if eid in esat_ids:
                                    if eid not in esat_eth_sats:
                                        esat_eth_sats[eid] = sub2
                                    else:
                                        # Mergeia: adiciona linhas novas
                                        # (port-map etc.) antes do exit.
                                        existente = esat_eth_sats[eid]
                                        # Set de linhas ja presentes
                                        ja = {l.strip()
                                              for l in existente}
                                        novas = []
                                        for l2 in sub2[1:]:
                                            ls = l2.strip()
                                            if (ls and ls != "exit"
                                                    and ls not in ja):
                                                novas.append(l2)
                                                ja.add(ls)
                                        if novas:
                                            # Insere antes do ultimo exit
                                            if (existente
                                                    and existente[-1]
                                                    .lstrip().startswith(
                                                        "exit")):
                                                esat_eth_sats[eid] = (
                                                    existente[:-1]
                                                    + novas
                                                    + [existente[-1]])
                                            else:
                                                esat_eth_sats[eid] = (
                                                    existente + novas)
                                kk = sub2_end
                                continue
                            kk += 1
                        k = sat_end
                        continue
                    # port-topology { port X to esat-N/... }
                    if rx_port_topo_hdr.match(ln):
                        pt_st, pt_end = self._swap_extract_stanza(
                            sys_stanza, k)
                        for ln3 in pt_st[1:]:
                            m_pt = re.match(
                                r"^\s*port\s+\S+\s+to\s+esat-(\d+)/",
                                ln3)
                            if m_pt and m_pt.group(1) in esat_ids:
                                eid = m_pt.group(1)
                                lst = esat_port_topo.setdefault(
                                    eid, [])
                                if ln3.strip() not in [
                                        x.strip() for x in lst]:
                                    lst.append(ln3)
                        k = pt_end
                        continue
                    k += 1
                i_sys = sys_end

        # \u2500\u2500 7.2) Filtra esat ports: manter apenas as que t\u00eam
        # configura\u00e7\u00e3o real (description) OU s\u00e3o referenciadas como
        # `port esat-...`/`sap esat-...` em alguma interface/servi\u00e7o
        # extra\u00eddo. \u201cport esat\u201d sem corpo \u00e9 descartado.
        esat_usados = set()
        # Refer\u00eancias em rb_interfaces e protocolos
        for _nm, _stz in rb_interfaces:
            for _ln in _stz:
                m_pe2 = re.match(
                    r"^\s*(?:port|sap)\s+(esat-\S+)", _ln)
                if m_pe2:
                    esat_usados.add(m_pe2.group(1).split(":")[0])
        for _path, _lst in proto_path_ifaces.items():
            for _nm, _stz in _lst:
                for _ln in _stz:
                    m_pe2 = re.match(
                        r"^\s*(?:port|sap)\s+(esat-\S+)", _ln)
                    if m_pe2:
                        esat_usados.add(
                            m_pe2.group(1).split(":")[0])
        # Refer\u00eancias em servi\u00e7os (extras + ifaces + saps)
        for _svc in servicos:
            for _src in (
                    list(_svc.get("extras_lines") or []),
                    *(list(s) for s in (
                        _svc.get("extras_stanzas") or [])),
                    *(list(s) for _, s in (
                        _svc.get("ifaces") or [])),
                    *(list(s) for _, s in (
                        _svc.get("saps") or []))):
                for _ln in _src:
                    m_pe2 = re.match(
                        r"^\s*(?:port|sap)\s+(esat-\S+)", _ln)
                    if m_pe2:
                        esat_usados.add(
                            m_pe2.group(1).split(":")[0])

        def _esat_tem_config(stz):
            """Considera que a porta esat tem configuração real apenas
            se houver alguma linha significativa além de
            `shutdown`/`no shutdown`/`exit` e do bloco `ethernet`
            VAZIO. Portas no padrão:
                port esat-X/Y/Z
                    shutdown
                    ethernet
                    exit
                exit
            são consideradas SEM configuração (devem ser ignoradas).
            """
            i = 1
            n_st = len(stz)
            while i < n_st:
                s = stz[i].strip()
                if (not s or s == "exit" or s == "shutdown"
                        or s == "no shutdown"):
                    i += 1
                    continue
                if s == "ethernet":
                    # Verifica se o bloco ethernet tem filhos
                    # significativos (qualquer coisa que não seja
                    # apenas `exit`).
                    j = i + 1
                    eth_tem_filho = False
                    while j < n_st:
                        s2 = stz[j].strip()
                        if not s2:
                            j += 1
                            continue
                        if s2 == "exit":
                            j += 1
                            break
                        eth_tem_filho = True
                        j += 1
                    if eth_tem_filho:
                        return True
                    i = j
                    continue
                # Qualquer outra linha (description, mtu, lag, etc.)
                # conta como configuração real.
                return True
            return False

        port_stanzas_filt = []
        for _p, _stz in port_stanzas:
            if _p.startswith("esat-"):
                if _p in esat_usados or _esat_tem_config(_stz):
                    port_stanzas_filt.append((_p, _stz))
                continue
            port_stanzas_filt.append((_p, _stz))
        port_stanzas = port_stanzas_filt

        # \u2500\u2500 7.3) Transforma stanza de porta f\u00edsica n\u00e3o-esat:
        # \u2022 remove linhas `no otu`
        # \u2022 insere `shutdown` logo ap\u00f3s o header (antes da
        #   description)
        # \u2022 renomeia o header da porta DE para a porta PARA
        #   correspondente (mapeamento posicional do xlsx)
        def _transforma_porta_fisica(p_de, stz):
            if not stz:
                return p_de, stz
            p_para = de_to_para_port.get(p_de, p_de)
            hdr = stz[0]
            ind = len(hdr) - len(hdr.lstrip(" "))
            base = " " * ind
            inner = " " * (ind + 4)
            new_hdr = f"{base}port {p_para}"
            # Detecta se a description aponta para um ESAT
            # (ex.: "Ligacao AGG01.ULAMA-ESAT-1 | ..."). Se sim,
            # remove `no shutdown` para deixar a porta uplink em
            # shutdown ate o satelite ser ativado.
            aponta_esat = False
            for ln in stz[1:]:
                m_d = re.match(r'^\s*description\s+"?(.*?)"?\s*$', ln)
                if m_d and re.search(
                        r"esat[-_/\s]?\d", m_d.group(1), re.IGNORECASE):
                    aponta_esat = True
                    break
            body = []
            for ln in stz[1:]:
                if ln.strip() == "no otu":
                    continue
                if aponta_esat and ln.strip() == "no shutdown":
                    continue
                body.append(ln)
            # Insere `shutdown` (com indent inner) logo ap\u00f3s o header
            new_stz = [new_hdr, inner + "shutdown"] + body
            return p_para, new_stz

        port_stanzas_xform = []
        for _p, _stz in port_stanzas:
            if _p.startswith("esat-"):
                # Esat ports: removemos `no otu` mas N\u00c3O renomeamos
                # nem inserimos shutdown (j\u00e1 v\u00eam no padr\u00e3o esperado).
                _stz_clean = [_stz[0]] + [
                    ln for ln in _stz[1:] if ln.strip() != "no otu"]
                port_stanzas_xform.append((_p, _stz_clean))
            else:
                port_stanzas_xform.append(
                    _transforma_porta_fisica(_p, _stz))
        port_stanzas = port_stanzas_xform

        # ── 8) Renderiza em hierarquia SR-OS ─────────────────
        def reindent(stanza, base):
            if not stanza:
                return []
            header = stanza[0]
            orig = len(header) - len(header.lstrip(" "))
            out = []
            for line in stanza:
                if not line.strip():
                    out.append("")
                    continue
                if line[:orig] == " " * orig:
                    rest = line[orig:]
                else:
                    rest = line.lstrip()
                out.append(base + rest)
            return out

        def _render_bgp_vprn_full(bgp_stanza, peer_ips, base_indent):
            """Para VPRN: copia a stanza `bgp` INTEGRAL preservando
            todos os parametros do nivel bgp (damping, multi-path,
            rib-management, etc.) e todos os neighbors soltos. APENAS
            descarta os blocos `group "X" {...}` cujos `neighbor <ip>`
            n\u00e3o contenham nenhum IP de `peer_ips`. Groups com pelo
            menos um neighbor casado s\u00e3o copiados INTEGRALMENTE.

            Se NENHUM group/neighbor solto/local-address da stanza
            casar com `peer_ips`, retorna [] (omite o bgp por
            completo) — assim VPRNs sem BGP relacionado a porta
            alvo nao aparecem com `bgp` vazio no output.
            """
            ips_set = set(peer_ips or [])
            if not ips_set:
                return []
            # Pre-check: existe ALGUM neighbor (em group ou solto) ou
            # local-address dentro do bgp que case com peer_ips? Se
            # nao, omite o bgp por completo.
            tem_match = False
            for ln in bgp_stanza:
                m_nb = re.match(r"^\s*neighbor\s+(\S+)", ln)
                if m_nb and m_nb.group(1) in ips_set:
                    tem_match = True
                    break
                m_la = re.match(r"^\s*local-address\s+(\S+)", ln)
                if m_la and m_la.group(1) in ips_set:
                    tem_match = True
                    break
            if not tem_match:
                return []
            blines = bgp_stanza
            base_ind = (
                len(blines[0]) - len(blines[0].lstrip(" ")))
            out_lines = [base_indent + "bgp"]
            child_indent = base_indent + "    "
            j = 1
            while j < len(blines):
                ln = blines[j]
                if not ln.strip():
                    j += 1
                    continue
                ind_ln = len(ln) - len(ln.lstrip(" "))
                strip_ln = ln.lstrip()
                # `exit` final do bgp
                if (ind_ln == base_ind
                        and strip_ln.startswith("exit")):
                    j += 1
                    continue
                # Filho direto do bgp
                if ind_ln == base_ind + 4:
                    if strip_ln.startswith("group "):
                        st, end = self._swap_extract_stanza(
                            blines, j)
                        # Verifica se o group cont\u00e9m algum neighbor
                        # com IP em peer_ips OU body com `local-address
                        # <ip>` em peer_ips (proprio ip da interface).
                        manter = False
                        if ips_set:
                            for ln2 in st:
                                m_nb = re.match(
                                    r"^\s*neighbor\s+(\S+)", ln2)
                                if m_nb and m_nb.group(1) in ips_set:
                                    manter = True
                                    break
                                m_la = re.match(
                                    r'^\s*local-address\s+(\S+)',
                                    ln2)
                                if m_la and m_la.group(1) in ips_set:
                                    manter = True
                                    break
                        if manter:
                            out_lines.extend(
                                reindent(st, child_indent))
                        j = end
                        continue
                    # Demais filhos diretos (params, neighbors soltos,
                    # multi-path, rib-management, etc.) \u2192 copia integral.
                    sub, sub_end = self._swap_extract_stanza(
                        blines, j)
                    if (sub_end - j) > 1:
                        out_lines.extend(
                            reindent(sub, child_indent))
                        j = sub_end
                    else:
                        out_lines.append(child_indent + strip_ln)
                        j += 1
                    continue
                j += 1
            out_lines.append(base_indent + "exit")
            return out_lines

        def _render_bgp_filtrado(bgp_stanza, nbs_lst, base_indent,
                                 grupos_no_para=None):
            """Renderiza o bgp preservando o body original (damping,
            multi-path, enable-peer-tracking, rapid-withdrawal, etc.),
            mas mantendo apenas os groups que têm pelo menos um
            neighbor casado em `nbs_lst`. Para cada group casado,
            se o group J\u00c1 EXISTE no PARA (`grupos_no_para`), emite
            apenas os neighbors casados (sem repetir os par\u00e2metros
            do group). Se o group N\u00c3O EXISTE, preserva os par\u00e2metros
            do group (family, peer-as, auth-key, etc.) junto com os
            neighbors casados.
            `nbs_lst` é a lista [(group_header_or_None, nb_stanza)]
            já filtrada por _filtra_bgp.
            """
            grupos_no_para = grupos_no_para or set()
            # Indices úteis: por header strip → lista de neighbor stanzas
            grupos_alvo = {}
            soltos = []
            for grp_hdr, nb_stanza in nbs_lst:
                if grp_hdr is None:
                    soltos.append(nb_stanza)
                else:
                    grupos_alvo.setdefault(grp_hdr.strip(), []).append(
                        nb_stanza)

            blines = bgp_stanza
            base_ind = len(blines[0]) - len(blines[0].lstrip(" "))
            out_lines = [base_indent + "bgp"]
            child_indent = base_indent + "    "

            j = 1
            while j < len(blines):
                ln = blines[j]
                if not ln.strip():
                    j += 1
                    continue
                ind_ln = len(ln) - len(ln.lstrip(" "))
                strip_ln = ln.lstrip()
                # `exit` final do bgp
                if (ind_ln == base_ind
                        and strip_ln.startswith("exit")):
                    j += 1
                    continue
                # Filho direto do bgp (indent = base_ind + 4)
                if ind_ln == base_ind + 4:
                    if strip_ln.startswith("group "):
                        grp_key = strip_ln
                        st, end = self._swap_extract_stanza(blines, j)
                        if grp_key in grupos_alvo:
                            nbs_alvo = grupos_alvo[grp_key]
                            ip_set = set()
                            for nb in nbs_alvo:
                                m = re.match(
                                    r"^\s*neighbor\s+(\S+)", nb[0])
                                if m:
                                    ip_set.add(m.group(1))
                            # Extrai o nome do group ("X") para checar
                            # se ja existe no PARA.
                            m_gn = re.match(
                                r'^\s*group\s+"([^"]+)"', grp_key)
                            grp_nome = (
                                m_gn.group(1) if m_gn else "")
                            grp_existe_para = (
                                grp_nome in grupos_no_para)
                            # Renderiza group:
                            #  - se existe no PARA: somente neighbors
                            #  - se nao existe: params + neighbors
                            out_lines.append(child_indent + grp_key)
                            grp_inner_ind = child_indent + "    "
                            grp_base = (len(st[0])
                                        - len(st[0].lstrip(" ")))
                            k = 1
                            while k < len(st):
                                ln2 = st[k]
                                if not ln2.strip():
                                    k += 1
                                    continue
                                ind2 = len(ln2) - len(ln2.lstrip(" "))
                                strip2 = ln2.lstrip()
                                if (ind2 == grp_base
                                        and strip2.startswith("exit")):
                                    k += 1
                                    continue
                                # Filho direto do group
                                if ind2 == grp_base + 4:
                                    if strip2.startswith("neighbor "):
                                        m_nb = re.match(
                                            r"^\s*neighbor\s+(\S+)",
                                            ln2)
                                        sub, sub_end = (
                                            self._swap_extract_stanza(
                                                st, k))
                                        if (m_nb
                                                and m_nb.group(1)
                                                in ip_set):
                                            out_lines.extend(
                                                reindent(sub,
                                                         grp_inner_ind))
                                        k = sub_end
                                        continue
                                    # Param do grupo (peer-as, family,
                                    # auth-key, etc.) ou stanza filho:
                                    # se group J\u00c1 existe no PARA, pula
                                    # (n\u00e3o precisa repetir param).
                                    if grp_existe_para:
                                        # mas preciso pular a stanza
                                        # filha completa, n\u00e3o s\u00f3 a linha
                                        k2 = k + 1
                                        while (k2 < len(st)
                                               and not st[k2]
                                               .strip()):
                                            k2 += 1
                                        if (k2 < len(st)
                                                and (len(st[k2])
                                                     - len(st[k2]
                                                           .lstrip(
                                                               " ")))
                                                > ind2):
                                            _, sub_end = (
                                                self
                                                ._swap_extract_stanza(
                                                    st, k))
                                            k = sub_end
                                            continue
                                        k += 1
                                        continue
                                    # checa se tem filhos
                                    k2 = k + 1
                                    while (k2 < len(st)
                                           and not st[k2].strip()):
                                        k2 += 1
                                    if (k2 < len(st)
                                            and (len(st[k2])
                                                 - len(st[k2].lstrip(" ")))
                                            > ind2):
                                        sub, sub_end = (
                                            self._swap_extract_stanza(
                                                st, k))
                                        out_lines.extend(
                                            reindent(sub, grp_inner_ind))
                                        k = sub_end
                                        continue
                                    out_lines.append(
                                        grp_inner_ind + strip2)
                                    k += 1
                                    continue
                                k += 1
                            out_lines.append(child_indent + "exit")
                        j = end
                        continue
                    if strip_ln.startswith("neighbor "):
                        # neighbor solto direto no bgp (sem group)
                        m_nb = re.match(r"^\s*neighbor\s+(\S+)", ln)
                        sub, sub_end = self._swap_extract_stanza(
                            blines, j)
                        if any(re.match(r"^\s*neighbor\s+" + re.escape(
                                m_nb.group(1) if m_nb else "??"),
                                nb[0]) for nb in soltos):
                            out_lines.extend(reindent(sub, child_indent))
                        j = sub_end
                        continue
                    # Param do bgp (damping, rapid-withdrawal, etc.)
                    # ou stanza (multi-path, rib-management,
                    # next-hop-resolution, best-path-selection, etc.):
                    # DESCARTAMOS, exceto `no shutdown` (que e
                    # idempotente e necessario para ativar groups
                    # recem-criados). O usuario nao quer parametros
                    # globais do bgp no SWAP, apenas os groups com
                    # neighbors casados.
                    if strip_ln == "no shutdown":
                        out_lines.append(child_indent + strip_ln)
                        j += 1
                        continue
                    # Pula a stanza/linha sem emitir
                    j2 = j + 1
                    while j2 < len(blines) and not blines[j2].strip():
                        j2 += 1
                    if (j2 < len(blines)
                            and (len(blines[j2])
                                 - len(blines[j2].lstrip(" ")))
                            > ind_ln):
                        _, sub_end = self._swap_extract_stanza(
                            blines, j)
                        j = sub_end
                        continue
                    j += 1
                    continue
                j += 1
            out_lines.append(base_indent + "exit")
            return out_lines

        # Pre-coleta de lsp-templates (e paths referenciados) que
        # existem no cfg_de mas N\u00c3O no cfg_para. Roda sempre,
        # independente de haver interfaces MPLS sendo emitidas.
        # Apenas defini\u00e7\u00f5es reais (n\u00e3o refer\u00eancias) s\u00e3o coletadas:
        # uma defini\u00e7\u00e3o tem a pr\u00f3xima linha n\u00e3o-vazia mais
        # indentada que o pr\u00f3prio header.
        lsp_extras_global = []   # [(nome, stanza)]
        paths_extras_global = []  # [(nome, stanza)]
        if cfg_text:
            de_lines_g = cfg_text.splitlines()
            rx_lsp_g = re.compile(
                r'^\s*lsp-template\s+"([^"]+)"')
            rx_pth_g = re.compile(
                r'^\s*path\s+"([^"]+)"')
            para_lsp_names_g = set()
            para_path_names_g = set()
            if cfg_para:
                for ln_p in cfg_para.splitlines():
                    m_l = rx_lsp_g.match(ln_p)
                    if m_l:
                        # so conta como defini\u00e7\u00e3o se a pr\u00f3xima linha
                        # n\u00e3o-vazia for mais indentada
                        idx_p = cfg_para.splitlines().index(ln_p)
                        # otimiza\u00e7\u00e3o: n\u00e3o ha como pegar idx aqui de
                        # forma confi\u00e1vel; preferimos varrer
                        # corretamente mais abaixo
                        para_lsp_names_g.add(m_l.group(1))
                    m_p = rx_pth_g.match(ln_p)
                    if m_p:
                        para_path_names_g.add(m_p.group(1))

            def _eh_definicao(linhas, idx):
                ind = (len(linhas[idx])
                       - len(linhas[idx].lstrip(" ")))
                jj = idx + 1
                while jj < len(linhas) and not linhas[jj].strip():
                    jj += 1
                if jj >= len(linhas):
                    return False
                return (len(linhas[jj])
                        - len(linhas[jj].lstrip(" "))) > ind

            seen_lsp = set()
            for i_de, ln in enumerate(de_lines_g):
                m = rx_lsp_g.match(ln)
                if not m:
                    continue
                nome_lsp = m.group(1)
                if (nome_lsp in seen_lsp
                        or nome_lsp in para_lsp_names_g):
                    continue
                if not _eh_definicao(de_lines_g, i_de):
                    continue
                st_lsp, _ = self._swap_extract_stanza(
                    de_lines_g, i_de)
                if not st_lsp:
                    continue
                seen_lsp.add(nome_lsp)
                lsp_extras_global.append((nome_lsp, st_lsp))
                # default-path "X" → buscar path "X" no cfg_de
                seen_pth_local = {p[0] for p in paths_extras_global}
                for ln2 in st_lsp:
                    m_dp = re.match(
                        r'^\s*default-path\s+"([^"]+)"', ln2)
                    if not m_dp:
                        continue
                    nome_pth = m_dp.group(1)
                    if (nome_pth in para_path_names_g
                            or nome_pth in seen_pth_local):
                        continue
                    rx_pth_de = re.compile(
                        rf'^\s*path\s+"'
                        rf'{re.escape(nome_pth)}"')
                    for j_de, ln3 in enumerate(de_lines_g):
                        if rx_pth_de.match(ln3):
                            if not _eh_definicao(
                                    de_lines_g, j_de):
                                continue
                            st_pth, _ = (
                                self._swap_extract_stanza(
                                    de_lines_g, j_de))
                            if st_pth:
                                paths_extras_global.append(
                                    (nome_pth, st_pth))
                                seen_pth_local.add(nome_pth)
                            break

        # ── Filtra lsp-templates e paths para manter apenas os
        # efetivamente usados pela configuracao que esta sendo
        # salva (port_stanzas, proto_path_ifaces_rb, buckets,
        # servicos). Um lsp-template e "usado" se seu nome aparece
        # como referencia (entre aspas) em alguma stanza coletada.
        # Um path e "usado" se aparece em alguma stanza coletada
        # OU se for `default-path` de um lsp-template mantido.
        if lsp_extras_global or paths_extras_global:
            ref_lines = []
            _seen_ids = set()

            def _coletar(obj):
                if obj is None:
                    return
                if isinstance(obj, str):
                    ref_lines.append(obj)
                    return
                # Protecao contra ciclos em estruturas aninhadas.
                oid = id(obj)
                if oid in _seen_ids:
                    return
                _seen_ids.add(oid)
                if isinstance(obj, dict):
                    for v in obj.values():
                        _coletar(v)
                elif isinstance(obj, (list, tuple, set)):
                    for v in obj:
                        _coletar(v)

            _coletar(port_stanzas)
            _coletar(proto_path_ifaces)
            for _svc in servicos:
                _coletar(_svc.get("ifaces"))
                _coletar(_svc.get("saps"))
                _coletar(_svc.get("extras_lines"))
                _coletar(_svc.get("extras_stanzas"))
                _coletar(_svc.get("bgp_full"))
                _coletar(_svc.get("statics_full"))

            ref_text = "\n".join(ref_lines)

            # Filtra lsp-templates: mantem so os nomes referenciados.
            lsp_extras_global = [
                (nm, st) for (nm, st) in lsp_extras_global
                if f'"{nm}"' in ref_text]

            # Coleta nomes de paths referenciados pelos lsp-templates
            # mantidos (via default-path "X").
            paths_referenciados = set()
            rx_dp = re.compile(r'^\s*default-path\s+"([^"]+)"')
            for _nm, _st in lsp_extras_global:
                for ln in _st:
                    m_dp = rx_dp.match(ln)
                    if m_dp:
                        paths_referenciados.add(m_dp.group(1))

            # Filtra paths: mantem se aparece no ref_text OU se e
            # default-path de um lsp-template mantido.
            paths_extras_global = [
                (nm, st) for (nm, st) in paths_extras_global
                if (nm in paths_referenciados
                    or f'"{nm}"' in ref_text)]

        out = ["exit all", "configure"]
        num_blocos = 0

        # Separa portas f\u00edsicas n\u00e3o-esat das portas esat (ordena\u00e7\u00e3o
        # exigida: 1.1 portas f\u00edsicas \u2192 1.2 software-repository \u2192
        # 1.3 eth-sat \u2192 1.4 port-topology \u2192 1.5 no shutdown nas
        # uplinks \u2192 1.6 portas esat \u2192 1.7 demais).
        port_stanzas_fisicas = [
            (p, st) for p, st in port_stanzas
            if not p.startswith("esat-")]
        port_stanzas_esat = [
            (p, st) for p, st in port_stanzas
            if p.startswith("esat-")]

        # 8.1 Ports f\u00edsicas (n\u00e3o-esat) primeiro
        for p, stanza in port_stanzas_fisicas:
            out.extend(reindent(stanza, "    "))
            num_blocos += 1

        # 8.1.x Blocos esat-N (sw-repo \u2192 eth-sat \u2192 port-topology
        # \u2192 no shutdown das uplinks \u2192 portas esat).
        for eid in esat_ids:
            sw = esat_sw_repos.get(eid)
            es = esat_eth_sats.get(eid)
            pt = esat_port_topo.get(eid)
            # 1.2 software-repository
            if sw:
                # Remove `secondary-location ...` (sera reconfigurado
                # no upgrade conforme localizacao primaria).
                sw = [ln for ln in sw
                      if not re.match(
                          r"^\s*secondary-location\b", ln)]
                out.append("    system")
                out.extend(reindent(sw, "        "))
                out.append("    exit")
                num_blocos += 1
            # 1.3 eth-sat N create
            if es:
                # Reordena: move `no shutdown` para DEPOIS do
                # ultimo `port-map` (padrao SR-OS para satelites).
                no_sd_lines = [ln for ln in es
                               if ln.strip() == "no shutdown"]
                es_sem_nosd = [ln for ln in es
                               if ln.strip() != "no shutdown"]
                if no_sd_lines:
                    # Acha indice da ultima linha `port-map`.
                    last_pm = -1
                    for idx, ln in enumerate(es_sem_nosd):
                        if ln.lstrip().startswith("port-map "):
                            last_pm = idx
                    if last_pm >= 0:
                        # Insere `no shutdown` (com indent dos demais
                        # filhos) logo apos o ultimo port-map.
                        ind_pm = (
                            len(es_sem_nosd[last_pm])
                            - len(es_sem_nosd[last_pm].lstrip(" ")))
                        es_sem_nosd.insert(
                            last_pm + 1,
                            " " * ind_pm + "no shutdown")
                        es = es_sem_nosd
                    # Se nao ha port-map, mantem original (es inalterado).
                out.append("    system")
                out.append("        satellite")
                out.extend(reindent(es, "            "))
                out.append("        exit")
                out.append("    exit")
                num_blocos += 1
            # 1.4 port-topology
            if pt:
                out.append("    system")
                out.append("        port-topology")
                for ln in pt:
                    out.append("            " + ln.strip())
                out.append("        exit")
                out.append("    exit")
                num_blocos += 1
            # 1.5 no shutdown nas portas f\u00edsicas uplink (origem do
            # port-topology: `port X to esat-N/u? create`).
            if pt:
                seen_up = []
                for ln in pt:
                    m_up = re.match(
                        r'^\s*port\s+(\S+)\s+to\s+esat-', ln)
                    if m_up and m_up.group(1) not in seen_up:
                        seen_up.append(m_up.group(1))
                for up in seen_up:
                    out.append(
                        f"/configure port {up} no shutdown")
                    num_blocos += 1
            # 1.6 Portas esat-N/... deste sat\u00e9lite (j\u00e1 filtradas)
            for p, stanza in port_stanzas_esat:
                if not p.startswith(f"esat-{eid}/"):
                    continue
                out.extend(reindent(stanza, "    "))
                num_blocos += 1

        # Esat ports cujo eid n\u00e3o aparece em esat_ids (caso raro): emite
        # ao final do bloco esat geral.
        eids_set = {eid for eid in esat_ids}
        for p, stanza in port_stanzas_esat:
            m_eid = re.match(r"^esat-(\d+)/", p)
            if m_eid and m_eid.group(1) not in eids_set:
                out.extend(reindent(stanza, "    "))
                num_blocos += 1

        # 8.2 LAGs
        for lid, stanza in lag_stanzas:
            out.extend(reindent(stanza, "    "))
            num_blocos += 1

        # 8.3 Router Base + protocolos + bgp/static por peer-IP
        # IMPORTANT: filtra proto_path_ifaces para conter apenas paths
        # de escopo router-base. Paths cujo header come\u00e7a com
        # vprn/ies/vpls/etc s\u00e3o emitidos dentro do pr\u00f3prio servi\u00e7o
        # (n\u00e3o devem aparecer sob "router Base").
        rx_path_svc_emit = re.compile(
            r'^(vprn|ies|vpls|epipe|ipipe|apipe|fpipe|cpipe)\s+\S+',
            re.IGNORECASE)
        proto_path_ifaces_rb = {
            p: lst for p, lst in proto_path_ifaces.items()
            if not (p and rx_path_svc_emit.match(p[0].strip()))
        }
        tem_rb = (bool(rb_interfaces) or bool(proto_path_ifaces_rb)
                  or bool(rb_bgp_neighbors) or bool(rb_static_routes))
        if tem_rb:
            out.append("    router Base")
            for nome, stanza in rb_interfaces:
                out.extend(reindent(stanza, "        "))
                num_blocos += 1

            # Static routes (formato unificado)
            for st in rb_static_routes:
                out.extend(reindent(st, "        "))
                num_blocos += 1

            # BGP filtrado
            if rb_bgp_neighbors:
                # Usa o renderer compartilhado (PARA-aware): se o
                # group ja existe no cfg_para, emite somente os
                # neighbors casados; sen\u00e3o, emite o group completo
                # com seus par\u00e2metros + neighbors casados.
                bgp_full_rb = (rb_bgp_raw_holder[0]
                               if rb_bgp_raw_holder else None)
                if bgp_full_rb:
                    out.extend(_render_bgp_filtrado(
                        bgp_full_rb, rb_bgp_neighbors, "        ",
                        grupos_no_para=bgp_grupos_no_para.get(
                            "rb", set())))
                    num_blocos += len(rb_bgp_neighbors)
                else:
                    # Fallback: emite minimal (mesmo formato anterior)
                    out.append("        bgp")
                    grupos = {}
                    for grp_hdr, nb_stanza in rb_bgp_neighbors:
                        key = grp_hdr.strip() if grp_hdr else None
                        grupos.setdefault(key, []).append(nb_stanza)
                    for grp_key, nbs in grupos.items():
                        if grp_key:
                            out.append("            " + grp_key)
                            for nb in nbs:
                                out.extend(reindent(
                                    nb, "                "))
                                num_blocos += 1
                            out.append("            exit")
                        else:
                            for nb in nbs:
                                out.extend(reindent(
                                    nb, "            "))
                                num_blocos += 1
                    out.append("        exit")

            # Protocolos com interfaces
            ordem_proto = ["ospf", "ospf3", "isis", "mpls", "rsvp",
                           "ldp", "pim"]

            def proto_sort_key(path):
                # path[0] e o parent ("router Base"); path[1] e o
                # protocolo mais externo.
                proto_chain = path[1:] if path else ()
                top = (proto_chain[0].split()[0].lower()
                       if proto_chain else "")
                try:
                    return (ordem_proto.index(top), proto_chain)
                except ValueError:
                    return (len(ordem_proto), proto_chain)

            for path in sorted(proto_path_ifaces_rb.keys(),
                               key=proto_sort_key):
                ifaces = proto_path_ifaces_rb[path]
                # path[0] = "router Base" (ja emitido); usa apenas o
                # restante (cadeia de protocolos) para emitir headers.
                proto_chain = path[1:] if path else ()
                indent_lvl = "        "
                for header_str in proto_chain:
                    out.append(indent_lvl + header_str)
                    indent_lvl += "    "

                # MPLS: antes das interfaces, injeta entradas
                # `if-attribute > admin-group "X" value N` faltantes
                # no PARA (extra\u00eddas do cfg_de). Ap\u00f3s as interfaces,
                # injeta `lsp-template`s faltantes (com seus paths).
                top = (proto_chain[0].split()[0].lower()
                       if proto_chain else "")
                ag_extras = []   # [(nome, valor_str)]
                lsp_extras = []  # [(nome, stanza)]
                paths_extras = []  # [(nome, stanza)] — paths refs por lsp
                if top == "mpls":
                    # 1) admin-groups referenciados nas interfaces MPLS
                    ags_ref = set()
                    rx_ag = re.compile(
                        r'^\s*admin-group\s+"([^"]+)"')
                    rx_inc = re.compile(
                        r'^\s*include\s+"([^"]+)"')
                    for _nm, st in ifaces:
                        for ln in st:
                            m = rx_ag.match(ln)
                            if m:
                                ags_ref.add(m.group(1))
                    # `include "X"` em lsp-templates mantidos
                    # (X = nome de if-attribute > admin-group)
                    for _nm, st in lsp_extras_global:
                        for ln in st:
                            m = rx_inc.match(ln)
                            if m:
                                ags_ref.add(m.group(1))
                    # admin-groups j\u00e1 existentes no PARA
                    ags_no_para = set()
                    if cfg_para:
                        rx_ag_def = re.compile(
                            r'^\s*admin-group\s+"([^"]+)"\s+value\s+\d+')
                        for ln in cfg_para.splitlines():
                            m = rx_ag_def.match(ln)
                            if m:
                                ags_no_para.add(m.group(1))
                    # busca defini\u00e7\u00e3o no cfg_de
                    if cfg_text:
                        rx_ag_def = re.compile(
                            r'^\s*admin-group\s+"([^"]+)"\s+value\s+(\d+)')
                        for ln in cfg_text.splitlines():
                            m = rx_ag_def.match(ln)
                            if m and m.group(1) in ags_ref \
                                    and m.group(1) not in ags_no_para:
                                ag_extras.append(
                                    (m.group(1), m.group(2)))
                    # dedup
                    ag_extras = list(dict.fromkeys(ag_extras))

                    # 2) lsp-templates e paths j\u00e1 pre-coletados
                    # globalmente em lsp_extras_global /
                    # paths_extras_global. Apenas referencia aqui.
                    lsp_extras = list(lsp_extras_global)
                    paths_extras = list(paths_extras_global)
                    # Marca como j\u00e1 emitidos para evitar duplica\u00e7\u00e3o
                    # no fallback p\u00f3s-loop.
                    lsp_extras_global = []
                    paths_extras_global = []

                    # Renderiza if-attribute faltante (antes das ifaces)
                    if ag_extras:
                        out.append(indent_lvl + "if-attribute")
                        for _nm, _val in ag_extras:
                            out.append(
                                indent_lvl + "    "
                                + f'admin-group "{_nm}" value {_val}')
                            num_blocos += 1
                        out.append(indent_lvl + "exit")
                    # Paths referenciados (DEVEM vir antes do
                    # lsp-template que os referencia).
                    for _nm, st in paths_extras:
                        out.extend(reindent(st, indent_lvl))
                        num_blocos += 1

                for nome, stanza in ifaces:
                    out.extend(reindent(stanza, indent_lvl))
                    num_blocos += 1

                # MPLS: lsp-templates faltantes (depois das ifaces,
                # antes do exit do mpls).
                if top == "mpls" and lsp_extras:
                    for _nm, st in lsp_extras:
                        out.extend(reindent(st, indent_lvl))
                        num_blocos += 1

                indent_lvl = indent_lvl[:-4]
                for _ in proto_chain:
                    out.append(indent_lvl + "exit")
                    indent_lvl = (indent_lvl[:-4]
                                  if len(indent_lvl) >= 4 else "")
            # Antes de fechar o bloco router Base: se sobraram
            # lsp-templates/paths que n\u00e3o foram emitidos via
            # proto_path_ifaces (porque n\u00e3o havia interface MPLS
            # da porta-alvo), injeta um sub-bloco mpls aqui dentro
            # \u2014 evita criar um SEGUNDO `router Base` separado.
            if lsp_extras_global or paths_extras_global:
                out.append("        mpls")
                # if-attribute para `include "X"` em lsp-templates
                ag_extras_fb = self._swap_compute_ag_extras_de_lsp(
                    lsp_extras_global, cfg_text, cfg_para)
                if ag_extras_fb:
                    out.append("            if-attribute")
                    for _nm, _val in ag_extras_fb:
                        out.append(
                            "                "
                            + f'admin-group "{_nm}" value {_val}')
                        num_blocos += 1
                    out.append("            exit")
                for _nm, st in paths_extras_global:
                    out.extend(reindent(st, "            "))
                    num_blocos += 1
                for _nm, st in lsp_extras_global:
                    out.extend(reindent(st, "            "))
                    num_blocos += 1
                out.append("        exit")
                lsp_extras_global = []
                paths_extras_global = []
            out.append("    exit")

        # Fallback: se sobraram lsp-templates (por n\u00e3o ter havido
        # path mpls em proto_path_ifaces E tem_rb foi False), emite
        # um bloco sint\u00e9tico router Base > mpls com os
        # lsp-templates e paths faltantes.
        if lsp_extras_global or paths_extras_global:
            out.append("    router Base")
            out.append("        mpls")
            ag_extras_fb2 = self._swap_compute_ag_extras_de_lsp(
                lsp_extras_global, cfg_text, cfg_para)
            if ag_extras_fb2:
                out.append("            if-attribute")
                for _nm, _val in ag_extras_fb2:
                    out.append(
                        "                "
                        + f'admin-group "{_nm}" value {_val}')
                    num_blocos += 1
                out.append("            exit")
            for _nm, st in paths_extras_global:
                out.extend(reindent(st, "            "))
                num_blocos += 1
            for _nm, st in lsp_extras_global:
                out.extend(reindent(st, "            "))
                num_blocos += 1
            out.append("        exit")
            out.append("    exit")
            lsp_extras_global = []
            paths_extras_global = []

        # 8.3.1 (movido) \u2014 blocos `system` para esat-N agora s\u00e3o
        # emitidos no in\u00edcio (ver 8.1.x), antes das interfaces, para
        # respeitar a ordem de provisionamento exigida.

        # 8.4 Serviços
        if servicos:
            # Verifica se o servi\u00e7o (por TIPO e ID, ex.: `ies 1009169`,
            # `vprn 5000`) j\u00e1 existe no cfg_para. Se sim, NAO descarta
            # \u2014 marca svc["existe_no_para"]=True para que, na emiss\u00e3o,
            # o header seja encurtado para apenas `<type> <id>` (sem
            # `name`, `customer` nem `create`), pois o servi\u00e7o j\u00e1
            # existe e estamos apenas adicionando interfaces/SAPs nele.
            if cfg_para:
                rx_svc_id_para = re.compile(
                    r'^\s*(vpls|vprn|ies|epipe|ipipe|apipe|fpipe|cpipe|'
                    r'mirror-dest|pw-template)\s+(\S+)\b')
                ids_no_para = set()
                for ln in cfg_para.splitlines():
                    m = rx_svc_id_para.match(ln)
                    if m:
                        ids_no_para.add((m.group(1).lower(), m.group(2)))
                for svc in servicos:
                    chave_id = (svc["type"], svc["id"])
                    svc["existe_no_para"] = chave_id in ids_no_para
            else:
                for svc in servicos:
                    svc["existe_no_para"] = False
        if servicos:
            # 8.3.5 Customers: extrai `customer <ID>` dos headers dos
            # servicos, verifica se ja existe no cfg_para e, se nao,
            # copia a stanza completa do cfg_de para o .txt (antes do
            # bloco `service`, pois servicos referenciam customer).
            cust_ids = []
            rx_cust = re.compile(r"\bcustomer\s+(\d+)\b")
            for svc in servicos:
                # Servi\u00e7o que j\u00e1 existe no PARA n\u00e3o precisa do customer
                # (ja foi criado junto com o servi\u00e7o original).
                if svc.get("existe_no_para"):
                    continue
                m_c = rx_cust.search(svc.get("header", ""))
                if m_c and m_c.group(1) not in cust_ids:
                    cust_ids.append(m_c.group(1))
            cust_stanzas = []
            if cust_ids:
                cfg_de_lines = (cfg_text.splitlines()
                                if cfg_text else [])
                cfg_para_lines = (cfg_para.splitlines()
                                  if cfg_para else [])
                rx_cust_hdr_de = {
                    cid: re.compile(
                        rf'^\s*customer\s+{re.escape(cid)}\s+'
                        rf'(?:name\s+"[^"]*"\s+)?create\s*$')
                    for cid in cust_ids
                }
                # Customer ja existente no PARA: header igual.
                cust_no_para = set()
                for ln in cfg_para_lines:
                    for cid, rxh in rx_cust_hdr_de.items():
                        if rxh.match(ln):
                            cust_no_para.add(cid)
                for cid in cust_ids:
                    if cid in cust_no_para:
                        continue
                    rxh = rx_cust_hdr_de[cid]
                    for i_de, ln in enumerate(cfg_de_lines):
                        if rxh.match(ln):
                            st, _ = self._swap_extract_stanza(
                                cfg_de_lines, i_de)
                            if st:
                                cust_stanzas.append((cid, st))
                            break
            if cust_stanzas:
                out.append("    service")
                for _cid, st in cust_stanzas:
                    out.extend(reindent(st, "        "))
                    num_blocos += 1
                out.append("    exit")

            # 8.3.6 dist-cpu-protection: varre as stanzas dos servi\u00e7os
            # (ifaces/saps/extras) atr\u00e1s de `dist-cpu-protection "X"`,
            # verifica se a policy existe no cfg_para e, se n\u00e3o, copia
            # a stanza completa do cfg_de para o .txt.
            dcp_refs = []
            rx_dcp = re.compile(r'^\s*dist-cpu-protection\s+"([^"]+)"')
            for svc in servicos:
                fontes = []
                fontes.extend(svc.get("extras_lines") or [])
                for st in (svc.get("extras_stanzas") or []):
                    fontes.extend(st)
                for _nm, st in (svc.get("ifaces") or []):
                    fontes.extend(st)
                for _nm, st in (svc.get("saps") or []):
                    fontes.extend(st)
                for ln in fontes:
                    m = rx_dcp.match(ln)
                    if m and m.group(1) not in dcp_refs:
                        dcp_refs.append(m.group(1))
            dcp_stanzas = []
            if dcp_refs:
                rx_dcp_para = {
                    nm: re.compile(
                        rf'^\s*policy\s+"{re.escape(nm)}"')
                    for nm in dcp_refs
                }
                # j\u00e1 no PARA?
                dcp_no_para = set()
                if cfg_para:
                    in_dcp_blk = False
                    for ln in cfg_para.splitlines():
                        s = ln.strip()
                        if s.startswith("dist-cpu-protection"):
                            in_dcp_blk = True
                            continue
                        if in_dcp_blk:
                            for nm, rxh in rx_dcp_para.items():
                                if rxh.match(ln):
                                    dcp_no_para.add(nm)
                            if s == "exit":
                                in_dcp_blk = False
                # extrai do cfg_de
                if cfg_text:
                    de_lines = cfg_text.splitlines()
                    in_dcp_blk = False
                    for i_de, ln in enumerate(de_lines):
                        s = ln.strip()
                        if s.startswith("dist-cpu-protection"):
                            in_dcp_blk = True
                            continue
                        if not in_dcp_blk:
                            continue
                        if s == "exit":
                            in_dcp_blk = False
                            continue
                        for nm in dcp_refs:
                            if nm in dcp_no_para:
                                continue
                            if any(d[0] == nm for d in dcp_stanzas):
                                continue
                            if rx_dcp_para[nm].match(ln):
                                st, _ = self._swap_extract_stanza(
                                    de_lines, i_de)
                                if st:
                                    dcp_stanzas.append((nm, st))
            if dcp_stanzas:
                out.append("    system")
                out.append("        security")
                out.append("            dist-cpu-protection")
                for _nm, st in dcp_stanzas:
                    out.extend(reindent(st, "                "))
                    num_blocos += 1
                out.append("            exit")
                out.append("        exit")
                out.append("    exit")

            out.append("    service")
            for svc in servicos:
                hdr = svc["header"].strip()
                # Se o servi\u00e7o JA EXISTE no PARA, encurta o header
                # para apenas `<type> <id>` (sem name/customer/create).
                if svc.get("existe_no_para"):
                    hdr = f'{svc["type"]} {svc["id"]}'
                out.append("        " + hdr)
                chave = (svc["type"], svc["id"])
                svc_orig = svc_full.get(chave)
                # Caminha a stanza ORIGINAL preservando a ordem dos
                # filhos diretos (description, autonomous-system,
                # route-distinguisher, vrf-target, auto-bind-tunnel,
                # mvpn, pim, ospf, interface, bgp, no shutdown ...).
                # Substitui:
                #  - interface "X" → vers\u00e3o de svc["ifaces"]
                #    (filtrada por porta + p\u00f3s-processada)
                #  - sap "X"       → vers\u00e3o de svc["saps"]
                #  - bgp           → renderer filtrado por peer-IP
                #  - static-route* → vers\u00e3o filtrada
                # `no shutdown` \u00e9 movido para o fim.
                if svc_orig:
                    # Dedup: mantem a PRIMEIRA ocorr\u00eancia (mesma
                    # interface pode aparecer em svc["ifaces"] e em
                    # proto_path_ifaces; preferimos a vers\u00e3o do body
                    # do servi\u00e7o que tende a ser mais completa).
                    ifaces_dict = {}
                    for nm, st in svc["ifaces"]:
                        if nm not in ifaces_dict:
                            ifaces_dict[nm] = st
                    saps_dict = {}
                    for sid, st in svc["saps"]:
                        if sid not in saps_dict:
                            saps_dict[sid] = st
                    base_ind_svc = (
                        len(svc_orig[0])
                        - len(svc_orig[0].lstrip(" ")))
                    child_ind = base_ind_svc + 4

                    # Coleta filhos diretos do servi\u00e7o em buckets
                    # categorizados. Dump SR-OS pode emitir o mesmo
                    # vprn N em v\u00e1rios blocos (Configuration vs.
                    # BGP vs. Interface), e a ordem em que esses
                    # blocos aparecem no dump n\u00e3o coincide com a
                    # ordem canonica Nokia. Reordenamos.
                    bk_geral = []      # description, autonomous-system, etc.
                    bk_mcast = []      # mvpn, igmp, pim, msdp
                    bk_iface = []      # interface "X"
                    bk_sap = []        # sap, spoke-sdp, mesh-sdp
                    bk_static = None   # placeholder (emitido uma vez)
                    bk_bgp = None      # placeholder (emitido uma vez)
                    bk_proto = []      # ospf, isis, ldp, rsvp
                    no_shutdown_pendente = False

                    def _categoria_filho(strip):
                        # Retorna (bucket, prioridade_dentro). N\u00e3o
                        # usado para sap/iface/bgp/static (tratados
                        # separadamente).
                        s0 = strip.split()[0] if strip else ""
                        if s0 in ("mvpn", "igmp", "pim", "msdp",
                                  "dhcp", "dhcp6", "ipsec"):
                            return "mcast"
                        if s0 in ("ospf", "ospf3", "isis",
                                  "ldp", "rsvp"):
                            return "proto"
                        return "geral"

                    j = 1
                    while j < len(svc_orig):
                        ln = svc_orig[j]
                        if not ln.strip():
                            j += 1
                            continue
                        ind_ln = len(ln) - len(ln.lstrip(" "))
                        if (ind_ln == base_ind_svc
                                and ln.lstrip().startswith("exit")):
                            break
                        if ind_ln != child_ind:
                            j += 1
                            continue
                        strip_ln = ln.lstrip()
                        if strip_ln == "exit":
                            j += 1
                            continue
                        # stanza ou linha solta?
                        j2 = j + 1
                        while (j2 < len(svc_orig)
                               and not svc_orig[j2].strip()):
                            j2 += 1
                        eh_stanza = (
                            j2 < len(svc_orig)
                            and (len(svc_orig[j2])
                                 - len(svc_orig[j2].lstrip(" ")))
                            > ind_ln)
                        if eh_stanza:
                            sub, sub_end = self._swap_extract_stanza(
                                svc_orig, j)
                        else:
                            sub = [ln]
                            sub_end = j + 1

                        m_if = re.match(
                            r'^\s*interface\s+"([^"]+)"', ln)
                        m_sap_l = re.match(
                            r'^\s*sap\s+(\S+)', ln)

                        if m_if:
                            nm_if = m_if.group(1)
                            if nm_if in ifaces_dict:
                                bk_iface.append(
                                    (nm_if, ifaces_dict[nm_if]))
                                # consome para n\u00e3o duplicar se
                                # aparecer em outro bloco
                                ifaces_dict.pop(nm_if, None)
                            j = sub_end
                            continue
                        if m_sap_l:
                            sid_l = m_sap_l.group(1)
                            if sid_l in saps_dict:
                                bk_sap.append(
                                    (sid_l, saps_dict[sid_l]))
                                saps_dict.pop(sid_l, None)
                            j = sub_end
                            continue
                        if (strip_ln.startswith("spoke-sdp")
                                or strip_ln.startswith("mesh-sdp")):
                            bk_sap.append((strip_ln, sub))
                            j = sub_end
                            continue
                        if (strip_ln == "bgp"
                                or strip_ln.startswith("bgp ")):
                            if bk_bgp is None:
                                bk_bgp = sub  # ref para bgp_full fallback
                            j = sub_end
                            continue
                        if (strip_ln.startswith("static-route")
                                or strip_ln.startswith("static-routes")):
                            if bk_static is None:
                                bk_static = True  # marca presen\u00e7a
                            j = sub_end
                            continue
                        if strip_ln == "no shutdown":
                            no_shutdown_pendente = True
                            j = sub_end
                            continue
                        # Catch-all classificado
                        cat = _categoria_filho(strip_ln)
                        if cat == "mcast":
                            bk_mcast.append(sub)
                        elif cat == "proto":
                            bk_proto.append(sub)
                        else:
                            bk_geral.append(sub)
                        j = sub_end

                    # Adiciona ifaces/saps que existem em
                    # svc["ifaces"]/svc["saps"] mas n\u00e3o apareceram
                    # no svc_orig (caso o walker original tenha
                    # capturado iface fora do bloco principal).
                    for nm_if, st in ifaces_dict.items():
                        bk_iface.append((nm_if, st))
                    for sid_l, st in saps_dict.items():
                        bk_sap.append((sid_l, st))

                    # ---- Emiss\u00e3o em ordem can\u00f4nica Nokia ----
                    # 1) Geral (description, autonomous-system,
                    #    route-distinguisher, vrf-target,
                    #    auto-bind-tunnel, etc.)
                    for st in bk_geral:
                        out.extend(reindent(st, "            "))
                        num_blocos += 1
                    # 2) Interfaces
                    for _nm, st in bk_iface:
                        out.extend(reindent(st, "            "))
                        num_blocos += 1
                    # 3) SAPs (e spoke-sdp/mesh-sdp)
                    for _sid, st in bk_sap:
                        out.extend(reindent(st, "            "))
                        num_blocos += 1
                    # 4) Static routes filtradas
                    for st_sr in svc_static_routes.get(chave, []):
                        out.extend(reindent(st_sr, "            "))
                        num_blocos += 1
                    # 5) BGP (vprn) — copia INTEGRAL da stanza bgp,
                    #    descartando apenas os groups cujos neighbors
                    #    nao tenham IP em peers_por_escopo[chave].
                    bgp_full_v = svc.get("bgp_full") or bk_bgp
                    if bgp_full_v:
                        bgp_lines_vprn = _render_bgp_vprn_full(
                            bgp_full_v,
                            peers_por_escopo.get(chave, []),
                            "            ")
                        if bgp_lines_vprn:
                            out.extend(bgp_lines_vprn)
                            num_blocos += 1
                    # 6) Multicast (mvpn, igmp, pim, msdp) APOS bgp
                    for st in bk_mcast:
                        out.extend(reindent(st, "            "))
                        num_blocos += 1
                    # 7) Protocolos (ospf, isis, ldp, rsvp)
                    # Para vprn: filtra ospf/ospf3/isis para manter
                    # apenas substanzas referentes a interfaces alvo
                    # (descartando area inteiras se ficarem vazias).
                    target_ifs_proto = {nm for nm, _ in bk_iface}
                    for st in bk_proto:
                        s0 = (st[0].lstrip().split() or [""])[0]
                        if s0 in ("ospf", "ospf3", "isis"):
                            stf = self._swap_filter_proto_for_targets(
                                st, target_ifs_proto)
                            if not stf:
                                continue
                            out.extend(reindent(stf, "            "))
                        else:
                            out.extend(reindent(st, "            "))
                        num_blocos += 1
                    # 8) no shutdown final
                    if no_shutdown_pendente:
                        out.append("            no shutdown")
                    out.append("        exit")
                    continue
                # Fallback (svc_orig ausente): comportamento antigo.
                # 1) Config geral do serviço (extras), exceto no shutdown
                #    que vai para o fim. Também separa spoke-sdp/mesh-sdp
                #    que devem aparecer DEPOIS dos SAPs (ordem do dump
                #    SR-OS).
                tem_no_shutdown = False
                for ln in svc["extras_lines"]:
                    if ln.strip() == "no shutdown":
                        tem_no_shutdown = True
                        continue
                    out.append("            " + ln)
                    num_blocos += 1
                extras_pre = []
                extras_post = []
                for st in svc["extras_stanzas"]:
                    hdr0 = st[0].strip() if st else ""
                    if (hdr0.startswith("spoke-sdp")
                            or hdr0.startswith("mesh-sdp")):
                        extras_post.append(st)
                    else:
                        extras_pre.append(st)
                # vprn: filtra ospf/ospf3/isis em extras_pre para
                # manter apenas substanzas referentes a interfaces
                # alvo (mesma logica do path principal).
                target_ifs_proto_fb = (
                    {nm for nm, _ in svc["ifaces"]}
                    if svc["type"] == "vprn" else set())
                extras_pre_filt = []
                for st in extras_pre:
                    if not st:
                        continue
                    s0 = (st[0].lstrip().split() or [""])[0]
                    if (svc["type"] == "vprn"
                            and s0 in ("ospf", "ospf3", "isis")):
                        stf = self._swap_filter_proto_for_targets(
                            st, target_ifs_proto_fb)
                        if not stf:
                            continue
                        extras_pre_filt.append(stf)
                    else:
                        extras_pre_filt.append(st)
                for st in extras_pre_filt:
                    out.extend(reindent(st, "            "))
                    num_blocos += 1
                # 2) Interfaces e SAPs
                for nome, stanza in svc["ifaces"]:
                    out.extend(reindent(stanza, "            "))
                    num_blocos += 1
                for sap_id, stanza in svc["saps"]:
                    out.extend(reindent(stanza, "            "))
                    num_blocos += 1
                # 2.5) spoke-sdp/mesh-sdp DEPOIS dos SAPs
                for st in extras_post:
                    out.extend(reindent(st, "            "))
                    num_blocos += 1
                # 3) Static routes filtradas (vprn) — ANTES do BGP
                for st in svc_static_routes.get(chave, []):
                    out.extend(reindent(st, "            "))
                    num_blocos += 1
                # 4) BGP (vprn) — copia INTEGRAL da stanza bgp,
                #    descartando apenas os groups cujos neighbors
                #    nao tenham IP em peers_por_escopo[chave].
                bgp_full_v = svc.get("bgp_full")
                if bgp_full_v:
                    bgp_lines_vprn = _render_bgp_vprn_full(
                        bgp_full_v,
                        peers_por_escopo.get(chave, []),
                        "            ")
                    if bgp_lines_vprn:
                        out.extend(bgp_lines_vprn)
                        num_blocos += 1
                # 5) no shutdown final
                if tem_no_shutdown:
                    out.append("            no shutdown")
                out.append("        exit")
            out.append("    exit")

        out.append("exit all")
        texto_final = "\n".join(out) + "\n"

        # Pós-processamento global do texto:
        # • Substituição DE→PARA em TODAS as instâncias (port
        #   header, port em iface, sap, port-topology, etc.)
        # • Dentro de strings entre aspas duplas (descriptions e
        #   nomes de interface), o PARA fica em MAIUSCULO
        #   (ex.: 2/1/c3/1 → 2/1/C3/1).
        # • Remove a palavra "Base" de "router Base".
        # Ordena por tamanho decrescente para evitar substituição
        # parcial (ex.: 1/1/1 vs 1/1/10).
        if de_to_para_port:
            chaves_ord = sorted(
                de_to_para_port.keys(), key=len, reverse=True)

            # 1) Dentro de aspas duplas: PARA em MAIUSCULO.
            # Fronteira esquerda mais permissiva (so impede colagem
            # com outro digito ou barra) para casar dentro de
            # descriptions tipo "Tengige2/1/6:1526.3".
            def _upper_in_quotes(match):
                s = match.group(0)
                for _de_p in chaves_ord:
                    _para_p = de_to_para_port[_de_p]
                    if _de_p == _para_p:
                        continue
                    s = re.sub(
                        r"(?<![\d/])" + re.escape(_de_p) + r"(?![\w/-])",
                        _para_p.upper(), s)
                return s

            texto_final = re.sub(
                r'"[^"]*"', _upper_in_quotes, texto_final)

            # 2) Resto do texto (fora das aspas): mantem lowercase.
            for _de_p in chaves_ord:
                _para_p = de_to_para_port[_de_p]
                if _de_p == _para_p:
                    continue
                texto_final = re.sub(
                    r"(?<![\w/-])" + re.escape(_de_p) + r"(?![\w/-])",
                    _para_p, texto_final)
        # Remove o sufixo "Base" do header `router Base` (mantendo
        # apenas `router`). Aceita qualquer indentação.
        texto_final = re.sub(
            r"^([ \t]*)router\s+Base(\s*)$",
            r"\1router\2",
            texto_final, flags=re.MULTILINE)

        # ── 11) Bloco de SHUTDOWN para o roteador DE ─────────
        # Gera comandos `/configure ... shutdown` (uma linha cada)
        # para portas, LAGs, interfaces, static-routes e BGP
        # neighbors envolvidos no SWAP — para serem executados no
        # roteador DE apos o PARA assumir.
        #
        # Antes do builder, computa estruturas de apoio
        # reaproveitadas tanto pelo shutdown quanto pelos blocos
        # de testes (a frente):
        #
        #   bgp_peers_por_escopo: {escopo -> set(neighbor_ip)}
        #     "rb"             -> neighbors do BGP do router Base
        #     ("vprn", svc_id) -> neighbors do BGP da VPRN
        #
        #   proto_por_iface: {(escopo..., iface_name) -> set(proto)}
        #     ex.: ("rb", "TO_X")           -> {"ospf","mpls"}
        #          ("vprn","1009041","X")   -> {"ospf"}
        #
        #   ospf_iface_shutdowns: lista de comandos prontos
        #     '/configure router ospf <id> area <area> interface "<n>" shutdown'
        #     '/configure service vprn <sid> ospf <id> area <area> interface "<n>" shutdown'
        #     (analogo para ospf3 / IES). Emitido ANTES do shutdown
        #     da interface L3 para o protocolo retirar a iface da SPF
        #     de forma graciosa antes do down fisico.
        rx_nb_ip = re.compile(r"^\s*neighbor\s+(\S+)")
        bgp_peers_por_escopo = {"rb": set()}
        for _grp, nb_st in (rb_bgp_neighbors or []):
            if nb_st:
                m_ip = rx_nb_ip.match(nb_st[0])
                if m_ip:
                    bgp_peers_por_escopo["rb"].add(m_ip.group(1))
        for chave, nbs in (svc_bgp_neighbors or {}).items():
            sset = bgp_peers_por_escopo.setdefault(chave, set())
            for _grp, nb_st in nbs:
                if nb_st:
                    m_ip = rx_nb_ip.match(nb_st[0])
                    if m_ip:
                        sset.add(m_ip.group(1))

        rx_proto_path = re.compile(
            r'^(ospf3?|isis|mpls|rsvp|ldp|pim)(?:\s+(\S+))?\s*$')
        rx_area_path = re.compile(r'^area\s+(\S+)')
        proto_por_iface = {}
        ospf_iface_shutdowns = []
        _vistos_osh = set()
        for path, lst in (proto_path_ifaces or {}).items():
            if not path:
                continue
            head0 = path[0].strip()
            m_sc = rx_path_svc.match(head0)
            if m_sc:
                svc_t = m_sc.group(1).lower()
                svc_i = m_sc.group(2)
                escopo_pref = (svc_t, svc_i)
            else:
                svc_t = svc_i = None
                escopo_pref = ("rb",)
            proto_nome = None
            proto_inst = None
            area_id = None
            for h in path:
                hs = h.strip()
                m_p = rx_proto_path.match(hs)
                if m_p and proto_nome is None:
                    proto_nome = m_p.group(1).lower()
                    proto_inst = m_p.group(2)  # pode ser None p/ ex 'mpls'
                    continue
                m_a = rx_area_path.match(hs)
                if m_a:
                    area_id = m_a.group(1)
            if not proto_nome:
                continue
            for _nome, _stz in lst:
                key = escopo_pref + (_nome,)
                proto_por_iface.setdefault(key, set()).add(proto_nome)
                # Comando de shutdown OSPF/OSPF3 (so faz sentido com area)
                if proto_nome in ("ospf", "ospf3") and area_id:
                    if escopo_pref == ("rb",):
                        prefix = "/configure router"
                    elif svc_t in ("vprn", "ies"):
                        prefix = (f"/configure service "
                                  f"{svc_t} {svc_i}")
                    else:
                        continue
                    inst = f" {proto_inst}" if proto_inst else ""
                    cmd = (f'{prefix} {proto_nome}{inst} '
                           f'area {area_id} interface "{_nome}" shutdown')
                    if cmd not in _vistos_osh:
                        _vistos_osh.add(cmd)
                        ospf_iface_shutdowns.append(cmd)

        texto_shutdown_de = self._swap_montar_bloco_shutdown_de(
            portas=portas,
            lags_alvo=lags_alvo,
            rb_interfaces=rb_interfaces,
            rb_static_routes=rb_static_routes,
            rb_bgp_neighbors=rb_bgp_neighbors,
            servicos=servicos,
            svc_static_routes=svc_static_routes,
            svc_bgp_neighbors=svc_bgp_neighbors,
            ospf_iface_shutdowns=ospf_iface_shutdowns,
        )

        # ── 12) Blocos de TESTES (DE e PARA) ─────────────────
        # Mesma coleta de interfaces/SAPs/peers ja extraida acima e
        # reaproveitada para gerar comandos de validacao (ping +
        # show ...) tanto para o roteador DE (portas/SAPs originais)
        # quanto para o roteador PARA (com substituicao DE -> PARA
        # nos nomes que carregam o slot/MDA da porta).
        # As estruturas bgp_peers_por_escopo / proto_por_iface ja
        # foram computadas acima (na etapa 11) e sao reutilizadas.

        texto_testes_de = self._swap_montar_bloco_testes(
            rb_interfaces=rb_interfaces,
            servicos=servicos,
            bgp_peers_por_escopo=bgp_peers_por_escopo,
            proto_por_iface=proto_por_iface,
            de_to_para_port=None,
        )
        texto_testes_para = self._swap_montar_bloco_testes(
            rb_interfaces=rb_interfaces,
            servicos=servicos,
            bgp_peers_por_escopo=bgp_peers_por_escopo,
            proto_por_iface=proto_por_iface,
            de_to_para_port=de_to_para_port or None,
        )

        # ── 13) Blocos de FALLBACK (DE e PARA) ───────────────
        # Extrai descriptions ORIGINAIS de portas/LAGs direto do
        # cfg_text (config bruto do DE) — port_stanzas/lag_stanzas
        # ja foram transformados (renomeados DE->PARA), entao nao
        # servem aqui.
        rx_desc_in = re.compile(r'^\s*description\s+(.+?)\s*$')
        port_descs_de = {}
        for _p in (portas or []):
            rx_phdr = re.compile(
                rf'^(\s*)port\s+{re.escape(_p)}\s*$')
            linhas_cfg = (cfg_text or "").split("\n")
            for i, ln in enumerate(linhas_cfg):
                m_h = rx_phdr.match(ln)
                if not m_h:
                    continue
                stz_p, _ = self._swap_extract_stanza(linhas_cfg, i)
                for sub in stz_p[1:]:
                    md = rx_desc_in.match(sub)
                    if md:
                        port_descs_de[_p] = md.group(1).strip()
                        break
                break
        lag_descs_de = {}
        for _lid in (lags_alvo or []):
            rx_lhdr = re.compile(
                rf'^(\s*)lag\s+{re.escape(str(_lid))}\s*$')
            linhas_cfg = (cfg_text or "").split("\n")
            for i, ln in enumerate(linhas_cfg):
                m_h = rx_lhdr.match(ln)
                if not m_h:
                    continue
                stz_l, _ = self._swap_extract_stanza(linhas_cfg, i)
                for sub in stz_l[1:]:
                    md = rx_desc_in.match(sub)
                    if md:
                        lag_descs_de[str(_lid)] = md.group(1).strip()
                        break
                break

        texto_fallback_de = self._swap_montar_bloco_fallback_de(
            portas=portas,
            lags_alvo=lags_alvo,
            rb_interfaces=rb_interfaces,
            rb_static_routes=rb_static_routes,
            rb_bgp_neighbors=rb_bgp_neighbors,
            servicos=servicos,
            svc_static_routes=svc_static_routes,
            svc_bgp_neighbors=svc_bgp_neighbors,
            port_descs=port_descs_de,
            lag_descs=lag_descs_de,
        )

        # Para o PARA, derive lista de portas-PARA do mapa.
        portas_para_fb = []
        for _de_p in (portas or []):
            _pp = (de_to_para_port or {}).get(_de_p, _de_p)
            if _pp not in portas_para_fb:
                portas_para_fb.append(_pp)
        texto_fallback_para = self._swap_montar_bloco_fallback_para(
            portas_para=portas_para_fb,
            lags_alvo=lags_alvo,
            rb_interfaces=rb_interfaces,
            servicos=servicos,
            de_to_para_port=de_to_para_port or None,
        )

        return (texto_final, lags_alvo, num_blocos, texto_shutdown_de,
                texto_testes_de, texto_testes_para,
                texto_fallback_de, texto_fallback_para)

    def _swap_montar_bloco_shutdown_de(
            self, portas, lags_alvo, rb_interfaces, rb_static_routes,
            rb_bgp_neighbors, servicos, svc_static_routes,
            svc_bgp_neighbors, ospf_iface_shutdowns=None):
        """Monta o bloco de comandos a serem executados no roteador DE
        para desabilitar tudo que sera assumido pelo roteador PARA.

        Cada comando e emitido em uma unica linha no estilo classic CLI
        do SR-OS (`/configure ...`). O bloco e envolto por `exit all`
        no inicio e no fim.

        Ignora rotas default (0.0.0.0/0 e ::/0) por seguranca.

        `ospf_iface_shutdowns` (opcional): lista de comandos prontos
        do tipo
            '/configure router ospf <id> area <a> interface "<n>" shutdown'
            '/configure service vprn <sid> ospf <id> area <a> '
                'interface "<n>" shutdown'
        Esses comandos sao emitidos ANTES do shutdown da interface L3
        correspondente (e antes do shutdown do BGP), para que o
        protocolo de roteamento retire a iface da SPF de forma
        graciosa antes do down fisico/logico.
        """
        rx_sr_entry = re.compile(
            r"^\s*(?:static-route-entry|static-route)\s+(\S+)")
        rx_nh = re.compile(r"^\s*next-hop\s+(\S+)")
        rx_grp = re.compile(r'^\s*group\s+"?([^"\s]+)"?')
        rx_nb = re.compile(r"^\s*neighbor\s+(\S+)")

        DEFAULT_PREFIXES = ("0.0.0.0/0", "::/0", "::0/0", "::0")

        def _extrair_sr(stanza):
            """Extrai pares (prefix, next_hop) de uma stanza static-route(s).

            Suporta ambos os formatos:
              - `static-route-entry <prefix>` ... `next-hop <ip>` ...
              - `static-route <prefix> next-hop <ip>` (linha unica)
              - `static-routes` ... `route <prefix> next-hop <ip>` ...

            Pares com prefix default (0.0.0.0/0, ::/0) sao descartados.
            """
            pares = []
            atual_prefix = None
            for ln in stanza:
                m_sr = rx_sr_entry.match(ln)
                if m_sr:
                    atual_prefix = m_sr.group(1)
                    # Tenta achar next-hop inline na mesma linha
                    m_inl = re.search(r"\bnext-hop\s+(\S+)", ln)
                    if m_inl and atual_prefix not in DEFAULT_PREFIXES:
                        pares.append((atual_prefix, m_inl.group(1)))
                    continue
                # `route <prefix> next-hop <ip>` dentro de static-routes
                m_route = re.match(
                    r"^\s*route\s+(\S+)\s+next-hop\s+(\S+)", ln)
                if m_route:
                    pfx = m_route.group(1)
                    if pfx not in DEFAULT_PREFIXES:
                        pares.append((pfx, m_route.group(2)))
                    continue
                m_h = rx_nh.match(ln)
                if m_h and atual_prefix:
                    pfx = atual_prefix
                    nh = m_h.group(1)
                    if pfx not in DEFAULT_PREFIXES:
                        pares.append((pfx, nh))
            # dedup mantendo ordem
            vistos = set()
            unicos = []
            for par in pares:
                if par in vistos:
                    continue
                vistos.add(par)
                unicos.append(par)
            return unicos

        def _grp_nb_de_par(par):
            """Recebe (group_header, neighbor_stanza) e devolve
            (group_name_or_None, neighbor_ip_or_None).
            """
            grp_hdr, nb_st = par
            grp = None
            if grp_hdr:
                m_g = rx_grp.match(grp_hdr)
                if m_g:
                    grp = m_g.group(1)
            nb_ip = None
            if nb_st:
                first = nb_st[0] if isinstance(nb_st, list) else str(nb_st)
                m_n = rx_nb.match(first)
                if m_n:
                    nb_ip = m_n.group(1)
            return grp, nb_ip

        out = ["exit all"]

        # Indexa comandos OSPF por (escopo, nome_iface) para emitir
        # logo após o shutdown da interface L3 correspondente. Assim
        # cada interface fica agrupada com seu OSPF.
        #   chave router base : ("router", None)
        #   chave servico     : (svc_type, svc_id)  ex.: ("vprn","7650")
        ospf_por_iface = {}
        rx_ospf_rb = re.compile(
            r'^/configure\s+router\s+ospf3?\s+\S+\s+area\s+\S+\s+'
            r'interface\s+"([^"]+)"\s+shutdown\s*$')
        rx_ospf_svc = re.compile(
            r'^/configure\s+service\s+(vprn|ies)\s+(\S+)\s+'
            r'ospf3?\s+\S+\s+area\s+\S+\s+'
            r'interface\s+"([^"]+)"\s+shutdown\s*$')
        ospf_orfaos = []  # comandos OSPF que nao casaram com nenhuma iface
        for cmd in (ospf_iface_shutdowns or []):
            m_svc = rx_ospf_svc.match(cmd)
            if m_svc:
                chave = (m_svc.group(1), m_svc.group(2))
                nome_if = m_svc.group(3)
                ospf_por_iface.setdefault((chave, nome_if), []).append(cmd)
                continue
            m_rb = rx_ospf_rb.match(cmd)
            if m_rb:
                chave = ("router", None)
                nome_if = m_rb.group(1)
                ospf_por_iface.setdefault((chave, nome_if), []).append(cmd)
                continue
            ospf_orfaos.append(cmd)

        # 1) LAGs primeiro (shutdown + remove description)
        for lid in lags_alvo or []:
            out.append(f"/configure lag {lid} shutdown")
            out.append(f"/configure lag {lid} no description")

        # 2) Portas: shutdown + remove description
        for p in portas or []:
            out.append(f"/configure port {p} shutdown")
            out.append(f"/configure port {p} no description")

        # 3) Interfaces do router Base
        for nome, _stz in rb_interfaces or []:
            out.append(
                f'/configure router interface "{nome}" shutdown')
            for cmd in ospf_por_iface.pop((("router", None), nome), []):
                out.append(cmd)

        # 4) Static-routes do router Base
        for st in rb_static_routes or []:
            for pfx, nh in _extrair_sr(st):
                out.append(
                    f"/configure router static-route-entry {pfx} "
                    f"next-hop {nh} shutdown")

        # 5) BGP neighbors do router Base
        for par in rb_bgp_neighbors or []:
            grp, nb_ip = _grp_nb_de_par(par)
            if not nb_ip:
                continue
            if grp:
                out.append(
                    f'/configure router bgp group "{grp}" '
                    f"neighbor {nb_ip} shutdown")
            else:
                out.append(
                    f"/configure router bgp neighbor {nb_ip} shutdown")

        # 6) Servicos: interfaces + (vprn) static-routes + bgp
        for svc in servicos or []:
            svc_type = svc.get("type", "")
            svc_id = svc.get("id", "")
            for nome, _stz in svc.get("ifaces", []):
                out.append(
                    f"/configure service {svc_type} {svc_id} "
                    f'interface "{nome}" shutdown')
                for cmd in ospf_por_iface.pop(
                        ((svc_type, svc_id), nome), []):
                    out.append(cmd)
            chave = (svc_type, svc_id)
            if svc_type == "vprn":
                for st in (svc_static_routes or {}).get(chave, []):
                    for pfx, nh in _extrair_sr(st):
                        out.append(
                            f"/configure service vprn {svc_id} "
                            f"static-route-entry {pfx} "
                            f"next-hop {nh} shutdown")
                for par in (svc_bgp_neighbors or {}).get(chave, []):
                    grp, nb_ip = _grp_nb_de_par(par)
                    if not nb_ip:
                        continue
                    if grp:
                        out.append(
                            f"/configure service vprn {svc_id} "
                            f'bgp group "{grp}" '
                            f"neighbor {nb_ip} shutdown")
                    else:
                        out.append(
                            f"/configure service vprn {svc_id} "
                            f"bgp neighbor {nb_ip} shutdown")

        # Comandos OSPF que nao casaram com nenhuma interface
        # emitida acima (defensivo): vai tudo no final, antes do
        # exit all, para nao perder nada.
        for resto in ospf_por_iface.values():
            for cmd in resto:
                out.append(cmd)
        for cmd in ospf_orfaos:
            out.append(cmd)

        out.append("exit all")
        return "\n".join(out) + "\n"

    # ─────────────────────────────────────────────────────────────
    #  FALLBACK DE — desfaz o shutdown e restaura descriptions
    # ─────────────────────────────────────────────────────────────
    def _swap_montar_bloco_fallback_de(
            self, portas, lags_alvo, rb_interfaces, rb_static_routes,
            rb_bgp_neighbors, servicos, svc_static_routes,
            svc_bgp_neighbors, port_descs=None, lag_descs=None):
        """Bloco de FALLBACK no roteador DE: reverte o shutdown
        aplicado e restaura as descriptions originais de
        portas/LAGs (que foram removidas no bloco de shutdown).
        """
        rx_sr_entry = re.compile(
            r"^\s*(?:static-route-entry|static-route)\s+(\S+)")
        rx_nh = re.compile(r"^\s*next-hop\s+(\S+)")
        rx_grp = re.compile(r'^\s*group\s+"?([^"\s]+)"?')
        rx_nb = re.compile(r"^\s*neighbor\s+(\S+)")
        DEFAULT_PREFIXES = ("0.0.0.0/0", "::/0", "::0/0", "::0")

        def _extrair_sr(stanza):
            pares = []
            atual_prefix = None
            for ln in stanza:
                m_sr = rx_sr_entry.match(ln)
                if m_sr:
                    atual_prefix = m_sr.group(1)
                    m_inl = re.search(r"\bnext-hop\s+(\S+)", ln)
                    if m_inl and atual_prefix not in DEFAULT_PREFIXES:
                        pares.append((atual_prefix, m_inl.group(1)))
                    continue
                m_route = re.match(
                    r"^\s*route\s+(\S+)\s+next-hop\s+(\S+)", ln)
                if m_route:
                    pfx = m_route.group(1)
                    if pfx not in DEFAULT_PREFIXES:
                        pares.append((pfx, m_route.group(2)))
                    continue
                m_h = rx_nh.match(ln)
                if m_h and atual_prefix:
                    pfx = atual_prefix
                    if pfx not in DEFAULT_PREFIXES:
                        pares.append((pfx, m_h.group(1)))
            vistos = set(); unicos = []
            for par in pares:
                if par in vistos: continue
                vistos.add(par); unicos.append(par)
            return unicos

        def _grp_nb_de_par(par):
            grp_hdr, nb_st = par
            grp = None
            if grp_hdr:
                m_g = rx_grp.match(grp_hdr)
                if m_g: grp = m_g.group(1)
            nb_ip = None
            if nb_st:
                first = (nb_st[0] if isinstance(nb_st, list)
                         else str(nb_st))
                m_n = rx_nb.match(first)
                if m_n: nb_ip = m_n.group(1)
            return grp, nb_ip

        port_descs = port_descs or {}
        lag_descs = lag_descs or {}
        out = ["exit all"]

        # 1) LAGs: no shutdown + restaura description
        for lid in lags_alvo or []:
            out.append(f"/configure lag {lid} no shutdown")
            d = lag_descs.get(str(lid)) or lag_descs.get(lid)
            if d:
                out.append(f"/configure lag {lid} description {d}")

        # 2) Portas: no shutdown + restaura description
        for p in portas or []:
            out.append(f"/configure port {p} no shutdown")
            d = port_descs.get(p)
            if d:
                out.append(f"/configure port {p} description {d}")

        # 3) Interfaces do router Base
        for nome, _stz in rb_interfaces or []:
            out.append(
                f'/configure router interface "{nome}" no shutdown')

        # 4) Static-routes do router Base
        for st in rb_static_routes or []:
            for pfx, nh in _extrair_sr(st):
                out.append(
                    f"/configure router static-route-entry {pfx} "
                    f"next-hop {nh} no shutdown")

        # 5) BGP neighbors do router Base
        for par in rb_bgp_neighbors or []:
            grp, nb_ip = _grp_nb_de_par(par)
            if not nb_ip: continue
            if grp:
                out.append(
                    f'/configure router bgp group "{grp}" '
                    f"neighbor {nb_ip} no shutdown")
            else:
                out.append(
                    f"/configure router bgp neighbor {nb_ip} "
                    "no shutdown")

        # 6) Servicos
        for svc in servicos or []:
            svc_type = svc.get("type", "")
            svc_id = svc.get("id", "")
            for nome, _stz in svc.get("ifaces", []):
                out.append(
                    f"/configure service {svc_type} {svc_id} "
                    f'interface "{nome}" no shutdown')
            chave = (svc_type, svc_id)
            if svc_type == "vprn":
                for st in (svc_static_routes or {}).get(chave, []):
                    for pfx, nh in _extrair_sr(st):
                        out.append(
                            f"/configure service vprn {svc_id} "
                            f"static-route-entry {pfx} "
                            f"next-hop {nh} no shutdown")
                for par in (svc_bgp_neighbors or {}).get(chave, []):
                    grp, nb_ip = _grp_nb_de_par(par)
                    if not nb_ip: continue
                    if grp:
                        out.append(
                            f"/configure service vprn {svc_id} "
                            f'bgp group "{grp}" '
                            f"neighbor {nb_ip} no shutdown")
                    else:
                        out.append(
                            f"/configure service vprn {svc_id} "
                            f"bgp neighbor {nb_ip} no shutdown")

        out.append("exit all")
        return "\n".join(out) + "\n"

    # ─────────────────────────────────────────────────────────────
    #  FALLBACK PARA — shutdown de portas/LAG/interfaces criadas
    # ─────────────────────────────────────────────────────────────
    def _swap_montar_bloco_fallback_para(
            self, portas_para, lags_alvo, rb_interfaces, servicos,
            de_to_para_port=None):
        """Bloco de FALLBACK no roteador PARA: shutdown das portas,
        LAGs e interfaces que foram criadas/movidas para o PARA pelo
        SWAP. Nao desfaz a configuracao (apenas isola o trafego).

        `de_to_para_port`: usado para substituir nomes de interface
        (slot/MDA do DE -> PARA) caso o nome contenha referencia ao
        port. Para o SWAP atual, os nomes de interface sao copiados
        do DE sem substituicao no header — entao na pratica este
        mapa nao altera nada e `rb_interfaces` ja tem os nomes
        corretos.
        """
        out = ["exit all"]

        # 1) LAGs
        for lid in lags_alvo or []:
            out.append(f"/configure lag {lid} shutdown")

        # 2) Portas (do PARA)
        for p in portas_para or []:
            out.append(f"/configure port {p} shutdown")

        # 3) Interfaces do router Base
        for nome, _stz in rb_interfaces or []:
            out.append(
                f'/configure router interface "{nome}" shutdown')

        # 4) Servicos
        for svc in servicos or []:
            svc_type = svc.get("type", "")
            svc_id = svc.get("id", "")
            for nome, _stz in svc.get("ifaces", []):
                out.append(
                    f"/configure service {svc_type} {svc_id} "
                    f'interface "{nome}" shutdown')

        out.append("exit all")
        return "\n".join(out) + "\n"

    # ─────────────────────────────────────────────────────────────
    #  Script_SWAP — Bloco de TESTES (validacao por interface)
    # ─────────────────────────────────────────────────────────────
    def _swap_montar_bloco_testes(self, rb_interfaces, servicos,
                                  bgp_peers_por_escopo=None,
                                  proto_por_iface=None,
                                  de_to_para_port=None):
        """Monta o bloco de comandos de validacao (ping + show ...)
        para o roteador DE ou PARA, reaproveitando a mesma coleta de
        interfaces/SAPs/peers que monta o .txt de configuracao.

        Para cada interface da router Base e de cada servico
        (vpls/epipe/ies/vprn) extraidos, gera:
          - ping no proprio IP (v4 e v6) e no peer (v4 e v6)
          - show router (ospf/mpls/rsvp/ldp/pim/arp) na router Base
          - show service id <id> base [+ fdb detail no vpls]
          - show service sap-using sap <SAP>
          - show router [<vprn_id>] arp "<iface>"
          - show lag <N> description (se a iface estiver em LAG)
          - show router [<vprn_id>] route-table summary [ipv6]
          - show router [<vprn_id>] bgp summary neighbor <peer>
            (somente quando ha BGP neighbor configurado p/ aquele peer)
          - show router [<vprn_id>] ospf all neighbor [detail]
            (somente quando a iface participa de OSPF no escopo)

        Parametros adicionais (todos opcionais; se nao informados, o
        comportamento e equivalente a "emitir tudo"):

          - `bgp_peers_por_escopo`: dict {escopo -> set(neighbor_ip)}
              escopo "rb"             -> neighbors do BGP do router Base
              escopo ("vprn", svc_id) -> neighbors do BGP da VPRN
            Define se um peer-IP da iface tem realmente BGP no escopo.

          - `proto_por_iface`: dict {(escopo..., iface_name) -> set(proto)}
              ex.: ("rb", "TO_X")           -> {"ospf", "mpls", ...}
                   ("vprn", "1009041", "X") -> {"ospf"}
            Define quais protocolos referenciam a iface naquele escopo
            (usado para emitir 'show router ospf ...' so quando aplicavel).

        Quando `de_to_para_port` (dict {de_port: para_port}) for
        informado, aplica substituicao DE -> PARA nos nomes/SAPs que
        carregam o slot/MDA da porta (ex.: "Tengige1/1/c3/4:1502.685"
        -> "Tengige2/1/c3/4:1502.685").
        """
        bgp_peers_por_escopo = bgp_peers_por_escopo or {}
        proto_por_iface = proto_por_iface or {}

        def _iface_tem_proto(escopo_key, nome, *protos):
            """True se a iface tem QUALQUER um dos protos no escopo.
            `escopo_key` e tupla: ("rb",) ou ("vprn", sid)."""
            ps = proto_por_iface.get(escopo_key + (nome,)) or set()
            return any(p in ps for p in protos)

        def _bgp_peer_ativo(escopo, peer_ip):
            """True se o peer-IP esta na lista de BGP neighbors do escopo."""
            if not peer_ip:
                return False
            return peer_ip in (bgp_peers_por_escopo.get(escopo) or set())

        def _bgp_peers_na_subnet(escopo, addr_str, plen, is_v6=False):
            """Retorna a lista de peer-IPs do escopo que pertencem a
            mesma sub-rede da iface. Util quando a iface nao e /30
            (ex.: /29) e o peer nao e simplesmente o 'outro host'."""
            try:
                net = ipaddress.ip_interface(
                    f"{addr_str}/{plen}").network
            except Exception:
                return []
            out = []
            for ip in (bgp_peers_por_escopo.get(escopo) or set()):
                try:
                    pip = ipaddress.ip_address(ip)
                except Exception:
                    continue
                if pip in net and str(pip) != addr_str:
                    out.append(str(pip))
            return out

        def _peer_ip(addr_str, plen, is_v6=False):
            try:
                iface = ipaddress.ip_interface(f"{addr_str}/{plen}")
            except Exception:
                return None
            net = iface.network
            # /31 (IPv4) ou /127 (IPv6): apenas 2 enderecos, ambos hosts
            if (not is_v6 and plen == 31) or (is_v6 and plen == 127):
                for h in net:
                    if str(h) != addr_str:
                        return str(h)
                return None
            # caso geral: primeiro host != self
            try:
                for h in net.hosts():
                    if str(h) != addr_str:
                        return str(h)
            except Exception:
                return None
            return None

        def _extract_addrs(stanza):
            """Devolve (v4, v4_peer, v4_plen, v6, v6_peer, v6_plen)."""
            v4 = v4p = v6 = v6p = None
            v4plen = v6plen = None
            for ln in stanza:
                if v4 is None:
                    m = re.search(
                        r"^\s*address\s+(\d+\.\d+\.\d+\.\d+)/(\d+)\b",
                        ln,
                    )
                    if m:
                        v4 = m.group(1)
                        v4plen = int(m.group(2))
                        v4p = _peer_ip(v4, v4plen, is_v6=False)
                        continue
                if v6 is None:
                    m6 = re.search(
                        r"^\s*address\s+([0-9a-fA-F:]+)/(\d+)\b",
                        ln,
                    )
                    if m6 and ":" in m6.group(1):
                        v6 = m6.group(1)
                        v6plen = int(m6.group(2))
                        v6p = _peer_ip(v6, v6plen, is_v6=True)
            return (v4, v4p, v4plen, v6, v6p, v6plen)

        def _bgp_peers_da_iface(escopo, v4, v4p, v4plen, v6, v6p, v6plen):
            """Conjunto de peer-IPs do escopo associados a iface
            (combina /30 peer + neighbors na mesma sub-rede). Usado
            para emitir 'show router [<sid>] bgp summary neighbor X'
            apenas quando ha peer BGP de fato relacionado."""
            peers = set()
            if v4p and _bgp_peer_ativo(escopo, v4p):
                peers.add(v4p)
            if v6p and _bgp_peer_ativo(escopo, v6p):
                peers.add(v6p)
            if v4 and v4plen is not None:
                for p in _bgp_peers_na_subnet(escopo, v4, v4plen, False):
                    peers.add(p)
            if v6 and v6plen is not None:
                for p in _bgp_peers_na_subnet(escopo, v6, v6plen, True):
                    peers.add(p)
            return peers

        def _extract_sap_e_lag(stanza):
            """Extrai SAP id e LAG id (se houver) da iface IES/VPRN."""
            for ln in stanza:
                m = re.match(r"^\s*sap\s+(\S+)", ln)
                if not m:
                    continue
                sap_id = m.group(1)
                m_l = re.match(r"^lag-(\d+)", sap_id)
                lag_n = m_l.group(1) if m_l else None
                return sap_id, lag_n
            return None, None

        def _is_network_iface(stanza):
            """True se a 1a linha for `network-interface "..."` (VPRN CSC)."""
            return bool(stanza and re.match(
                r'^\s*network-interface\s+"', stanza[0]))

        out = []

        def _add(s=""):
            out.append(s)

        # ── ROUTER BASE ─────────────────────────────────────────
        if rb_interfaces:
            _add("# ===== router Base =====")
            for nome, stanza in rb_interfaces:
                v4, v4p, v4pl, v6, v6p, v6pl = _extract_addrs(stanza)
                tem_ospf = _iface_tem_proto(("rb",), nome, "ospf", "ospf3")
                tem_mpls = _iface_tem_proto(("rb",), nome, "mpls")
                tem_rsvp = _iface_tem_proto(("rb",), nome, "rsvp")
                tem_ldp = _iface_tem_proto(("rb",), nome, "ldp")
                tem_pim = _iface_tem_proto(("rb",), nome, "pim")
                bgp_peers_iface = _bgp_peers_da_iface(
                    "rb", v4, v4p, v4pl, v6, v6p, v6pl)
                _add("")
                _add(f'#interface "{nome}"')
                _add("#")
                if v4:
                    _add(f"ping {v4} count 50 rapid")
                if v4p:
                    _add(f"ping {v4p} count 50 rapid")
                if v6:
                    _add(f"ping {v6} count 50 rapid")
                if v6p:
                    _add(f"ping {v6p} count 50 rapid")
                if tem_ospf:
                    _add(f'show router ospf all neighbor "{nome}"')
                    _add(f'show router ospf all neighbor "{nome}" detail')
                if tem_mpls:
                    _add(f'show router mpls interface "{nome}" detail')
                if tem_rsvp:
                    _add(f'show router rsvp interface "{nome}"')
                    _add(f'show router rsvp interface "{nome}" detail')
                if tem_ldp:
                    _add(f'show router ldp interface "{nome}"')
                    _add(f'show router ldp interface "{nome}" detail')
                if tem_pim:
                    _add(f'show router pim neighbor "{nome}"')
                    _add(f'show router pim neighbor "{nome}" detail')
                for _bp in sorted(bgp_peers_iface):
                    _add(f"show router bgp summary neighbor {_bp}")
                _add(f'show router arp "{nome}"')
                _add("#")

        # ── SERVICOS ────────────────────────────────────────────
        for svc in (servicos or []):
            tipo = (svc.get("type") or "").lower()
            sid = svc.get("id") or ""

            if tipo == "vpls":
                _add("")
                _add(f"# ===== VPLS {sid} =====")
                _add(f"show service id {sid} base")
                _add(f"show service id {sid} fdb detail")
                continue

            if tipo == "epipe":
                _add("")
                _add(f"# ===== EPIPE {sid} =====")
                _add(f"show service id {sid} base")
                continue

            if tipo in ("ipipe", "apipe", "fpipe", "cpipe"):
                _add("")
                _add(f"# ===== {tipo.upper()} {sid} =====")
                _add(f"show service id {sid} base")
                continue

            if tipo == "ies":
                # IES nao tem instancia de routing propria: BGP/OSPF
                # quando existem para a iface estao no router Base.
                for nome, stanza in svc.get("ifaces", []):
                    v4, v4p, v4pl, v6, v6p, v6pl = _extract_addrs(stanza)
                    sap_id, lag_n = _extract_sap_e_lag(stanza)
                    is_net = _is_network_iface(stanza)
                    tem_ospf = _iface_tem_proto(
                        ("rb",), nome, "ospf", "ospf3")
                    bgp_peers_iface = _bgp_peers_da_iface(
                        "rb", v4, v4p, v4pl, v6, v6p, v6pl)
                    _add("")
                    _add(f"#ies {sid}")
                    _add(f'#interface "{nome}"')
                    _add("#")
                    if v4:
                        _add(f"ping {v4} count 50 rapid")
                    if v4p:
                        _add(f"ping {v4p} count 50 rapid")
                    if v6:
                        _add(f"ping {v6} count 50 rapid")
                    if v6p:
                        _add(f"ping {v6p} count 50 rapid")
                    _add(f"show service id {sid} base")
                    if sap_id:
                        _add(f"show service sap-using sap {sap_id}")
                    if not is_net:
                        _add(f'show router arp "{nome}"')
                    if lag_n:
                        _add(f"show lag {lag_n} description")
                    _add("show router route-table summary")
                    if v6:
                        _add("show router route-table summary ipv6")
                    for _bp in sorted(bgp_peers_iface):
                        _add(f"show router bgp summary neighbor {_bp}")
                    if tem_ospf:
                        _add("show router ospf all neighbor")
                        _add("show router ospf all neighbor detail")
                    _add("#")
                continue

            if tipo == "vprn":
                escopo = ("vprn", sid)
                for nome, stanza in svc.get("ifaces", []):
                    v4, v4p, v4pl, v6, v6p, v6pl = _extract_addrs(stanza)
                    sap_id, lag_n = _extract_sap_e_lag(stanza)
                    is_net = _is_network_iface(stanza)
                    tem_ospf = _iface_tem_proto(
                        escopo, nome, "ospf", "ospf3")
                    bgp_peers_iface = _bgp_peers_da_iface(
                        escopo, v4, v4p, v4pl, v6, v6p, v6pl)
                    _add("")
                    _add(f"#vprn {sid}")
                    _add(f'#interface "{nome}"')
                    _add("#")
                    if v4:
                        _add(f"ping router {sid} {v4} count 50 rapid")
                    if v4p:
                        _add(f"ping router {sid} {v4p} count 50 rapid")
                    if v6:
                        _add(f"ping router {sid} {v6} count 50 rapid")
                    if v6p:
                        _add(f"ping router {sid} {v6p} count 50 rapid")
                    _add(f"show service id {sid} base")
                    if sap_id:
                        _add(f"show service sap-using sap {sap_id}")
                    if not is_net:
                        _add(f'show router {sid} arp "{nome}"')
                    if lag_n:
                        _add(f"show lag {lag_n} description")
                    _add(f"show router {sid} route-table summary")
                    if v6:
                        _add(f"show router {sid} route-table summary ipv6")
                    for _bp in sorted(bgp_peers_iface):
                        _add(
                            f"show router {sid} bgp summary "
                            f"neighbor {_bp}")
                    if tem_ospf:
                        _add(f"show router {sid} ospf all neighbor")
                        _add(
                            f"show router {sid} ospf all neighbor "
                            f"detail")
                    _add("#")
                continue

        texto = "\n".join(out)

        # ── Substituicao DE -> PARA (modo PARA) ─────────────────
        # Substitui ocorrencias de cada porta DE pelo equivalente
        # PARA, com word boundary permissiva (mesmo padrao usado
        # na geracao do .txt de config). Cobre nomes tipo
        # "Tengige1/1/c3/4:1502.685" e SAPs "1/1/c3/4:1502.685".
        if de_to_para_port:
            chaves_ord = sorted(
                de_to_para_port.keys(), key=len, reverse=True)
            for de_p in chaves_ord:
                para_p = de_to_para_port[de_p]
                if not para_p or de_p == para_p:
                    continue
                texto = re.sub(
                    r"(?<![\w/-])" + re.escape(de_p) + r"(?![\w/-])",
                    para_p, texto)

            # Pass adicional: dentro de nomes de interface entre aspas
            # (ex.: '#interface "Tengige1/1/1"' e
            # 'show router [<id>] arp "Tengige1/1/1"'), a porta fica
            # COLADA a um prefixo alfanumerico (Tengige, 100ge, etc.),
            # entao o lookbehind acima nao casa. Aqui aplicamos a
            # substituicao SEM lookbehind, restrita a essas linhas, e
            # uppercase no trecho substituido (padrao Nokia: X1/2/C6/1).
            rx_quoted_iface = re.compile(
                r'(#interface\s+"|show\s+router(?:\s+\S+)?\s+arp\s+")'
                r'([^"]+)(")')

            def _rename_iface(m):
                prefix, name, suffix = m.group(1), m.group(2), m.group(3)
                novo = name
                for de_p in chaves_ord:
                    para_p = de_to_para_port[de_p]
                    if not para_p or de_p == para_p:
                        continue
                    pat = re.compile(re.escape(de_p), re.IGNORECASE)
                    novo = pat.sub(para_p.upper(), novo)
                return prefix + novo + suffix

            texto = rx_quoted_iface.sub(_rename_iface, texto)

        return texto

    def mostrar_upgrade(self):
        self._limpar_content()
        if not self.page_upgrade:
            self.page_upgrade = tk.Frame(self.content_frame, bg=COR_CINZA_FUNDO)
            self.page_upgrade.pack(fill=tk.BOTH, expand=True)

            # Container horizontal para relógios + campos
            topo_frame = tk.Frame(self.page_upgrade, bg=COR_CINZA_FUNDO, height=130)
            topo_frame.pack(pady=(4, 2), padx=24, fill=tk.X)
            topo_frame.pack_propagate(False)


            # Relógios à esquerda (fundo preto, letra vermelha, textos atualizados)
            clock_frame = tk.Frame(topo_frame, bg="#000000", bd=2, relief="groove",
                                   highlightbackground=COR_CINZA_BORDA, highlightcolor=COR_CINZA_BORDA, highlightthickness=1)
            clock_frame.pack(side=tk.LEFT, anchor="nw")

            tk.Label(clock_frame, text="Horário local", font=("Segoe UI", 7, "bold"), bg="#000000", fg="#FF2222").pack(anchor="w", padx=6, pady=(2, 0))
            self.upg_clock_brasilia = tk.Label(clock_frame, text="--:--:--", font=("Consolas", 12, "bold"), bg="#000000", fg="#FF2222")
            self.upg_clock_brasilia.pack(anchor="w", padx=6)

            tk.Label(clock_frame, text="Tempo do upgrade", font=("Segoe UI", 7, "bold"), bg="#000000", fg="#FF2222").pack(anchor="w", padx=6, pady=(2, 0))
            self.upg_clock_elapsed = tk.Label(clock_frame, text="00:00:00", font=("Consolas", 12, "bold"), bg="#000000", fg="#FF2222")
            self.upg_clock_elapsed.pack(anchor="w", padx=6, pady=(0, 2))

            self._upgrade_start_time = None
            self._upgrade_clock_running = False
            self._upgrade_timer_running = False

            def _atualizar_relogios():
                # Hora de Brasília (UTC-3)
                agora = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=-3)))
                self.upg_clock_brasilia.config(text=agora.strftime("%H:%M:%S"))
                # Tempo do upgrade (cronômetro)
                if self._upgrade_timer_running and self._upgrade_start_time:
                    delta = datetime.datetime.now() - self._upgrade_start_time
                    total_s = int(delta.total_seconds())
                    h, rem = divmod(total_s, 3600)
                    m, s = divmod(rem, 60)
                    self.upg_clock_elapsed.config(text=f"{h:02d}:{m:02d}:{s:02d}")
                self.root.after(1000, _atualizar_relogios)

            # Inicia relógio local imediatamente
            self._upgrade_timer_running = False
            self._upgrade_start_time = None
            _atualizar_relogios()

            # Função para iniciar o cronômetro do upgrade
            def iniciar_cronometro_upgrade():
                self._upgrade_start_time = datetime.datetime.now()
                self._upgrade_timer_running = True
            self._iniciar_cronometro_upgrade = iniciar_cronometro_upgrade

            # Frame com borda para título e campos (centralizado sobre as caixas Log SSH e Console)
            campos_borda = tk.Frame(topo_frame, bg=COR_CINZA_FUNDO, bd=1, relief="groove", highlightbackground=COR_CINZA_BORDA, highlightcolor=COR_CINZA_BORDA, highlightthickness=1)
            campos_borda.place(relx=0.5, rely=0, anchor="n")

            # Título junto dos campos
            label = tk.Label(campos_borda, text="Upgrade de software - 7x50",
                            font=("Segoe UI", 11, "bold"), bg=COR_CINZA_FUNDO, fg=COR_AZUL_NOKIA)
            label.pack(pady=(2, 1), anchor="n")

            # Frame dos campos dentro da borda
            campos_frame = tk.Frame(campos_borda, bg=COR_CINZA_FUNDO)
            campos_frame.pack(pady=(0, 2), anchor="n")


            # Hostname com botão Ir (compacto)
            tk.Label(campos_frame, text="Hostname:", font=("Segoe UI", 8), bg=COR_CINZA_FUNDO, anchor="w", width=11).grid(row=0, column=0, sticky="w", pady=0)
            # Frame para Entry e botão lado a lado
            hostname_container = tk.Frame(campos_frame, bg=COR_CINZA_FUNDO)
            hostname_container.grid(row=0, column=1, sticky="w", pady=0)
            # Entry de hostname (agora sem botão Ir)
            self.upg_hostname = tk.Entry(hostname_container, font=("Segoe UI", 8), width=14, relief="solid", bd=1)
            self.upg_hostname.pack(side=tk.LEFT, fill=tk.X, expand=True)
            # Permitir Enter para acionar lógica e iniciar cronômetro
            def on_hostname_return(event=None):
                self._iniciar_cronometro_upgrade()
                self._upgrade_resolver_e_conectar(event)
            self.upg_hostname.bind("<Return>", on_hostname_return)
            # Botão SSH ao lado do hostname
            self.btn_upg_ssh = tk.Button(
                hostname_container, text="\U0001F5A5", font=("Segoe UI", 8),
                bg=COR_AZUL_NOKIA, fg=COR_BRANCO, bd=0, relief="flat",
                activebackground=COR_AZUL_CLARO, activeforeground=COR_BRANCO,
                width=3, command=self._upgrade_ssh_conectar,
                cursor="hand2",
            )
            self.btn_upg_ssh.pack(side=tk.LEFT, padx=(4, 0))

            # IP (readonly, fundo escuro, compacto)
            tk.Label(campos_frame, text="IP:", font=("Segoe UI", 8), bg=COR_CINZA_FUNDO, anchor="w", width=11).grid(row=1, column=0, sticky="w", pady=0)
            self.upg_ip = tk.Entry(campos_frame, font=("Segoe UI", 8), width=14, state="readonly", readonlybackground="#E5E7EB", fg="#1E293B")
            self.upg_ip.grid(row=1, column=1, sticky="w", pady=0)

            # Chassis (readonly, fundo escuro, compacto)
            tk.Label(campos_frame, text="Chassis:", font=("Segoe UI", 8), bg=COR_CINZA_FUNDO, anchor="w", width=11).grid(row=2, column=0, sticky="w", pady=0)
            self.upg_roteador = tk.Entry(campos_frame, font=("Segoe UI", 8), width=14, state="readonly", readonlybackground="#E5E7EB", fg="#1E293B")
            self.upg_roteador.grid(row=2, column=1, sticky="w", pady=0)

            # Versão de software atual (readonly, fundo escuro, compacto)
            tk.Label(campos_frame, text="Versão atual:", font=("Segoe UI", 8), bg=COR_CINZA_FUNDO, anchor="w", width=11).grid(row=3, column=0, sticky="w", pady=0)
            self.upg_versao = tk.Entry(campos_frame, font=("Segoe UI", 8), width=14, state="readonly", readonlybackground="#E5E7EB", fg="#1E293B")
            self.upg_versao.grid(row=3, column=1, sticky="w", pady=0)


            # Frame à direita para campos: Satelitte, Versão Satelitte, Rede de controle
            right_frame = tk.Frame(campos_frame, bg=COR_CINZA_FUNDO)
            right_frame.grid(row=0, column=3, rowspan=7, padx=(24,8), pady=0, sticky="n")

            # Satelitte (readonly, compacto)
            tk.Label(right_frame, text="Satelitte:", font=("Segoe UI", 8), bg=COR_CINZA_FUNDO, anchor="w").grid(row=0, column=0, sticky="w", pady=0)
            self.upg_satelitte = tk.Entry(right_frame, font=("Segoe UI", 8), width=14, state="readonly", readonlybackground="#E5E7EB", fg="#1E293B")
            self.upg_satelitte.grid(row=0, column=1, sticky="w", pady=0, padx=(4,0))

            # Versão Satelitte (readonly, compacto)
            tk.Label(right_frame, text="Versão Sat.:", font=("Segoe UI", 8), bg=COR_CINZA_FUNDO, anchor="w").grid(row=1, column=0, sticky="w", pady=0)
            self.upg_control = tk.Entry(right_frame, font=("Segoe UI", 8), width=14, state="readonly", readonlybackground="#E5E7EB", fg="#1E293B")
            self.upg_control.grid(row=1, column=1, sticky="w", pady=0, padx=(4,0))

            # Rede de controle (readonly, compacto)
            tk.Label(right_frame, text="Rede ctrl.:", font=("Segoe UI", 8), bg=COR_CINZA_FUNDO, anchor="w").grid(row=2, column=0, sticky="w", pady=0)
            self.upg_rede_controle = tk.Entry(right_frame, font=("Segoe UI", 8), width=14, state="readonly", readonlybackground="#E5E7EB", fg="#1E293B")
            self.upg_rede_controle.grid(row=2, column=1, sticky="w", pady=0, padx=(4,0))

            # Ctrl ativa (readonly, compacto)
            tk.Label(right_frame, text="Ctrl ativa:", font=("Segoe UI", 8), bg=COR_CINZA_FUNDO, anchor="w").grid(row=3, column=0, sticky="w", pady=0)
            self.upg_ctrl_ativa = tk.Entry(right_frame, font=("Segoe UI", 8), width=14, state="readonly", readonlybackground="#E5E7EB", fg="#1E293B")
            self.upg_ctrl_ativa.grid(row=3, column=1, sticky="w", pady=0, padx=(4,0))


            # Caixas de log maiores, alinhadas, mesmo tamanho, com borda e título
            log_frame = tk.Frame(self.page_upgrade, bg=COR_CINZA_FUNDO)
            log_frame.pack(fill=tk.BOTH, expand=True, padx=24, pady=(4, 8))


            log_font = ("Lucida Console", 7)

            # Usar grid para garantir tamanhos idênticos
            log_frame.columnconfigure(0, weight=1, uniform="log")
            log_frame.columnconfigure(1, weight=1, uniform="log")
            log_frame.rowconfigure(0, weight=1)

            # Widget com borda e título Log SSH
            log1_widget = tk.Frame(log_frame, bg=COR_CINZA_FUNDO, bd=2, relief="groove", highlightbackground=COR_CINZA_BORDA, highlightcolor=COR_CINZA_BORDA, highlightthickness=2)
            log1_widget.grid(row=0, column=0, sticky="nsew", padx=(0,8), pady=0)
            tk.Label(log1_widget, text="Log SSH", font=("Segoe UI", 12, "bold"), bg=COR_CINZA_FUNDO, fg=COR_AZUL_NOKIA).pack(anchor="w", padx=8, pady=(4,0))
            log1_inner = tk.Frame(log1_widget, bg=COR_CINZA_FUNDO)
            log1_inner.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0,8))
            log1_scroll = tk.Scrollbar(log1_inner, orient="vertical", bg="#333", troughcolor="#222", activebackground="#555")
            log1_scroll.pack(side=tk.RIGHT, fill=tk.Y)
            self.upg_log1 = tk.Text(
                log1_inner, bg="#111", fg="#00FF00",
                font=log_font, insertbackground="#00FF00", wrap="none",
                bd=0, relief="flat", yscrollcommand=log1_scroll.set
            )
            self.upg_log1.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
            log1_scroll.config(command=self.upg_log1.yview)

            # Widget com borda e título Console
            log2_widget = tk.Frame(log_frame, bg=COR_CINZA_FUNDO, bd=2, relief="groove", highlightbackground=COR_CINZA_BORDA, highlightcolor=COR_CINZA_BORDA, highlightthickness=2)
            log2_widget.grid(row=0, column=1, sticky="nsew", padx=(8,0), pady=0)
            tk.Label(log2_widget, text="Console", font=("Segoe UI", 12, "bold"), bg=COR_CINZA_FUNDO, fg=COR_AZUL_NOKIA).pack(anchor="w", padx=8, pady=(4,0))
            log2_inner = tk.Frame(log2_widget, bg=COR_CINZA_FUNDO)
            log2_inner.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0,8))
            log2_scroll = tk.Scrollbar(log2_inner, orient="vertical", bg="#333", troughcolor="#222", activebackground="#555")
            log2_scroll.pack(side=tk.RIGHT, fill=tk.Y)
            self.upg_log2 = tk.Text(
                log2_inner, bg="#111", fg="#00FF00",
                font=log_font, insertbackground="#00FF00", wrap="none",
                bd=0, relief="flat", yscrollcommand=log2_scroll.set
            )
            self.upg_log2.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
            log2_scroll.config(command=self.upg_log2.yview)

            # Campo de entrada de comando para Console
            console_cmd_frame = tk.Frame(log2_widget, bg="#111")
            console_cmd_frame.pack(fill=tk.X, padx=8, pady=(0, 4))
            tk.Label(console_cmd_frame, text=">", font=log_font, bg="#111", fg="#00FF00").pack(side=tk.LEFT)
            self.console_cmd_entry = tk.Entry(
                console_cmd_frame, font=log_font, bg="#111", fg="#00FF00",
                insertbackground="#00FF00", relief="flat", bd=0,
            )
            self.console_cmd_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(2, 0))
            self.console_cmd_entry.bind("<Return>", self._console_enviar_cmd)

            # Inicializar estado da sessão
            self._rc_shell = None
            self._rc_jump_client = None
            self._rc_reader_running = False
            self._rc_first_cmd_sent = False
            self._rc_fallback_to_ssh = False

            # Inicializar emulador VT100 para Console
            self._vt_screen = pyte.Screen(200, 50)
            self._vt_stream = pyte.Stream(self._vt_screen)
            self._vt_screen.set_mode(pyte.modes.LNM)
            self._vt_tags = set()

        else:
            self.page_upgrade.pack(fill=tk.BOTH, expand=True)

    def _upgrade_log(self, msg):
        # Intercepta erros NOKIA
        erro = detectar_erro_nokia(msg)
        if erro and not getattr(self, '_suppress_nokia_error_box', False):
            self._exibir_caixa_erro_nokia(erro)
        self.upg_log1.insert(tk.END, msg + "\n")
        self.upg_log1.see(tk.END)
        self.upg_log1.update_idletasks()

    def _console_enviar_cmd(self, event=None):
        """Envia comando digitado pelo usuário para o shell da rede de controle."""
        cmd = self.console_cmd_entry.get()
        self.console_cmd_entry.delete(0, tk.END)
        if not hasattr(self, '_rc_shell') or self._rc_shell is None:
            self._upgrade_log2("[ERRO] Sessão de rede de controle não está ativa.")
            return
        try:
            self._rc_shell.send(cmd + "\n")
            # No primeiro comando após conexão, enviar # + login + senha + show commands
            if not self._rc_first_cmd_sent:
                self._rc_first_cmd_sent = True
                def _login_sequence():
                    time.sleep(2)
                    self._rc_shell.send("#\n")
                    time.sleep(1)
                    self._rc_shell.send("93191142\n")
                    time.sleep(2)
                    self._rc_shell.send("X%aA5&z3\n")
                    time.sleep(5)

                    self._rc_shell.send("environment no more\n")
                    time.sleep(1)

                    try:
                        _qtd_sat_pre = int(self.upg_satelitte.get().strip())
                    except Exception:
                        _qtd_sat_pre = 0
                    if _qtd_sat_pre != 0:
                        self._rc_shell.send("show system satellite\n")
                        time.sleep(3)
                    self._rc_shell.send("show sfm\n")
                    time.sleep(3)
                    self._rc_shell.send("show card state\n")
                    time.sleep(5)

                    # Banner 3D flutuante de confirmação antes do upgrade
                    import queue as _banner_q_mod
                    banner_q = _banner_q_mod.Queue()
                    def _mostrar_banner():
                        dlg = tk.Toplevel(self.root)
                        dlg.title("⚠ CONFIRMAÇÃO DE UPGRADE ⚠")
                        dlg.resizable(False, False)
                        dlg.overrideredirect(False)
                        dlg.attributes("-topmost", True)

                        # Frame exterior com efeito 3D forte
                        outer = tk.Frame(dlg, bg="#1a1a2e", bd=6, relief="raised")
                        outer.pack(fill=tk.BOTH, expand=True)
                        mid = tk.Frame(outer, bg="#16213e", bd=4, relief="ridge")
                        mid.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)
                        inner = tk.Frame(mid, bg="#0f3460", bd=3, relief="groove")
                        inner.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)

                        # Título com cores fortes
                        tk.Label(
                            inner, text="⚠  ATENÇÃO — UPGRADE DE SOFTWARE  ⚠",
                            font=("Segoe UI", 14, "bold"), bg="#0f3460", fg="#FF4444",
                        ).pack(padx=30, pady=(18, 6))

                        # Separador visual
                        tk.Frame(inner, bg="#e94560", height=3).pack(fill=tk.X, padx=20, pady=6)

                        # Perguntas com checkboxes
                        perguntas = [
                            "Tudo pronto para o Upgrade?",
                            "Time de campo de sobre-aviso?",
                            "OMR autorizou?",
                            "NOC autorizou?",
                        ]
                        check_vars = []
                        check_labels = []
                        for p in perguntas:
                            var = tk.IntVar(value=0)
                            check_vars.append(var)
                            cb = tk.Checkbutton(
                                inner, text=p,
                                variable=var,
                                font=("Segoe UI", 12, "bold"), bg="#0f3460",
                                fg="#FF4444", selectcolor="#0f3460",
                                activebackground="#0f3460", activeforeground="#FF4444",
                                anchor="w", bd=0, highlightthickness=0,
                            )
                            cb.pack(padx=40, pady=3, anchor="w")
                            check_labels.append(cb)

                        # Separador
                        tk.Frame(inner, bg="#e94560", height=3).pack(fill=tk.X, padx=20, pady=8)

                        # Aviso de reinicialização
                        tk.Label(
                            inner,
                            text="⚠  Entenda que ao dar OK o roteador será reiniciado!",
                            font=("Segoe UI", 11, "bold"), bg="#0f3460", fg="#FFD700",
                            wraplength=480,
                        ).pack(padx=30, pady=(4, 6))

                        tk.Label(
                            inner, text="Deseja prosseguir?",
                            font=("Segoe UI", 13, "bold"), bg="#0f3460", fg="#FFFFFF",
                        ).pack(padx=30, pady=(4, 12))

                        # Botões
                        bf = tk.Frame(inner, bg="#0f3460")
                        bf.pack(pady=(0, 18))
                        btn_sim = tk.Button(
                            bf, text="SIM — Prosseguir", font=("Segoe UI", 11, "bold"),
                            bg="#555555", fg="#999999", activebackground="#555555",
                            width=18, height=2, bd=3, relief="raised",
                            state=tk.DISABLED,
                            command=lambda: (dlg.destroy(), banner_q.put(True)),
                        )
                        btn_sim.pack(side=tk.LEFT, padx=10)
                        tk.Button(
                            bf, text="NÃO — Cancelar", font=("Segoe UI", 11, "bold"),
                            bg="#dc3545", fg="white", activebackground="#c82333",
                            width=18, height=2, bd=3, relief="raised",
                            command=lambda: (dlg.destroy(), banner_q.put(False)),
                        ).pack(side=tk.LEFT, padx=10)

                        def _atualizar_checks(*_args):
                            todos_ok = True
                            for i, var in enumerate(check_vars):
                                if var.get():
                                    check_labels[i].config(fg="#00FF88", activeforeground="#00FF88")
                                else:
                                    check_labels[i].config(fg="#FF4444", activeforeground="#FF4444")
                                    todos_ok = False
                            if todos_ok:
                                btn_sim.config(state=tk.NORMAL, bg="#28a745", fg="white", activebackground="#218838")
                            else:
                                btn_sim.config(state=tk.DISABLED, bg="#555555", fg="#999999", activebackground="#555555")

                        for var in check_vars:
                            var.trace_add("write", _atualizar_checks)

                        dlg.update_idletasks()
                        w = dlg.winfo_reqwidth()
                        h = dlg.winfo_reqheight()
                        x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
                        y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
                        dlg.geometry(f"+{x}+{y}")
                        dlg.focus_set()
                        dlg.grab_set()
                    self.root.after(0, _mostrar_banner)
                    resposta = banner_q.get()
                    if resposta:
                        self._upgrade_log2("[OK] Upgrade confirmado pelo usuário.")

                        # Executar reboot dos satélites (se qtd > 0)
                        try:
                            qtd_sat_str = self.upg_satelitte.get().strip()
                            qtd_sat_reboot = int(qtd_sat_str) if qtd_sat_str.isdigit() else 0
                        except Exception:
                            qtd_sat_reboot = 0

                        if qtd_sat_reboot > 0:
                            for sat_id in range(1, qtd_sat_reboot + 1):
                                cmd_sat_reboot = f"admin satellite eth-sat {sat_id} reboot upgrade now"
                                self._upgrade_log2(f"[CMD] {cmd_sat_reboot}")
                                self._rc_shell.send(cmd_sat_reboot + "\n")
                                time.sleep(5)

                        # Executar admin reboot upgrade now
                        self._upgrade_log2("[CMD] admin reboot upgrade now")
                        self._rc_shell.send("admin reboot upgrade now\n")
                        self._upgrade_log2("[INFO] Roteador em processo de reload. Mantendo sessão ativa para capturar output...")

                        # Iniciar keepalive + reconexão no Log SSH (caixa de fundo preto)
                        threading.Thread(target=self._keepalive_e_reconectar, daemon=True).start()
                    else:
                        self._upgrade_log2("[INFO] Upgrade cancelado pelo usuário.")
                threading.Thread(target=_login_sequence, daemon=True).start()
        except Exception as e:
            self._upgrade_log2(f"[ERRO] Falha ao enviar comando: {e}")

    def _rc_reader_loop(self):
        """Thread que lê continuamente as respostas do shell da rede de controle."""
        _login_buffer = ""
        while self._rc_reader_running:
            try:
                if self._rc_shell and self._rc_shell.recv_ready():
                    data = self._rc_shell.recv(65535).decode("utf-8", errors="replace")
                    if data:
                        if getattr(self, '_rc_fallback_to_ssh', False):
                            self.root.after(0, self._upgrade_log, _limpar_ansi(data).strip())
                        else:
                            self.root.after(0, self._upgrade_console_raw, data)
                        # Auto-detect login/password prompts
                        _login_buffer += data
                        # Keep only last 500 chars to avoid unbounded growth
                        if len(_login_buffer) > 500:
                            _login_buffer = _login_buffer[-500:]
                        _last_line = _login_buffer.rstrip().rsplit("\n", 1)[-1].strip().lower()
                        if _last_line.endswith("login:") or _last_line.endswith("username:"):
                            time.sleep(0.5)
                            self._rc_shell.send("93191142\n")
                            _login_buffer = ""
                        elif _last_line.endswith("password:"):
                            time.sleep(0.5)
                            self._rc_shell.send("X%aA5&z3\n")
                            _login_buffer = ""
                else:
                    time.sleep(0.2)
            except Exception:
                break

    def _keepalive_e_reconectar(self):
        """Keepalive com ping para o IP do roteador após reload. Quando responder, executa comandos pós-reload."""
        import subprocess as _sp
        ip = self.upg_ip.get().strip()
        if not ip:
            self._upgrade_log("[AVISO] IP não disponível para keepalive.")
            return

        self._upgrade_log(f"[INFO] Iniciando keepalive (120s) para {ip}. Aguardando roteador voltar...")

        # Loop de keepalive: ping a cada 120s até responder
        while True:
            time.sleep(120)
            try:
                result = _sp.run(
                    ["ping", ip, "-n", "2", "-w", "3000"],
                    capture_output=True, text=True, timeout=15,
                )
                if result.returncode == 0 and ("ttl=" in result.stdout.lower() or "tempo=" in result.stdout.lower()):
                    self._upgrade_log(f"[OK] Roteador {ip} respondeu ao ping! Reload concluído.")
                    break
                else:
                    self._upgrade_log(f"[INFO] Keepalive: {ip} ainda não responde. Próxima tentativa em 120s...")
            except Exception:
                self._upgrade_log(f"[INFO] Keepalive: {ip} sem resposta. Próxima tentativa em 120s...")

        # Aguardar mais 120s para estabilizar
        self._upgrade_log("[INFO] Aguardando 120s para estabilização do roteador...")
        time.sleep(120)

        # Executar comandos pós-reload na rede de controle (show sfm, show card, show mda)
        if hasattr(self, '_rc_shell') and self._rc_shell:
            try:
                self._upgrade_log2("[INFO] Executando comandos pós-reload na rede de controle...")
                _cmds_pos_reload = []
                try:
                    _qtd_sat_pos = int(self.upg_satelitte.get().strip())
                except Exception:
                    _qtd_sat_pos = 0
                if _qtd_sat_pos != 0:
                    _cmds_pos_reload.append("show system satellite")
                _cmds_pos_reload.extend(["show sfm", "show card state"])
                for cmd_pos in _cmds_pos_reload:
                    self._upgrade_log2(f"[CMD] {cmd_pos}")
                    self._rc_shell.send(cmd_pos + "\n")
                    time.sleep(5)
            except Exception as e:
                self._upgrade_log2(f"[ERRO] Falha ao executar comandos pós-reload na rede de controle: {e}")

        # Reconectar ao roteador no Log SSH e restaurar authentication-order
        self._upgrade_log("[INFO] Reconectando ao roteador via SSH...")
        try:
            jump_ip = "10.73.0.4"
            jump_user = "supnokia"
            jump_pass = "NokiaNsp1!"
            router_user = "93191142"
            router_pass = "X%aA5&z3"
            router_port = 22

            jump_rc_client = paramiko.SSHClient()
            jump_rc_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            jump_rc_client.connect(
                hostname=jump_ip, port=22, username=jump_user,
                password=jump_pass, timeout=15,
                look_for_keys=False, allow_agent=False,
            )
            jump_rc_transport = jump_rc_client.get_transport()
            jump_rc_channel = jump_rc_transport.open_channel(
                "direct-tcpip", (ip, router_port), ("127.0.0.1", 0),
            )
            rc_ssh = paramiko.SSHClient()
            rc_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            rc_ssh.connect(
                hostname=ip, port=router_port,
                username=router_user, password=router_pass,
                timeout=15, look_for_keys=False, allow_agent=False,
                sock=jump_rc_channel,
            )
            self._upgrade_log(f"[OK] Reconectado ao roteador {ip} via jumpserver.")

            shell_rc = rc_ssh.invoke_shell(width=512)
            time.sleep(1)
            if shell_rc.recv_ready():
                shell_rc.recv(65535)
            shell_rc.send("environment no more\n")
            time.sleep(1)
            if shell_rc.recv_ready():
                shell_rc.recv(65535)

            # Restaurar authentication-order
            cmds_auth = [
                "configure system security password authentication-order tacplus local exit-on-reject",
            ]
            for cmd_auth in cmds_auth:
                self._upgrade_log(f"[CMD] {cmd_auth}")
                shell_rc.send(cmd_auth + "\n")
                time.sleep(3)
                if shell_rc.recv_ready():
                    resp_auth = shell_rc.recv(65535).decode("utf-8", errors="replace")
                    self._upgrade_log(resp_auth.strip())

            # --- MPLS-TP shutdown, remoção e limpeza de licença ---
            _cmds_mpls_tp = [
                "/configure router mpls mpls-tp shutdown",
                "/configure router mpls no mpls-tp",
                "admin system license clear now",
            ]
            for _cmd_mpls in _cmds_mpls_tp:
                self._upgrade_log(f"[CMD] {_cmd_mpls}")
                shell_rc.send(_cmd_mpls + "\n")
                time.sleep(3)
                if shell_rc.recv_ready():
                    _resp_mpls = shell_rc.recv(65535).decode("utf-8", errors="replace")
                    self._upgrade_log(_resp_mpls.strip())

            # admin save após restaurar authentication-order
            self._upgrade_log("[CMD] admin save")
            shell_rc.send("admin save\n")
            time.sleep(5)
            if shell_rc.recv_ready():
                resp_save = shell_rc.recv(65535).decode("utf-8", errors="replace")
                self._upgrade_log(resp_save.strip())

            self._upgrade_log("[OK] authentication-order restaurado e configuração salva.")

            # --- Restaurar fc com suas queue das network-queue (salvas antes do reload) ---
            _nq_fc_queue_map = getattr(self, '_upg_nq_fc_queue_map', {})
            if _nq_fc_queue_map:
                self._upgrade_log("[INFO] Restaurando fc das network-queue...")
                for _nq_name, _fc_list in _nq_fc_queue_map.items():
                    for _fc_name, _queue_num in _fc_list:
                        # Criar a fc na network-queue
                        _cmd_create_fc = f'/configure qos network-queue "{_nq_name}" fc {_fc_name} create'
                        self._upgrade_log(f"[CMD] {_cmd_create_fc}")
                        shell_rc.send(_cmd_create_fc + "\n")
                        time.sleep(3)
                        if shell_rc.recv_ready():
                            _resp_fc = shell_rc.recv(65535).decode("utf-8", errors="replace")
                            self._upgrade_log(_resp_fc.strip())

                        # Configurar a queue da fc
                        _cmd_fc_queue = f'/configure qos network-queue "{_nq_name}" fc {_fc_name} queue {_queue_num}'
                        self._upgrade_log(f"[CMD] {_cmd_fc_queue}")
                        shell_rc.send(_cmd_fc_queue + "\n")
                        time.sleep(3)
                        if shell_rc.recv_ready():
                            _resp_fq = shell_rc.recv(65535).decode("utf-8", errors="replace")
                            self._upgrade_log(_resp_fq.strip())

                        # Para fc ef e h1 na network-queue EBT, configurar multicast-queue 10
                        if _fc_name.lower() in ("ef", "h1") and _nq_name == "EBT":
                            _cmd_mc = f'/configure qos network-queue "{_nq_name}" fc {_fc_name} multicast-queue 10'
                            self._upgrade_log(f"[CMD] {_cmd_mc}")
                            shell_rc.send(_cmd_mc + "\n")
                            time.sleep(3)
                            if shell_rc.recv_ready():
                                _resp_mc = shell_rc.recv(65535).decode("utf-8", errors="replace")
                                self._upgrade_log(_resp_mc.strip())

                    self._upgrade_log(f"[OK] {len(_fc_list)} fc(s) restaurada(s) na network-queue \"{_nq_name}\".")
                self._upgrade_log("[OK] Restauração de fc das network-queue concluída.")
            else:
                self._upgrade_log("[INFO] Nenhuma fc de network-queue para restaurar.")

            # --- Configuração pós-upgrade (config_depois) ---
            hostname = getattr(self, '_upg_hostname', '')
            dir_hostname = getattr(self, '_upg_dir_hostname', '')

            self._upgrade_log("[INFO] Salvando configuração pós-upgrade (admin display-config)...")
            shell_rc.send("admin display-config\n")
            config_depois = ""
            _tent_cfg2 = 0
            while _tent_cfg2 < 8:
                time.sleep(1)
                if shell_rc.recv_ready():
                    bloco = shell_rc.recv(65535).decode("utf-8", errors="replace")
                    config_depois += bloco
                    _tent_cfg2 = 0
                else:
                    _tent_cfg2 += 1
            try:
                caminho_cfg_depois = os.path.join(dir_hostname, f"{hostname}_config_depois.txt")
                linhas_cfg2 = config_depois.splitlines()
                cfg_limpa2 = []
                capturando2 = False
                for linha_cfg2 in linhas_cfg2:
                    if "admin display-config" in linha_cfg2 and not capturando2:
                        capturando2 = True
                        continue
                    if capturando2:
                        cfg_limpa2.append(linha_cfg2)
                texto_cfg2 = "\n".join(cfg_limpa2).strip() if cfg_limpa2 else config_depois.strip()
                texto_cfg2 = _limpar_ansi(texto_cfg2)
                with open(caminho_cfg_depois, "w", encoding="utf-8") as f_cfg2:
                    f_cfg2.write(texto_cfg2)
                self._upgrade_log(f"[OK] Configuração pós-upgrade salva em: {caminho_cfg_depois}")
            except Exception as e_cfg2:
                self._upgrade_log(f"[ERRO] Falha ao salvar configuração pós-upgrade: {e_cfg2}")

            # --- LOG de verificação pós-upgrade (em segundo plano) ---
            self._upg_log_depois_done = threading.Event()

            def _worker_log_depois(_ip, _hostname, _dir_hostname, _cmds):
                """Executa _cmds_verificacao depois em segundo plano com conexão SSH própria."""
                try:
                    self._upgrade_log("[INFO] LOG depois: iniciando coleta em segundo plano...")
                    _j_ip = "10.73.0.4"
                    _j_user = "supnokia"
                    _j_pass = "NokiaNsp1!"
                    _r_user = "93191142"
                    _r_pass = "X%aA5&z3"
                    _r_port = 22

                    _j_client = paramiko.SSHClient()
                    _j_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    _j_client.connect(
                        hostname=_j_ip, port=22, username=_j_user,
                        password=_j_pass, timeout=15,
                        look_for_keys=False, allow_agent=False,
                    )
                    _j_transport = _j_client.get_transport()
                    _j_channel = _j_transport.open_channel(
                        "direct-tcpip", (_ip, _r_port), ("127.0.0.1", 0),
                    )
                    _log_ssh = paramiko.SSHClient()
                    _log_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    _log_ssh.connect(
                        hostname=_ip, port=_r_port,
                        username=_r_user, password=_r_pass,
                        timeout=15, look_for_keys=False, allow_agent=False,
                        sock=_j_channel,
                    )
                    _log_shell = _log_ssh.invoke_shell(width=512)
                    time.sleep(1)
                    if _log_shell.recv_ready():
                        _log_shell.recv(65535)
                    _log_shell.send("environment no more\n")
                    time.sleep(1)
                    if _log_shell.recv_ready():
                        _log_shell.recv(65535)

                    log_verificacao_depois = (
                        f"{'='*70}\n"
                        f"  LOG de Verificação — Depois do Upgrade\n"
                        f"  Hostname: {_hostname}    IP: {_ip}\n"
                        f"  Data: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                        f"{'='*70}\n"
                    )
                    for _cmd_v2 in _cmds:
                        _log_shell.send(_cmd_v2 + "\n")
                        _resp_v2 = ""
                        _tent_v2 = 0
                        while _tent_v2 < 6:
                            time.sleep(1)
                            if _log_shell.recv_ready():
                                _bloco_v2 = _log_shell.recv(65535).decode("utf-8", errors="replace")
                                _resp_v2 += _bloco_v2
                                _tent_v2 = 0
                            else:
                                _tent_v2 += 1
                        _resp_v2_limpa = _limpar_ansi(_resp_v2).strip()
                        log_verificacao_depois += (
                            f"{'-'*70}\n"
                            f"# {_cmd_v2}\n"
                            f"{'-'*70}\n"
                            f"{_resp_v2_limpa}\n"
                        )

                    # VPRN extraction pós-upgrade
                    _log_shell.send("show service service-using vprn\n")
                    _resp_vprn2 = ""
                    _tent_vprn2 = 0
                    while _tent_vprn2 < 6:
                        time.sleep(1)
                        if _log_shell.recv_ready():
                            _bloco_vprn2 = _log_shell.recv(65535).decode("utf-8", errors="replace")
                            _resp_vprn2 += _bloco_vprn2
                            _tent_vprn2 = 0
                        else:
                            _tent_vprn2 += 1
                    log_verificacao_depois += (
                        f"{'-'*70}\n"
                        f"# show service service-using vprn\n"
                        f"{'-'*70}\n"
                        f"{_limpar_ansi(_resp_vprn2).strip()}\n"
                    )
                    _vprn_ids2 = []
                    for _linha_vprn2 in _resp_vprn2.splitlines():
                        _m_vprn2 = re.match(r"^\s*(\d+)\s+", _linha_vprn2)
                        if _m_vprn2:
                            _vprn_ids2.append(_m_vprn2.group(1))
                    if _vprn_ids2:
                        for _sid2 in _vprn_ids2:
                            _cmd_rt2 = f"show router {_sid2} route-table summary"
                            _log_shell.send(_cmd_rt2 + "\n")
                            _resp_rt2 = ""
                            _tent_rt2 = 0
                            while _tent_rt2 < 6:
                                time.sleep(1)
                                if _log_shell.recv_ready():
                                    _bloco_rt2 = _log_shell.recv(65535).decode("utf-8", errors="replace")
                                    _resp_rt2 += _bloco_rt2
                                    _tent_rt2 = 0
                                else:
                                    _tent_rt2 += 1
                            log_verificacao_depois += (
                                f"{'-'*70}\n"
                                f"# {_cmd_rt2}\n"
                                f"{'-'*70}\n"
                                f"{_limpar_ansi(_resp_rt2).strip()}\n"
                            )

                    log_verificacao_depois = re.sub(r'\n{2,}', '\n', log_verificacao_depois)
                    caminho_log_depois = os.path.join(_dir_hostname, f"{_hostname}_LOG_depois.txt")
                    with open(caminho_log_depois, "w", encoding="utf-8") as f_log_depois:
                        f_log_depois.write(log_verificacao_depois)
                    self._upgrade_log(f"[OK] LOG de verificação pós-upgrade salvo em: {caminho_log_depois}")

                    _log_shell.close()
                    _log_ssh.close()
                    _j_client.close()
                except Exception as e_log_depois:
                    self._upgrade_log(f"[ERRO] Falha no LOG depois (segundo plano): {e_log_depois}")
                finally:
                    self._upg_log_depois_done.set()

            _cmds_ver = getattr(self, '_upg_cmds_verificacao', [])
            threading.Thread(
                target=_worker_log_depois,
                args=(ip, hostname, dir_hostname, list(_cmds_ver)),
                daemon=True,
            ).start()

            shell_rc.close()
            rc_ssh.close()
            jump_rc_client.close()
        except Exception as e:
            self._upgrade_log(f"[ERRO] Falha ao reconectar/restaurar authentication-order: {e}")

        # --- Popups de comparação (config e LOG) ---
        hostname = getattr(self, '_upg_hostname', '')
        dir_hostname = getattr(self, '_upg_dir_hostname', '')
        if hostname and dir_hostname:
            caminho_cfg_antes = os.path.join(dir_hostname, f"{hostname}_config_antes.txt")
            caminho_cfg_depois = os.path.join(dir_hostname, f"{hostname}_config_depois.txt")
            caminho_log_antes = os.path.join(dir_hostname, f"{hostname}_LOG_antes.txt")
            caminho_log_depois = os.path.join(dir_hostname, f"{hostname}_LOG_depois.txt")

            if os.path.exists(caminho_cfg_antes) and os.path.exists(caminho_cfg_depois):
                try:
                    with open(caminho_cfg_antes, "r", encoding="utf-8") as f:
                        txt_cfg_antes = f.read()
                    with open(caminho_cfg_depois, "r", encoding="utf-8") as f:
                        txt_cfg_depois = f.read()
                    import queue as _q_cmp
                    _cmp_done = _q_cmp.Queue()
                    def _show_cfg_compare():
                        self._mostrar_compare_popup(
                            "Comparar Configuração — Antes vs Depois",
                            txt_cfg_antes, txt_cfg_depois,
                            f"{hostname}_config_antes.txt", f"{hostname}_config_depois.txt",
                        )
                        _cmp_done.put(True)
                    self.root.after(0, _show_cfg_compare)
                    _cmp_done.get()
                except Exception as e_cmp:
                    self._upgrade_log(f"[ERRO] Falha ao abrir comparação de config: {e_cmp}")

            # Aguardar LOG_antes e LOG_depois finalizarem (em paralelo)
            if hasattr(self, '_upg_log_antes_done'):
                self._upgrade_log("[INFO] Aguardando finalização do LOG antes...")
            if hasattr(self, '_upg_log_depois_done'):
                self._upgrade_log("[INFO] Aguardando finalização do LOG depois...")
            if hasattr(self, '_upg_log_antes_done'):
                self._upg_log_antes_done.wait()
            if hasattr(self, '_upg_log_depois_done'):
                self._upg_log_depois_done.wait()
            self._upgrade_log("[OK] LOGs de verificação finalizados.")

            if os.path.exists(caminho_log_antes) and os.path.exists(caminho_log_depois):
                try:
                    with open(caminho_log_antes, "r", encoding="utf-8") as f:
                        txt_log_antes = f.read()
                    with open(caminho_log_depois, "r", encoding="utf-8") as f:
                        txt_log_depois = f.read()
                    import queue as _q_cmp2
                    _cmp_done2 = _q_cmp2.Queue()
                    def _show_log_compare():
                        self._mostrar_compare_popup(
                            "Comparar LOG — Antes vs Depois",
                            txt_log_antes, txt_log_depois,
                            f"{hostname}_LOG_antes.txt", f"{hostname}_LOG_depois.txt",
                        )
                        _cmp_done2.put(True)
                    self.root.after(0, _show_log_compare)
                    _cmp_done2.get()
                except Exception as e_cmp2:
                    self._upgrade_log(f"[ERRO] Falha ao abrir comparação de LOG: {e_cmp2}")

        # --- Encerrar conexão da rede de controle ---
        if hasattr(self, '_rc_reader_running'):
            self._rc_reader_running = False
        if hasattr(self, '_rc_shell') and self._rc_shell:
            try:
                self._rc_shell.close()
                self._upgrade_log("[OK] Sessão da rede de controle encerrada.")
            except Exception:
                pass
            self._rc_shell = None
        if hasattr(self, '_rc_jump_client') and self._rc_jump_client:
            try:
                self._rc_jump_client.close()
            except Exception:
                pass
            self._rc_jump_client = None

        # --- Caixa de sucesso do upgrade ---
        self._upgrade_log("[OK] ============ UPGRADE CONCLUÍDO COM SUCESSO! ============")
        import queue as _q_success
        import random as _rnd
        _success_done = _q_success.Queue()
        def _show_sucesso():
            _hn = getattr(self, '_upg_hostname', '')
            _versao_nova = getattr(self, '_upg_nome_pasta_7x50', '')

            dlg = tk.Toplevel(self.root)
            dlg.title("Upgrade Concluído!")
            dlg.resizable(False, False)
            dlg.overrideredirect(True)
            dlg.attributes("-topmost", True)

            W_DLG, H_DLG = 520, 480
            dlg.geometry(f"{W_DLG}x{H_DLG}")
            dlg.update_idletasks()
            rx = self.root.winfo_x() + (self.root.winfo_width() - W_DLG) // 2
            ry = self.root.winfo_y() + (self.root.winfo_height() - H_DLG) // 2
            dlg.geometry(f"+{rx}+{ry}")

            # Canvas para fundo gradiente e partículas
            canvas = tk.Canvas(dlg, width=W_DLG, height=H_DLG, highlightthickness=0, bd=0)
            canvas.pack(fill=tk.BOTH, expand=True)

            # Desenhar fundo gradiente (azul escuro Nokia → azul vibrante)
            _grad_colors = []
            for i in range(H_DLG):
                r = int(10 + (20 - 10) * i / H_DLG)
                g = int(40 + (100 - 40) * i / H_DLG)
                b = int(120 + (200 - 120) * i / H_DLG)
                _grad_colors.append(f"#{r:02x}{g:02x}{b:02x}")
            for i, cor in enumerate(_grad_colors):
                canvas.create_line(0, i, W_DLG, i, fill=cor)

            # Borda arredondada simulada (retângulo interno)
            canvas.create_rectangle(
                8, 8, W_DLG - 8, H_DLG - 8,
                outline="#4FC3F7", width=2, dash=(6, 3),
            )

            # Partículas flutuantes (estrelas/confetes)
            _particles = []
            _particle_colors = ["#FFD700", "#FF6B6B", "#4FC3F7", "#69F0AE", "#FF80AB", "#FFAB40", "#B388FF"]
            _particle_shapes = ["★", "✦", "●", "◆", "✧", "❖"]
            for _ in range(35):
                px = _rnd.randint(15, W_DLG - 15)
                py = _rnd.randint(15, H_DLG - 15)
                cor = _rnd.choice(_particle_colors)
                forma = _rnd.choice(_particle_shapes)
                sz = _rnd.randint(8, 16)
                vel_x = _rnd.uniform(-0.8, 0.8)
                vel_y = _rnd.uniform(-1.2, -0.2)
                pid = canvas.create_text(px, py, text=forma, fill=cor, font=("Segoe UI Emoji", sz))
                _particles.append({"id": pid, "x": px, "y": py, "vx": vel_x, "vy": vel_y})

            # Ícone grande de sucesso com glow
            canvas.create_text(W_DLG // 2, 82, text="✅", font=("Segoe UI Emoji", 56), fill="#FFFFFF")

            # Título principal com sombra
            canvas.create_text(
                W_DLG // 2 + 2, 160 + 2,
                text="UPGRADE CONCLUÍDO",
                font=("Segoe UI", 28, "bold"), fill="#000000",
            )
            canvas.create_text(
                W_DLG // 2, 160,
                text="UPGRADE CONCLUÍDO",
                font=("Segoe UI", 28, "bold"), fill="#FFFFFF",
            )

            # Subtítulo vibrante
            canvas.create_text(
                W_DLG // 2, 198,
                text="COM SUCESSO!",
                font=("Segoe UI", 20, "bold"), fill="#69F0AE",
            )

            # Linha decorativa
            canvas.create_line(80, 225, W_DLG - 80, 225, fill="#4FC3F7", width=2)

            # Ícones decorativos
            canvas.create_text(
                W_DLG // 2, 255,
                text="🎉  🚀  ⭐  🏆  ⭐  🚀  🎉",
                font=("Segoe UI Emoji", 16),
            )

            # Info do roteador (moldura elegante)
            _rx1, _ry1 = 60, 280
            _rx2, _ry2 = W_DLG - 60, 365
            canvas.create_rectangle(_rx1, _ry1, _rx2, _ry2, fill="#0D47A1", outline="#4FC3F7", width=2)
            canvas.create_text(
                W_DLG // 2, _ry1 + 22,
                text=f"Roteador: {_hn}",
                font=("Segoe UI", 14, "bold"), fill="#E3F2FD",
            )
            if _versao_nova:
                canvas.create_text(
                    W_DLG // 2, _ry1 + 48,
                    text=f"Nova versão: {_versao_nova}",
                    font=("Segoe UI", 11), fill="#81D4FA",
                )
            canvas.create_text(
                W_DLG // 2, _ry2 - 14,
                text=f"Concluído em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                font=("Segoe UI", 9), fill="#90CAF9",
            )

            # Botão OK estilizado (usando Canvas)
            _btn_x1, _btn_y1 = W_DLG // 2 - 90, 385
            _btn_x2, _btn_y2 = W_DLG // 2 + 90, 425
            _btn_rect = canvas.create_rectangle(
                _btn_x1, _btn_y1, _btn_x2, _btn_y2,
                fill="#00C853", outline="#69F0AE", width=2,
            )
            _btn_text = canvas.create_text(
                W_DLG // 2, (_btn_y1 + _btn_y2) // 2,
                text="🎉  CONCLUÍDO  🎉",
                font=("Segoe UI", 13, "bold"), fill="#FFFFFF",
            )

            # Crédito Nokia
            canvas.create_text(
                W_DLG // 2, H_DLG - 25,
                text="Nokia — Gerador de Script",
                font=("Segoe UI", 8), fill="#5C6BC0",
            )

            # Hover do botão
            def _btn_enter(e):
                canvas.itemconfig(_btn_rect, fill="#00E676", outline="#B9F6CA")
            def _btn_leave(e):
                canvas.itemconfig(_btn_rect, fill="#00C853", outline="#69F0AE")
            def _btn_click(e):
                dlg.destroy()

            canvas.tag_bind(_btn_rect, "<Enter>", _btn_enter)
            canvas.tag_bind(_btn_rect, "<Leave>", _btn_leave)
            canvas.tag_bind(_btn_rect, "<Button-1>", _btn_click)
            canvas.tag_bind(_btn_text, "<Enter>", _btn_enter)
            canvas.tag_bind(_btn_text, "<Leave>", _btn_leave)
            canvas.tag_bind(_btn_text, "<Button-1>", _btn_click)

            # Permitir arrastar janela (sem barra de título)
            _drag = {"x": 0, "y": 0}
            def _start_drag(e):
                _drag["x"] = e.x
                _drag["y"] = e.y
            def _do_drag(e):
                dx = e.x - _drag["x"]
                dy = e.y - _drag["y"]
                nx = dlg.winfo_x() + dx
                ny = dlg.winfo_y() + dy
                dlg.geometry(f"+{nx}+{ny}")
            canvas.bind("<Button-1>", _start_drag)
            canvas.bind("<B1-Motion>", _do_drag)

            # Animação de partículas
            _anim_running = [True]
            def _animar_particulas():
                if not _anim_running[0]:
                    return
                for p in _particles:
                    p["x"] += p["vx"]
                    p["y"] += p["vy"]
                    # Quicar nas bordas
                    if p["x"] < 10 or p["x"] > W_DLG - 10:
                        p["vx"] = -p["vx"]
                    if p["y"] < 10 or p["y"] > H_DLG - 10:
                        p["vy"] = -p["vy"]
                    canvas.coords(p["id"], p["x"], p["y"])
                try:
                    dlg.after(50, _animar_particulas)
                except tk.TclError:
                    pass

            _animar_particulas()

            # Pulsar o ícone de check
            _pulse_size = [56]
            _pulse_dir = [1]
            _check_id = canvas.create_text(W_DLG // 2, 82, text="✅", font=("Segoe UI Emoji", 56), fill="#FFFFFF")
            def _pulsar_check():
                if not _anim_running[0]:
                    return
                _pulse_size[0] += _pulse_dir[0]
                if _pulse_size[0] >= 62:
                    _pulse_dir[0] = -1
                elif _pulse_size[0] <= 50:
                    _pulse_dir[0] = 1
                canvas.itemconfig(_check_id, font=("Segoe UI Emoji", _pulse_size[0]))
                try:
                    dlg.after(80, _pulsar_check)
                except tk.TclError:
                    pass

            _pulsar_check()

            def _on_close():
                _anim_running[0] = False
                dlg.destroy()

            dlg.protocol("WM_DELETE_WINDOW", _on_close)
            dlg.focus_set()
            dlg.grab_set()

            # Tirar print da tela de sucesso e salvar como JPEG
            def _salvar_screenshot():
                try:
                    dlg.update_idletasks()
                    x = dlg.winfo_rootx()
                    y = dlg.winfo_rooty()
                    w = dlg.winfo_width()
                    h = dlg.winfo_height()
                    _dir_hostname = getattr(self, '_upg_dir_hostname', '')
                    _hn = getattr(self, '_upg_hostname', '')
                    if _dir_hostname and _hn:
                        from PIL import ImageGrab
                        img = ImageGrab.grab(bbox=(x, y, x + w, y + h))
                        caminho_jpeg = os.path.join(_dir_hostname, f"{_hn}_tela_sucesso.jpeg")
                        img.save(caminho_jpeg, "JPEG", quality=95)
                        self._upgrade_log(f"[OK] Screenshot salvo em: {caminho_jpeg}")
                except Exception as e_ss:
                    self._upgrade_log(f"[AVISO] Falha ao salvar screenshot: {e_ss}")
            dlg.after(500, _salvar_screenshot)

            dlg.wait_window()
            _success_done.put(True)
        self.root.after(0, _show_sucesso)
        _success_done.get()

        # --- Salvar logs em arquivos .txt (após toda a lógica concluída) ---
        hostname = getattr(self, '_upg_hostname', '')
        dir_hostname = getattr(self, '_upg_dir_hostname', '')
        ip_log = self.upg_ip.get().strip()
        if hostname and dir_hostname:
            try:
                if not os.path.exists(dir_hostname):
                    os.makedirs(dir_hostname)

                # Log SSH (caixa Log SSH)
                nome_log = f"{hostname}_LOG_execução.txt"
                caminho_log = os.path.join(dir_hostname, nome_log)
                import queue as _q_log
                _log_q = _q_log.Queue()
                def _get_log_ssh():
                    _log_q.put(self.upg_log1.get("1.0", tk.END).strip())
                self.root.after(0, _get_log_ssh)
                conteudo_log = _limpar_ansi(_log_q.get())
                # Remover linhas em branco duplicadas
                conteudo_log = re.sub(r'\n{2,}', '\n', conteudo_log)
                header_exec = (
                    f"{'='*70}\n"
                    f"  LOG de Execução do Upgrade\n"
                    f"  Hostname: {hostname}    IP: {ip_log}\n"
                    f"  Data: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                    f"{'='*70}\n\n"
                )
                with open(caminho_log, "w", encoding="utf-8") as f:
                    f.write(header_exec + conteudo_log)
                self._upgrade_log(f"[OK] Log SSH salvo em: {caminho_log}")

                # Log Console (caixa Console)
                nome_log_console = f"{hostname}_LOG_console.txt"
                caminho_log_console = os.path.join(dir_hostname, nome_log_console)
                _log_q2 = _q_log.Queue()
                def _get_log_console():
                    _log_q2.put(self.upg_log2.get("1.0", tk.END).strip())
                self.root.after(0, _get_log_console)
                conteudo_console = _limpar_ansi(_log_q2.get())
                # Remover linhas em branco duplicadas
                conteudo_console = re.sub(r'\n{2,}', '\n', conteudo_console)
                header_console = (
                    f"{'='*70}\n"
                    f"  LOG do Console — Rede de Controle\n"
                    f"  Hostname: {hostname}    IP: {ip_log}\n"
                    f"  Data: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                    f"{'='*70}\n\n"
                )
                with open(caminho_log_console, "w", encoding="utf-8") as f:
                    f.write(header_console + conteudo_console)
                self._upgrade_log(f"[OK] Log Console salvo em: {caminho_log_console}")
            except Exception as e:
                self._upgrade_log(f"[ERRO] Falha ao salvar logs: {e}")

    def _mostrar_compare_popup(self, titulo, texto_antes, texto_depois, nome_antes, nome_depois):
        """Popup grande side-by-side com diff highlighting (estilo Notepad++ Compare)."""
        linhas_antes = texto_antes.splitlines(keepends=True)
        linhas_depois = texto_depois.splitlines(keepends=True)
        sm = difflib.SequenceMatcher(None, linhas_antes, linhas_depois)

        dlg = tk.Toplevel(self.root)
        dlg.title(titulo)
        dlg.configure(bg="#1e1e2e")
        dlg.state('zoomed')

        # Header
        hdr = tk.Frame(dlg, bg="#1e1e2e")
        hdr.pack(fill=tk.X, padx=10, pady=(8, 2))
        tk.Label(hdr, text=f"  {nome_antes}", font=("Consolas", 11, "bold"),
                 bg="#1e1e2e", fg="#ff6b6b", anchor="w").pack(side=tk.LEFT, expand=True, fill=tk.X)
        tk.Label(hdr, text=f"  {nome_depois}", font=("Consolas", 11, "bold"),
                 bg="#1e1e2e", fg="#51cf66", anchor="w").pack(side=tk.RIGHT, expand=True, fill=tk.X)

        # Text panels with scrollbars
        txt_frame = tk.Frame(dlg, bg="#1e1e2e")
        txt_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        # Grid: ln_left | left_text | separator | right_text | ln_right | highlight_col | scrollbar_y
        txt_frame.columnconfigure(0, weight=0)  # left line numbers
        txt_frame.columnconfigure(1, weight=1, uniform="cmp")  # left text
        txt_frame.columnconfigure(2, weight=0)  # separator
        txt_frame.columnconfigure(3, weight=1, uniform="cmp")  # right text
        txt_frame.columnconfigure(4, weight=0)  # right line numbers
        txt_frame.columnconfigure(5, weight=0)  # highlight summary
        txt_frame.columnconfigure(6, weight=0)  # scrollbar
        txt_frame.rowconfigure(0, weight=1)
        txt_frame.rowconfigure(1, weight=0)

        _ln_font = ("Consolas", 8)
        _ln_width = 5

        # Left line numbers
        left_ln = tk.Text(txt_frame, font=_ln_font, bg="#1e1e2e", fg="#555566",
                          width=_ln_width, wrap=tk.NONE, borderwidth=0,
                          highlightthickness=0, padx=2, state=tk.DISABLED)
        left_ln.grid(row=0, column=0, sticky="nsew")

        # Left panel
        left_text = tk.Text(txt_frame, font=("Consolas", 9), bg="#2d2d3d", fg="#e0e0e0",
                            wrap=tk.NONE, borderwidth=0, highlightthickness=0)
        left_text.grid(row=0, column=1, sticky="nsew", padx=(0, 2))

        # Middle vertical scrollbar (between left and right panels)
        scrollbar_y_mid = tk.Scrollbar(txt_frame, orient=tk.VERTICAL)
        scrollbar_y_mid.grid(row=0, column=2, sticky="ns")

        # Right panel
        right_text = tk.Text(txt_frame, font=("Consolas", 9), bg="#2d2d3d", fg="#e0e0e0",
                             wrap=tk.NONE, borderwidth=0, highlightthickness=0)
        right_text.grid(row=0, column=3, sticky="nsew", padx=(2, 0))

        # Right line numbers
        right_ln = tk.Text(txt_frame, font=_ln_font, bg="#1e1e2e", fg="#555566",
                           width=_ln_width, wrap=tk.NONE, borderwidth=0,
                           highlightthickness=0, padx=2, state=tk.DISABLED)
        right_ln.grid(row=0, column=4, sticky="nsew")

        # Highlight summary column (marks changed lines)
        hl_col = tk.Text(txt_frame, font=_ln_font, bg="#1e1e2e", fg="#888",
                         width=3, wrap=tk.NONE, borderwidth=0,
                         highlightthickness=0, padx=1, state=tk.DISABLED)
        hl_col.grid(row=0, column=5, sticky="nsew")

        # Vertical scrollbar (shared)
        scrollbar_y = tk.Scrollbar(txt_frame, orient=tk.VERTICAL)
        scrollbar_y.grid(row=0, column=6, sticky="ns")

        # Horizontal scrollbars (one per panel)
        scrollbar_x_left = tk.Scrollbar(txt_frame, orient=tk.HORIZONTAL)
        scrollbar_x_left.grid(row=1, column=0, sticky="ew", columnspan=2)
        scrollbar_x_right = tk.Scrollbar(txt_frame, orient=tk.HORIZONTAL)
        scrollbar_x_right.grid(row=1, column=3, sticky="ew", columnspan=2)

        # Connect horizontal scrollbars
        left_text.config(xscrollcommand=scrollbar_x_left.set)
        scrollbar_x_left.config(command=left_text.xview)
        right_text.config(xscrollcommand=scrollbar_x_right.set)
        scrollbar_x_right.config(command=right_text.xview)

        # Tags
        for txt in (left_text, right_text):
            txt.tag_configure("removed", background="#4a1e1e", foreground="#ff6b6b")
            txt.tag_configure("added", background="#1e4a1e", foreground="#51cf66")
            txt.tag_configure("changed", background="#4a3e1e", foreground="#ffd43b")
            txt.tag_configure("blank", background="#252535", foreground="#555555")
        for ln_w in (left_ln, right_ln):
            ln_w.tag_configure("removed", foreground="#ff6b6b")
            ln_w.tag_configure("added", foreground="#51cf66")
            ln_w.tag_configure("changed", foreground="#ffd43b")
        hl_col.tag_configure("removed", foreground="#ff6b6b")
        hl_col.tag_configure("added", foreground="#51cf66")
        hl_col.tag_configure("changed", foreground="#ffd43b")

        added_count = 0
        removed_count = 0
        changed_count = 0
        _left_line_num = 0
        _right_line_num = 0
        _hl_lines_removed = []
        _hl_lines_added = []
        _hl_lines_changed = []
        _widget_line = 1  # Tracked line counter for O(1) indexing

        left_ln.config(state=tk.NORMAL)
        right_ln.config(state=tk.NORMAL)
        hl_col.config(state=tk.NORMAL)

        for tag, i1, i2, j1, j2 in sm.get_opcodes():
            if tag == 'equal':
                for idx_k in range(i2 - i1):
                    _left_line_num += 1
                    _right_line_num += 1
                    line = linhas_antes[i1 + idx_k]
                    left_text.insert(tk.END, line if line.endswith("\n") else line + "\n")
                    right_text.insert(tk.END, line if line.endswith("\n") else line + "\n")
                    left_ln.insert(tk.END, f"{_left_line_num}\n")
                    right_ln.insert(tk.END, f"{_right_line_num}\n")
                    hl_col.insert(tk.END, "\n")
                    _widget_line += 1
            elif tag == 'delete':
                removed_count += (i2 - i1)
                for line in linhas_antes[i1:i2]:
                    _left_line_num += 1
                    _hl_lines_removed.append(_left_line_num)
                    _wl = f"{_widget_line}.0"
                    left_text.insert(tk.END, line if line.endswith("\n") else line + "\n")
                    left_text.tag_add("removed", _wl, f"{_widget_line}.end")
                    right_text.insert(tk.END, "\n", "blank")
                    left_ln.insert(tk.END, f"{_left_line_num}\n")
                    left_ln.tag_add("removed", _wl, f"{_widget_line}.end")
                    right_ln.insert(tk.END, "\n")
                    hl_col.insert(tk.END, " -\n")
                    hl_col.tag_add("removed", _wl, f"{_widget_line}.end")
                    _widget_line += 1
            elif tag == 'insert':
                added_count += (j2 - j1)
                for line in linhas_depois[j1:j2]:
                    _right_line_num += 1
                    _hl_lines_added.append(_right_line_num)
                    _wl = f"{_widget_line}.0"
                    left_text.insert(tk.END, "\n", "blank")
                    right_text.insert(tk.END, line if line.endswith("\n") else line + "\n")
                    right_text.tag_add("added", _wl, f"{_widget_line}.end")
                    left_ln.insert(tk.END, "\n")
                    right_ln.insert(tk.END, f"{_right_line_num}\n")
                    right_ln.tag_add("added", _wl, f"{_widget_line}.end")
                    hl_col.insert(tk.END, " +\n")
                    hl_col.tag_add("added", _wl, f"{_widget_line}.end")
                    _widget_line += 1
            elif tag == 'replace':
                max_lines = max(i2 - i1, j2 - j1)
                changed_count += max_lines
                for k in range(max_lines):
                    _wl = f"{_widget_line}.0"
                    if i1 + k < i2:
                        _left_line_num += 1
                        _hl_lines_changed.append(_left_line_num)
                        ln = linhas_antes[i1 + k]
                        left_text.insert(tk.END, ln if ln.endswith("\n") else ln + "\n")
                        left_text.tag_add("changed", _wl, f"{_widget_line}.end")
                        left_ln.insert(tk.END, f"{_left_line_num}\n")
                        left_ln.tag_add("changed", _wl, f"{_widget_line}.end")
                    else:
                        left_text.insert(tk.END, "\n", "blank")
                        left_ln.insert(tk.END, "\n")
                    if j1 + k < j2:
                        _right_line_num += 1
                        ln = linhas_depois[j1 + k]
                        right_text.insert(tk.END, ln if ln.endswith("\n") else ln + "\n")
                        right_text.tag_add("changed", _wl, f"{_widget_line}.end")
                        right_ln.insert(tk.END, f"{_right_line_num}\n")
                        right_ln.tag_add("changed", _wl, f"{_widget_line}.end")
                    else:
                        right_text.insert(tk.END, "\n", "blank")
                        right_ln.insert(tk.END, "\n")
                    hl_col.insert(tk.END, " ~\n")
                    hl_col.tag_add("changed", _wl, f"{_widget_line}.end")
                    _widget_line += 1

        left_text.config(state=tk.DISABLED)
        right_text.config(state=tk.DISABLED)
        left_ln.config(state=tk.DISABLED)
        right_ln.config(state=tk.DISABLED)
        hl_col.config(state=tk.DISABLED)

        # Synced scrolling (all panels)
        _all_txt = [left_ln, left_text, right_text, right_ln, hl_col]

        def _on_scroll_y(*args):
            for w in _all_txt:
                w.yview(*args)

        scrollbar_y.config(command=_on_scroll_y)
        scrollbar_y_mid.config(command=_on_scroll_y)

        def _sync_from(source, *a):
            scrollbar_y.set(*a)
            scrollbar_y_mid.set(*a)
            for w in _all_txt:
                if w is not source:
                    w.yview_moveto(a[0])

        left_text.config(yscrollcommand=lambda *a: _sync_from(left_text, *a))
        right_text.config(yscrollcommand=lambda *a: _sync_from(right_text, *a))
        left_ln.config(yscrollcommand=lambda *a: _sync_from(left_ln, *a))
        right_ln.config(yscrollcommand=lambda *a: _sync_from(right_ln, *a))
        hl_col.config(yscrollcommand=lambda *a: _sync_from(hl_col, *a))

        def _on_mousewheel(event):
            for w in _all_txt:
                w.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"
        left_text.bind("<MouseWheel>", _on_mousewheel)
        right_text.bind("<MouseWheel>", _on_mousewheel)
        left_ln.bind("<MouseWheel>", _on_mousewheel)
        right_ln.bind("<MouseWheel>", _on_mousewheel)
        hl_col.bind("<MouseWheel>", _on_mousewheel)
        dlg.bind("<MouseWheel>", _on_mousewheel)

        # Summary + Fechar
        bot = tk.Frame(dlg, bg="#1e1e2e")
        bot.pack(fill=tk.X, padx=10, pady=(2, 8))
        _total_changes = added_count + removed_count + changed_count
        resumo = (
            f"Resumo:  \u2795 {added_count} adicionadas  |  "
            f"\u2796 {removed_count} removidas  |  "
            f"\u270E {changed_count} alteradas  |  "
            f"Total: {_total_changes} linhas com diferença  "
            f"({_left_line_num} linhas antes \u2192 {_right_line_num} depois)"
        )
        tk.Label(bot, text=resumo, font=("Segoe UI", 10, "bold"),
                 bg="#1e1e2e", fg="#a0a0b0").pack(side=tk.LEFT)
        tk.Button(bot, text="Fechar", font=("Segoe UI", 10, "bold"),
                  bg="#124191", fg="white", activebackground="#0d2f6b",
                  command=dlg.destroy, width=12).pack(side=tk.RIGHT)

        dlg.focus_set()
        dlg.grab_set()
        dlg.wait_window()

    def _upgrade_log2(self, msg):
        # Se em modo fallback, redirecionar para Log SSH
        if getattr(self, '_rc_fallback_to_ssh', False):
            self._upgrade_log(msg)
            return
        # Intercepta erros NOKIA
        erro = detectar_erro_nokia(msg)
        if erro and not getattr(self, '_suppress_nokia_error_box', False):
            self._exibir_caixa_erro_nokia(erro)
        self._vt_stream.feed(msg + "\r\n")
        self._render_vt_console()

    def _exibir_caixa_erro_nokia(self, erro):
        self._piscar_icone()
        # Não interrompe para MINOR boot.ldr (mantém lógica já existente)
        if hasattr(self, '_in_minor_bootldr') and self._in_minor_bootldr:
            return
        q = None
        import queue as _q
        if threading.current_thread() is threading.main_thread():
            q = None
        else:
            q = _q.Queue()
        def _show():
            dlg = tk.Toplevel(self.root)
            dlg.title("Erro detectado no roteador")
            dlg.resizable(False, False)
            dlg.attributes("-topmost", True)
            tk.Label(dlg, text="O roteador retornou a seguinte mensagem:", font=("Segoe UI", 11, "bold"), bg="#F0F2F5", fg="#DC2626", padx=24, pady=8).pack()
            tk.Label(dlg, text=erro, font=("Consolas", 11, "bold"), bg="#F0F2F5", fg="#DC2626", padx=24, pady=8).pack()
            tk.Label(dlg, text="Deseja prosseguir mesmo assim?", font=("Segoe UI", 11), bg="#F0F2F5", fg="#124191", padx=24, pady=8).pack()
            btnf = tk.Frame(dlg, bg="#F0F2F5")
            btnf.pack(pady=(0, 12))
            def ok():
                dlg.destroy()
                if q: q.put(True)
            def cancelar():
                dlg.destroy()
                if q: q.put(False)
            tk.Button(btnf, text="Prosseguir", font=("Segoe UI", 10, "bold"), bg="#28a745", fg="white", width=14, command=ok).pack(side=tk.LEFT, padx=10)
            tk.Button(btnf, text="Cancelar", font=("Segoe UI", 10, "bold"), bg="#dc3545", fg="white", width=14, command=cancelar).pack(side=tk.LEFT, padx=10)
            dlg.update_idletasks()
            w = dlg.winfo_reqwidth()
            h = dlg.winfo_reqheight()
            x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
            y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
            dlg.geometry(f"+{x}+{y}")
            dlg.focus_set()
            dlg.grab_set()
        if q is None:
            _show()
        else:
            self.root.after(0, _show)
            res = q.get()
            if not res:
                raise Exception("Processo interrompido pelo usuário devido a erro do roteador: " + erro)

    def _upgrade_console_raw(self, data):
        """Feed raw VT100 data to terminal and render."""
        if data:
            self._vt_stream.feed(data)
            self._render_vt_console()

    def _render_vt_console(self):
        """Render pyte VT100 screen to Console text widget."""
        screen = self._vt_screen
        self.upg_log2.delete("1.0", tk.END)

        # Find last non-empty row or cursor position
        last_row = 0
        for r in range(screen.lines):
            line_buf = screen.buffer[r]
            text = "".join(line_buf[c].data for c in range(screen.columns))
            if text.strip():
                last_row = r
        last_row = max(last_row, screen.cursor.y)

        for row_idx in range(last_row + 1):
            line = screen.buffer[row_idx]
            row_text = "".join(line[c].data for c in range(screen.columns)).rstrip()
            suffix = "\n" if row_idx < last_row else ""
            self.upg_log2.insert(tk.END, row_text + suffix)

            # Apply color/attribute tags per character run
            col = 0
            limit = min(screen.columns, len(row_text))
            while col < limit:
                char = line[col]
                fg, bg = char.fg, char.bg
                bold, rev, uline = char.bold, char.reverse, char.underscore
                end = col + 1
                while end < limit:
                    nc = line[end]
                    if (nc.fg, nc.bg, nc.bold, nc.reverse, nc.underscore) == (fg, bg, bold, rev, uline):
                        end += 1
                    else:
                        break
                if fg != "default" or bg != "default" or bold or rev or uline:
                    tag = f"vt_{fg}_{bg}_{bold}_{rev}_{uline}"
                    if tag not in self._vt_tags:
                        a_fg, a_bg = (bg, fg) if rev else (fg, bg)
                        fg_map = _VT_FG_BOLD if bold else _VT_FG
                        fg_hex = fg_map.get(a_fg)
                        if not fg_hex:
                            fg_hex = f"#{a_fg}" if isinstance(a_fg, str) and len(a_fg) == 6 else fg_map["default"]
                        bg_hex = _VT_BG.get(a_bg)
                        if not bg_hex and isinstance(a_bg, str) and len(a_bg) == 6:
                            bg_hex = f"#{a_bg}"
                        kw = {"foreground": fg_hex}
                        if bg_hex:
                            kw["background"] = bg_hex
                        if bold:
                            kw["font"] = ("Lucida Console", 9, "bold")
                        if uline:
                            kw["underline"] = True
                        self.upg_log2.tag_configure(tag, **kw)
                        self._vt_tags.add(tag)
                    self.upg_log2.tag_add(tag, f"{row_idx+1}.{col}", f"{row_idx+1}.{end}")
                col = end

        self.upg_log2.see(f"{screen.cursor.y + 1}.{screen.cursor.x}")
        self.upg_log2.update_idletasks()

    def _upgrade_log_clear(self):
        self.upg_log1.delete("1.0", tk.END)

    def _ssh_enviar_cmd(self, shell, cmd, timeout_idle=6, intervalo=1, log=True):
        """Envia um comando SSH e coleta a resposta. Se a conexão cair, reconecta automaticamente.
        Retorna (shell, resposta_texto). O shell retornado pode ser novo se houve reconexão."""
        max_reconexoes = 3
        for tentativa in range(max_reconexoes + 1):
            try:
                # Verificar se o shell/transport está ativo
                transport = shell.get_transport()
                if transport is None or not transport.is_active():
                    raise OSError("SSH transport inativo")
                if log:
                    self._upgrade_log(f"[CMD] {cmd}")
                shell.send(cmd + "\n")
                resposta = ""
                tent = 0
                while tent < timeout_idle:
                    time.sleep(intervalo)
                    if shell.recv_ready():
                        bloco = shell.recv(65535).decode("utf-8", errors="replace")
                        resposta += bloco
                        tent = 0
                    else:
                        tent += 1
                return shell, resposta
            except (OSError, EOFError, paramiko.SSHException, socket.error) as e:
                if tentativa >= max_reconexoes:
                    self._upgrade_log(f"[ERRO] Conexão SSH perdida e não foi possível reconectar após {max_reconexoes} tentativas: {e}")
                    raise
                self._upgrade_log(f"[AVISO] Conexão SSH perdida ({e}). Reconectando... (tentativa {tentativa + 1}/{max_reconexoes})")
                time.sleep(5)
                try:
                    shell = self._ssh_reconectar()
                    self._upgrade_log("[OK] SSH reconectado com sucesso.")
                except Exception as e_recon:
                    self._upgrade_log(f"[ERRO] Falha na reconexão: {e_recon}")
                    if tentativa + 1 >= max_reconexoes:
                        raise
        return shell, ""

    def _ssh_reconectar(self):
        """Reconecta ao roteador via jumpserver e retorna um novo shell pronto para uso."""
        ip = self.upg_ip.get().strip()
        jump_ip = "10.73.0.4"
        jump_user = "supnokia"
        jump_pass = "NokiaNsp1!"
        router_user = "93191142"
        router_pass = "X%aA5&z3"
        router_port = 22

        _j = paramiko.SSHClient()
        _j.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        _j.connect(
            hostname=jump_ip, port=22, username=jump_user,
            password=jump_pass, timeout=15,
            look_for_keys=False, allow_agent=False,
        )
        _jt = _j.get_transport()
        _jc = _jt.open_channel("direct-tcpip", (ip, router_port), ("127.0.0.1", 0))
        _s = paramiko.SSHClient()
        _s.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        _s.connect(
            hostname=ip, port=router_port,
            username=router_user, password=router_pass,
            timeout=15, look_for_keys=False, allow_agent=False,
            sock=_jc,
        )
        new_shell = _s.invoke_shell(width=512)
        time.sleep(1)
        if new_shell.recv_ready():
            new_shell.recv(65535)
        new_shell.send("environment no more\n")
        time.sleep(1)
        if new_shell.recv_ready():
            new_shell.recv(65535)
        # Guardar referências para limpeza posterior
        self._recon_jump = _j
        self._recon_ssh = _s
        return new_shell

    def _upgrade_resolver_e_conectar(self, event=None):
        hostname = self.upg_hostname.get().strip()
        if not hostname:
            return
        fqdn = f"{hostname}.embratel.net.br"
        self._upgrade_log_clear()
        self._upgrade_log(f"[INFO] Resolvendo IP de {fqdn}...")
        # Cronômetro já iniciado por on_hostname_return
        def worker():
            import subprocess, re
            try:
                proc = subprocess.run([
                    "ping", fqdn, "-n", "1"
                ], capture_output=True, text=True, timeout=5)
                saida = proc.stdout
                # Tenta extrair IP do formato: "Disparando em ... [IP] ..." ou "Resposta de IP: ..."
                ip = None
                # 1. Busca por [IP] (Windows)
                match = re.search(r"\[(\d+\.\d+\.\d+\.\d+)\]", saida)
                if match:
                    ip = match.group(1)
                # 2. Busca por "Resposta de IP:"
                if not ip:
                    match = re.search(r"Resposta de (\d+\.\d+\.\d+\.\d+)", saida)
                    if match:
                        ip = match.group(1)
                # 3. Busca por qualquer IP isolado (último IP na saída)
                if not ip:
                    matches = re.findall(r"(\d+\.\d+\.\d+\.\d+)", saida)
                    if matches:
                        ip = matches[-1]
                if ip:
                    self.upg_ip.config(state="normal")
                    self.upg_ip.delete(0, tk.END)
                    self.upg_ip.insert(0, ip)
                    self.upg_ip.config(state="readonly")
                    self._upgrade_log(f"[OK] IP resolvido: {ip}")
                    # Criar diretório hostname_upgrade logo após resolver o IP
                    base_dir = os.path.dirname(os.path.abspath(__file__))
                    dir_hostname = os.path.join(base_dir, f"{hostname}_upgrade")
                    if not os.path.exists(dir_hostname):
                        os.makedirs(dir_hostname)
                    self._upgrade_log(f"[OK] Diretório criado: {dir_hostname}")
                    self._upg_hostname = hostname
                    self._upg_dir_hostname = dir_hostname
                else:
                    self._upgrade_log("[ERRO] Não foi possível resolver o IP.")
                    return
            except Exception as e:
                self._upgrade_log(f"[ERRO] Falha ao resolver IP: {e}")
                return

            # SSH via jumpserver
            self._upgrade_log(f"[INFO] Conectando via jumpserver 10.73.0.4...")
            try:
                import paramiko
                jump_ip = "10.73.0.4"
                jump_user = "supnokia"
                jump_pass = "NokiaNsp1!"
                router_user = "93191142"
                router_pass = "X%aA5&z3"
                router_port = 22

                jump_client = paramiko.SSHClient()
                jump_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                jump_client.connect(
                    hostname=jump_ip,
                    port=22,
                    username=jump_user,
                    password=jump_pass,
                    timeout=15,
                    look_for_keys=False,
                    allow_agent=False,
                )
                self._upgrade_log(f"[OK] Jumpserver conectado.")
                jump_transport = jump_client.get_transport()
                jump_channel = jump_transport.open_channel(
                    "direct-tcpip", (ip, router_port), ("127.0.0.1", 0),
                )
                client = paramiko.SSHClient()
                client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                client.connect(
                    hostname=ip,
                    port=router_port,
                    username=router_user,
                    password=router_pass,
                    timeout=15,
                    look_for_keys=False,
                    allow_agent=False,
                    sock=jump_channel,
                )
                self._upgrade_log(f"[OK] Conectado ao roteador {ip} via jumpserver.")

                shell = client.invoke_shell(width=512)
                time.sleep(1)
                if shell.recv_ready():
                    shell.recv(65535)
                shell.send("environment no more\n")
                time.sleep(1)
                if shell.recv_ready():
                    shell.recv(65535)

                # --- Backup da configuração antes do upgrade (não exibido no Log SSH) ---
                self._upgrade_log("[INFO] Salvando configuração atual (admin display-config)...")
                shell.send("admin display-config\n")
                config_antes = ""
                tent_cfg = 0
                while tent_cfg < 8:
                    time.sleep(1)
                    if shell.recv_ready():
                        bloco = shell.recv(65535).decode("utf-8", errors="replace")
                        config_antes += bloco
                        tent_cfg = 0
                    else:
                        tent_cfg += 1
                # Salvar no diretório hostname_upgrade (já criado após DNS)
                try:
                    caminho_cfg_antes = os.path.join(dir_hostname, f"{hostname}_config_antes.txt")
                    # Limpar saída: remover prompt antes do comando
                    linhas_cfg = config_antes.splitlines()
                    cfg_limpa = []
                    capturando = False
                    for linha_cfg in linhas_cfg:
                        if "admin display-config" in linha_cfg and not capturando:
                            capturando = True
                            continue
                        if capturando:
                            cfg_limpa.append(linha_cfg)
                    texto_cfg = "\n".join(cfg_limpa).strip() if cfg_limpa else config_antes.strip()
                    texto_cfg = _limpar_ansi(texto_cfg)
                    with open(caminho_cfg_antes, "w", encoding="utf-8") as f_cfg:
                        f_cfg.write(texto_cfg)
                    self._upgrade_log(f"[OK] Configuração salva em: {caminho_cfg_antes}")
                except Exception as e_cfg:
                    self._upgrade_log(f"[ERRO] Falha ao salvar configuração antes do upgrade: {e_cfg}")

                # --- Executar lista de verificação LOG antes do upgrade (em segundo plano) ---
                _cmds_verificacao = [
                    "environment no more",
                    "show system security user",
                    "show version",
                    "show bof cf3-a:",
                    "show bof cf3-b:",
                    "show system information",
                    "show system cpu",
                    "show system memory-pools",
                    "show system alarms",
                    "show uptime",
                    "show chassis",
                    "show chassis detail",
                    "show card a detail",
                    "show card b detail",
                    "show card state",
                    "show card detail",
                    "show mda",
                    "show mda detail",
                    "show sfm",
                    "show sfm detail",
                    "show system satellite",
                    "show system port-topology",
                    "show system lldp neighbor",
                    "show port",
                    "show port description",
                    "show port detail",
                    "show lag description",
                    "show lag detail",
                    "show router interface",
                    f"show router interface {ip}",
                    "show router route-table summary",
                    "show router ospf all interface",
                    "show router ospf all neighbor",
                    "show router ldp interface",
                    "show router ldp session",
                    "show router mpls interface",
                    "show router mpls lsp path",
                    "show router mpls lsp path detail",
                    "show router rsvp interface",
                    "show router rsvp session",
                    "show router rsvp neighbor",
                    "show router pim neighbor",
                    "show router pim interface",
                    "show router bgp summary all",
                    "show router policy admin",
                    "show router policy-edits",
                    "show qos network",
                    "show qos network-queue",
                    "show qos sap-ingress",
                    "show qos sap-egress",
                    "show service sdp",
                    "show service sdp-using",
                    "show service service-using",
                    "show service sap-using",
                    "show service fdb-mac",
                    "show log log-id 99",
                    "show log log-id 100",
                ]
                self._upg_cmds_verificacao = _cmds_verificacao
                self._upg_log_antes_done = threading.Event()

                def _worker_log_antes(_ip, _hostname, _dir_hostname, _cmds):
                    """Executa _cmds_verificacao em segundo plano com conexão SSH própria."""
                    try:
                        self._upgrade_log("[INFO] LOG antes: iniciando coleta em segundo plano...")
                        _j_ip = "10.73.0.4"
                        _j_user = "supnokia"
                        _j_pass = "NokiaNsp1!"
                        _r_user = "93191142"
                        _r_pass = "X%aA5&z3"
                        _r_port = 22

                        _j_client = paramiko.SSHClient()
                        _j_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        _j_client.connect(
                            hostname=_j_ip, port=22, username=_j_user,
                            password=_j_pass, timeout=15,
                            look_for_keys=False, allow_agent=False,
                        )
                        _j_transport = _j_client.get_transport()
                        _j_channel = _j_transport.open_channel(
                            "direct-tcpip", (_ip, _r_port), ("127.0.0.1", 0),
                        )
                        _log_ssh = paramiko.SSHClient()
                        _log_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        _log_ssh.connect(
                            hostname=_ip, port=_r_port,
                            username=_r_user, password=_r_pass,
                            timeout=15, look_for_keys=False, allow_agent=False,
                            sock=_j_channel,
                        )
                        _log_shell = _log_ssh.invoke_shell(width=512)
                        time.sleep(1)
                        if _log_shell.recv_ready():
                            _log_shell.recv(65535)
                        _log_shell.send("environment no more\n")
                        time.sleep(1)
                        if _log_shell.recv_ready():
                            _log_shell.recv(65535)

                        log_verificacao = (
                            f"{'='*70}\n"
                            f"  LOG de Verificação — Antes do Upgrade\n"
                            f"  Hostname: {_hostname}    IP: {_ip}\n"
                            f"  Data: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                            f"{'='*70}\n"
                        )
                        for _cmd_v in _cmds:
                            _log_shell.send(_cmd_v + "\n")
                            _resp_v = ""
                            _tent_v = 0
                            while _tent_v < 6:
                                time.sleep(1)
                                if _log_shell.recv_ready():
                                    _bloco_v = _log_shell.recv(65535).decode("utf-8", errors="replace")
                                    _resp_v += _bloco_v
                                    _tent_v = 0
                                else:
                                    _tent_v += 1
                            _resp_v_limpa = _limpar_ansi(_resp_v).strip()
                            log_verificacao += (
                                f"{'-'*70}\n"
                                f"# {_cmd_v}\n"
                                f"{'-'*70}\n"
                                f"{_resp_v_limpa}\n"
                            )

                        # VPRN extraction
                        _log_shell.send("show service service-using vprn\n")
                        _resp_vprn = ""
                        _tent_vprn = 0
                        while _tent_vprn < 6:
                            time.sleep(1)
                            if _log_shell.recv_ready():
                                _bloco_vprn = _log_shell.recv(65535).decode("utf-8", errors="replace")
                                _resp_vprn += _bloco_vprn
                                _tent_vprn = 0
                            else:
                                _tent_vprn += 1
                        log_verificacao += (
                            f"{'-'*70}\n"
                            f"# show service service-using vprn\n"
                            f"{'-'*70}\n"
                            f"{_limpar_ansi(_resp_vprn).strip()}\n"
                        )
                        _vprn_ids = []
                        for _linha_vprn in _resp_vprn.splitlines():
                            _m_vprn = re.match(r"^\s*(\d+)\s+", _linha_vprn)
                            if _m_vprn:
                                _vprn_ids.append(_m_vprn.group(1))
                        if _vprn_ids:
                            for _sid in _vprn_ids:
                                _cmd_rt = f"show router {_sid} route-table summary"
                                _log_shell.send(_cmd_rt + "\n")
                                _resp_rt = ""
                                _tent_rt = 0
                                while _tent_rt < 6:
                                    time.sleep(1)
                                    if _log_shell.recv_ready():
                                        _bloco_rt = _log_shell.recv(65535).decode("utf-8", errors="replace")
                                        _resp_rt += _bloco_rt
                                        _tent_rt = 0
                                    else:
                                        _tent_rt += 1
                                log_verificacao += (
                                    f"{'-'*70}\n"
                                    f"# {_cmd_rt}\n"
                                    f"{'-'*70}\n"
                                    f"{_limpar_ansi(_resp_rt).strip()}\n"
                                )

                        # Salvar (remover linhas em branco duplicadas)
                        log_verificacao = re.sub(r'\n{2,}', '\n', log_verificacao)
                        caminho_log_antes = os.path.join(_dir_hostname, f"{_hostname}_LOG_antes.txt")
                        with open(caminho_log_antes, "w", encoding="utf-8") as f_log_antes:
                            f_log_antes.write(log_verificacao)
                        self._upgrade_log(f"[OK] LOG de verificação (antes) salvo em: {caminho_log_antes}")

                        _log_shell.close()
                        _log_ssh.close()
                        _j_client.close()
                    except Exception as e_log_antes:
                        self._upgrade_log(f"[ERRO] Falha no LOG antes (segundo plano): {e_log_antes}")
                    finally:
                        self._upg_log_antes_done.set()

                threading.Thread(
                    target=_worker_log_antes,
                    args=(ip, hostname, dir_hostname, list(_cmds_verificacao)),
                    daemon=True,
                ).start()

                # --- show system information ---
                self._upgrade_log(f"[CMD] show system information")
                shell.send("show system information\n")
                saida = ""
                tentativas = 0
                while tentativas < 6:
                    time.sleep(1)
                    if shell.recv_ready():
                        bloco = shell.recv(65535).decode("utf-8", errors="replace")
                        saida += bloco
                        self._upgrade_log(bloco.strip())
                        tentativas = 0
                    else:
                        tentativas += 1
                # Extrair System Type e Version
                tipo = ""
                versao = ""
                for linha in saida.splitlines():
                    l = linha.strip().lower()
                    if l.startswith("system type"):
                        tipo = linha.split(":",1)[-1].strip()
                    if l.startswith("system version"):
                        versao = linha.split(":",1)[-1].strip()
                self.upg_roteador.config(state="normal")
                self.upg_roteador.delete(0, tk.END)
                self.upg_roteador.insert(0, tipo)
                self.upg_roteador.config(state="readonly")
                self.upg_versao.config(state="normal")
                self.upg_versao.delete(0, tk.END)
                self.upg_versao.insert(0, versao)
                self.upg_versao.config(state="readonly")

                # --- show system satellite ---
                self._upgrade_log(f"[CMD] show system satellite")
                shell.send("show system satellite\n")
                saida2 = ""
                tentativas = 0
                while tentativas < 6:
                    time.sleep(1)
                    if shell.recv_ready():
                        bloco = shell.recv(65535).decode("utf-8", errors="replace")
                        saida2 += bloco
                        self._upgrade_log(bloco.strip())
                        tentativas = 0
                    else:
                        tentativas += 1

                # Extrair quantidade de satélites pelo campo "No. of Satellites:"
                qtd = None
                for linha in saida2.splitlines():
                    m = re.search(r"No\. of Satellites:\s*(\d+)", linha)
                    if m:
                        qtd = m.group(1)
                        break
                # Se não encontrar, mantém fallback antigo (contar linhas)
                if qtd is None:
                    qtd_fallback = 0
                    for linha in saida2.splitlines():
                        if re.search(r"Satellite\s+Id", linha):
                            idx = saida2.splitlines().index(linha)
                            for l2 in saida2.splitlines()[idx+1:]:
                                if l2.strip() and re.match(r"\d+", l2.strip()):
                                    qtd_fallback += 1
                                else:
                                    break
                            break
                    qtd = str(qtd_fallback)
                if not qtd or not str(qtd).isdigit():
                    qtd = "0"
                self.upg_satelitte.config(state="normal")
                self.upg_satelitte.delete(0, tk.END)
                self.upg_satelitte.insert(0, str(qtd))
                self.upg_satelitte.config(state="readonly")

                # --- show system satellite eth-sat 1 ---
                self._upgrade_log(f"[CMD] show system satellite eth-sat 1")
                shell.send("show system satellite eth-sat 1\n")
                saida3 = ""
                tentativas = 0
                while tentativas < 6:
                    time.sleep(1)
                    if shell.recv_ready():
                        bloco = shell.recv(65535).decode("utf-8", errors="replace")
                        saida3 += bloco
                        self._upgrade_log(bloco.strip())
                        tentativas = 0
                    else:
                        tentativas += 1

                # Extrair versão de software do resultado (padrão B-20.9.R4)
                versao_sat = ""
                padrao_versao = re.compile(r"[A-Z]-\d+\.\d+\.R\d+", re.IGNORECASE)
                for linha in saida3.splitlines():
                    l = linha.strip().lower()
                    # Tenta encontrar linha com "software version" ou similar
                    if "software version" in l or "versão de software" in l:
                        # Extrai após ':'
                        partes = linha.split(":",1)
                        if len(partes) > 1:
                            possivel = partes[1].strip()
                            m = padrao_versao.search(possivel)
                            if m:
                                versao_sat = m.group(0)
                                break
                self.upg_control.config(state="normal")
                self.upg_control.delete(0, tk.END)
                self.upg_control.insert(0, versao_sat)
                self.upg_control.config(state="readonly")

                # --- show router 8083 arp ---
                self._upgrade_log(f"[CMD] show router 8083 arp")
                shell.send("show router 8083 arp\n")
                saida4 = ""
                tentativas = 0
                while tentativas < 6:
                    time.sleep(1)
                    if shell.recv_ready():
                        bloco = shell.recv(65535).decode("utf-8", errors="replace")
                        saida4 += bloco
                        self._upgrade_log(bloco.strip())
                        tentativas = 0
                    else:
                        tentativas += 1

                # Extrair IP da linha com "Dyn[I]"
                rede_controle = ""
                for linha in saida4.splitlines():
                    if "Dyn[I]" in linha:
                        m = re.search(r"(\d+\.\d+\.\d+\.\d+)", linha)
                        if m:
                            rede_controle = m.group(1)
                            break
                if not rede_controle:
                    self._upgrade_log("[AVISO] Não foi possível extrair IP de 'show router 8083 arp'.")
                    import queue
                    resultado_q = queue.Queue()
                    def _pedir_ip():
                        val = simpledialog.askstring(
                            "Rede de controle",
                            "Não foi possível obter o IP da rede de controle.\nDigite o IP manualmente:",
                            parent=self.root
                        )
                        resultado_q.put(val)
                    self.root.after(0, _pedir_ip)
                    ip_manual = resultado_q.get()
                    if ip_manual and ip_manual.strip():
                        rede_controle = ip_manual.strip()
                        self._upgrade_log(f"[OK] IP informado manualmente: {rede_controle}")
                    else:
                        self._upgrade_log("[INFO] IP da rede de controle não informado.")
                self.upg_rede_controle.config(state="normal")
                self.upg_rede_controle.delete(0, tk.END)
                self.upg_rede_controle.insert(0, rede_controle)
                self.upg_rede_controle.config(state="readonly")

                # --- show card (extrair controladora ativa) ---
                self._upgrade_log(f"[CMD] show card")
                shell.send("show card\n")
                saida_card = ""
                tentativas = 0
                while tentativas < 6:
                    time.sleep(1)
                    if shell.recv_ready():
                        bloco = shell.recv(65535).decode("utf-8", errors="replace")
                        saida_card += bloco
                        self._upgrade_log(bloco.strip())
                        tentativas = 0
                    else:
                        tentativas += 1
                ctrl_ativa = ""
                for linha in saida_card.splitlines():
                    if "up/active" in linha.lower():
                        m_ctrl = re.match(r'^\s*([ABab])\s', linha)
                        if m_ctrl:
                            ctrl_ativa = m_ctrl.group(1).upper()
                            break
                self.upg_ctrl_ativa.config(state="normal")
                self.upg_ctrl_ativa.delete(0, tk.END)
                self.upg_ctrl_ativa.insert(0, ctrl_ativa)
                self.upg_ctrl_ativa.config(state="readonly")
                if ctrl_ativa:
                    self._upgrade_log(f"[OK] Controladora ativa: {ctrl_ativa}")
                else:
                    self._upgrade_log("[AVISO] Não foi possível detectar a controladora ativa.")

                # --- Se há satélites, solicitar diretórios e criar pastas no roteador ---
                qtd_sat = int(qtd) if qtd and qtd.isdigit() else 0
                if qtd_sat != 0:
                    import queue as _q

                    # Mensagem personalizada antes de cada askdirectory
                    msg_sat = "Agora me diz onde estão os arquivos do TiMOS do Satelitte"
                    msg_7x50 = "Agora me diz onde estão os arquivos do 7x50"
                    def _msg_box(msg):
                        mb = tk.Toplevel(self.root)
                        mb.title("Selecionar diretório")
                        mb.resizable(False, False)
                        mb.attributes("-topmost", True)
                        mb.configure(bg="#F0F2F5")
                        tk.Label(mb, text=msg, font=("Segoe UI", 11, "bold"), bg="#F0F2F5", fg="#124191", padx=24, pady=18).pack()
                        tk.Button(mb, text="OK", font=("Segoe UI", 10, "bold"), bg="#124191", fg="white",
                                  activebackground="#0d2f6b", command=mb.destroy, width=10).pack(pady=(0, 12))
                        mb.update_idletasks()
                        w = mb.winfo_reqwidth()
                        h = mb.winfo_reqheight()
                        x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
                        y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
                        mb.geometry(f"+{x}+{y}")
                        mb.focus_set()
                        mb.grab_set()
                        mb.wait_window()

                    self._piscar_icone()
                    _msg_box(msg_sat)
                    dir_sat_q = _q.Queue()
                    def _pedir_dir_sat():
                        self.root.update_idletasks()
                        val = filedialog.askdirectory(
                            title="Selecionar diretório do TiMOS do Satélite",
                            parent=self.root,
                        )
                        dir_sat_q.put(val)
                    self.root.after(0, _pedir_dir_sat)
                    dir_sat = dir_sat_q.get()
                    if dir_sat:
                        _confirm_sat_q = _q.Queue()
                        def _confirmar_dir_sat():
                            resp = messagebox.askyesno(
                                "Confirmar diretório",
                                f"Diretório selecionado:\n{dir_sat}\n\nEstá correto?",
                                parent=self.root,
                            )
                            _confirm_sat_q.put(resp)
                        self.root.after(0, _confirmar_dir_sat)
                        if not _confirm_sat_q.get():
                            dir_sat = ""

                    self._piscar_icone()
                    _msg_box(msg_7x50)
                    dir_7x50_q = _q.Queue()
                    def _pedir_dir_7x50():
                        self.root.update_idletasks()
                        val = filedialog.askdirectory(
                            title="Selecionar diretório do TiMOS do 7x50",
                            parent=self.root,
                        )
                        dir_7x50_q.put(val)
                    self.root.after(0, _pedir_dir_7x50)
                    dir_7x50 = dir_7x50_q.get()
                    if dir_7x50:
                        _confirm_7x50_q = _q.Queue()
                        def _confirmar_dir_7x50():
                            resp = messagebox.askyesno(
                                "Confirmar diretório",
                                f"Diretório selecionado:\n{dir_7x50}\n\nEstá correto?",
                                parent=self.root,
                            )
                            _confirm_7x50_q.put(resp)
                        self.root.after(0, _confirmar_dir_7x50)
                        if not _confirm_7x50_q.get():
                            dir_7x50 = ""

                    if dir_sat:
                        nome_pasta_sat = os.path.basename(dir_sat)
                        conteudo_sat = os.listdir(dir_sat)
                        self._upgrade_log(f"[INFO] Pasta satélite: {nome_pasta_sat}")
                        self._upgrade_log(f"[INFO] Conteúdo: {conteudo_sat}")

                        # Reconectar ao roteador para criar diretórios
                        jump_client2 = paramiko.SSHClient()
                        jump_client2.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        jump_client2.connect(
                            hostname=jump_ip, port=22, username=jump_user,
                            password=jump_pass, timeout=15,
                            look_for_keys=False, allow_agent=False,
                        )
                        jump_transport2 = jump_client2.get_transport()
                        jump_channel2 = jump_transport2.open_channel(
                            "direct-tcpip", (ip, router_port), ("127.0.0.1", 0),
                        )
                        client2 = paramiko.SSHClient()
                        client2.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        client2.connect(
                            hostname=ip, port=router_port,
                            username=router_user, password=router_pass,
                            timeout=15, look_for_keys=False, allow_agent=False,
                            sock=jump_channel2,
                        )
                        shell_md = client2.invoke_shell(width=512)
                        time.sleep(1)
                        if shell_md.recv_ready():
                            shell_md.recv(65535)
                        shell_md.send("environment no more\n")
                        _tent_drain = 0
                        while _tent_drain < 3:
                            time.sleep(1)
                            if shell_md.recv_ready():
                                shell_md.recv(65535)
                                _tent_drain = 0
                            else:
                                _tent_drain += 1

                        # Verificar diretórios existentes via checagem direta por diretório
                        def _dir_existe_no_roteador(shell, cf, nome_pasta):
                            """Verifica se um diretório existe no roteador usando file dir direto."""
                            cmd = f"file dir {cf}\\{nome_pasta}"
                            self._upgrade_log(f"[CMD] {cmd}")
                            shell.send(cmd + "\n")
                            resp = ""
                            tent = 0
                            while tent < 6:
                                time.sleep(1)
                                if shell.recv_ready():
                                    bloco = shell.recv(65535).decode("utf-8", errors="replace")
                                    resp += bloco
                                    tent = 0
                                else:
                                    tent += 1
                            resp_lower = resp.lower()
                            # Se contém "not found", "error", "invalid" → não existe
                            if "not found" in resp_lower or "error" in resp_lower or "invalid" in resp_lower or "bad" in resp_lower:
                                self._upgrade_log(f"[INFO] {cf}\\{nome_pasta} não existe.")
                                return False
                            # Se contém "<DIR>" ou "file" ou "bytes" → existe
                            self._upgrade_log(f"[INFO] {cf}\\{nome_pasta} já existe no roteador.")
                            return True

                        # Criar diretórios do satélite (somente se não existem)
                        for cf in ("cf3-a:", "cf3-b:"):
                            if _dir_existe_no_roteador(shell_md, cf, nome_pasta_sat):
                                self._upgrade_log(f"[SKIP] {cf}\\{nome_pasta_sat} já existe, não será criado.")
                            else:
                                cmd_md = f"file md {cf}\\{nome_pasta_sat}"
                                self._upgrade_log(f"[CMD] {cmd_md}")
                                shell_md.send(cmd_md + "\n")
                                time.sleep(2)
                                if shell_md.recv_ready():
                                    resp_md = shell_md.recv(65535).decode("utf-8", errors="replace")
                                    self._upgrade_log(resp_md.strip())

                        if dir_7x50:
                            nome_pasta_7x50 = os.path.basename(dir_7x50)
                            conteudo_7x50 = os.listdir(dir_7x50)
                            self._upgrade_log(f"[INFO] Pasta 7x50: {nome_pasta_7x50}")
                            self._upgrade_log(f"[INFO] Conteúdo: {conteudo_7x50}")

                            # Criar diretórios do 7x50 (somente se não existem)
                            for cf in ("cf3-a:", "cf3-b:"):
                                if _dir_existe_no_roteador(shell_md, cf, nome_pasta_7x50):
                                    self._upgrade_log(f"[SKIP] {cf}\\{nome_pasta_7x50} já existe, não será criado.")
                                else:
                                    cmd_md = f"file md {cf}\\{nome_pasta_7x50}"
                                    self._upgrade_log(f"[CMD] {cmd_md}")
                                    shell_md.send(cmd_md + "\n")
                                    time.sleep(2)
                                    if shell_md.recv_ready():
                                        resp_md = shell_md.recv(65535).decode("utf-8", errors="replace")
                                        self._upgrade_log(resp_md.strip())

                        shell_md.close()
                        client2.close()
                        jump_client2.close()
                        self._upgrade_log("[OK] Verificação/criação de diretórios concluída.")

                        # --- Verificar arquivos já existentes no roteador ---
                        self._upgrade_log("[INFO] Verificando arquivos existentes no roteador...")
                        jump_dir_client = paramiko.SSHClient()
                        jump_dir_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        jump_dir_client.connect(
                            hostname=jump_ip, port=22, username=jump_user,
                            password=jump_pass, timeout=15,
                            look_for_keys=False, allow_agent=False,
                        )
                        jump_dir_transport = jump_dir_client.get_transport()
                        jump_dir_channel = jump_dir_transport.open_channel(
                            "direct-tcpip", (ip, router_port), ("127.0.0.1", 0),
                        )
                        dir_ssh = paramiko.SSHClient()
                        dir_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        dir_ssh.connect(
                            hostname=ip, port=router_port,
                            username=router_user, password=router_pass,
                            timeout=15, look_for_keys=False, allow_agent=False,
                            sock=jump_dir_channel,
                        )
                        shell_dir = dir_ssh.invoke_shell(width=512)
                        time.sleep(1)
                        if shell_dir.recv_ready():
                            shell_dir.recv(65535)
                        shell_dir.send("environment no more\n")
                        time.sleep(1)
                        if shell_dir.recv_ready():
                            shell_dir.recv(65535)

                        # Mapear arquivos existentes: {(cf, pasta): set(nomes)}
                        arquivos_existentes = {}
                        pastas_verificar = []
                        pastas_verificar.append(("cf3-a:", nome_pasta_sat))
                        pastas_verificar.append(("cf3-b:", nome_pasta_sat))
                        if dir_7x50:
                            pastas_verificar.append(("cf3-a:", nome_pasta_7x50))
                            pastas_verificar.append(("cf3-b:", nome_pasta_7x50))

                        for cf, pasta in pastas_verificar:
                            cmd_dir = f"file dir {cf}\\{pasta}"
                            self._upgrade_log(f"[CMD] {cmd_dir}")
                            shell_dir.send(cmd_dir + "\n")
                            saida_dir = ""
                            tentativas_dir = 0
                            while tentativas_dir < 6:
                                time.sleep(1)
                                if shell_dir.recv_ready():
                                    bloco = shell_dir.recv(65535).decode("utf-8", errors="replace")
                                    saida_dir += bloco
                                    tentativas_dir = 0
                                else:
                                    tentativas_dir += 1
                            # Extrair nomes de arquivo da saída do file dir
                            nomes_remoto = set()
                            for linha in saida_dir.splitlines():
                                linha_s = linha.strip()
                                # Linhas de arquivo típicas: "04/02/2026 10:30  12345678  arquivo.tim"
                                # ou formato Nokia: tamanho  nome
                                partes = linha_s.split()
                                if len(partes) >= 3:
                                    possivel_nome = partes[-1]
                                    # Ignorar linhas de cabeçalho/totais/diretório
                                    if possivel_nome in (".", "..", "<DIR>", "bytes", "free."):
                                        continue
                                    if possivel_nome.endswith("\\") or possivel_nome.endswith("/"):
                                        continue
                                    # Verificar se há um campo numérico (tamanho) antes do nome
                                    for p in partes[:-1]:
                                        if p.replace(",", "").isdigit() and len(p.replace(",", "")) > 0:
                                            nomes_remoto.add(possivel_nome)
                                            break
                            arquivos_existentes[(cf, pasta)] = nomes_remoto
                            if nomes_remoto:
                                self._upgrade_log(f"[INFO] {cf}\\{pasta}: {len(nomes_remoto)} arquivo(s) encontrado(s)")
                            else:
                                self._upgrade_log(f"[INFO] {cf}\\{pasta}: vazio ou não listável")

                        shell_dir.close()
                        dir_ssh.close()
                        jump_dir_client.close()

                        # Armazenar informações para uso posterior
                        self._upg_dir_sat = dir_sat
                        self._upg_dir_7x50 = dir_7x50
                        self._upg_nome_pasta_sat = nome_pasta_sat
                        self._upg_nome_pasta_7x50 = nome_pasta_7x50 if dir_7x50 else ""
                        self._upg_conteudo_sat = conteudo_sat
                        self._upg_conteudo_7x50 = conteudo_7x50 if dir_7x50 else []

                        # --- Transferência SFTP e verificação de arquivos ---
                        # Coletar todos os arquivos locais
                        arquivos_sat_local = []
                        for arq in conteudo_sat:
                            local_path = os.path.join(dir_sat, arq)
                            if os.path.isfile(local_path):
                                arquivos_sat_local.append((local_path, arq, nome_pasta_sat))
                        arquivos_7x50_local = []
                        if dir_7x50:
                            for arq in conteudo_7x50:
                                local_path = os.path.join(dir_7x50, arq)
                                if os.path.isfile(local_path):
                                    arquivos_7x50_local.append((local_path, arq, nome_pasta_7x50))

                        # Construir lista completa para verificação (TODOS os arquivos)
                        todos_remotos = []
                        for local_path, nome_arq, pasta in arquivos_sat_local:
                            for cf in ("cf3-a:", "cf3-b:"):
                                todos_remotos.append((cf, pasta, nome_arq))
                        for local_path, nome_arq, pasta in arquivos_7x50_local:
                            for cf in ("cf3-a:", "cf3-b:"):
                                todos_remotos.append((cf, pasta, nome_arq))

                        # Filtrar apenas arquivos faltantes para transferência (somente cf3-a: via SFTP)
                        arquivos_transferir = []  # (local_path, nome_arq, pasta)
                        arquivos_copiar_b = []    # (nome_arq, pasta) - faltantes em cf3-b:
                        for local_path, nome_arq, pasta in arquivos_sat_local:
                            existentes_a = arquivos_existentes.get(("cf3-a:", pasta), set())
                            if nome_arq not in existentes_a:
                                arquivos_transferir.append((local_path, nome_arq, pasta))
                            else:
                                self._upgrade_log(f"[SKIP] {nome_arq} já existe em cf3-a:\\{pasta}")
                            existentes_b = arquivos_existentes.get(("cf3-b:", pasta), set())
                            if nome_arq not in existentes_b:
                                arquivos_copiar_b.append((nome_arq, pasta))
                            else:
                                self._upgrade_log(f"[SKIP] {nome_arq} já existe em cf3-b:\\{pasta}")
                        for local_path, nome_arq, pasta in arquivos_7x50_local:
                            existentes_a = arquivos_existentes.get(("cf3-a:", pasta), set())
                            if nome_arq not in existentes_a:
                                arquivos_transferir.append((local_path, nome_arq, pasta))
                            else:
                                self._upgrade_log(f"[SKIP] {nome_arq} já existe em cf3-a:\\{pasta}")
                            existentes_b = arquivos_existentes.get(("cf3-b:", pasta), set())
                            if nome_arq not in existentes_b:
                                arquivos_copiar_b.append((nome_arq, pasta))
                            else:
                                self._upgrade_log(f"[SKIP] {nome_arq} já existe em cf3-b:\\{pasta}")

                        if not arquivos_transferir and not arquivos_copiar_b:
                            self._upgrade_log("[OK] Todos os arquivos já existem no roteador. Nenhuma transferência necessária.")

                        # Calcular tamanho total em bytes para progresso real
                        total_bytes = 0
                        for local_path, nome_arq, pasta in arquivos_transferir:
                            total_bytes += os.path.getsize(local_path)

                        # Criar janela de progresso 3D com Canvas para barras animadas
                        progress_q = _q.Queue()
                        def _criar_progress():
                            dlg = tk.Toplevel(self.root)
                            dlg.title("Transferência SFTP")
                            dlg.resizable(False, False)
                            dlg.configure(bg="#D4D0C8", bd=3, relief="raised")
                            dlg.transient(self.root)
                            dlg.grab_set()

                            inner = tk.Frame(dlg, bg="#D4D0C8", bd=2, relief="sunken")
                            inner.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

                            # --- Barra de transferência ---
                            tk.Label(inner, text="Transferência de arquivos:",
                                     font=("Segoe UI", 9, "bold"), bg="#D4D0C8").pack(anchor="w", padx=8, pady=(8, 2))
                            lbl_arq = tk.Label(inner, text="", font=("Segoe UI", 8), bg="#D4D0C8", fg="#333")
                            lbl_arq.pack(anchor="w", padx=8)
                            bar_canvas1 = tk.Canvas(inner, height=22, bg="#808080", bd=1, relief="sunken", highlightthickness=0)
                            bar_canvas1.pack(fill=tk.X, padx=8, pady=(0, 8))

                            # --- Barra de verificação ---
                            tk.Label(inner, text="Verificação de arquivos:",
                                     font=("Segoe UI", 9, "bold"), bg="#D4D0C8").pack(anchor="w", padx=8, pady=(4, 2))
                            bar_canvas2 = tk.Canvas(inner, height=22, bg="#808080", bd=1, relief="sunken", highlightthickness=0)
                            bar_canvas2.pack(fill=tk.X, padx=8, pady=(0, 8))

                            dlg.update_idletasks()
                            w = dlg.winfo_reqwidth()
                            if w < 460:
                                w = 460
                            h = dlg.winfo_reqheight()
                            dlg.geometry(f"{w}x{h}")
                            dlg.update_idletasks()
                            rx = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
                            ry = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
                            dlg.geometry(f"+{rx}+{ry}")

                            progress_q.put({
                                "dlg": dlg, "inner": inner,
                                "bar_canvas1": bar_canvas1, "bar_canvas2": bar_canvas2,
                                "lbl_arq": lbl_arq,
                            })
                        self.root.after(0, _criar_progress)
                        pw = progress_q.get()

                        _bytes_enviados = [0]  # mutable para callback
                        _anim_phase = [0]  # fase da animação

                        def _atualizar_barra_transf(pct, nome_arq_atual=""):
                            def _upd():
                                c = pw["bar_canvas1"]
                                c.delete("all")
                                cw = c.winfo_width()
                                ch = c.winfo_height()
                                if cw < 2:
                                    return
                                fill_w = int(cw * pct / 100)
                                # Barra principal azul
                                if fill_w > 0:
                                    c.create_rectangle(0, 0, fill_w, ch, fill="#000080", outline="")
                                    # Efeito de brilho animado (faixa diagonal)
                                    phase = _anim_phase[0] % (cw + 40)
                                    stripe_x = phase - 20
                                    if stripe_x < fill_w:
                                        x1 = max(0, stripe_x)
                                        x2 = min(fill_w, stripe_x + 30)
                                        if x2 > x1:
                                            c.create_rectangle(x1, 0, x2, ch, fill="#3366CC", outline="")
                                    _anim_phase[0] += 8
                                # Texto percentual centrado na barra
                                txt = f"{int(pct)}%"
                                cx = cw // 2
                                cy = ch // 2
                                # Sombra para legibilidade
                                c.create_text(cx + 1, cy + 1, text=txt, fill="#000000",
                                              font=("Segoe UI", 9, "bold"))
                                c.create_text(cx, cy, text=txt, fill="#FFFFFF",
                                              font=("Segoe UI", 9, "bold"))
                                if nome_arq_atual:
                                    pw["lbl_arq"].config(text=nome_arq_atual)
                            self.root.after(0, _upd)

                        def _atualizar_barra_check(pct):
                            def _upd():
                                c = pw["bar_canvas2"]
                                c.delete("all")
                                cw = c.winfo_width()
                                ch = c.winfo_height()
                                if cw < 2:
                                    return
                                fill_w = int(cw * pct / 100)
                                if fill_w > 0:
                                    c.create_rectangle(0, 0, fill_w, ch, fill="#008000", outline="")
                                    # Efeito de brilho animado
                                    phase = _anim_phase[0] % (cw + 40)
                                    stripe_x = phase - 20
                                    if stripe_x < fill_w:
                                        x1 = max(0, stripe_x)
                                        x2 = min(fill_w, stripe_x + 30)
                                        if x2 > x1:
                                            c.create_rectangle(x1, 0, x2, ch, fill="#33CC66", outline="")
                                    _anim_phase[0] += 8
                                txt = f"{int(pct)}%"
                                cx = cw // 2
                                cy = ch // 2
                                c.create_text(cx + 1, cy + 1, text=txt, fill="#000000",
                                              font=("Segoe UI", 9, "bold"))
                                c.create_text(cx, cy, text=txt, fill="#FFFFFF",
                                              font=("Segoe UI", 9, "bold"))
                            self.root.after(0, _upd)

                        # Helper: dialog de erro com opção de prosseguir ou refazer
                        def _perguntar_erro(titulo, mensagem, obs_text=None):
                            """Exibe dialog de erro na GUI thread. Retorna 'prosseguir' ou 'refazer'."""
                            resp_q = _q.Queue()
                            def _dlg():
                                dlg = tk.Toplevel(self.root)
                                dlg.title(titulo)
                                dlg.resizable(False, False)
                                dlg.configure(bg="#FFF3CD", bd=3, relief="raised")
                                dlg.transient(self.root)
                                dlg.grab_set()
                                tk.Label(
                                    dlg, text=mensagem,
                                    font=("Segoe UI", 10), bg="#FFF3CD", fg="#856404",
                                    justify="left", wraplength=500,
                                ).pack(padx=20, pady=(16, 4))
                                if obs_text:
                                    tk.Label(
                                        dlg, text=obs_text,
                                        font=("Segoe UI", 9, "italic"), bg="#FFF3CD", fg="#664d03",
                                        justify="left", wraplength=500,
                                    ).pack(padx=20, pady=(2, 10))
                                bf = tk.Frame(dlg, bg="#FFF3CD")
                                bf.pack(pady=(0, 14))
                                tk.Button(bf, text="Refazer", font=("Segoe UI", 9, "bold"), width=12,
                                          command=lambda: (dlg.destroy(), resp_q.put("refazer"))).pack(side=tk.LEFT, padx=6)
                                tk.Button(bf, text="Prosseguir", font=("Segoe UI", 9), width=12,
                                          command=lambda: (dlg.destroy(), resp_q.put("prosseguir"))).pack(side=tk.LEFT, padx=6)
                                dlg.update_idletasks()
                                w = dlg.winfo_reqwidth()
                                h = dlg.winfo_reqheight()
                                x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
                                y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
                                dlg.geometry(f"+{x}+{y}")
                                dlg.focus_set()
                            self.root.after(0, _dlg)
                            return resp_q.get()

                        # Helper: dialog de MINOR com opção OK (prosseguir) ou Não (interromper)
                        def _perguntar_minor(titulo, mensagem):
                            """Exibe dialog MINOR na GUI thread. Retorna True (OK) ou False (Não)."""
                            resp_q = _q.Queue()
                            def _dlg():
                                dlg = tk.Toplevel(self.root)
                                dlg.title(titulo)
                                dlg.resizable(False, False)
                                dlg.configure(bg="#FFF8DC", bd=3, relief="raised")
                                dlg.transient(self.root)
                                dlg.grab_set()
                                tk.Label(
                                    dlg, text=mensagem,
                                    font=("Segoe UI", 10), bg="#FFF8DC", fg="#856404",
                                    justify="left", wraplength=520,
                                ).pack(padx=20, pady=(16, 12))
                                bf = tk.Frame(dlg, bg="#FFF8DC")
                                bf.pack(pady=(0, 14))
                                tk.Button(bf, text="OK", font=("Segoe UI", 9, "bold"), width=12,
                                          command=lambda: (dlg.destroy(), resp_q.put(True))).pack(side=tk.LEFT, padx=6)
                                tk.Button(bf, text="Não", font=("Segoe UI", 9), width=12,
                                          command=lambda: (dlg.destroy(), resp_q.put(False))).pack(side=tk.LEFT, padx=6)
                                dlg.update_idletasks()
                                w = dlg.winfo_reqwidth()
                                h = dlg.winfo_reqheight()
                                x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
                                y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
                                dlg.geometry(f"+{x}+{y}")
                                dlg.focus_set()
                            self.root.after(0, _dlg)
                            return resp_q.get()

                        # Helper: dialog de erro grave com opção OK ou Não (interromper)
                        def _perguntar_erro_sync(titulo, mensagem):
                            """Exibe dialog de erro na GUI thread. Retorna True (OK) ou False (Não/interromper)."""
                            resp_q = _q.Queue()
                            def _dlg():
                                dlg = tk.Toplevel(self.root)
                                dlg.title(titulo)
                                dlg.resizable(False, False)
                                dlg.configure(bg="#F8D7DA", bd=3, relief="raised")
                                dlg.transient(self.root)
                                dlg.grab_set()
                                tk.Label(
                                    dlg, text=mensagem,
                                    font=("Segoe UI", 10), bg="#F8D7DA", fg="#721c24",
                                    justify="left", wraplength=520,
                                ).pack(padx=20, pady=(16, 12))
                                bf = tk.Frame(dlg, bg="#F8D7DA")
                                bf.pack(pady=(0, 14))
                                tk.Button(bf, text="OK — Prosseguir", font=("Segoe UI", 9, "bold"), width=16,
                                          command=lambda: (dlg.destroy(), resp_q.put(True))).pack(side=tk.LEFT, padx=6)
                                tk.Button(bf, text="Não — Interromper", font=("Segoe UI", 9), width=16,
                                          command=lambda: (dlg.destroy(), resp_q.put(False))).pack(side=tk.LEFT, padx=6)
                                dlg.update_idletasks()
                                w = dlg.winfo_reqwidth()
                                h = dlg.winfo_reqheight()
                                x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
                                y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
                                dlg.geometry(f"+{x}+{y}")
                                dlg.focus_set()
                            self.root.after(0, _dlg)
                            return resp_q.get()

                        # Conectar SFTP via jumpserver com buffers otimizados
                        if arquivos_transferir:
                            self._upgrade_log(f"[INFO] {len(arquivos_transferir)} arquivo(s) para transferir...")
                            self._upgrade_log("[INFO] Conectando SFTP via jumpserver...")
                            jump_sftp_client = paramiko.SSHClient()
                            jump_sftp_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                            jump_sftp_client.connect(
                                hostname=jump_ip, port=22, username=jump_user,
                                password=jump_pass, timeout=15,
                                look_for_keys=False, allow_agent=False,
                            )
                            jump_sftp_transport = jump_sftp_client.get_transport()
                            jump_sftp_transport.set_keepalive(30)
                            jump_sftp_channel = jump_sftp_transport.open_channel(
                                "direct-tcpip", (ip, router_port), ("127.0.0.1", 0),
                            )
                            sftp_ssh = paramiko.SSHClient()
                            sftp_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                            sftp_ssh.connect(
                                hostname=ip, port=router_port,
                                username=router_user, password=router_pass,
                                timeout=15, look_for_keys=False, allow_agent=False,
                                sock=jump_sftp_channel,
                            )
                            sftp_transport = sftp_ssh.get_transport()
                            sftp_transport.default_window_size = 2147483647
                            sftp_transport.default_max_packet_size = 32768
                            sftp_transport.packetizer.REKEY_BYTES = pow(2, 40)
                            sftp_transport.packetizer.REKEY_PACKETS = pow(2, 40)
                            sftp_transport.use_compression(False)
                            sftp_transport.set_keepalive(30)
                            sftp = sftp_ssh.open_sftp()
                            sftp.get_channel().settimeout(300)
                            sftp.get_channel().in_window_size = 2147483647
                            sftp.get_channel().out_window_size = 2147483647
                            sftp.get_channel().in_max_packet_size = 32768
                            sftp.get_channel().out_max_packet_size = 32768
                            self._upgrade_log("[OK] SFTP conectado (buffers otimizados).")

                            # Callback de progresso por bytes
                            _last_update = [0.0]
                            def _sftp_callback(bytes_so_far, total_file_bytes):
                                _bytes_enviados[0] += bytes_so_far - getattr(_sftp_callback, '_prev', 0)
                                _sftp_callback._prev = bytes_so_far
                                now = time.time()
                                if now - _last_update[0] >= 0.08:
                                    _last_update[0] = now
                                    if total_bytes > 0:
                                        pct = min(_bytes_enviados[0] * 100 / total_bytes, 100)
                                        _atualizar_barra_transf(pct, _sftp_callback._nome)

                            # Transferir apenas arquivos faltantes para cf3-a: via SFTP (um por um)
                            for local_path, nome_arq, pasta in arquivos_transferir:
                                remote_path = f"/cf3-a:/{pasta}/{nome_arq}"
                                label_arq = f"{nome_arq} → cf3-a:\\{pasta}\\"
                                while True:
                                    self._upgrade_log(f"[SFTP] {label_arq}")
                                    _sftp_callback._prev = 0
                                    _sftp_callback._nome = label_arq
                                    try:
                                        # file_size for confirm=True; buffer_size=32768 for max throughput
                                        fsize = os.path.getsize(local_path)
                                        sftp.put(local_path, remote_path, callback=_sftp_callback, confirm=True)
                                        self._upgrade_log(f"[OK] {nome_arq} transferido ({fsize:,} bytes).")
                                        break
                                    except Exception as e:
                                        self._upgrade_log(f"[ERRO] Falha ao transferir {nome_arq}: {e}")
                                        acao = _perguntar_erro(
                                            "Erro na transferência SFTP",
                                            f"Falha ao transferir {nome_arq}:\n{e}\n\nDeseja prosseguir ou refazer a ação?",
                                        )
                                        if acao == "prosseguir":
                                            break

                            _atualizar_barra_transf(100, "Concluído")
                            sftp.close()
                            sftp_ssh.close()
                            jump_sftp_client.close()
                            self._upgrade_log("[OK] Transferência SFTP para cf3-a: concluída.")
                        else:
                            if not arquivos_copiar_b:
                                _atualizar_barra_transf(100, "Todos os arquivos já existem no roteador")
                            else:
                                _atualizar_barra_transf(100, "Nenhum arquivo para transferir via SFTP")

                        # Copiar arquivos de cf3-a: para cf3-b: via SSH file copy
                        if arquivos_copiar_b:
                            self._upgrade_log(f"[INFO] Copiando {len(arquivos_copiar_b)} arquivo(s) de cf3-a: para cf3-b:...")
                            jump_cp_client = paramiko.SSHClient()
                            jump_cp_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                            jump_cp_client.connect(
                                hostname=jump_ip, port=22, username=jump_user,
                                password=jump_pass, timeout=15,
                                look_for_keys=False, allow_agent=False,
                            )
                            jump_cp_transport = jump_cp_client.get_transport()
                            jump_cp_channel = jump_cp_transport.open_channel(
                                "direct-tcpip", (ip, router_port), ("127.0.0.1", 0),
                            )
                            cp_ssh = paramiko.SSHClient()
                            cp_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                            cp_ssh.connect(
                                hostname=ip, port=router_port,
                                username=router_user, password=router_pass,
                                timeout=15, look_for_keys=False, allow_agent=False,
                                sock=jump_cp_channel,
                            )
                            shell_cp = cp_ssh.invoke_shell(width=512)
                            time.sleep(1)
                            if shell_cp.recv_ready():
                                shell_cp.recv(65535)
                            shell_cp.send("environment no more\n")
                            time.sleep(1)
                            if shell_cp.recv_ready():
                                shell_cp.recv(65535)

                            for nome_arq, pasta in arquivos_copiar_b:
                                cmd_cp = f"file copy cf3-a:\\{pasta}\\{nome_arq} cf3-b:\\{pasta}\\{nome_arq} force"
                                while True:
                                    self._upgrade_log(f"[CMD] {cmd_cp}")
                                    shell_cp.send(cmd_cp + "\n")
                                    resp_cp = ""
                                    tent_cp = 0
                                    while tent_cp < 10:
                                        time.sleep(2)
                                        if shell_cp.recv_ready():
                                            bloco = shell_cp.recv(65535).decode("utf-8", errors="replace")
                                            resp_cp += bloco
                                            tent_cp = 0
                                        else:
                                            tent_cp += 1
                                    self._upgrade_log(resp_cp.strip())
                                    if "copied" in resp_cp.lower() or "1 file" in resp_cp.lower():
                                        self._upgrade_log(f"[OK] {nome_arq} copiado para cf3-b:")
                                        break
                                    elif "fail" in resp_cp.lower() or "error" in resp_cp.lower():
                                        self._upgrade_log(f"[ERRO] Falha ao copiar {nome_arq} para cf3-b:")
                                        acao = _perguntar_erro(
                                            "Erro na cópia de arquivo",
                                            f"Falha ao copiar {nome_arq} para cf3-b:\n\nDeseja prosseguir ou refazer a ação?",
                                        )
                                        if acao == "prosseguir":
                                            break
                                        # else: refazer → loop again
                                    else:
                                        break

                            shell_cp.close()
                            cp_ssh.close()
                            jump_cp_client.close()
                            self._upgrade_log("[OK] Cópia para cf3-b: concluída.")

                        # Verificação de arquivos (file version check)
                        self._upgrade_log("[INFO] Iniciando verificação de arquivos...")
                        jump_chk_client = paramiko.SSHClient()
                        jump_chk_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        jump_chk_client.connect(
                            hostname=jump_ip, port=22, username=jump_user,
                            password=jump_pass, timeout=15,
                            look_for_keys=False, allow_agent=False,
                        )
                        jump_chk_transport = jump_chk_client.get_transport()
                        jump_chk_channel = jump_chk_transport.open_channel(
                            "direct-tcpip", (ip, router_port), ("127.0.0.1", 0),
                        )
                        chk_ssh = paramiko.SSHClient()
                        chk_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        chk_ssh.connect(
                            hostname=ip, port=router_port,
                            username=router_user, password=router_pass,
                            timeout=15, look_for_keys=False, allow_agent=False,
                            sock=jump_chk_channel,
                        )
                        shell_chk = chk_ssh.invoke_shell(width=512)
                        time.sleep(1)
                        if shell_chk.recv_ready():
                            shell_chk.recv(65535)
                        shell_chk.send("environment no more\n")
                        time.sleep(1)
                        if shell_chk.recv_ready():
                            shell_chk.recv(65535)

                        # Filtrar .txt da lista de verificação
                        remotos_check = [(cf, pasta, nome_arq) for cf, pasta, nome_arq in todos_remotos
                                         if not nome_arq.lower().endswith(".txt")]
                        total_checks = len(remotos_check)

                        chk_count = 0
                        _interromper_check = False
                        _boot_ldr_minors_chk = []  # Coletar MINORs de boot.ldr para merge
                        self._suppress_nokia_error_box = True  # Suprimir caixa genérica durante checks
                        for cf, pasta, nome_arq in remotos_check:
                            if _interromper_check:
                                break
                            cmd_chk = f"file version check {cf}\\{pasta}\\{nome_arq}"
                            while True:
                                self._upgrade_log(f"[CMD] {cmd_chk}")
                                shell_chk.send(cmd_chk + "\n")
                                resp_chk = ""
                                tent_chk = 0
                                while tent_chk < 10:
                                    time.sleep(2)
                                    if shell_chk.recv_ready():
                                        bloco = shell_chk.recv(65535).decode("utf-8", errors="replace")
                                        resp_chk += bloco
                                        tent_chk = 0
                                    else:
                                        tent_chk += 1
                                self._upgrade_log(resp_chk.strip())
                                # Verificar MINOR:
                                if "minor:" in resp_chk.lower():
                                    minor_info = ""
                                    for l in resp_chk.splitlines():
                                        if "minor:" in l.lower():
                                            minor_info = l.strip()
                                            break
                                    # Se boot.ldr, acumular para merge em vez de mostrar individualmente
                                    if "boot.ldr" in nome_arq.lower():
                                        _boot_ldr_minors_chk.append((f"{cf}\\{pasta}\\{nome_arq}", minor_info))
                                        break
                                    else:
                                        ok = _perguntar_minor(
                                            "MINOR detectado",
                                            f"O roteador reportou MINOR no file version check de:\n"
                                            f"{cf}\\{pasta}\\{nome_arq}\n\n"
                                            f"{minor_info}\n\n"
                                            f"Deseja prosseguir?",
                                        )
                                        if ok:
                                            break
                                        else:
                                            _interromper_check = True
                                            self._upgrade_log("[INFO] Verificação interrompida pelo usuário.")
                                            break
                                elif "fail" in resp_chk.lower() or "error" in resp_chk.lower():
                                    self._upgrade_log(f"[ERRO] Falha no version check de {nome_arq}")
                                    acao = _perguntar_erro(
                                        "Erro na verificação de arquivo",
                                        f"Falha no file version check de:\n{cf}\\{pasta}\\{nome_arq}\n\nDeseja prosseguir ou refazer a ação?",
                                    )
                                    if acao == "prosseguir":
                                        break
                                    # else: refazer → loop again
                                else:
                                    break
                            chk_count += 1
                            if total_checks > 0:
                                _atualizar_barra_check(chk_count * 100 / total_checks)
                        self._suppress_nokia_error_box = False  # Restaurar caixa genérica

                        # Exibir diálogo único com todos os MINORs de boot.ldr das pastas
                        if _boot_ldr_minors_chk and not _interromper_check:
                            _v_atual = re.search(r'-(\d+)\.', versao)
                            _v_dest = re.search(r'-(\d+)\.', nome_pasta_7x50 if dir_7x50 else '')
                            msg_esperado = ""
                            if _v_atual and _v_dest and _v_atual.group(1) == '20' and _v_dest.group(1) == '24':
                                msg_esperado = "\n\nEsta falha é um comportamento esperado, pois a versão 20 não lê o boot loader da versão 24. Pode prosseguir."
                            detalhes_chk = "\n".join(
                                f"{caminho} — {info}" for caminho, info in _boot_ldr_minors_chk
                            )
                            ok_chk = _perguntar_minor(
                                "MINOR detectado — boot.ldr",
                                f"O roteador reportou MINOR no file version check de:\n"
                                f"{detalhes_chk}\n\n"
                                f"Versão atual: {versao}\n"
                                f"Versão destino: {nome_pasta_7x50 if dir_7x50 else '7x50'}{msg_esperado}\n\n"
                                f"Deseja prosseguir?",
                            )
                            if not ok_chk:
                                _interromper_check = True
                                self._upgrade_log("[INFO] Verificação interrompida pelo usuário.")

                        _atualizar_barra_check(100)

                        # Exibir conteúdo dos diretórios após finalizar check
                        dirs_listar = []
                        dirs_listar.append(f"cf3-a:\\{nome_pasta_sat}")
                        dirs_listar.append(f"cf3-b:\\{nome_pasta_sat}")
                        if dir_7x50:
                            dirs_listar.append(f"cf3-a:\\{nome_pasta_7x50}")
                            dirs_listar.append(f"cf3-b:\\{nome_pasta_7x50}")
                        for dir_path in dirs_listar:
                            cmd_dir_list = f"file dir {dir_path}"
                            self._upgrade_log(f"[CMD] {cmd_dir_list}")
                            shell_chk.send(cmd_dir_list + "\n")
                            resp_dir_list = ""
                            tent_dir = 0
                            while tent_dir < 6:
                                time.sleep(1)
                                if shell_chk.recv_ready():
                                    bloco = shell_chk.recv(65535).decode("utf-8", errors="replace")
                                    resp_dir_list += bloco
                                    tent_dir = 0
                                else:
                                    tent_dir += 1
                            self._upgrade_log(resp_dir_list.strip())

                        shell_chk.close()
                        chk_ssh.close()
                        jump_chk_client.close()
                        self._upgrade_log("[OK] Verificação de arquivos concluída.")

                        # Fechar janela de progresso
                        time.sleep(1)
                        self.root.after(0, pw["dlg"].destroy)

                        # --- Configurar software-repository dos satélites ---
                        self._upgrade_log("[INFO] Configurando software-repository dos satélites...")
                        jump_cfg_client = paramiko.SSHClient()
                        jump_cfg_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        jump_cfg_client.connect(
                            hostname=jump_ip, port=22, username=jump_user,
                            password=jump_pass, timeout=15,
                            look_for_keys=False, allow_agent=False,
                        )
                        jump_cfg_transport = jump_cfg_client.get_transport()
                        jump_cfg_channel = jump_cfg_transport.open_channel(
                            "direct-tcpip", (ip, router_port), ("127.0.0.1", 0),
                        )
                        cfg_ssh = paramiko.SSHClient()
                        cfg_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        cfg_ssh.connect(
                            hostname=ip, port=router_port,
                            username=router_user, password=router_pass,
                            timeout=15, look_for_keys=False, allow_agent=False,
                            sock=jump_cfg_channel,
                        )
                        shell_cfg = cfg_ssh.invoke_shell(width=512)
                        time.sleep(1)
                        if shell_cfg.recv_ready():
                            shell_cfg.recv(65535)
                        shell_cfg.send("environment no more\n")
                        time.sleep(1)
                        if shell_cfg.recv_ready():
                            shell_cfg.recv(65535)

                        for esat_id in range(1, qtd_sat + 1):
                            esat_name = f"esat-{esat_id}"
                            # Extrair Primary Location atual
                            cmd_show = f'show system software-repository "{esat_name}"'
                            self._upgrade_log(f"[CMD] {cmd_show}")
                            shell_cfg.send(cmd_show + "\n")
                            saida_repo = ""
                            tent_repo = 0
                            while tent_repo < 6:
                                time.sleep(1)
                                if shell_cfg.recv_ready():
                                    bloco = shell_cfg.recv(65535).decode("utf-8", errors="replace")
                                    saida_repo += bloco
                                    tent_repo = 0
                                else:
                                    tent_repo += 1
                            self._upgrade_log(saida_repo.strip())

                            primary_loc = ""
                            for linha in saida_repo.splitlines():
                                if "primary location" in linha.lower():
                                    m_loc = re.search(r"(cf3:\S+)", linha)
                                    if m_loc:
                                        primary_loc = m_loc.group(1)
                                    break
                            if primary_loc:
                                self._upgrade_log(f"[INFO] {esat_name} Primary Location atual: {primary_loc}")
                            else:
                                self._upgrade_log(f"[AVISO] Não foi possível extrair Primary Location de {esat_name}")

                            # Configurar Primary Location atual como Secondary Location
                            if primary_loc:
                                cmd_sec = f'/configure system software-repository "{esat_name}" secondary-location {primary_loc}'
                                self._upgrade_log(f"[CMD] {cmd_sec}")
                                shell_cfg.send(cmd_sec + "\n")
                                time.sleep(3)
                                if shell_cfg.recv_ready():
                                    resp_cfg = shell_cfg.recv(65535).decode("utf-8", errors="replace")
                                    self._upgrade_log(resp_cfg.strip())

                            # Configurar novo Primary Location com o diretório do satélite
                            new_primary = f"cf3:\\{nome_pasta_sat}"
                            cmd_pri = f'/configure system software-repository "{esat_name}" primary-location {new_primary}'
                            self._upgrade_log(f"[CMD] {cmd_pri}")
                            shell_cfg.send(cmd_pri + "\n")
                            time.sleep(3)
                            if shell_cfg.recv_ready():
                                resp_cfg = shell_cfg.recv(65535).decode("utf-8", errors="replace")
                                self._upgrade_log(resp_cfg.strip())

                        shell_cfg.close()
                        cfg_ssh.close()
                        jump_cfg_client.close()
                        self._upgrade_log("[OK] Configuração software-repository concluída.")

                        # --- Configurar BOF primary-image e secondary-image ---
                        if dir_7x50:
                            self._upgrade_log("[INFO] Configurando BOF primary/secondary image...")
                            jump_bof_client = paramiko.SSHClient()
                            jump_bof_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                            jump_bof_client.connect(
                                hostname=jump_ip, port=22, username=jump_user,
                                password=jump_pass, timeout=15,
                                look_for_keys=False, allow_agent=False,
                            )
                            jump_bof_transport = jump_bof_client.get_transport()
                            jump_bof_channel = jump_bof_transport.open_channel(
                                "direct-tcpip", (ip, router_port), ("127.0.0.1", 0),
                            )
                            bof_ssh = paramiko.SSHClient()
                            bof_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                            bof_ssh.connect(
                                hostname=ip, port=router_port,
                                username=router_user, password=router_pass,
                                timeout=15, look_for_keys=False, allow_agent=False,
                                sock=jump_bof_channel,
                            )
                            shell_bof = bof_ssh.invoke_shell(width=512)
                            time.sleep(1)
                            if shell_bof.recv_ready():
                                shell_bof.recv(65535)
                            shell_bof.send("environment no more\n")
                            time.sleep(1)
                            if shell_bof.recv_ready():
                                shell_bof.recv(65535)

                            # Extrair primary-image atual do BOF
                            self._upgrade_log("[CMD] show bof")
                            shell_bof.send("show bof\n")
                            saida_bof = ""
                            tent_bof = 0
                            while tent_bof < 6:
                                time.sleep(1)
                                if shell_bof.recv_ready():
                                    bloco = shell_bof.recv(65535).decode("utf-8", errors="replace")
                                    saida_bof += bloco
                                    tent_bof = 0
                                else:
                                    tent_bof += 1
                            self._upgrade_log(saida_bof.strip())

                            primary_image = ""
                            for linha in saida_bof.splitlines():
                                if "primary-image" in linha.lower() and "secondary" not in linha.lower():
                                    partes = linha.split(":", 1)
                                    if len(partes) > 1:
                                        primary_image = partes[0].split()[-1] + ":" + partes[1].strip()
                                    break

                            if primary_image:
                                self._upgrade_log(f"[INFO] BOF primary-image atual: {primary_image}")

                                # Configurar primary-image atual como secondary-image
                                cmd_sec_img = f"bof secondary-image {primary_image}"
                                self._upgrade_log(f"[CMD] {cmd_sec_img}")
                                shell_bof.send(cmd_sec_img + "\n")
                                time.sleep(3)
                                if shell_bof.recv_ready():
                                    resp_bof = shell_bof.recv(65535).decode("utf-8", errors="replace")
                                    self._upgrade_log(resp_bof.strip())
                            else:
                                self._upgrade_log("[AVISO] Não foi possível extrair primary-image do BOF.")

                            # Configurar novo primary-image com o diretório 7x50
                            new_primary_img = f"cf3:\\{nome_pasta_7x50}"
                            cmd_pri_img = f"bof primary-image {new_primary_img}"
                            self._upgrade_log(f"[CMD] {cmd_pri_img}")
                            shell_bof.send(cmd_pri_img + "\n")
                            time.sleep(3)
                            if shell_bof.recv_ready():
                                resp_bof = shell_bof.recv(65535).decode("utf-8", errors="replace")
                                self._upgrade_log(resp_bof.strip())

                            shell_bof.close()
                            bof_ssh.close()
                            jump_bof_client.close()
                            self._upgrade_log("[OK] Configuração BOF concluída.")

                        # --- Copiar boot.ldr, admin save, bof save, sync ---
                        self._upgrade_log("[INFO] Executando comandos finais de preparação...")
                        jump_fin_client = paramiko.SSHClient()
                        jump_fin_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        jump_fin_client.connect(
                            hostname=jump_ip, port=22, username=jump_user,
                            password=jump_pass, timeout=15,
                            look_for_keys=False, allow_agent=False,
                        )
                        jump_fin_transport = jump_fin_client.get_transport()
                        jump_fin_channel = jump_fin_transport.open_channel(
                            "direct-tcpip", (ip, router_port), ("127.0.0.1", 0),
                        )
                        fin_ssh = paramiko.SSHClient()
                        fin_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        fin_ssh.connect(
                            hostname=ip, port=router_port,
                            username=router_user, password=router_pass,
                            timeout=15, look_for_keys=False, allow_agent=False,
                            sock=jump_fin_channel,
                        )
                        shell_fin = fin_ssh.invoke_shell(width=512)
                        time.sleep(1)
                        if shell_fin.recv_ready():
                            shell_fin.recv(65535)
                        shell_fin.send("environment no more\n")
                        time.sleep(1)
                        if shell_fin.recv_ready():
                            shell_fin.recv(65535)

                        # Copiar boot.ldr para cf3-a:\ e cf3-b:\ (aguardar conclusão)
                        if dir_7x50:
                            for cf_dest in ("cf3-a:", "cf3-b:"):
                                cmd_cp_boot = f"file copy cf3:\\{nome_pasta_7x50}\\boot.ldr {cf_dest}\\ force"
                                self._upgrade_log(f"[CMD] {cmd_cp_boot}")
                                shell_fin.send(cmd_cp_boot + "\n")
                                resp_cp_boot = ""
                                tent_cp_boot = 0
                                while tent_cp_boot < 15:
                                    time.sleep(2)
                                    if shell_fin.recv_ready():
                                        bloco = shell_fin.recv(65535).decode("utf-8", errors="replace")
                                        resp_cp_boot += bloco
                                        tent_cp_boot = 0
                                    else:
                                        tent_cp_boot += 1
                                self._upgrade_log(resp_cp_boot.strip())

                            # Verificar boot.ldr copiado (aguardar conclusão + MINOR handling)
                            boot_minor_msgs = []  # Coletar MINORs de ambos cf3
                            self._suppress_nokia_error_box = True  # Suprimir caixa genérica durante check boot.ldr
                            for cf_dest in ("cf3-a:", "cf3-b:"):
                                cmd_chk_boot = f"file version check {cf_dest}\\boot.ldr"
                                self._upgrade_log(f"[CMD] {cmd_chk_boot}")
                                shell_fin.send(cmd_chk_boot + "\n")
                                resp_chk_boot = ""
                                tent_chk_boot = 0
                                while tent_chk_boot < 10:
                                    time.sleep(2)
                                    if shell_fin.recv_ready():
                                        bloco = shell_fin.recv(65535).decode("utf-8", errors="replace")
                                        resp_chk_boot += bloco
                                        tent_chk_boot = 0
                                    else:
                                        tent_chk_boot += 1
                                self._upgrade_log(resp_chk_boot.strip())
                                if "minor:" in resp_chk_boot.lower():
                                    minor_info = ""
                                    for l in resp_chk_boot.splitlines():
                                        if "minor:" in l.lower():
                                            minor_info = l.strip()
                                            break
                                    boot_minor_msgs.append((cf_dest, minor_info))

                            # Exibir uma única caixa de diálogo com os MINORs de ambos cf3
                            if boot_minor_msgs:
                                _v_atual = re.search(r'-(\d+)\.', versao)
                                _v_dest = re.search(r'-(\d+)\.', nome_pasta_7x50)
                                msg_esperado = ""
                                if _v_atual and _v_dest and _v_atual.group(1) == '20' and _v_dest.group(1) == '24':
                                    msg_esperado = "\n\nEsta falha é um comportamento esperado, pois a versão 20 não lê o boot loader da versão 24. Pode prosseguir."
                                detalhes = "\n".join(
                                    f"{cf}\\boot.ldr — {info}" for cf, info in boot_minor_msgs
                                )
                                _perguntar_minor(
                                    "MINOR detectado — boot.ldr",
                                    f"O roteador reportou MINOR no file version check de:\n"
                                    f"{detalhes}\n\n"
                                    f"Versão atual: {versao}\n"
                                    f"Versão destino: {nome_pasta_7x50}{msg_esperado}",
                                )
                            self._suppress_nokia_error_box = False  # Restaurar caixa genérica

                        # Remover authentication-order antes do admin save
                        cmd_no_auth = "configure system security password no authentication-order"
                        shell_fin, resp_no_auth = self._ssh_enviar_cmd(shell_fin, cmd_no_auth)
                        self._upgrade_log(_limpar_ansi(resp_no_auth).strip())

                        # --- Lógica port-weight-speed (logo após remoção do authentication-order) ---
                        self._upgrade_log("[INFO] Verificando port-weight-speed...")
                        shell_fin, _resp_pws = self._ssh_enviar_cmd(shell_fin, "admin display-config | match context all port-weight-speed")
                        _resp_pws_limpa = _limpar_ansi(_resp_pws)
                        self._upgrade_log(_resp_pws_limpa.strip())

                        # Extrair LAGs que têm port-weight-speed
                        _lags_pws = set()
                        for _linha_pws in _resp_pws_limpa.splitlines():
                            _m_lag = re.search(r'lag[- ](\d+)', _linha_pws, re.IGNORECASE)
                            if _m_lag:
                                _lags_pws.add(_m_lag.group(1))
                        if _lags_pws:
                            self._upgrade_log(f"[INFO] LAGs com port-weight-speed: {sorted(_lags_pws, key=int)}")
                            for _lag_num in sorted(_lags_pws, key=int):
                                _cmd_flow = f"show lag {_lag_num} flow-distribution"
                                shell_fin, _resp_flow = self._ssh_enviar_cmd(shell_fin, _cmd_flow)
                                _resp_flow_limpa = _limpar_ansi(_resp_flow)
                                self._upgrade_log(_resp_flow_limpa.strip())

                                _ports_config = []
                                for _linha_flow in _resp_flow_limpa.splitlines():
                                    # Match port id (e.g. 1/1/1 or esat-1/1/26) followed by Hash-Weight value
                                    _m_port_hw = re.match(r'^\s*((?:esat-)?\d+/\d+/\S+)\s+(\d+)', _linha_flow)
                                    if _m_port_hw:
                                        _port_id = _m_port_hw.group(1)
                                        try:
                                            _hw_val_atual = int(_m_port_hw.group(2))
                                        except ValueError:
                                            continue
                                        if _hw_val_atual == 100:
                                            _ports_config.append((_port_id, 10))
                                        elif _hw_val_atual == 10:
                                            _ports_config.append((_port_id, 1))

                                if _ports_config:
                                    _cmd_no_pws = f"/configure lag {_lag_num} no port-weight-speed"
                                    shell_fin, _ = self._ssh_enviar_cmd(shell_fin, _cmd_no_pws, timeout_idle=3)

                                    for _port_id, _hw_val in _ports_config:
                                        _cmd_hw = f"/configure lag {_lag_num} port {_port_id} hash-weight {_hw_val}"
                                        shell_fin, _resp_hw = self._ssh_enviar_cmd(shell_fin, _cmd_hw, timeout_idle=3)
                                        self._upgrade_log(_limpar_ansi(_resp_hw).strip())

                                    self._upgrade_log(f"[INFO] {len(_ports_config)} port(s) configurada(s) no lag {_lag_num}")
                                else:
                                    self._upgrade_log(f"[INFO] Nenhuma port com hash-weight 100 ou 10 encontrada no lag {_lag_num}")

                            # Exibir flow-distribution final pelo comando admin display-config | match context all port-weight-speed
                            shell_fin, _resp_pws_final = self._ssh_enviar_cmd(shell_fin, "admin display-config | match context all port-weight-speed")
                            self._upgrade_log(_limpar_ansi(_resp_pws_final).strip())
                        else:
                            self._upgrade_log("[INFO] Nenhum LAG com port-weight-speed encontrado.")
                        self._upgrade_log("[OK] Verificação port-weight-speed concluída.")

                        # --- Lógica QoS egress-hsmda (extrair fc e queue das network-queue, remover fc) ---
                        self._upgrade_log("[INFO] Verificando QoS egress-hsmda...")
                        shell_fin, _resp_hsmda = self._ssh_enviar_cmd(shell_fin, "admin display-config | match context all egress-hsmda")
                        _resp_hsmda_limpa = _limpar_ansi(_resp_hsmda)
                        self._upgrade_log(_resp_hsmda_limpa.strip())

                        # Extrair (network-queue, fc, queue) do output do egress-hsmda
                        # Formato real do "match context all egress-hsmda":
                        #   /configure qos network-queue "EBT" fc af egress-hsmda
                        #   /configure qos network-queue "EBT" fc af egress-hsmda queue 3
                        # A queue aparece APÓS "egress-hsmda" na mesma linha
                        _nq_fc_queue_map = {}  # {nq_name: [(fc_name, queue_num), ...]}
                        # Primeiro: coletar as fc com queue do egress-hsmda (linhas com "egress-hsmda queue N")
                        for _linha_hsmda in _resp_hsmda_limpa.splitlines():
                            _l_h = _linha_hsmda.strip()
                            _m_fc_q = re.search(
                                r'network-queue\s+"([^"]+)"\s+fc\s+(\S+)\s+egress-hsmda\s+queue\s+(\d+)',
                                _l_h, re.IGNORECASE
                            )
                            if _m_fc_q:
                                _nq_name = _m_fc_q.group(1)
                                _fc_name = _m_fc_q.group(2)
                                _queue_num = _m_fc_q.group(3)
                                if _nq_name not in _nq_fc_queue_map:
                                    _nq_fc_queue_map[_nq_name] = []
                                if (_fc_name, _queue_num) not in _nq_fc_queue_map[_nq_name]:
                                    _nq_fc_queue_map[_nq_name].append((_fc_name, _queue_num))

                        # Se não encontrou nenhuma queue nas linhas do egress-hsmda, tentar
                        # extrair as network-queues e fc apenas pelo nome e buscar a queue separadamente
                        if not _nq_fc_queue_map:
                            _nqs_com_hsmda = set()
                            for _linha_hsmda in _resp_hsmda_limpa.splitlines():
                                _l_h = _linha_hsmda.strip()
                                _m_nq_fc = re.search(
                                    r'network-queue\s+"([^"]+)"\s+fc\s+(\S+)\s+egress-hsmda',
                                    _l_h, re.IGNORECASE
                                )
                                if _m_nq_fc:
                                    _nqs_com_hsmda.add((_m_nq_fc.group(1), _m_nq_fc.group(2)))
                            # Para cada (nq, fc), buscar a queue com comando dedicado
                            for _nq_name_h, _fc_name_h in _nqs_com_hsmda:
                                _cmd_fc_q = f'admin display-config | match context all "network-queue \\"{_nq_name_h}\\" fc {_fc_name_h} queue"'
                                shell_fin, _resp_fc_q = self._ssh_enviar_cmd(shell_fin, _cmd_fc_q)
                                _resp_fc_q_limpa = _limpar_ansi(_resp_fc_q)
                                self._upgrade_log(_resp_fc_q_limpa.strip())
                                _m_q = re.search(
                                    r'network-queue\s+"' + re.escape(_nq_name_h) + r'"\s+fc\s+' + re.escape(_fc_name_h) + r'\s+queue\s+(\d+)',
                                    _resp_fc_q_limpa, re.IGNORECASE
                                )
                                if _m_q:
                                    if _nq_name_h not in _nq_fc_queue_map:
                                        _nq_fc_queue_map[_nq_name_h] = []
                                    if (_fc_name_h, _m_q.group(1)) not in _nq_fc_queue_map[_nq_name_h]:
                                        _nq_fc_queue_map[_nq_name_h].append((_fc_name_h, _m_q.group(1)))

                        for _nq_name, _fc_list in _nq_fc_queue_map.items():
                            self._upgrade_log(f"[INFO] network-queue \"{_nq_name}\": {len(_fc_list)} fc(s) encontrada(s): {_fc_list}")

                        # Armazenar para restauração pós-reload
                        self._upg_nq_fc_queue_map = _nq_fc_queue_map

                        # Agora remover TODAS as fc de cada network-queue
                        if _nq_fc_queue_map:
                            for _nq_name, _fc_list in _nq_fc_queue_map.items():
                                for _fc_name, _queue_num in _fc_list:
                                    _cmd_no_fc = f'/configure qos network-queue "{_nq_name}" no fc {_fc_name}'
                                    shell_fin, _resp_no_fc = self._ssh_enviar_cmd(shell_fin, _cmd_no_fc, timeout_idle=3)
                                    self._upgrade_log(_limpar_ansi(_resp_no_fc).strip())
                                self._upgrade_log(f"[OK] {len(_fc_list)} fc(s) removida(s) da network-queue \"{_nq_name}\".")
                        else:
                            self._upgrade_log("[INFO] Nenhuma network-queue com egress-hsmda encontrada.")
                        self._upgrade_log("[OK] Verificação QoS egress-hsmda concluída.")

                        # admin save
                        shell_fin, resp_fin = self._ssh_enviar_cmd(shell_fin, "admin save", timeout_idle=8, intervalo=2)
                        self._upgrade_log(_limpar_ansi(resp_fin).strip())

                        # bof save
                        shell_fin, resp_fin = self._ssh_enviar_cmd(shell_fin, "bof save", timeout_idle=8, intervalo=2)
                        self._upgrade_log(_limpar_ansi(resp_fin).strip())

                        # Satellite sync-boot-env (se qtd_sat >= 1)
                        if qtd_sat >= 1:
                            for sat_id in range(1, qtd_sat + 1):
                                cmd_sat_sync = f"admin satellite eth-sat {sat_id} sync-boot-env"
                                shell_fin, resp_fin = self._ssh_enviar_cmd(shell_fin, cmd_sat_sync, timeout_idle=8, intervalo=2)
                                self._upgrade_log(_limpar_ansi(resp_fin).strip())

                        # admin redundancy synchronize config (aguardar OK do roteador)
                        self._upgrade_log("[CMD] admin redundancy synchronize config")
                        shell_fin.send("admin redundancy synchronize config\n")
                        saida_sync_cfg = ""
                        tent_sync_cfg = 0
                        while tent_sync_cfg < 30:
                            time.sleep(2)
                            if shell_fin.recv_ready():
                                bloco = shell_fin.recv(65535).decode("utf-8", errors="replace")
                                saida_sync_cfg += bloco
                                self._upgrade_log(bloco.strip())
                                tent_sync_cfg = 0
                                # Sair do loop ao detectar conclusão do comando
                                if "completed" in saida_sync_cfg.lower():
                                    break
                            else:
                                tent_sync_cfg += 1
                        if "fail" in saida_sync_cfg.lower() or "error" in saida_sync_cfg.lower():
                            self._upgrade_log("[ERRO] Falha no admin redundancy synchronize config")
                            ok_sync_cfg = _perguntar_erro_sync(
                                "Erro — admin redundancy synchronize config",
                                f"O roteador reportou erro no admin redundancy synchronize config.\n\n"
                                f"Deseja prosseguir ou interromper?\n"
                                f"Se interromper, acesse a caixa pelo botão SSH ao lado do hostname e verifique.",
                            )
                            if not ok_sync_cfg:
                                self._upgrade_log("[INFO] Processo interrompido pelo usuário.")
                                self._upgrade_log("[INFO] Acesse a caixa pelo botão SSH ao lado do hostname e verifique.")
                                shell_fin.close()
                                fin_ssh.close()
                                jump_fin_client.close()
                                raise Exception("Processo interrompido pelo usuário no admin redundancy synchronize config.")
                        self._upgrade_log("[OK] admin redundancy synchronize config concluído.")

                        # admin redundancy synchronize boot-env (aguardar OK do roteador)
                        self._upgrade_log("[CMD] admin redundancy synchronize boot-env")
                        shell_fin.send("admin redundancy synchronize boot-env\n")
                        saida_sync = ""
                        tent_sync = 0
                        while tent_sync < 30:
                            time.sleep(2)
                            if shell_fin.recv_ready():
                                bloco = shell_fin.recv(65535).decode("utf-8", errors="replace")
                                saida_sync += bloco
                                self._upgrade_log(bloco.strip())
                                tent_sync = 0
                                # Sair do loop ao detectar conclusão do comando
                                if "completed" in saida_sync.lower():
                                    break
                            else:
                                tent_sync += 1
                        if "fail" in saida_sync.lower() or "error" in saida_sync.lower():
                            self._upgrade_log("[ERRO] Falha no admin redundancy synchronize boot-env")
                            ok_sync_boot = _perguntar_erro_sync(
                                "Erro — admin redundancy synchronize boot-env",
                                f"O roteador reportou erro no admin redundancy synchronize boot-env.\n\n"
                                f"Deseja prosseguir ou interromper?\n"
                                f"Se interromper, acesse a caixa pelo botão SSH ao lado do hostname e verifique.",
                            )
                            if not ok_sync_boot:
                                self._upgrade_log("[INFO] Processo interrompido pelo usuário.")
                                self._upgrade_log("[INFO] Acesse a caixa pelo botão SSH ao lado do hostname e verifique.")
                                shell_fin.close()
                                fin_ssh.close()
                                jump_fin_client.close()
                                raise Exception("Processo interrompido pelo usuário no admin redundancy synchronize boot-env.")
                        self._upgrade_log("[OK] Sincronização boot-env concluída.")

                        shell_fin.close()
                        fin_ssh.close()
                        jump_fin_client.close()
                        self._upgrade_log("[OK] Comandos finais concluídos.")

                shell.close()
                client.close()
                jump_client.close()
                self._upgrade_log(f"[OK] SSH finalizado.")

                # --- Rede de controle via jumpserver + ssh + telnet router 8083 ---
                if rede_controle:
                    # Resolver hostname vizinho (ímpar↔par)
                    m_num = re.search(r"(\D+)(\d+)(.*)", hostname)
                    if m_num:
                        prefixo = m_num.group(1)
                        num = int(m_num.group(2))
                        sufixo = m_num.group(3)
                        if num % 2 == 0:
                            vizinho_num = num - 1
                        else:
                            vizinho_num = num + 1
                        vizinho_host = f"{prefixo}{vizinho_num:02d}{sufixo}"
                    else:
                        vizinho_host = hostname
                    vizinho_fqdn = f"{vizinho_host}.embratel.net.br"
                    self._upgrade_log2(f"[INFO] Resolvendo IP do vizinho {vizinho_fqdn}...")
                    vizinho_ip = None
                    try:
                        proc2 = subprocess.run(
                            ["ping", vizinho_fqdn, "-n", "1"],
                            capture_output=True, text=True, timeout=5,
                        )
                        m_ip = re.search(r"\[(\d+\.\d+\.\d+\.\d+)\]", proc2.stdout)
                        if m_ip:
                            vizinho_ip = m_ip.group(1)
                        if not vizinho_ip:
                            m_ip = re.search(r"Resposta de (\d+\.\d+\.\d+\.\d+)", proc2.stdout)
                            if m_ip:
                                vizinho_ip = m_ip.group(1)
                        if not vizinho_ip:
                            ips_found = re.findall(r"(\d+\.\d+\.\d+\.\d+)", proc2.stdout)
                            if ips_found:
                                vizinho_ip = ips_found[-1]
                    except Exception as e:
                        self._upgrade_log2(f"[ERRO] Falha ao resolver vizinho: {e}")
                    if not vizinho_ip:
                        self._upgrade_log2(f"[ERRO] Não foi possível resolver IP de {vizinho_fqdn}")
                    else:
                        self._upgrade_log2(f"[OK] IP vizinho resolvido: {vizinho_host} → {vizinho_ip}")

                    if vizinho_ip:
                        self._upgrade_log2(f"[INFO] Conectando ao jumpserver 10.73.0.4 para rede de controle...")
                        try:
                            jump1_client = paramiko.SSHClient()
                            jump1_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                            jump1_client.connect(
                                hostname="10.73.0.4",
                                port=22,
                                username="supnokia",
                                password="NokiaNsp1!",
                                timeout=15,
                                look_for_keys=False,
                                allow_agent=False,
                            )
                            self._upgrade_log2(f"[OK] Jumpserver 10.73.0.4 conectado.")

                            shell2 = jump1_client.invoke_shell(term="vt100", width=512)
                            time.sleep(1)
                            if shell2.recv_ready():
                                banner2 = shell2.recv(65535).decode("utf-8", errors="replace")
                                self._upgrade_console_raw(banner2)
                            self._upgrade_log2(f"[INFO] Terminal negociado como vt100")

                            # SSH para o vizinho
                            self._upgrade_log2(f"[CMD] ssh 93191142@{vizinho_ip}")
                            shell2.send(f"ssh 93191142@{vizinho_ip}\n")
                            # Aguardar prompt de senha ou yes/no
                            time.sleep(3)
                            resp = ""
                            while shell2.recv_ready():
                                resp += shell2.recv(65535).decode("utf-8", errors="replace")
                            self._upgrade_console_raw(resp)
                            if "yes/no" in resp.lower():
                                shell2.send("yes\n")
                                time.sleep(2)
                                resp = ""
                                while shell2.recv_ready():
                                    resp += shell2.recv(65535).decode("utf-8", errors="replace")
                                self._upgrade_console_raw(resp)
                            if "password" in resp.lower():
                                shell2.send("X%aA5&z3\n")
                                time.sleep(3)
                                if shell2.recv_ready():
                                    resp = shell2.recv(65535).decode("utf-8", errors="replace")
                                    self._upgrade_console_raw(resp)
                            self._upgrade_log2(f"[OK] SSH {vizinho_ip} conectado.")

                            # Telnet router 8083 para rede de controle
                            self._upgrade_log2(f"[CMD] telnet router 8083 {rede_controle}")
                            shell2.send(f"telnet router 8083 {rede_controle}\n")
                            time.sleep(5)
                            resp = ""
                            while shell2.recv_ready():
                                resp += shell2.recv(65535).decode("utf-8", errors="replace")
                            self._upgrade_console_raw(resp)

                            # Login engenharia
                            self._upgrade_log2(f"[CMD] engenharia (login)")
                            shell2.send("engenharia\n")
                            time.sleep(3)
                            resp = ""
                            while shell2.recv_ready():
                                resp += shell2.recv(65535).decode("utf-8", errors="replace")
                            self._upgrade_console_raw(resp)

                            self._upgrade_log2(f"[CMD] engenharia (senha)")
                            shell2.send("engenharia\n")
                            time.sleep(3)
                            resp = ""
                            while shell2.recv_ready():
                                resp += shell2.recv(65535).decode("utf-8", errors="replace")
                            self._upgrade_console_raw(resp)

                            # Comando vt100
                            self._upgrade_log2(f"[CMD] vt100")
                            shell2.send("vt100\n")
                            time.sleep(2)
                            if shell2.recv_ready():
                                resp = shell2.recv(65535).decode("utf-8", errors="replace")
                                self._upgrade_console_raw(resp)

                            self._upgrade_log2(f"[OK] Rede de controle conectado. Sessão mantida aberta.")

                            # Manter referências para uso posterior
                            self._rc_shell = shell2
                            self._rc_jump_client = jump1_client
                            self._rc_first_cmd_sent = False

                            # Iniciar thread de leitura contínua do shell
                            self._rc_reader_running = True
                            threading.Thread(target=self._rc_reader_loop, daemon=True).start()
                        except Exception as e:
                            self._upgrade_log2(f"[ERRO] Rede de controle: {e}")
                            self._upgrade_log(f"[AVISO] Falha na rede de controle. Usando sessão SSH (Log SSH) como fallback...")
                            self._rc_fallback_to_ssh = True
                            # Conectar ao roteador via jumpserver para fallback
                            try:
                                _fb_jump = paramiko.SSHClient()
                                _fb_jump.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                                _fb_jump.connect(
                                    hostname="10.73.0.4", port=22,
                                    username="supnokia", password="NokiaNsp1!",
                                    timeout=15, look_for_keys=False, allow_agent=False,
                                )
                                _fb_transport = _fb_jump.get_transport()
                                _fb_channel = _fb_transport.open_channel(
                                    "direct-tcpip", (ip, 22), ("127.0.0.1", 0),
                                )
                                _fb_ssh = paramiko.SSHClient()
                                _fb_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                                _fb_ssh.connect(
                                    hostname=ip, port=22,
                                    username="93191142", password="X%aA5&z3",
                                    timeout=15, look_for_keys=False, allow_agent=False,
                                    sock=_fb_channel,
                                )
                                _fb_shell = _fb_ssh.invoke_shell(width=512)
                                time.sleep(1)
                                if _fb_shell.recv_ready():
                                    _fb_shell.recv(65535)
                                _fb_shell.send("environment no more\n")
                                time.sleep(1)
                                if _fb_shell.recv_ready():
                                    _fb_shell.recv(65535)
                                self._rc_shell = _fb_shell
                                self._rc_jump_client = _fb_jump
                                self._rc_first_cmd_sent = False
                                self._rc_reader_running = True
                                threading.Thread(target=self._rc_reader_loop, daemon=True).start()
                                self._upgrade_log(f"[OK] Fallback SSH conectado ao roteador {ip}. Comandos da rede de controle serão executados no Log SSH.")
                            except Exception as e_fb:
                                self._upgrade_log(f"[ERRO] Falha ao conectar fallback SSH: {e_fb}")
                else:
                    self._upgrade_log2("[INFO] Sem IP de rede de controle, conexão não executada.")

                self._upgrade_log(f"[OK] Finalizado.")

            except Exception as e:
                self._upgrade_log(f"[ERRO] {e}")

        threading.Thread(target=worker, daemon=True).start()

    def _upgrade_ssh_conectar(self):
        """Abre PuTTY para conectar ao roteador via jumpserver."""
        ip = self.upg_ip.get().strip()
        if not ip:
            messagebox.showwarning("Atenção", "Resolva o IP primeiro (pressione Enter no hostname).")
            return

        try:
            proxy_cmd = (
                'plink.exe -ssh -l supnokia -pw NokiaNsp1! 10.73.0.4 '
                '-nc %host:%port'
            )
            cmd = [
                'putty.exe',
                '-ssh', ip,
                '-l', '93191142',
                '-pw', 'X%aA5&z3',
                '-P', '22',
                '-proxycmd', proxy_cmd,
            ]
            subprocess.Popen(cmd)
            self._upgrade_log(f"[OK] PuTTY aberto para {ip} via jumpserver 10.73.0.4")
        except FileNotFoundError:
            messagebox.showerror(
                "Erro",
                "putty.exe ou plink.exe não encontrado.\n"
                "Verifique se o PuTTY está instalado e no PATH do sistema.",
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao abrir PuTTY:\n{e}")

    def _piscar_icone(self):
        if sys.platform != "win32":
            return
        try:
            hwnd = self.root.winfo_id()
            is_iconic = ctypes.windll.user32.IsIconic(hwnd)
            fg_win = ctypes.windll.user32.GetForegroundWindow()
            if is_iconic or fg_win != hwnd:
                FLASHW_ALL = 3
                FLASHW_TIMERNOFG = 12
                class FLASHWINFO(ctypes.Structure):
                    _fields_ = [("cbSize", ctypes.wintypes.UINT),
                                ("hwnd", ctypes.wintypes.HWND),
                                ("dwFlags", ctypes.wintypes.DWORD),
                                ("uCount", ctypes.wintypes.UINT),
                                ("dwTimeout", ctypes.wintypes.DWORD)]
                fi = FLASHWINFO(ctypes.sizeof(FLASHWINFO), hwnd, FLASHW_ALL | FLASHW_TIMERNOFG, 5, 0)
                ctypes.windll.user32.FlashWindowEx(ctypes.byref(fi))
        except Exception:
            pass

    def _limpar_content(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()




# --- Main ---
if __name__ == "__main__":
    root = tk.Tk()
    app = AppMaster(root)
    root.mainloop()

